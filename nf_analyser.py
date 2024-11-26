# -*- coding: utf-8 -*-

### TO DO ###
# Fix section 1 - Box file download to server
# Make subsection headings

#region 0) IMPORT PACKAGES.

import logging
from datetime import datetime
import time
import urllib.parse
import webbrowser
import os
import threading
from http.server import BaseHTTPRequestHandler, HTTPServer
import boxsdk
from boxsdk import OAuth2, Client
from boxsdk.object.file import File
from boxsdk.object.folder import Folder
import psutil
import signal
import sys
import subprocess
import pandas as pd
import shutil
import numpy as np
import re
import nibabel as nib
import matplotlib.pyplot as plt
import fnmatch
from collections import defaultdict
import io
import msoffcrypto
import openpyxl
import warnings
import pydicom
import random
import nibabel as nib
from skimage.metrics import structural_similarity as ssim
from plotnine import *
from scipy import stats
from statsmodels.stats.multitest import multipletests
from pingouin import rm_anova
import json
import glob
from PIL import Image
import hashlib
from nilearn.interfaces.fmriprep import load_confounds_strategy
import nibabel as nib
from nibabel.processing import resample_from_to
# import rpy2.robjects as ro
# from rpy2.robjects import pandas2ri
# from rpy2.robjects.packages import importr
import warnings
warnings.filterwarnings("ignore", category=DeprecationWarning)

#endregion

#region 1) DOWNLOAD BOX FILES TO SERVER.

def download_box_files():
    p_id = logged_input("Enter the participant's ID (e.g. P001).\n")
    working_dir = os.getcwd()
    p_id_folder = os.path.join(os.getcwd(), p_id)
    if not os.path.exists(p_id_folder):
        os.makedirs(p_id_folder)

    # Define the signal handler function
    def signal_handler(sig, frame):
        # Kill the process using port 8080
        for proc in psutil.process_iter():
            try:
                connections = proc.connections()
                for conn in connections:
                    if conn.laddr.port == 8080:
                        proc.kill()
            except psutil.AccessDenied:
                pass
        # Exit the script
        exit(0)
    # Register the signal handler function for SIGINT and SIGTSTP signals
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTSTP, signal_handler)
    # Set up OAuth2 credentials
    CLIENT_ID = 'hv3z8wjk584zopc89fgsc29ikb6m0emp'
    CLIENT_SECRET = 'IpbJPrsXb0LnhtJW36Z0bfXQFhIObgpH'
    REDIRECT_URI = 'http://localhost:8080'
    # Create the OAuth2 object
    oauth = OAuth2(
        client_id=CLIENT_ID,
        client_secret=CLIENT_SECRET,
        store_tokens=None,
    )
    # Create a simple HTTP server to handle the OAuth2 callback
    class CallbackHandler(BaseHTTPRequestHandler):
        def log_message(self, format, *args):
            # Suppress log messages from the web server. Note that when troubleshooting, try to comment out this function as it may be suppressing error messages that can shine a light on the issue. 
            pass
        def do_GET(self):
            if self.path.startswith('/'):
                # Extract the authorization code from the URL
                query = urllib.parse.urlparse(self.path).query
                params = urllib.parse.parse_qs(query)
                if 'code' in params:
                    authorization_code = params['code'][0]
                    # Exchange the authorization code for an access token
                    access_token, refresh_token = oauth.authenticate(
                        authorization_code)
                    # Shut down the HTTP server after the code is retrieved
                    self.send_response(200)
                    self.send_header('Content-type', 'text/html')
                    self.end_headers()
                    self.wfile.write(
                        b'<html><head><title>Authorization Successful</title></head><body><h1>Authorization Successful</h1><p>You can close this window now.</p></body></html>')
                    threading.Thread(target=self.server.shutdown).start()
                else:
                    self.send_response(400)
                    self.send_header('Content-type', 'text/html')
                    self.end_headers()
                    self.wfile.write(
                        b'<html><head><title>Bad Request</title></head><body><h1>Bad Request</h1></body></html>')
            else:
                self.send_response(400)
                self.send_header('Content-type', 'text/html')
                self.end_headers()
                self.wfile.write(
                    b'<html><head><title>Bad Request</title></head><body><h1>Bad Request</h1></body></html>')
    # Start the local web server to handle the OAuth2 callback
    server = HTTPServer(('localhost', 8080), CallbackHandler)
    server_thread = threading.Thread(target=server.serve_forever)
    server_thread.start()
    # Perform the authentication flow
    auth_url, _ = oauth.get_authorization_url(REDIRECT_URI)
    # Open the authorization URL in a web browser
    webbrowser.open(auth_url)
    # Wait for the user to complete the authorization flow
    print("Waiting for authorisation...")
    logged_input("Press Enter to continue once authorised...")
    # Shut down the local web server
    server.shutdown()
    # Create the Box client using the access token
    client = Client(oauth)
    # Get the folder ID by name
    folder_name = f'{p_id}'
    search_results = client.search().query(query=folder_name, type='folder')
    folder = next(
        (item for item in search_results if item.name == folder_name), None)
    # Define get_downloaded_files function
    def get_downloaded_files(save_directory):
        downloaded_files = set()
        for root, dirs, files in os.walk(save_directory):
            for file in files:
                downloaded_files.add(file)
        return downloaded_files
    # Define download_files_from_folder function
    def download_files_from_folder(folder, save_directory, downloaded_files):
        MAX_RETRY_ATTEMPTS = 3
        RETRY_DELAY_SECONDS = 5
        items = list(client.folder(folder.id).get_items(limit=1000, offset=0))
        for item in items:
            if isinstance(item, File):
                if item.name in downloaded_files:
                    continue
                file_path = f'{save_directory}/{item.name}'
                retry_attempts = 0
                while retry_attempts < MAX_RETRY_ATTEMPTS:
                    try:
                        with open(file_path, 'wb') as writeable_stream:
                            item.download_to(writeable_stream)
                            downloaded_files.add(item.name)
                            print(f"Downloaded: {item.name}")
                        break
                    except Exception as e:
                        print(f"An error occurred while downloading '{item.name}': {str(e)}")
                        print("Retrying...")
                        time.sleep(RETRY_DELAY_SECONDS)
                        retry_attempts += 1
                if retry_attempts == MAX_RETRY_ATTEMPTS:
                    print(f"Failed to download '{item.name}' after {MAX_RETRY_ATTEMPTS} attempts.")
            elif isinstance(item, Folder):
                subdirectory = f'{save_directory}/{item.name}'
                os.makedirs(subdirectory, exist_ok=True)
                download_files_from_folder(item, subdirectory, downloaded_files)
    # Specify the parent folder name
    parent_folder_name = f'{p_id}'
    # Retrieve the parent folder
    search_results = client.search().query(query=parent_folder_name, type='folder')
    parent_folder = next((item for item in search_results if item.name == parent_folder_name), None)
    if parent_folder:
        # Specify the save directory
        save_directory = f'/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/{p_id}'
        # Get the set of downloaded files
        downloaded_files = get_downloaded_files(save_directory)
        # Download files recursively from the parent folder and its subfolders
        MAX_RETRY_ATTEMPTS = 3
        RETRY_DELAY_SECONDS = 5
        while True:
            download_files_from_folder(parent_folder, save_directory, downloaded_files)
            specific_files = ['eCRF.xlsx', 'heuristic.py']
            for file in specific_files:
                file_path = f'/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/{file}'
                if file not in downloaded_files:
                    file_item = next((item for item in client.folder(parent_folder.parent.id).get_items() if item.name == file), None)
                    if file_item:
                        retry_attempts = 0
                        while retry_attempts < MAX_RETRY_ATTEMPTS:
                            try:
                                with open(file_path, 'wb') as writeable_stream:
                                    file_item.download_to(writeable_stream)
                                    downloaded_files.add(file)
                                    print(f"Downloaded: {file}.")
                                break
                            except Exception as e:
                                print(f"An error occurred while downloading {file}: {str(e)}")
                                print("Retrying...")
                                time.sleep(RETRY_DELAY_SECONDS)
                                retry_attempts += 1
                        if retry_attempts == MAX_RETRY_ATTEMPTS:
                            print(f"Failed to download {file} after {MAX_RETRY_ATTEMPTS} attempts.")
                    else:
                        print(f"{file} not found in parent folder.")
                else:
                    print(f"{file} already downloaded.")
            # Get the updated folder information to check if all files have been downloaded
            parent_folder_info = client.folder(parent_folder.id).get()
            item_collection = parent_folder_info["item_collection"]
            total_items = item_collection["total_count"]
            # Check if all files have been downloaded
            if len(downloaded_files) == total_items + 1:
                break  # Break the loop if all files have been downloaded
            # Check if the access token needs refreshing
            try:
                if oauth.access_token_expires_at - time.time() < 60:  # Refresh if token expires within 60 seconds
                    oauth.refresh(oauth.access_token, oauth.refresh_token)
                    # Update the Box client with the refreshed token
                    client = Client(oauth)
            except AttributeError:
                pass  # Ignore the AttributeError if the access_token_expires_at attribute is not present
                break
    else:
        print(f"Parent folder '{parent_folder_name}' not found.")
    # Wait for the web server thread to complete
    server_thread.join()

#endregion

#region 2) THERMOMETER ANALYSIS.

def thermometer_analysis():
    participants = ['P004', 'P006', 'P020', 'P030', 'P059', 'P078', 'P093', 'P094', 'P100', 'P107', 'P122', 'P125', 'P127', 'P128', 'P136', 'P145', 'P155', 'P199', 'P215', 'P216']
    restart = logged_input("\nWould you like to start the thermometer analysis from scratch? This will delete the entire analysis/thermometer_analysis folder. (y/n)\n")
    if restart == 'y':
        double_check = logged_input("\nAre you sure? (y/n)\n")
        if double_check == 'y':
            thermometer_analysis_folder = 'analysis/thermometer_analysis'
            print(f"Deleting analysis/thermometer_analysis folder...")
            shutil.rmtree(thermometer_analysis_folder)
        else:
            sys.exit()

    # Step 1: Create Directories.
    print("\n###### STEP 1: CREATE DIRECTORIES ######")
    analysis_folder = 'analysis'
    os.makedirs(analysis_folder, exist_ok=True)
    thermometer_analysis_folder = 'analysis/thermometer_analysis'
    os.makedirs(thermometer_analysis_folder, exist_ok=True)
    figs_folder = 'analysis/thermometer_analysis/figs'
    os.makedirs(figs_folder, exist_ok=True)
    print("Directories created.")

    # Step 2: Access Run 2 and 3 tbv_script thermometer files and extract relevant data into dataframe.
    print("\n###### STEP 2: EXTRACT TBV THERMOMETER DATA ######")
    def find_second_and_third_largest(files):
        sorted_files = sorted(files, key=lambda x: int(x.split('_')[-1].split('.')[0]), reverse=True)
        second_largest_path = os.path.join(folder_path, sorted_files[-2])
        third_largest_path = os.path.join(folder_path, sorted_files[-3])
        return second_largest_path, third_largest_path
    df = pd.DataFrame(columns=participants)
    for x in participants:
        folder_path = os.path.join(os.getcwd(), 'data', 'raw_data', f'{x}', 'data', 'neurofeedback', 'tbv_script', 'data')
        files = [f for f in os.listdir(folder_path) if os.path.isfile(os.path.join(folder_path, f))]
        if len(files) == 4:
            run2_path, run3_path = find_second_and_third_largest(files)
        else:
            print("Error: The folder should contain exactly 4 files.")
        def process_file(file_path):
            if os.path.exists(file_path):
                with open(file_path, 'r') as file:
                    lines = file.readlines()[12:]
                current_value = None
                value_counter = {}
                for line in lines:
                    values = line.strip().split(',')
                    if values[2] != current_value:
                        current_value = values[2]
                        value_counter[current_value] = value_counter.get(current_value, 0) + 1
                    condition = f"{current_value.lower()}{value_counter[current_value]:02d}"
                    if 'rest' in condition:
                        condition = condition.replace('rest', 'sub')
                    if 'guilt first' in condition:
                        condition = condition.replace('guilt first', 'guilt_first')
                    if 'guilt second' in condition:
                        condition = condition.replace('guilt second', 'guilt_second')
                    if 'indignation first' in condition:
                        condition = condition.replace('indignation first', 'indig_first')
                    if 'indignation second' in condition:
                        condition = condition.replace('indignation second', 'indig_second')
                    row_header = 'r' + values[1] + '_' + condition + '_' + 'vol' + values[3] + '_val'
                    feedback_lvl_header = 'r' + values[1] + '_' + condition + '_' + 'vol' + values[3] + '_lvl'
                    value = float(values[8])
                    feedback_lvl = float(values[9])
                    if row_header not in df.index:
                        df.loc[row_header] = [None] * len(participants)
                    df.at[row_header, f'{x}'] = value
                    if feedback_lvl_header not in df.index:
                        df.loc[feedback_lvl_header] = [None] * len(participants)
                    df.at[feedback_lvl_header, f'{x}'] = feedback_lvl
        process_file(run2_path)
        process_file(run3_path)
    therm_data_path = '/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/thermometer_analysis/therm_data.xlsx'
    df.to_excel(therm_data_path, index=True)
    print("TBV thermometer data extracted.")
    
    # Step 3: Access eCRF document and extract relevant data into dataframe.
    print("\n###### STEP 3: EXTRACT eCRF DATA ######")
    warnings.simplefilter("ignore", UserWarning)
    df_row_headers = ['dob', 'gender', 'handedness', 'exercise', 'education', 'work_status', 'panic', 'agoraphobia', 'social_anx', 'ocd', 'ptsd', 'gad', 'comorbid_anx', 'msm', 'psi_sociotropy', 'psi_autonomy', 'raads', 'panas_pos_vis_1', 'panas_neg_vis_1', 'qids_vis_1', 'gad_vis_1', 'rosenberg_vis_1', 'madrs_vis_1', 'pre_memory_intensity_guilt_1', 'pre_memory_intensity_guilt_2', 'pre_memory_intensity_indignation_1', 'pre_memory_intensity_indignation_2', 'intervention', 'techniques_guilt', 'techniques_indignation', 'perceived_success_guilt', 'perceived_success_indignation', 'post_memory_intensity_guilt_1', 'post_memory_intensity_guilt_2', 'post_memory_intensity_indignation_1', 'post_memory_intensity_indignation_2', 'rosenberg_vis_2', 'panas_pos_vis_3', 'panas_neg_vis_3', 'qids_vis_3', 'gad_vis_3', 'rosenberg_vis_3', 'madrs_vis_3']
    data_df = pd.DataFrame(index = df_row_headers)
    ecrf_file_path = '/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/eCRF.xlsx'
    password = 'SussexDepNF22'
    df_values_dict = {}
    location_dict = {
        'P004_vis_1_locations': {'dob': (77, 3), 'gender': (81, 3),  'handedness': (82, 3), 'exercise': (83, 3), 'education': (84, 3), 'work_status': (85, 3), 'panic': (132, 3), 'agoraphobia': (134, 3), 'social_anx': (135, 3), 'ocd': (137, 3), 'ptsd': (140, 3), 'gad': (141, 3), 'comorbid_anx': (142, 3), 'msm': (120, 3), 'psi_sociotropy': (151, 3), 'psi_autonomy': (152, 3), 'raads': (155, 3), 'panas_pos_vis_1': (161, 3), 'panas_neg_vis_1': (162, 3), 'qids_vis_1': (172, 3), 'gad_vis_1': (173, 3), 'rosenberg_vis_1': (174, 3), 'madrs_vis_1': (185, 3)},
        'P004_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 3), 'pre_memory_intensity_guilt_2': (43, 3), 'pre_memory_intensity_indignation_1': (49, 3), 'pre_memory_intensity_indignation_2': (54, 3), 'intervention': (78, 3), 'techniques_guilt': (84, 3), 'techniques_indignation': (85, 3), 'perceived_success_guilt': (86, 3), 'perceived_success_indignation': (87, 3), 'post_memory_intensity_guilt_1': (88, 3), 'post_memory_intensity_guilt 2': (92, 3), 'post_memory_intensity_indignation_1': (97, 3), 'post_memory_indignation_2': (101, 3), 'rosenberg_vis_2': (104, 3)},
        'P004_vis_3_locations': {'panas_pos_vis_3': (36, 3), 'panas_neg_vis_3': (37, 3), 'qids_vis_3': (47, 3), 'gad_vis_3': (48, 3), 'rosenberg_vis_3': (49, 3), 'madrs_vis_3': (60, 3)},
        
        'P006_vis_1_locations': {'dob': (77, 4), 'gender': (81, 4),  'handedness': (82, 4), 'exercise': (83, 4), 'education': (84, 4), 'work_status': (85, 4), 'panic': (132, 4), 'agoraphobia': (134, 4), 'social_anx': (135, 4), 'ocd': (137, 4), 'ptsd': (140, 4), 'gad': (141, 4), 'comorbid_anx': (142, 4), 'msm': (120, 4), 'psi_sociotropy': (151, 4), 'psi_autonomy': (152, 4), 'raads': (155, 4), 'panas_pos_vis_1': (161, 4), 'panas_neg_vis_1': (162, 4), 'qids_vis_1': (172, 4), 'gad_vis_1': (173, 4), 'rosenberg_vis_1': (174, 4), 'madrs_vis_1': (185, 4)},
        'P006_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 4), 'pre_memory_intensity_guilt_2': (43, 4), 'pre_memory_intensity_indignation_1': (49, 4), 'pre_memory_intensity_indignation_2': (54, 4), 'intervention': (78, 4), 'techniques_guilt': (84, 4), 'techniques_indignation': (85, 4), 'perceived_success_guilt': (86, 4), 'perceived_success_indignation': (87, 4), 'post_memory_intensity_guilt_1': (88, 4), 'post_memory_intensity_guilt 2': (92, 4), 'post_memory_intensity_indignation_1': (97, 4), 'post_memory_indignation_2': (101, 4), 'rosenberg_vis_2': (104, 4)},
        'P006_vis_3_locations': {'panas_pos_vis_3': (36, 4), 'panas_neg_vis_3': (37, 4), 'qids_vis_3': (47, 4), 'gad_vis_3': (48, 4), 'rosenberg_vis_3': (49, 4), 'madrs_vis_3': (60, 4)},
    
        'P020_vis_1_locations': {'dob': (77, 7), 'gender': (81, 7),  'handedness': (82, 7), 'exercise': (83, 7), 'education': (84, 7), 'work_status': (85, 7), 'panic': (132, 7), 'agoraphobia': (134, 7), 'social_anx': (135, 7), 'ocd': (137, 7), 'ptsd': (140, 7), 'gad': (141, 7), 'comorbid_anx': (142, 7), 'msm': (120, 7), 'psi_sociotropy': (151, 7), 'psi_autonomy': (152, 7), 'raads': (155, 7), 'panas_pos_vis_1': (161, 7), 'panas_neg_vis_1': (162, 7), 'qids_vis_1': (172, 7), 'gad_vis_1': (173, 7), 'rosenberg_vis_1': (174, 7), 'madrs_vis_1': (185, 7)},
        'P020_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 6), 'pre_memory_intensity_guilt_2': (43, 6), 'pre_memory_intensity_indignation_1': (49, 6), 'pre_memory_intensity_indignation_2': (54, 6), 'intervention': (78, 6), 'techniques_guilt': (84, 6), 'techniques_indignation': (85, 6), 'perceived_success_guilt': (86, 6), 'perceived_success_indignation': (87, 6), 'post_memory_intensity_guilt_1': (88, 6), 'post_memory_intensity_guilt 2': (92, 6), 'post_memory_intensity_indignation_1': (97, 6), 'post_memory_indignation_2': (101, 6), 'rosenberg_vis_2': (104, 6)},
        'P020_vis_3_locations': {'panas_pos_vis_3': (36, 6), 'panas_neg_vis_3': (37, 6), 'qids_vis_3': (47, 6), 'gad_vis_3': (48, 6), 'rosenberg_vis_3': (49, 6), 'madrs_vis_3': (60, 6)},
    
        'P030_vis_1_locations': {'dob': (77, 6), 'gender': (81, 6),  'handedness': (82, 6), 'exercise': (83, 6), 'education': (84, 6), 'work_status': (85, 6), 'panic': (132, 6), 'agoraphobia': (134, 6), 'social_anx': (135, 6), 'ocd': (137, 6), 'ptsd': (140, 6), 'gad': (141, 6), 'comorbid_anx': (142, 6), 'msm': (120, 6), 'psi_sociotropy': (151, 6), 'psi_autonomy': (152, 6), 'raads': (155, 6), 'panas_pos_vis_1': (161, 6), 'panas_neg_vis_1': (162, 6), 'qids_vis_1': (172, 6), 'gad_vis_1': (173, 6), 'rosenberg_vis_1': (174, 6), 'madrs_vis_1': (185, 6)},
        'P030_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 5), 'pre_memory_intensity_guilt_2': (43, 5), 'pre_memory_intensity_indignation_1': (49, 5), 'pre_memory_intensity_indignation_2': (54, 5), 'intervention': (78, 5), 'techniques_guilt': (84, 5), 'techniques_indignation': (85, 5), 'perceived_success_guilt': (86, 5), 'perceived_success_indignation': (87, 5), 'post_memory_intensity_guilt_1': (88, 5), 'post_memory_intensity_guilt 2': (92, 5), 'post_memory_intensity_indignation_1': (97, 5), 'post_memory_indignation_2': (101, 5), 'rosenberg_vis_2': (104, 5)},
        'P030_vis_3_locations': {'panas_pos_vis_3': (36, 5), 'panas_neg_vis_3': (37, 5), 'qids_vis_3': (47, 5), 'gad_vis_3': (48, 5), 'rosenberg_vis_3': (49, 5), 'madrs_vis_3': (60, 5)},
    
        'P059_vis_1_locations': {'dob': (77, 23), 'gender': (81, 23),  'handedness': (82, 23), 'exercise': (83, 23), 'education': (84, 23), 'work_status': (85, 23), 'panic': (132, 23), 'agoraphobia': (134, 23), 'social_anx': (135, 23), 'ocd': (137, 23), 'ptsd': (140, 23), 'gad': (141, 23), 'comorbid_anx': (142, 23), 'msm': (120, 23), 'psi_sociotropy': (151, 23), 'psi_autonomy': (152, 23), 'raads': (155, 23), 'panas_pos_vis_1': (161, 23), 'panas_neg_vis_1': (162, 23), 'qids_vis_1': (172, 23), 'gad_vis_1': (173, 23), 'rosenberg_vis_1': (174, 23), 'madrs_vis_1': (185, 23)},
        'P059_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 20), 'pre_memory_intensity_guilt_2': (43, 20), 'pre_memory_intensity_indignation_1': (49, 20), 'pre_memory_intensity_indignation_2': (54, 20), 'intervention': (78, 20), 'techniques_guilt': (84, 20), 'techniques_indignation': (85, 20), 'perceived_success_guilt': (86, 20), 'perceived_success_indignation': (87, 20), 'post_memory_intensity_guilt_1': (88, 20), 'post_memory_intensity_guilt 2': (92, 20), 'post_memory_intensity_indignation_1': (97, 20), 'post_memory_indignation_2': (101, 20), 'rosenberg_vis_2': (104, 20)},
        'P059_vis_3_locations': {'panas_pos_vis_3': (36, 19), 'panas_neg_vis_3': (37, 19), 'qids_vis_3': (47, 19), 'gad_vis_3': (48, 19), 'rosenberg_vis_3': (49, 19), 'madrs_vis_3': (60, 19)},
    
        'P078_vis_1_locations': {'dob': (77, 9), 'gender': (81, 9),  'handedness': (82, 9), 'exercise': (83, 9), 'education': (84, 9), 'work_status': (85, 9), 'panic': (132, 9), 'agoraphobia': (134, 9), 'social_anx': (135, 9), 'ocd': (137, 9), 'ptsd': (140, 9), 'gad': (141, 9), 'comorbid_anx': (142, 9), 'msm': (120, 9), 'psi_sociotropy': (151, 9), 'psi_autonomy': (152, 9), 'raads': (155, 9), 'panas_pos_vis_1': (161, 9), 'panas_neg_vis_1': (162, 9), 'qids_vis_1': (172, 9), 'gad_vis_1': (173, 9), 'rosenberg_vis_1': (174, 9), 'madrs_vis_1': (185, 9)},
        'P078_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 7), 'pre_memory_intensity_guilt_2': (43, 7), 'pre_memory_intensity_indignation_1': (49, 7), 'pre_memory_intensity_indignation_2': (54, 7), 'intervention': (78, 7), 'techniques_guilt': (84, 7), 'techniques_indignation': (85, 7), 'perceived_success_guilt': (86, 7), 'perceived_success_indignation': (87, 7), 'post_memory_intensity_guilt_1': (88, 7), 'post_memory_intensity_guilt 2': (92, 7), 'post_memory_intensity_indignation_1': (97, 7), 'post_memory_indignation_2': (101, 7), 'rosenberg_vis_2': (104, 7)},
        'P078_vis_3_locations': {'panas_pos_vis_3': (36, 7), 'panas_neg_vis_3': (37, 7), 'qids_vis_3': (47, 7), 'gad_vis_3': (48, 7), 'rosenberg_vis_3': (49, 7), 'madrs_vis_3': (60, 7)},
    
        'P093_vis_1_locations': {'dob': (77, 11), 'gender': (81, 11),  'handedness': (82, 11), 'exercise': (83, 11), 'education': (84, 11), 'work_status': (85, 11), 'panic': (132, 11), 'agoraphobia': (134, 11), 'social_anx': (135, 11), 'ocd': (137, 11), 'ptsd': (140, 11), 'gad': (141, 11), 'comorbid_anx': (142, 11), 'msm': (120, 11), 'psi_sociotropy': (151, 11), 'psi_autonomy': (152, 11), 'raads': (155, 11), 'panas_pos_vis_1': (161, 11), 'panas_neg_vis_1': (162, 11), 'qids_vis_1': (172, 11), 'gad_vis_1': (173, 11), 'rosenberg_vis_1': (174, 11), 'madrs_vis_1': (185, 11)},
        'P093_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 8), 'pre_memory_intensity_guilt_2': (43, 8), 'pre_memory_intensity_indignation_1': (49, 8), 'pre_memory_intensity_indignation_2': (54, 8), 'intervention': (78, 8), 'techniques_guilt': (84, 8), 'techniques_indignation': (85, 8), 'perceived_success_guilt': (86, 8), 'perceived_success_indignation': (87, 8), 'post_memory_intensity_guilt_1': (88, 8), 'post_memory_intensity_guilt 2': (92, 8), 'post_memory_intensity_indignation_1': (97, 8), 'post_memory_indignation_2': (101, 8), 'rosenberg_vis_2': (104, 8)},
        'P093_vis_3_locations': {'panas_pos_vis_3': (36, 8), 'panas_neg_vis_3': (37, 8), 'qids_vis_3': (47, 8), 'gad_vis_3': (48, 8), 'rosenberg_vis_3': (49, 8), 'madrs_vis_3': (60, 8)},
    
        'P094_vis_1_locations': {'dob': (77, 12), 'gender': (81, 12),  'handedness': (82, 12), 'exercise': (83, 12), 'education': (84, 12), 'work_status': (85, 12), 'panic': (132, 12), 'agoraphobia': (134, 12), 'social_anx': (135, 12), 'ocd': (137, 12), 'ptsd': (140, 12), 'gad': (141, 12), 'comorbid_anx': (142, 12), 'msm': (120, 12), 'psi_sociotropy': (151, 12), 'psi_autonomy': (152, 12), 'raads': (155, 12), 'panas_pos_vis_1': (161, 12), 'panas_neg_vis_1': (162, 12), 'qids_vis_1': (172, 12), 'gad_vis_1': (173, 12), 'rosenberg_vis_1': (174, 12), 'madrs_vis_1': (185, 12)},
        'P094_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 9), 'pre_memory_intensity_guilt_2': (43, 9), 'pre_memory_intensity_indignation_1': (49, 9), 'pre_memory_intensity_indignation_2': (54, 9), 'intervention': (78, 9), 'techniques_guilt': (84, 9), 'techniques_indignation': (85, 9), 'perceived_success_guilt': (86, 9), 'perceived_success_indignation': (87, 9), 'post_memory_intensity_guilt_1': (88, 9), 'post_memory_intensity_guilt 2': (92, 9), 'post_memory_intensity_indignation_1': (97, 9), 'post_memory_indignation_2': (101, 9), 'rosenberg_vis_2': (104, 9)},
        'P094_vis_3_locations': {'panas_pos_vis_3': (36, 9), 'panas_neg_vis_3': (37, 9), 'qids_vis_3': (47, 9), 'gad_vis_3': (48, 9), 'rosenberg_vis_3': (49, 9), 'madrs_vis_3': (60, 9)},
    
        'P100_vis_1_locations': {'dob': (77, 13), 'gender': (81, 13),  'handedness': (82, 13), 'exercise': (83, 13), 'education': (84, 13), 'work_status': (85, 13), 'panic': (132, 13), 'agoraphobia': (134, 13), 'social_anx': (135, 13), 'ocd': (137, 13), 'ptsd': (140, 13), 'gad': (141, 13), 'comorbid_anx': (142, 13), 'msm': (120, 13), 'psi_sociotropy': (151, 13), 'psi_autonomy': (152, 13), 'raads': (155, 13), 'panas_pos_vis_1': (161, 13), 'panas_neg_vis_1': (162, 13), 'qids_vis_1': (172, 13), 'gad_vis_1': (173, 13), 'rosenberg_vis_1': (174, 13), 'madrs_vis_1': (185, 13)},
        'P100_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 10), 'pre_memory_intensity_guilt_2': (43, 10), 'pre_memory_intensity_indignation_1': (49, 10), 'pre_memory_intensity_indignation_2': (54, 10), 'intervention': (78, 10), 'techniques_guilt': (84, 10), 'techniques_indignation': (85, 10), 'perceived_success_guilt': (86, 10), 'perceived_success_indignation': (87, 10), 'post_memory_intensity_guilt_1': (88, 10), 'post_memory_intensity_guilt 2': (92, 10), 'post_memory_intensity_indignation_1': (97, 10), 'post_memory_indignation_2': (101, 10), 'rosenberg_vis_2': (104, 10)},
        'P100_vis_3_locations': {'panas_pos_vis_3': (36, 10), 'panas_neg_vis_3': (37, 10), 'qids_vis_3': (47, 10), 'gad_vis_3': (48, 10), 'rosenberg_vis_3': (49, 10), 'madrs_vis_3': (60, 10)},
    
        'P107_vis_1_locations': {'dob': (77, 14), 'gender': (81, 14),  'handedness': (82, 14), 'exercise': (83, 14), 'education': (84, 14), 'work_status': (85, 14), 'panic': (132, 14), 'agoraphobia': (134, 14), 'social_anx': (135, 14), 'ocd': (137, 14), 'ptsd': (140, 14), 'gad': (141, 14), 'comorbid_anx': (142, 14), 'msm': (120, 14), 'psi_sociotropy': (151, 14), 'psi_autonomy': (152, 14), 'raads': (155, 14), 'panas_pos_vis_1': (161, 14), 'panas_neg_vis_1': (162, 14), 'qids_vis_1': (172, 14), 'gad_vis_1': (173, 14), 'rosenberg_vis_1': (174, 14), 'madrs_vis_1': (185, 14)},
        'P107_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 11), 'pre_memory_intensity_guilt_2': (43, 11), 'pre_memory_intensity_indignation_1': (49, 11), 'pre_memory_intensity_indignation_2': (54, 11), 'intervention': (78, 11), 'techniques_guilt': (84, 11), 'techniques_indignation': (85, 11), 'perceived_success_guilt': (86, 11), 'perceived_success_indignation': (87, 11), 'post_memory_intensity_guilt_1': (88, 11), 'post_memory_intensity_guilt 2': (92, 11), 'post_memory_intensity_indignation_1': (97, 11), 'post_memory_indignation_2': (101, 11), 'rosenberg_vis_2': (104, 11)},
        'P107_vis_3_locations': {'panas_pos_vis_3': (36, 11), 'panas_neg_vis_3': (37, 11), 'qids_vis_3': (47, 11), 'gad_vis_3': (48, 11), 'rosenberg_vis_3': (49, 11), 'madrs_vis_3': (60, 11)},
    
        'P122_vis_1_locations': {'dob': (77, 17), 'gender': (81, 17),  'handedness': (82, 17), 'exercise': (83, 17), 'education': (84, 17), 'work_status': (85, 17), 'panic': (132, 17), 'agoraphobia': (134, 17), 'social_anx': (135, 17), 'ocd': (137, 17), 'ptsd': (140, 17), 'gad': (141, 17), 'comorbid_anx': (142, 17), 'msm': (120, 17), 'psi_sociotropy': (151, 17), 'psi_autonomy': (152, 17), 'raads': (155, 17), 'panas_pos_vis_1': (161, 17), 'panas_neg_vis_1': (162, 17), 'qids_vis_1': (172, 17), 'gad_vis_1': (173, 17), 'rosenberg_vis_1': (174, 17), 'madrs_vis_1': (185, 17)},
        'P122_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 14), 'pre_memory_intensity_guilt_2': (43, 14), 'pre_memory_intensity_indignation_1': (49, 14), 'pre_memory_intensity_indignation_2': (54, 14), 'intervention': (78, 14), 'techniques_guilt': (84, 14), 'techniques_indignation': (85, 14), 'perceived_success_guilt': (86, 14), 'perceived_success_indignation': (87, 14), 'post_memory_intensity_guilt_1': (88, 14), 'post_memory_intensity_guilt 2': (92, 14), 'post_memory_intensity_indignation_1': (97, 14), 'post_memory_indignation_2': (101, 14), 'rosenberg_vis_2': (104, 14)},
        'P122_vis_3_locations': {'panas_pos_vis_3': (36, 14), 'panas_neg_vis_3': (37, 14), 'qids_vis_3': (47, 14), 'gad_vis_3': (48, 14), 'rosenberg_vis_3': (49, 14), 'madrs_vis_3': (60, 14)},
    
        'P125_vis_1_locations': {'dob': (77, 18), 'gender': (81, 18),  'handedness': (82, 18), 'exercise': (83, 18), 'education': (84, 18), 'work_status': (85, 18), 'panic': (132, 18), 'agoraphobia': (134, 18), 'social_anx': (135, 18), 'ocd': (137, 18), 'ptsd': (140, 18), 'gad': (141, 18), 'comorbid_anx': (142, 18), 'msm': (120, 18), 'psi_sociotropy': (151, 18), 'psi_autonomy': (152, 18), 'raads': (155, 18), 'panas_pos_vis_1': (161, 18), 'panas_neg_vis_1': (162, 18), 'qids_vis_1': (172, 18), 'gad_vis_1': (173, 18), 'rosenberg_vis_1': (174, 18), 'madrs_vis_1': (185, 18)},
        'P125_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 15), 'pre_memory_intensity_guilt_2': (43, 15), 'pre_memory_intensity_indignation_1': (49, 15), 'pre_memory_intensity_indignation_2': (54, 15), 'intervention': (78, 15), 'techniques_guilt': (84, 15), 'techniques_indignation': (85, 15), 'perceived_success_guilt': (86, 15), 'perceived_success_indignation': (87, 15), 'post_memory_intensity_guilt_1': (88, 15), 'post_memory_intensity_guilt 2': (92, 15), 'post_memory_intensity_indignation_1': (97, 15), 'post_memory_indignation_2': (101, 15), 'rosenberg_vis_2': (104, 15)},
        'P125_vis_3_locations': {'panas_pos_vis_3': (36, 15), 'panas_neg_vis_3': (37, 15), 'qids_vis_3': (47, 15), 'gad_vis_3': (48, 15), 'rosenberg_vis_3': (49, 15), 'madrs_vis_3': (60, 15)},
    
        'P127_vis_1_locations': {'dob': (77, 16), 'gender': (81, 16),  'handedness': (82, 16), 'exercise': (83, 16), 'education': (84, 16), 'work_status': (85, 16), 'panic': (132, 16), 'agoraphobia': (134, 16), 'social_anx': (135, 16), 'ocd': (137, 16), 'ptsd': (140, 16), 'gad': (141, 16), 'comorbid_anx': (142, 16), 'msm': (120, 16), 'psi_sociotropy': (151, 16), 'psi_autonomy': (152, 16), 'raads': (155, 16), 'panas_pos_vis_1': (161, 16), 'panas_neg_vis_1': (162, 16), 'qids_vis_1': (172, 16), 'gad_vis_1': (173, 16), 'rosenberg_vis_1': (174, 16), 'madrs_vis_1': (185, 16)},
        'P127_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 13), 'pre_memory_intensity_guilt_2': (43, 13), 'pre_memory_intensity_indignation_1': (49, 13), 'pre_memory_intensity_indignation_2': (54, 13), 'intervention': (78, 13), 'techniques_guilt': (84, 13), 'techniques_indignation': (85, 13), 'perceived_success_guilt': (86, 13), 'perceived_success_indignation': (87, 13), 'post_memory_intensity_guilt_1': (88, 13), 'post_memory_intensity_guilt 2': (92, 13), 'post_memory_intensity_indignation_1': (97, 13), 'post_memory_indignation_2': (101, 13), 'rosenberg_vis_2': (104, 13)},
        'P127_vis_3_locations': {'panas_pos_vis_3': (36, 13), 'panas_neg_vis_3': (37, 13), 'qids_vis_3': (47, 13), 'gad_vis_3': (48, 13), 'rosenberg_vis_3': (49, 13), 'madrs_vis_3': (60, 13)},
   
        'P128_vis_1_locations': {'dob': (77, 15), 'gender': (81, 15),  'handedness': (82, 15), 'exercise': (83, 15), 'education': (84, 15), 'work_status': (85, 15), 'panic': (132, 15), 'agoraphobia': (134, 15), 'social_anx': (135, 15), 'ocd': (137, 15), 'ptsd': (140, 15), 'gad': (141, 15), 'comorbid_anx': (142, 15), 'msm': (120, 15), 'psi_sociotropy': (151, 15), 'psi_autonomy': (152, 15), 'raads': (155, 15), 'panas_pos_vis_1': (161, 15), 'panas_neg_vis_1': (162, 15), 'qids_vis_1': (172, 15), 'gad_vis_1': (173, 15), 'rosenberg_vis_1': (174, 15), 'madrs_vis_1': (185, 15)},
        'P128_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 12), 'pre_memory_intensity_guilt_2': (43, 12), 'pre_memory_intensity_indignation_1': (49, 12), 'pre_memory_intensity_indignation_2': (54, 12), 'intervention': (78, 12), 'techniques_guilt': (84, 12), 'techniques_indignation': (85, 12), 'perceived_success_guilt': (86, 12), 'perceived_success_indignation': (87, 12), 'post_memory_intensity_guilt_1': (88, 12), 'post_memory_intensity_guilt 2': (92, 12), 'post_memory_intensity_indignation_1': (97, 12), 'post_memory_indignation_2': (101, 12), 'rosenberg_vis_2': (104, 12)},
        'P128_vis_3_locations': {'panas_pos_vis_3': (36, 12), 'panas_neg_vis_3': (37, 12), 'qids_vis_3': (47, 12), 'gad_vis_3': (48, 12), 'rosenberg_vis_3': (49, 12), 'madrs_vis_3': (60, 12)},
    
        'P136_vis_1_locations': {'dob': (77, 19), 'gender': (81, 19),  'handedness': (82, 19), 'exercise': (83, 19), 'education': (84, 19), 'work_status': (85, 19), 'panic': (132, 19), 'agoraphobia': (134, 19), 'social_anx': (135, 19), 'ocd': (137, 19), 'ptsd': (140, 19), 'gad': (141, 19), 'comorbid_anx': (142, 19), 'msm': (120, 19), 'psi_sociotropy': (151, 19), 'psi_autonomy': (152, 19), 'raads': (155, 19), 'panas_pos_vis_1': (161, 19), 'panas_neg_vis_1': (162, 19), 'qids_vis_1': (172, 19), 'gad_vis_1': (173, 19), 'rosenberg_vis_1': (174, 19), 'madrs_vis_1': (185, 19)},
        'P136_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 16), 'pre_memory_intensity_guilt_2': (43, 16), 'pre_memory_intensity_indignation_1': (49, 16), 'pre_memory_intensity_indignation_2': (54, 16), 'intervention': (78, 16), 'techniques_guilt': (84, 16), 'techniques_indignation': (85, 16), 'perceived_success_guilt': (86, 16), 'perceived_success_indignation': (87, 16), 'post_memory_intensity_guilt_1': (88, 16), 'post_memory_intensity_guilt 2': (92, 16), 'post_memory_intensity_indignation_1': (97, 16), 'post_memory_indignation_2': (101, 16), 'rosenberg_vis_2': (104, 16)},
        'P136_vis_3_locations': {'panas_pos_vis_3': (36, 16), 'panas_neg_vis_3': (37, 16), 'qids_vis_3': (47, 16), 'gad_vis_3': (48, 16), 'rosenberg_vis_3': (49, 16), 'madrs_vis_3': (60, 16)},
    
        'P145_vis_1_locations': {'dob': (77, 21), 'gender': (81, 21),  'handedness': (82, 21), 'exercise': (83, 21), 'education': (84, 21), 'work_status': (85, 21), 'panic': (132, 21), 'agoraphobia': (134, 21), 'social_anx': (135, 21), 'ocd': (137, 21), 'ptsd': (140, 21), 'gad': (141, 21), 'comorbid_anx': (142, 21), 'msm': (120, 21), 'psi_sociotropy': (151, 21), 'psi_autonomy': (152, 21), 'raads': (155, 21), 'panas_pos_vis_1': (161, 21), 'panas_neg_vis_1': (162, 21), 'qids_vis_1': (172, 21), 'gad_vis_1': (173, 21), 'rosenberg_vis_1': (174, 21), 'madrs_vis_1': (185, 21)},
        'P145_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 18), 'pre_memory_intensity_guilt_2': (43, 18), 'pre_memory_intensity_indignation_1': (49, 18), 'pre_memory_intensity_indignation_2': (54, 18), 'intervention': (78, 18), 'techniques_guilt': (84, 18), 'techniques_indignation': (85, 18), 'perceived_success_guilt': (86, 18), 'perceived_success_indignation': (87, 18), 'post_memory_intensity_guilt_1': (88, 18), 'post_memory_intensity_guilt 2': (92, 18), 'post_memory_intensity_indignation_1': (97, 18), 'post_memory_indignation_2': (101, 18), 'rosenberg_vis_2': (104, 18)},
        'P145_vis_3_locations': {'panas_pos_vis_3': (36, 17), 'panas_neg_vis_3': (37, 17), 'qids_vis_3': (47, 17), 'gad_vis_3': (48, 17), 'rosenberg_vis_3': (49, 17), 'madrs_vis_3': (60, 17)},
    
        'P155_vis_1_locations': {'dob': (77, 22), 'gender': (81, 22),  'handedness': (82, 22), 'exercise': (83, 22), 'education': (84, 22), 'work_status': (85, 22), 'panic': (132, 22), 'agoraphobia': (134, 22), 'social_anx': (135, 22), 'ocd': (137, 22), 'ptsd': (140, 22), 'gad': (141, 22), 'comorbid_anx': (142, 22), 'msm': (120, 22), 'psi_sociotropy': (151, 22), 'psi_autonomy': (152, 22), 'raads': (155, 22), 'panas_pos_vis_1': (161, 22), 'panas_neg_vis_1': (162, 22), 'qids_vis_1': (172, 22), 'gad_vis_1': (173, 22), 'rosenberg_vis_1': (174, 22), 'madrs_vis_1': (185, 22)},
        'P155_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 19), 'pre_memory_intensity_guilt_2': (43, 19), 'pre_memory_intensity_indignation_1': (49, 19), 'pre_memory_intensity_indignation_2': (54, 19), 'intervention': (78, 19), 'techniques_guilt': (84, 19), 'techniques_indignation': (85, 19), 'perceived_success_guilt': (86, 19), 'perceived_success_indignation': (87, 19), 'post_memory_intensity_guilt_1': (88, 19), 'post_memory_intensity_guilt 2': (92, 19), 'post_memory_intensity_indignation_1': (97, 19), 'post_memory_indignation_2': (101, 19), 'rosenberg_vis_2': (104, 19)},
        'P155_vis_3_locations': {'panas_pos_vis_3': (36, 18), 'panas_neg_vis_3': (37, 18), 'qids_vis_3': (47, 18), 'gad_vis_3': (48, 18), 'rosenberg_vis_3': (49, 18), 'madrs_vis_3': (60, 18)},
    
        'P199_vis_1_locations': {'dob': (77, 27), 'gender': (81, 27),  'handedness': (82, 27), 'exercise': (83, 27), 'education': (84, 27), 'work_status': (85, 27), 'panic': (132, 27), 'agoraphobia': (134, 27), 'social_anx': (135, 27), 'ocd': (137, 27), 'ptsd': (140, 27), 'gad': (141, 27), 'comorbid_anx': (142, 27), 'msm': (120, 27), 'psi_sociotropy': (151, 27), 'psi_autonomy': (152, 27), 'raads': (155, 27), 'panas_pos_vis_1': (161, 27), 'panas_neg_vis_1': (162, 27), 'qids_vis_1': (172, 27), 'gad_vis_1': (173, 27), 'rosenberg_vis_1': (174, 27), 'madrs_vis_1': (185, 27)},
        'P199_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 22), 'pre_memory_intensity_guilt_2': (43, 22), 'pre_memory_intensity_indignation_1': (49, 22), 'pre_memory_intensity_indignation_2': (54, 22), 'intervention': (78, 22), 'techniques_guilt': (84, 22), 'techniques_indignation': (85, 22), 'perceived_success_guilt': (86, 22), 'perceived_success_indignation': (87, 22), 'post_memory_intensity_guilt_1': (88, 22), 'post_memory_intensity_guilt 2': (92, 22), 'post_memory_intensity_indignation_1': (97, 22), 'post_memory_indignation_2': (101, 22), 'rosenberg_vis_2': (104, 22)},
        'P199_vis_3_locations': {'panas_pos_vis_3': (36, 21), 'panas_neg_vis_3': (37, 21), 'qids_vis_3': (47, 21), 'gad_vis_3': (48, 21), 'rosenberg_vis_3': (49, 21), 'madrs_vis_3': (60, 21)},

        'P215_vis_1_locations': {'dob': (77, 26), 'gender': (81, 26),  'handedness': (82, 26), 'exercise': (83, 26), 'education': (84, 26), 'work_status': (85, 26), 'panic': (132, 26), 'agoraphobia': (134, 26), 'social_anx': (135, 26), 'ocd': (137, 26), 'ptsd': (140, 26), 'gad': (141, 26), 'comorbid_anx': (142, 26), 'msm': (120, 26), 'psi_sociotropy': (151, 26), 'psi_autonomy': (152, 26), 'raads': (155, 26), 'panas_pos_vis_1': (161, 26), 'panas_neg_vis_1': (162, 26), 'qids_vis_1': (172, 26), 'gad_vis_1': (173, 26), 'rosenberg_vis_1': (174, 26), 'madrs_vis_1': (185, 26)},
        'P215_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 21), 'pre_memory_intensity_guilt_2': (43, 21), 'pre_memory_intensity_indignation_1': (49, 21), 'pre_memory_intensity_indignation_2': (54, 21), 'intervention': (78, 21), 'techniques_guilt': (84, 21), 'techniques_indignation': (85, 21), 'perceived_success_guilt': (86, 21), 'perceived_success_indignation': (87, 21), 'post_memory_intensity_guilt_1': (88, 21), 'post_memory_intensity_guilt 2': (92, 21), 'post_memory_intensity_indignation_1': (97, 21), 'post_memory_indignation_2': (101, 21), 'rosenberg_vis_2': (104, 21)},
        'P215_vis_3_locations': {'panas_pos_vis_3': (36, 20), 'panas_neg_vis_3': (37, 20), 'qids_vis_3': (47, 20), 'gad_vis_3': (48, 20), 'rosenberg_vis_3': (49, 20), 'madrs_vis_3': (60, 20)},

        'P216_vis_1_locations': {'dob': (77, 28), 'gender': (81, 28),  'handedness': (82, 28), 'exercise': (83, 28), 'education': (84, 28), 'work_status': (85, 28), 'panic': (132, 28), 'agoraphobia': (134, 28), 'social_anx': (135, 28), 'ocd': (137, 28), 'ptsd': (140, 28), 'gad': (141, 28), 'comorbid_anx': (142, 28), 'msm': (120, 28), 'psi_sociotropy': (151, 28), 'psi_autonomy': (152, 28), 'raads': (155, 28), 'panas_pos_vis_1': (161, 28), 'panas_neg_vis_1': (162, 28), 'qids_vis_1': (172, 28), 'gad_vis_1': (173, 28), 'rosenberg_vis_1': (174, 28), 'madrs_vis_1': (185, 28)},
        'P216_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 23), 'pre_memory_intensity_guilt_2': (43, 23), 'pre_memory_intensity_indignation_1': (49, 23), 'pre_memory_intensity_indignation_2': (54, 23), 'intervention': (78, 23), 'techniques_guilt': (84, 23), 'techniques_indignation': (85, 23), 'perceived_success_guilt': (86, 23), 'perceived_success_indignation': (87, 23), 'post_memory_intensity_guilt_1': (88, 23), 'post_memory_intensity_guilt 2': (92, 23), 'post_memory_intensity_indignation_1': (97, 23), 'post_memory_indignation_2': (101, 23), 'rosenberg_vis_2': (104, 23)},
        'P216_vis_3_locations': {'panas_pos_vis_3': (36, 22), 'panas_neg_vis_3': (37, 22), 'qids_vis_3': (47, 22), 'gad_vis_3': (48, 22), 'rosenberg_vis_3': (49, 22), 'madrs_vis_3': (60, 22)}
    }
    for x in participants:
        print(f'Extracting {x} data from eCRF.xlsx.')
        decrypted_workbook = io.BytesIO()
        with open(ecrf_file_path, 'rb') as file:
            office_file = msoffcrypto.OfficeFile(file)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted_workbook)
        workbook = openpyxl.load_workbook(decrypted_workbook, data_only = True)
        ecrf_sheet = workbook['Visit 1']
        vis_1_values = [ecrf_sheet.cell(row=row, column=column).value for (row, column) in location_dict[f'{x}_vis_1_locations'].values()]
        ecrf_sheet = workbook['Visit 2']
        vis_2_values = [ecrf_sheet.cell(row=row, column=column).value for (row, column) in location_dict[f'{x}_vis_2_locations'].values()]
        ecrf_sheet = workbook['Visit 3']
        vis_3_values = [ecrf_sheet.cell(row=row, column=column).value for (row, column) in location_dict[f'{x}_vis_3_locations'].values()]
        df_values_dict[f'{x}'] = vis_1_values + vis_2_values + vis_3_values
        for key, values in df_values_dict.items():
            data_df[key] = values
        ecrf_data_path = '/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/thermometer_analysis/ecrf_data.xlsx'
        data_df.to_excel(ecrf_data_path, index=True)
        workbook.close()
    warnings.resetwarnings()
    print("eCRF data extracted.")

    # Step 4: Organise data.
    print("\n###### STEP 4: FORMAT DATA ######")
    ecrf_data = pd.read_excel(ecrf_data_path, index_col='Unnamed: 0')
    therm_data = pd.read_excel(therm_data_path, index_col='Unnamed: 0')
    intervention_row = ecrf_data.loc['intervention', :]
    therm_data.loc['intervention', :] = intervention_row
    participants = ['P004', 'P006', 'P020', 'P030', 'P059', 'P078', 'P093', 'P094', 'P100', 'P107', 'P122', 'P125', 'P127', 'P128', 'P136', 'P145', 'P155', 'P199', 'P215', 'P216']
    therm_lvl_column = []
    therm_val_column = []
    participant_column = []
    condition_column = []
    intervention_column = []
    for participant in participants:
        if participant in therm_data.columns:
            guilt_lvls = therm_data.loc[(therm_data.index.str.contains('guilt') & therm_data.index.str.contains('lvl')), therm_data.columns.str.contains(f'{participant}')].values.flatten().tolist()
            indig_lvls = therm_data.loc[(therm_data.index.str.contains('indig') & therm_data.index.str.contains('lvl')), therm_data.columns.str.contains(f'{participant}')].values.flatten().tolist()
            guilt_vals = therm_data.loc[(therm_data.index.str.contains('guilt') & therm_data.index.str.contains('val')), therm_data.columns.str.contains(f'{participant}')].values.flatten().tolist()
            indig_vals = therm_data.loc[(therm_data.index.str.contains('indig') & therm_data.index.str.contains('val')), therm_data.columns.str.contains(f'{participant}')].values.flatten().tolist()
            therm_lvl_column.extend(guilt_lvls + indig_lvls)
            therm_val_column.extend(guilt_vals + indig_vals)
            participant_column += [participant] * (len(guilt_lvls) + len(indig_lvls))
            condition_column += (['guilt'] * len(guilt_lvls)) + (['indig'] * (len(guilt_lvls)))
            intervention_column += therm_data.loc['intervention', participant] * (len(guilt_lvls) + len(indig_lvls))
    columns = ['participant', 'condition','intervention', 'therm_lvl', 'therm_val']
    therm_df = pd.DataFrame(columns=columns)
    therm_df['participant'] = participant_column
    therm_df['condition'] = condition_column
    therm_df['intervention'] = intervention_column
    therm_df['therm_lvl'] = therm_lvl_column
    therm_df['therm_val'] = therm_val_column
    print("Data formatted.")

    # Step 5: Perform LMM of thermometer levels.
    print("\n###### STEP 5: LMM OF 0-10 THERMOMETER LEVELS ######")
    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(nortest)
    # therm_df <- as.data.frame(therm_df)
    # model <- lmer(therm_lvl~condition*intervention + (1|participant), data = therm_df)
    # residuals_model <- residuals(model)
    # ad_test_result <- ad.test(residuals_model)
    # print(ad_test_result)
    # if (ad_test_result$p.value > 0.05) {
    # print("LMM of means residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # } else {
    # print("LMM of means residuals do not meet normality assumptions.")
    # }
    # """
    # ro.globalenv['therm_df'] = pandas2ri.py2rpy(therm_df)
    # result = r(r_script)
    # print(result)
    print("LMM cannot be performed on server. Please run code instead on local Spyder software.")

    # Step 6: Plot histogram of thermometer level data.
    print("\n###### STEP 6: PLOT HISTOGRAM OF 0-10 THERMOMETER LEVELS ######")
    therm_lvl_hist = ggplot(therm_df) + \
        geom_histogram(aes(x='therm_lvl'), binwidth=1, fill='skyblue', color='black', alpha=1) + \
        theme_classic() + \
        xlab('Thermometer Level') + \
        scale_x_continuous(breaks=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10]) + \
        scale_y_continuous(expand=(0, 0)) + \
        labs(y='Count') + \
        ggtitle('Histogram of Thermometer Levels')
    therm_lvl_hist.save('analysis/thermometer_analysis/figs/therm_lvl_hist.png')
    print("Histogram of 0-10 thermometer levels plotted.")

    # Step 7: Plot mean thermometer levels.
    print("\n###### STEP 7: PLOT MEAN 0-10 THERMOMETER LEVELS ######")
    mean_therm_lvl_df = therm_df.groupby(['participant', 'condition', 'intervention'])['therm_lvl'].mean().reset_index()
    mean_therm_lvl_group_df = mean_therm_lvl_df.groupby(['condition', 'intervention']).agg(
        mean_therm_lvl=('therm_lvl', 'mean'),
        std_dev=('therm_lvl', 'std'),
        n=('therm_lvl', 'size')
    ).reset_index()
    mean_therm_lvl_group_df['std_error'] = mean_therm_lvl_group_df['std_dev'] / np.sqrt(mean_therm_lvl_group_df['n'])
    mean_therm_lvl_group_df = mean_therm_lvl_group_df.drop(columns=['std_dev'])
    mean_lvl_plot = (ggplot(mean_therm_lvl_group_df, aes(x='intervention', y='mean_therm_lvl', fill='condition')) +
        geom_bar(stat='identity', position='dodge') + 
        geom_errorbar(aes(ymin='mean_therm_lvl - std_error', ymax='mean_therm_lvl + std_error'), position=position_dodge(width=0.9), width=0.2) +
        theme_classic() +
        scale_fill_manual(values=['indianred', 'skyblue']) +
        labs(title="Mean Thermometer Levels for Guilt and Indignation in Interventions A and B.", x='Intervention', y='Mean Thermometer Level') +
        scale_y_continuous(expand=(0, 0), limits=[0,3.5])
        )
    mean_lvl_plot.save('analysis/thermometer_analysis/figs/mean_lvl_plot.png')
    print("Mean 0-10 thermometer levels plotted.")

    # Step 8: Plot proportion of volumes with thermometer levels greater than 0.
    print("\n###### STEP 8: PLOT PROPORTION OF 0-10 THERMOMETER LEVELS > 0 ######")
    prop_therm_lvl_df = therm_df.groupby(['participant', 'condition', 'intervention']).agg(
        prop=('therm_lvl', lambda x: (x > 0).mean())
    ).reset_index()
    prop_therm_lvl_group_df = prop_therm_lvl_df.groupby(['condition', 'intervention']).agg(
        prop=('prop', 'mean'),
        std_dev=('prop', 'std'),
        n=('prop', 'size')
    ).reset_index()
    prop_therm_lvl_group_df['std_error'] = prop_therm_lvl_group_df['std_dev'] / np.sqrt(prop_therm_lvl_group_df['n'])
    prop_therm_lvl_group_df = prop_therm_lvl_group_df.drop(columns=['std_dev'])
    prop_lvl_plot = (ggplot(prop_therm_lvl_group_df, aes(x='intervention', y='prop', fill='condition')) +
        geom_bar(stat='identity', position='dodge') + 
        geom_errorbar(aes(ymin='prop - std_error', ymax='prop + std_error'), position=position_dodge(width=0.9), width=0.2) +
        theme_classic() +
        scale_fill_manual(values=['indianred', 'skyblue']) +
        labs(title='Proportion Plot', x='Intervention', y='Proportion Volumes with Therm. Level > 0') +
        scale_y_continuous(expand=(0, 0), limits=[0,0.8], breaks=[0,0.1,0.2,0.3,0.4,0.5,0.6,0.7,0.8])
        )
    prop_lvl_plot.save('analysis/thermometer_analysis/figs/prop_lvl_plot.png')
    print("Proportion of 0-10 thermometer levels > 0 plotted.")

    # Step 9: Calculate expanded thermometer level values and plot histograms.
    print("\n###### STEP 9: CALCULATE EXPANDED THERMOMETER LEVELS ######")
    therm_df['therm_lvl_exp'] = therm_df['therm_val'].round(1)
    for x in therm_df.index:
        if (therm_df.loc[x, 'intervention'] == 'a') and (therm_df.loc[x, 'condition'] == 'guilt'):
            transformed_value = round(abs(therm_df.loc[x, 'therm_lvl_exp'] - 100) * 10)
            if therm_df.loc[x, 'therm_lvl_exp'] < 100:
                transformed_value *= -1
            therm_df.loc[x, 'therm_lvl_exp'] = transformed_value
        elif (therm_df.loc[x, 'intervention'] == 'a') and (therm_df.loc[x, 'condition'] == 'indig'):
            transformed_value = round(abs(therm_df.loc[x, 'therm_lvl_exp'] - 100) * 10)
            if therm_df.loc[x, 'therm_lvl_exp'] > 100:
                transformed_value *= -1
            therm_df.loc[x, 'therm_lvl_exp'] = transformed_value
        elif (therm_df.loc[x, 'intervention'] == 'b') and (therm_df.loc[x, 'condition'] == 'guilt'):
            transformed_value = round(abs(therm_df.loc[x, 'therm_lvl_exp'] - 100) * 10)
            if therm_df.loc[x, 'therm_lvl_exp'] > 100:
                transformed_value *= -1
            therm_df.loc[x, 'therm_lvl_exp'] = transformed_value
        elif (therm_df.loc[x, 'intervention'] == 'b') and (therm_df.loc[x, 'condition'] == 'indig'):
            transformed_value = round(abs(therm_df.loc[x, 'therm_lvl_exp'] - 100) * 10)
            if therm_df.loc[x, 'therm_lvl_exp'] < 100:
                transformed_value *= -1
            therm_df.loc[x, 'therm_lvl_exp'] = transformed_value
    conditions = ['guilt', 'indig']
    interventions = ['a', 'b']
    for condition in conditions:
        for intervention in interventions:
            filtered_df = therm_df[(therm_df['condition'] == condition) & (therm_df['intervention'] == intervention)]       
            therm_lvl_exp_hist = (ggplot(filtered_df, aes(x='therm_lvl_exp')) +
                geom_histogram(binwidth=1, fill='skyblue', color='black', alpha=1) +
                labs(title=f'Histogram of Expanded Thermometer Levels for {condition} in {intervention}', x="Expanded Thermometer Level", y="Count") +
                theme_classic() +
                geom_vline(xintercept=0, linetype='dashed', color='red') +
                geom_vline(xintercept=10, linetype='dashed', color='red') +
                scale_y_continuous(expand=(0, 0), limits=[0,200], breaks=[0,50,100,150,200]) +
                scale_x_continuous(expand=(0, 0), limits=[-30,30], breaks=[-30,-20,-10,0,10,20,30])
            )       
            therm_lvl_exp_hist.save(f'analysis/thermometer_analysis/figs/therm_lvl_exp_hist_{condition}_{intervention}.png')
    print("Expanded thermometer levels calculated and histograms plotted.")

    # Step 10: Correlation of 0-10 and expanded thermometer levels.
    print("\n###### STEP 10: CORRELATION OF 0-10 AND EXPANDED THERMOMETER LEVELS ######")
    mean_therm_lvl_exp_df = therm_df.groupby(['participant', 'condition', 'intervention'])['therm_lvl_exp'].mean().reset_index()
    mean_therm_lvl_exp_group_df = mean_therm_lvl_exp_df.groupby(['condition', 'intervention']).agg(
        mean_therm_lvl_exp=('therm_lvl_exp', 'mean'),
        std_dev=('therm_lvl_exp', 'std'),
        n=('therm_lvl_exp', 'size')
    ).reset_index()
    mean_therm_lvl_exp_group_df['std_error'] = mean_therm_lvl_exp_group_df['std_dev'] / np.sqrt(mean_therm_lvl_exp_group_df['n'])
    mean_therm_lvl_exp_group_df = mean_therm_lvl_exp_group_df.drop(columns=['std_dev'])    
    merged_df = pd.merge(mean_therm_lvl_df, mean_therm_lvl_exp_df, on=['participant', 'condition', 'intervention'])
    correlations = merged_df.groupby(['condition', 'intervention']).apply(
        lambda group: group['therm_lvl'].corr(group['therm_lvl_exp'])
    ).reset_index(name='correlation')
    print(correlations)
    for condition in merged_df['condition'].unique():
        for intervention in merged_df['intervention'].unique():
            filtered_df = merged_df[(merged_df['condition'] == condition) & (merged_df['intervention'] == intervention)]
            correlation = correlations[(correlations['condition'] == condition) & (correlations['intervention'] == intervention)]['correlation'].values[0]
            correlation_plot = (
                ggplot(filtered_df, aes(x='therm_lvl', y='therm_lvl_exp'))
                + geom_point()
                + geom_smooth(method='lm', color='blue')
                + theme_classic()
                + labs(
                    title=f'Scatter Plot for {condition}, Intervention {intervention}\nCorrelation: {correlation:.2f}',
                    x='0-10 Thermometer Level',
                    y='Expanded Thermometer Level'
                )
                + scale_x_continuous(breaks=[0, 1, 2, 3, 4, 5, 6, 7, 8, 9, 10])
            )
            correlation_plot.save(f'analysis/thermometer_analysis/figs/correlation_plot_{condition}_{intervention}.png')
    print("Correlation of 0-10 and expanded thermometer levels completed.")

    # Step 11: Perform one-sample t-tests on expanded thermometer level data with prior participant pooling.
    print("\n###### STEP 11: ONE-SAMPLE T-TESTS OF EXPANDED THERMOMETER LEVELS ######")
    t_test_results = mean_therm_lvl_exp_df.groupby(['condition', 'intervention']).apply(
        lambda group: pd.Series({
            'shapiro_p_val': stats.shapiro(group['therm_lvl_exp']).pvalue,
            'test_type': 'para' if stats.shapiro(group['therm_lvl_exp']).pvalue > 0.05 else 'non-para',
            't_stat_or_w_stat': stats.ttest_1samp(group['therm_lvl_exp'], 0).statistic if stats.shapiro(group['therm_lvl_exp']).pvalue > 0.05 else stats.wilcoxon(group['therm_lvl_exp'] - 0).statistic,
            'p_val': stats.ttest_1samp(group['therm_lvl_exp'], 0).pvalue if stats.shapiro(group['therm_lvl_exp']).pvalue > 0.05 else stats.wilcoxon(group['therm_lvl_exp'] - 0).pvalue
        })
    ).reset_index()
    print(t_test_results)
    print("One-sample t-tests of expanded thermometer levels completed.")
        
    # Step 12: Perform one-sample t-tests on expanded thermometer level data for each participant.
    print("\n###### STEP 12: ONE-SAMPLE T-TESTS OF EXPANDED THERMOMETER LEVELS PER PARTICIPANT ######")
    t_test_results = therm_df.groupby(['participant', 'condition', 'intervention']).apply(
        lambda group: pd.Series({
            'shapiro_p_val': stats.shapiro(group['therm_lvl_exp']).pvalue,
            'test_type': 'para' if stats.shapiro(group['therm_lvl_exp']).pvalue > 0.05 else 'non-para',
            't_stat_or_w_stat': (
                stats.ttest_1samp(group['therm_lvl_exp'], 0).statistic 
                if stats.shapiro(group['therm_lvl_exp']).pvalue > 0.05 
                else stats.wilcoxon(group['therm_lvl_exp'] - 0).statistic
            ),
            'p_val': (
                stats.ttest_1samp(group['therm_lvl_exp'], 0).pvalue 
                if stats.shapiro(group['therm_lvl_exp']).pvalue > 0.05 
                else stats.wilcoxon(group['therm_lvl_exp'] - 0).pvalue
            )
        })
    ).reset_index()
    p_values = t_test_results['p_val']
    adjusted_p_values = multipletests(p_values, method='bonferroni')[1]
    t_test_results['p_val_adj'] = adjusted_p_values
    def add_asterisks(p_val):
        if p_val < 0.001:
            return '***'
        elif p_val < 0.01:
            return '**'
        elif p_val < 0.05:
            return '*'
        else:
            return ''
    t_test_results['significance_unadjusted'] = t_test_results['p_val'].apply(add_asterisks)
    t_test_results['significance_adjusted'] = t_test_results['p_val_adj'].apply(add_asterisks)
    print(t_test_results)
    print("Per participant one-sample t-tests of expanded thermometer levels completed.")

    # Step 13: Perform LMM on expanded thermometer level data.
    print("\n###### STEP 13: LMM OF EXPANDED THERMOMETER LEVELS ######")
    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # grdevices = importr('grDevices')
    # clubSandwich = importr('clubSandwich')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(nortest)
    # library(clubSandwich)
    # therm_df <- as.data.frame(therm_df)
    # model <- lmer(therm_lvl_exp~condition*intervention + (1|participant), data = therm_df)
    # robust_se <- vcovCR(model, type = "CR0")
    # residuals_model <- residuals(model)
    # ad_test_result <- ad.test(residuals_model)
    # print(ad_test_result)
    # png(filename="lmm_residuals_qqplot.png")
    # qqnorm(residuals_model)
    # qqline(residuals_model, col = 'red')
    # dev.off()
    # print("Robust Standard Errors:")
    # print(robust_se)
    # t_crit_condition <- qt(0.975, df = 6698)
    # t_crit_intervention <- qt(0.975, df = 19)
    # print(t_crit_condition)
    # print(t_crit_intervention)
    # if (ad_test_result$p.value > 0.05) {
    # print("LMM of means residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print("Model Summary:")
    # print(summary(model))
    # print("Parameter Estimates:")
    # estimates <- fixef(model)
    # print(estimates)
    # } else {
    # print("LMM of means residuals do not meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # print("Model Summary:")
    # print(summary(model))
    # print("Parameter Estimates:")
    # estimates <- fixef(model)
    # print(estimates)
    # }
    # """
    # ro.globalenv['therm_df'] = pandas2ri.py2rpy(therm_df)
    # result = r(r_script)
    # print(result)
    print("LMM cannot be performed on server. Please run code instead on local Spyder software.")

    # Step 14: Plot mean expanded thermometer levels.
    print("\n###### STEP 14: PLOT MEAN EXPANDED THERMOMETER LEVELS ######")
    guilt_a_rse = 0.4659416
    indig_a_rse = np.sqrt(0.4659416**2 + 2.647840**2)
    guilt_b_rse = np.sqrt(0.4659416**2 + 1.0043692**2)
    indig_b_rse = np.sqrt(0.4659416**2 + 2.647840**2 + 1.0043692**2 + 3.442264**2)
    rse_column = [guilt_a_rse, indig_a_rse, guilt_b_rse, indig_b_rse]
    mean_therm_lvl_exp_df = therm_df.groupby(['participant', 'condition', 'intervention'])['therm_lvl_exp'].mean().reset_index()
    mean_therm_lvl_exp_group_df = mean_therm_lvl_exp_df.groupby(['condition', 'intervention']).agg(
        mean_therm_lvl_exp=('therm_lvl_exp', 'mean'),
        std_dev=('therm_lvl_exp', 'std'),
        n=('therm_lvl_exp', 'size')
    ).reset_index()
    mean_therm_lvl_exp_group_df['std_error'] = mean_therm_lvl_exp_group_df['std_dev'] / np.sqrt(mean_therm_lvl_exp_group_df['n'])
    mean_therm_lvl_exp_group_df = mean_therm_lvl_exp_group_df.drop(columns=['std_dev'])
    mean_therm_lvl_exp_group_df['rse'] = rse_column
    mean_lvl_exp_plot = (ggplot(mean_therm_lvl_exp_group_df, aes(x='intervention', y='mean_therm_lvl_exp', fill='condition')) +
        geom_bar(stat='identity', position='dodge') + 
        geom_errorbar(aes(ymin='mean_therm_lvl_exp - std_error', ymax='mean_therm_lvl_exp + std_error'), position=position_dodge(width=0.9), width=0.2) +
        theme_classic() +
        scale_fill_manual(values=['indianred', 'skyblue']) +
        labs(title="Mean Expanded Thermometer Levels for Guilt and Indignation in Interventions A and B.", x='Intervention', y='Mean Expanded Thermometer Level') +
        scale_y_continuous(expand=(0, 0), limits=[-2.5,3.5], breaks=[-2.0,-1.0,0.0,1.0,2.0,3.0]) +
        scale_x_discrete(labels={'a': 'A', 'b': 'B'}) +
        geom_hline(yintercept=0, linetype='solid', color='black', size=0.5)
        )
    mean_lvl_exp_plot.save('analysis/thermometer_analysis/figs/mean_lvl_exp_plot.png')
    print("Mean expanded thermomter levels plotted.")

    # Step 15: Plot mean expanded thermometer level for each participant.
    print("\n###### STEP 15: PLOT MEAN EXPANDED THERMOMTER LEVELS PER PARTICIPANT ######")
    a_participants = ['P004', 'P006', 'P100', 'P128', 'P122', 'P125', 'P136', 'P145', 'P215', 'P216']
    b_participants = ['P020', 'P030', 'P059', 'P078', 'P093', 'P094', 'P107', 'P127', 'P155', 'P199']
    a_participant_means = []
    a_participant_std_errors = []
    b_participant_means = []
    b_participant_std_errors = []
    for participant in a_participants:
        guilt_a_participant_rows = therm_df[(therm_df['condition'] == 'guilt') & (therm_df['intervention'] == 'a') & (therm_df['participant'] == participant)]
        guilt_a_mean = guilt_a_participant_rows['therm_lvl_exp'].mean()
        a_participant_means.append(guilt_a_mean)
        guilt_a_std_error = guilt_a_participant_rows['therm_lvl_exp'].std() / np.sqrt(len(guilt_a_participant_rows['therm_lvl_exp']))
        a_participant_std_errors.append(guilt_a_std_error)
        indig_a_participant_rows = therm_df[(therm_df['condition'] == 'indig') & (therm_df['intervention'] == 'a') & (therm_df['participant'] == participant)]
        indig_a_mean = indig_a_participant_rows['therm_lvl_exp'].mean()
        a_participant_means.append(indig_a_mean)
        indig_a_std_error = indig_a_participant_rows['therm_lvl_exp'].std() / np.sqrt(len(indig_a_participant_rows['therm_lvl_exp']))
        a_participant_std_errors.append(indig_a_std_error)
    for participant in b_participants:
        guilt_b_participant_rows = therm_df[(therm_df['condition'] == 'guilt') & (therm_df['intervention'] == 'b') & (therm_df['participant'] == participant)]
        guilt_b_mean = guilt_b_participant_rows['therm_lvl_exp'].mean()
        b_participant_means.append(guilt_b_mean)
        guilt_b_std_error = guilt_b_participant_rows['therm_lvl_exp'].std() / np.sqrt(len(guilt_b_participant_rows['therm_lvl_exp']))
        b_participant_std_errors.append(guilt_b_std_error)
        indig_b_participant_rows = therm_df[(therm_df['condition'] == 'indig') & (therm_df['intervention'] == 'b') & (therm_df['participant'] == participant)]
        indig_b_mean = indig_b_participant_rows['therm_lvl_exp'].mean()
        b_participant_means.append(indig_b_mean)
        indig_b_std_error = indig_b_participant_rows['therm_lvl_exp'].std() / np.sqrt(len(indig_b_participant_rows['therm_lvl_exp']))
        b_participant_std_errors.append(indig_b_std_error)
    plot_data = pd.DataFrame({
        'Condition': ['Guilt', 'Indignation'] * len(a_participants) + ['Guilt', 'Indignation'] * len(b_participants),
        'Participants': sum([[p, p] for p in a_participants + b_participants], []),
        'Mean': a_participant_means + b_participant_means,
        'Std_Error': a_participant_std_errors + b_participant_std_errors
    })
    plot_data['Participants'] = pd.Categorical(plot_data['Participants'], categories=a_participants + b_participants, ordered=True)
    participant_mean_lvl_exp_plot = (ggplot(plot_data, aes(x='Participants', y='Mean', fill='Condition')) +
        geom_bar(stat='identity', position='dodge') + 
        geom_errorbar(aes(ymin='Mean - Std_Error', ymax='Mean + Std_Error'), position=position_dodge(width=0.9), width=0.2) +
        theme_classic() +
        scale_fill_manual(values=['indianred', 'skyblue']) +
        labs(title='Mean Expanded Thermometer Levels for Guilt and Indignation in Interventions A and B.', y='Mean Expanded Thermometer Level') +
        theme(axis_text_x=element_text(rotation=45, hjust=1)) +
        scale_y_continuous(expand=(0, 0), limits=[-8, 8], breaks=[-6, -4, -2, 0, 2, 4, 6]) +
        geom_hline(yintercept=0, linetype='solid', color='black', size=0.5) +
        geom_vline(xintercept=10.5, linetype='dotted', color='black', size=0.8)
        )
    participant_mean_lvl_exp_plot = participant_mean_lvl_exp_plot + annotate("text", x=5, y=-7, label="Int. A", size=16, color="black") + \
            annotate("text", x=15, y=-7, label="Int. B", size=16, color="black")
    participant_mean_lvl_exp_plot.save('analysis/thermometer_analysis/figs/participant_mean_lvl_exp_plot.png')
    print("Per participant mean expanded thermometer levels plotted.")

    # Step 16: Run t-tests between 1st and 2nd halves of each run.
    print("\n###### STEP 16: T-TESTS OF RUN-STARTS AND RUN-ENDS ######")
    def split_and_compute_means(df):
        midpoint = len(df) // 2
        first_half_mean = df.iloc[:midpoint]['therm_lvl_exp'].mean()  
        second_half_mean = df.iloc[midpoint:]['therm_lvl_exp'].mean()  
        return pd.Series({'first_half_mean': first_half_mean, 'second_half_mean': second_half_mean})
    therm_halves_df = therm_df.groupby(['participant', 'condition', 'intervention']).apply(split_and_compute_means).reset_index()
    therm_halves_long_df = pd.melt(therm_halves_df, id_vars=['participant', 'condition', 'intervention'], 
                                value_vars=['first_half_mean', 'second_half_mean'], 
                                var_name='half', value_name='therm_lvl_exp')
    mean_therm_lvl_exp_group_df = therm_halves_long_df.groupby(['condition', 'intervention', 'half']).agg(
        mean_therm_lvl_exp=('therm_lvl_exp', 'mean'),
        std_dev=('therm_lvl_exp', 'std'),
        n=('therm_lvl_exp', 'size')
    ).reset_index()
    mean_therm_lvl_exp_group_df['std_error'] = mean_therm_lvl_exp_group_df['std_dev'] / np.sqrt(mean_therm_lvl_exp_group_df['n'])
    mean_therm_lvl_exp_group_df = mean_therm_lvl_exp_group_df.drop(columns=['std_dev'])
    interventions = ['a', 'b']
    conditions = ['guilt', 'indig']
    for intervention in interventions:
        for condition in conditions:
            subset = therm_halves_long_df[(therm_halves_long_df['condition'] == condition) & (therm_halves_long_df['intervention'] == intervention)]
            first_half = subset[subset['half'] == 'first_half_mean']['therm_lvl_exp']
            second_half = subset[subset['half'] == 'second_half_mean']['therm_lvl_exp']
            shapiro_first = stats.shapiro(first_half)
            shapiro_second = stats.shapiro(second_half)
            print(f"Shapiro-Wilk Test for {condition}, {intervention} - First Half: p-value = {shapiro_first.pvalue}")
            print(f"Shapiro-Wilk Test for {condition}, {intervention} - Second Half: p-value = {shapiro_second.pvalue}")
            if shapiro_first.pvalue > 0.05 and shapiro_second.pvalue > 0.05:
                _, p_value = stats.ttest_rel(first_half, second_half)
                test_type = "Paired T-Test"
            else:
                _, p_value = stats.wilcoxon(first_half, second_half)
                test_type = "Wilcoxon Signed-Rank Test"
            print(f"{test_type} for {condition}, {intervention}:")
            print(f"P-value: {p_value}\n")
    print("T-tests of run-starts and run-ends completed.")

    # Step 17: Plot mean expanded thermometer levels for 1st and 2nd halves of each run.
    print("\n###### STEP 17: PLOT MEAN EXPANDED THERMOMETER LEVELS FOR RUN-START AND RUN-END ######")
    run_list = ['Intervention A\n+ Guilt', 'Intervention A\n+ Indig.', 'Intervention B\n+ Guilt', 'Intervention B\n+ Indig']
    start_list = [mean_therm_lvl_exp_group_df['mean_therm_lvl_exp'].iloc[0], mean_therm_lvl_exp_group_df['mean_therm_lvl_exp'].iloc[4], mean_therm_lvl_exp_group_df['mean_therm_lvl_exp'].iloc[2], mean_therm_lvl_exp_group_df['mean_therm_lvl_exp'].iloc[6]]
    end_list = [mean_therm_lvl_exp_group_df['mean_therm_lvl_exp'].iloc[1], mean_therm_lvl_exp_group_df['mean_therm_lvl_exp'].iloc[5], mean_therm_lvl_exp_group_df['mean_therm_lvl_exp'].iloc[3], mean_therm_lvl_exp_group_df['mean_therm_lvl_exp'].iloc[7]]
    start_std_error_list = [mean_therm_lvl_exp_group_df['std_error'].iloc[0], mean_therm_lvl_exp_group_df['std_error'].iloc[4], mean_therm_lvl_exp_group_df['std_error'].iloc[2], mean_therm_lvl_exp_group_df['std_error'].iloc[6]]
    end_std_error_list = [mean_therm_lvl_exp_group_df['std_error'].iloc[1], mean_therm_lvl_exp_group_df['std_error'].iloc[5], mean_therm_lvl_exp_group_df['std_error'].iloc[3], mean_therm_lvl_exp_group_df['std_error'].iloc[7]]
    plot_data = pd.DataFrame({
        'run': run_list * 2,
        'mean': start_list + end_list,
        'fill': ['Start'] * len(run_list) + ['End'] * len(run_list),
        'std_error': start_std_error_list + end_std_error_list
        })
    plot_data['fill'] = pd.Categorical(plot_data['fill'], categories=['Start', 'End'], ordered=True)
    run_startend_means_plot = (ggplot(plot_data, aes(x='run', y='mean', fill='fill')) +
        geom_bar(stat='identity', position='dodge', width=0.8) +
        geom_errorbar(aes(ymin='mean - std_error', ymax='mean + std_error'), position=position_dodge(width=0.8), width=0.2) +
        theme_classic() +
        theme(axis_text_x=element_text()) +  
        scale_y_continuous(expand=(0, 0), limits=[-4,4], breaks=[-4.0, -2.0, 0.0, 2.0, 4.0]) +
        labs(x='Run Type', y='Mean Expanded Thermometer Level', fill='Run Stage') +
        ggtitle('Mean Expanded Thermometer Levels for Run Start vs End') + 
        scale_fill_manual(values=['indianred', 'skyblue']) +
        geom_hline(yintercept=0, linetype='solid', color='black', size=0.5))
    run_startend_means_plot.save('analysis/thermometer_analysis/figs/run_startend_means_plot.png')
    print("Mean expanded thermometer levels for run-start and run-end plotted.")

    # Step 18: Calculate proportion of volumes for 1st and 2nd halves of each run where expanded thermometer level > 0, and plot.
    print("\n###### STEP 18: PROPORTION OF EXPANDED THERMOMETER LEVELS > 0 FOR RUN-START AND RUN-END ######")
    def split_and_compute_proportion(df):
        midpoint = len(df) // 2
        first_half_proportion = (df.iloc[:midpoint]['therm_lvl_exp'] > 0).mean()
        second_half_proportion = (df.iloc[midpoint:]['therm_lvl_exp'] > 0).mean()
        return pd.Series({'first_half_proportion': first_half_proportion, 'second_half_proportion': second_half_proportion})
    therm_halves_df = therm_df.groupby(['participant', 'condition', 'intervention']).apply(split_and_compute_proportion).reset_index()
    therm_halves_long_df = pd.melt(therm_halves_df, id_vars=['participant', 'condition', 'intervention'], 
                                value_vars=['first_half_proportion', 'second_half_proportion'], 
                                var_name='half', value_name='proportion')
    proportion_group_df = therm_halves_long_df.groupby(['condition', 'intervention', 'half']).agg(
        mean_proportion=('proportion', 'mean'),
        std_dev=('proportion', 'std'),
        n=('proportion', 'size')
    ).reset_index()
    proportion_group_df['std_error'] = proportion_group_df['std_dev'] / np.sqrt(proportion_group_df['n'])
    proportion_group_df = proportion_group_df.drop(columns=['std_dev'])
    run_list = ['Intervention A\n+ Guilt', 'Intervention A\n+ Indig.', 'Intervention B\n+ Guilt', 'Intervention B\n+ Indig']
    start_list = [proportion_group_df['mean_proportion'].iloc[0], proportion_group_df['mean_proportion'].iloc[4], proportion_group_df['mean_proportion'].iloc[2], proportion_group_df['mean_proportion'].iloc[6]]
    end_list = [proportion_group_df['mean_proportion'].iloc[1], proportion_group_df['mean_proportion'].iloc[5], proportion_group_df['mean_proportion'].iloc[3], proportion_group_df['mean_proportion'].iloc[7]]
    start_std_error_list = [proportion_group_df['std_error'].iloc[0], proportion_group_df['std_error'].iloc[4], proportion_group_df['std_error'].iloc[2], proportion_group_df['std_error'].iloc[6]]
    end_std_error_list = [proportion_group_df['std_error'].iloc[1], proportion_group_df['std_error'].iloc[5], proportion_group_df['std_error'].iloc[3], proportion_group_df['std_error'].iloc[7]]
    plot_data = pd.DataFrame({
        'run': run_list * 2,
        'mean': start_list + end_list,
        'fill': ['Start'] * len(run_list) + ['End'] * len(run_list),
        'std_error': start_std_error_list + end_std_error_list
        })
    plot_data['fill'] = pd.Categorical(plot_data['fill'], categories=['Start', 'End'], ordered=True)
    run_startend_prop_plot = (ggplot(plot_data, aes(x='run', y='mean', fill='fill')) +
        geom_bar(stat='identity', position='dodge', width=0.8) +
        geom_errorbar(aes(ymin='mean - std_error', ymax='mean + std_error'), position=position_dodge(width=0.8), width=0.2) +
        theme_classic() +
        theme(axis_text_x=element_text()) +  
        scale_y_continuous(expand=(0, 0), limits=[0,0.75]) +
        labs(x='Run Type', y='Mean Expanded Thermometer Level', fill='Run Stage') +
        ggtitle('Mean Expanded Thermometer Levels for Run Start vs End') + 
        scale_fill_manual(values=['indianred', 'skyblue']) +
        geom_hline(yintercept=0, linetype='solid', color='black', size=0.5))
    run_startend_prop_plot.save('analysis/thermometer_analysis/figs/run_startend_prop_plot.png')
    print("Proportion of expanded thermometer levels > 0 for run-start and run-end plotted.")

    # Step 19: Generate TMS Metric and plot.
    print("\n###### STEP 19: GENERATE TMS METRIC ######")
    tms_df = therm_df.groupby(['participant', 'condition', 'intervention'])['therm_lvl_exp'].agg(
        mean_therm_lvl_exp='mean',
        std_dev='std'
    ).reset_index()
    tms_df['cv'] = np.where(tms_df['mean_therm_lvl_exp'] != 0, tms_df['std_dev'] / tms_df['mean_therm_lvl_exp'], 0)
    tms_df['raw_tms'] = tms_df['mean_therm_lvl_exp'] * (1 - tms_df['cv'])
    min_tms = tms_df['raw_tms'].min()
    max_tms = tms_df['raw_tms'].max()
    tms_df['tms'] = (tms_df['raw_tms'] - min_tms) / (max_tms - min_tms)
    tms_group_df = tms_df.groupby(['condition', 'intervention']).agg(
        tms=('tms', 'mean'),
        std_dev=('tms', 'std'),
        n=('tms', 'size')
    ).reset_index()
    tms_group_df['std_error'] = tms_group_df['std_dev'] / np.sqrt(tms_group_df['n'])
    tms_group_df = tms_group_df.drop(columns=['std_dev'])   
    tms_plot = (ggplot(tms_group_df, aes(x='intervention', y='tms', fill='condition')) +
        geom_bar(stat='identity', position='dodge') + 
        geom_errorbar(aes(ymin='tms - std_error', ymax='tms + std_error'), position=position_dodge(width=0.9), width=0.2) +
        theme_classic() +
        scale_fill_manual(values=['indianred', 'skyblue']) +
        labs(title="TMS for Guilt and Indignation in Interventions A and B.", x='Intervention', y='TMS') +
        scale_y_continuous(expand=(0, 0), limits=[0,1]) +
        scale_x_discrete(labels={'a': 'A', 'b': 'B'}) +
        geom_hline(yintercept=0, linetype='solid', color='black', size=0.5)
        )
    tms_plot.save('analysis/thermometer_analysis/figs/tms_plot.png')
    print("TMS metric generated.")

    # Step 20: Perform correlations and Bland-Altman plots of TMS and mean vs. perceived success rating.
    print("\n###### STEP 20: CORRELATIONS OF TMS AND MEAN VS PERCEIVED SUCCESS RATING ######")
    perceived_success_guilt_values = ecrf_data.loc['perceived_success_guilt', :].tolist()
    perceived_success_indignation_values = ecrf_data.loc['perceived_success_indignation', :].tolist()
    interleaved_values = []
    max_len = max(len(perceived_success_guilt_values), len(perceived_success_indignation_values))
    for i in range(max_len):
        if i < len(perceived_success_guilt_values):
            interleaved_values.append(perceived_success_guilt_values[i])
        if i < len(perceived_success_indignation_values):
            interleaved_values.append(perceived_success_indignation_values[i])
    tms_df['perceived_success'] = interleaved_values
    correlation = tms_df['tms'].corr(tms_df['perceived_success'])
    print(f'Correlation between TMS and Perceived Success Ratings: {correlation}')
    tms_perceived_corr_plot = (ggplot(tms_df, aes(x='tms', y='perceived_success')) +
            geom_point() +
            geom_smooth(method='lm', color='blue') +
            theme_classic() +  
            labs(title=f'Scatter Plot of TMS vs Perceived Success Ratings\nCorrelation: {correlation}', 
                x='TMS', 
                y='Perceived Success Rating'))
    tms_perceived_corr_plot.save('analysis/thermometer_analysis/figs/tms_perceived_corr_plot.png')
    correlation = tms_df['mean_therm_lvl_exp'].corr(tms_df['perceived_success'])
    print(f'Correlation between mean therm_lvl_exp and Perceived Success Ratings: {correlation}')
    mean_perceived_corr_plot = (ggplot(tms_df, aes(x='mean_therm_lvl_exp', y='perceived_success')) +
            geom_point() +
            geom_smooth(method='lm', color='blue') +
            theme_classic() +  
            labs(title=f'Scatter Plot of mean therm_lvl_exp vs Perceived Success Ratings\nCorrelation: {correlation}', 
                x='Mean Expanded Thermometer Level', 
                y='Perceived Success Rating'))
    mean_perceived_corr_plot.save('analysis/thermometer_analysis/figs/mean_perceived_corr_plot.png')
    tms_df['perceived_success_norm'] = (tms_df['perceived_success'] - tms_df['perceived_success'].min()) / (tms_df['perceived_success'].max() - tms_df['perceived_success'].min())
    tms_df['mean'] = tms_df[['tms', 'perceived_success_norm']].mean(axis=1)
    tms_df['difference'] = tms_df['tms'] - tms_df['perceived_success_norm']
    mean_diff = tms_df['difference'].mean()
    std_diff = tms_df['difference'].std()
    loa_upper = mean_diff + 1.96 * std_diff
    loa_lower = mean_diff - 1.96 * std_diff
    tms_bland_altman_plot = (
        ggplot(tms_df, aes(x='mean', y='difference')) +
        geom_point(alpha=0.5) +
        geom_hline(yintercept=mean_diff, color='blue', linetype='dashed', size=1) +
        geom_hline(yintercept=loa_upper, color='red', linetype='dashed', size=1) +
        geom_hline(yintercept=loa_lower, color='red', linetype='dashed', size=1) +
        theme_classic() +
        scale_y_continuous(expand=(0, 0), limits=[-0.5,1]) +
        labs(
            title='Bland-Altman Plot of TMS vs Normalised Perceived Success',
            x='Mean of TMS and Normalised Perceived Success',
            y='Difference (TMS - Normalised Perceived Success)'
        )
    )
    tms_bland_altman_plot.save('analysis/thermometer_analysis/figs/tms_bland_altman_plot.png')
    tms_df['mean_therm_lvl_exp_norm'] = (tms_df['mean_therm_lvl_exp'] - tms_df['mean_therm_lvl_exp'].min()) / (tms_df['mean_therm_lvl_exp'].max() - tms_df['mean_therm_lvl_exp'].min())
    tms_df['mean'] = tms_df[['mean_therm_lvl_exp_norm', 'perceived_success_norm']].mean(axis=1)
    tms_df['difference'] = tms_df['mean_therm_lvl_exp_norm'] - tms_df['perceived_success_norm']
    mean_diff = tms_df['difference'].mean()
    std_diff = tms_df['difference'].std()
    loa_upper = mean_diff + 1.96 * std_diff
    loa_lower = mean_diff - 1.96 * std_diff
    mean_bland_altman_plot = (
        ggplot(tms_df, aes(x='mean', y='difference')) +
        geom_point(alpha=0.5) +
        geom_hline(yintercept=mean_diff, color='blue', linetype='dashed', size=1) +
        geom_hline(yintercept=loa_upper, color='red', linetype='dashed', size=1) +
        geom_hline(yintercept=loa_lower, color='red', linetype='dashed', size=1) +
        theme_classic() +
        scale_y_continuous(expand=(0, 0), limits=[-0.5,1]) +
        labs(
            title='Bland-Altman Plot of Normalised Mean vs Perceived Success',
            x='Mean of Normalised Mean and Perceived Success',
            y='Difference (Normalised Mean - Normalised Perceived Success)'
        )
    )
    mean_bland_altman_plot.save('analysis/thermometer_analysis/figs/mean_bland_altman_plot.png') 
    print("Correlations of TMS and mean expanded thermometer levels versus perceived success rating completed.")

    # Step 21: Plot timeseries of 0-10 thermometer levels for neurofeedback runs.
    print("\n###### STEP 21: PLOT NEUROFEEDBACK RUN TIMESERIES OF 0-10 THERMOMETER LEVELS ######")
    def insert_zero_sections(df):
        new_data = []
        for (participant, condition), group in df.groupby(['participant', 'condition']):
            section_size = len(group) // 4
            start_zeros = pd.DataFrame({
                'participant': [participant] * 30,
                'condition': [condition] * 30,
                'intervention': group['intervention'].iloc[0],
                'therm_lvl': [0] * 30,
                'therm_val': [0] * 30,
                'therm_lvl_exp': [0] * 30
                })
            new_data.append(start_zeros)
            for i in range(4):
                start_idx = i * section_size
                end_idx = (i + 1) * section_size if i < 3 else len(group)
                section = group.iloc[start_idx:end_idx]
                new_data.append(section)
                zeros_section = pd.DataFrame({
                    'participant': [participant] * 10,
                    'condition': [condition] * 10,
                    'intervention': section['intervention'].iloc[0],
                    'therm_lvl': [0] * 10,
                    'therm_val': [0] * 10,
                    'therm_lvl_exp': [0] * 10
                })
                new_data.append(zeros_section)
        return pd.concat(new_data, ignore_index=True)
    modified_therm_df = insert_zero_sections(therm_df)
    modified_therm_df['data_point'] = modified_therm_df.groupby(['participant', 'condition', 'intervention']).cumcount() + 1
    mean_therm_df = (
        modified_therm_df
        .groupby(['data_point', 'condition', 'intervention'], as_index=False)
        .agg(mean_therm_lvl=('therm_lvl', 'mean')))
    def custom_labeller(value):
        mapping = {'a': 'Intervention A', 'b': 'Intervention B'}
        return mapping.get(value, value)
    therm_timeseries_plot = (
        ggplot(mean_therm_df, aes(x='data_point', y='mean_therm_lvl', color='condition')) +
        geom_line(size=0.5) +
        geom_smooth(aes(color='condition'), method='lm', se=False, size=0.8, linetype='dashed') +
        facet_wrap('~ intervention', labeller=labeller(intervention=custom_labeller)) +
        labs(
            title="0-10 Thermometer Level across Participants by Intervention and Condition",
            x="Volume",
            y="Thermometer Level",
            color="Condition",
            linetype="Intervention"
        ) +
        theme_classic() +
        scale_y_continuous(limits=[0,8], breaks=[0,1,2,3,4,5,6,7,8]) +
        scale_x_continuous(limits=[0,250], breaks=[0,50,100,150,200,250]) +
        scale_color_manual(values={'guilt': 'indianred', 'indig': 'skyblue'}, labels={'guilt': 'Guilt', 'indig': 'Indig'}))
    therm_timeseries_plot.save('therm_timeseries_plot.png', width=14, height=6, dpi=300)
    mean_therm_exp_df = (
        modified_therm_df
        .groupby(['data_point', 'condition', 'intervention'], as_index=False)
        .agg(mean_therm_lvl_exp=('therm_lvl_exp', 'mean')))
    therm_exp_timeseries_plot = (
        ggplot(mean_therm_exp_df, aes(x='data_point', y='mean_therm_lvl_exp', color='condition')) +
        geom_line(size=0.5) +
        geom_smooth(aes(color='condition'), method='lm', se=False, size=0.8, linetype='dashed') +
        facet_wrap('~ intervention', labeller=labeller(intervention=custom_labeller)) +
        labs(
            title="Expanded Thermometer Level across Participants by Intervention and Condition",
            x="Volume",
            y="Expanded Thermometer Level",
            color="Condition",
            linetype="Intervention"
        ) +
        theme_classic() +
        scale_y_continuous(limits=[-9,9], breaks=[-8,-6,-4,-2,0,2,4,6,8]) +
        scale_x_continuous(limits=[0,250], breaks=[0,50,100,150,200,250]) +
        scale_color_manual(values={'guilt': 'indianred', 'indig': 'skyblue'}, labels={'guilt': 'Guilt', 'indig': 'Indig'}))
    therm_exp_timeseries_plot.save('therm_exp_timeseries_plot.png', width=14, height=6, dpi=300)

#endregion

#region 3) CLINICAL ANALYSIS.

def clinical_analysis():
    participants = ['P004', 'P006', 'P020', 'P030', 'P059', 'P078', 'P093', 'P094', 'P100', 'P107', 'P122', 'P125', 'P127', 'P128', 'P136', 'P145', 'P155', 'P199', 'P215', 'P216']
    runs = ['run01', 'run02', 'run03', 'run04']
    restart = logged_input("\nWould you like to start the clinical analysis from scratch? This will remove all files from the analysis/clinical_analysis folder. (y/n)\n")
    if restart == 'y':
        double_check = logged_input("\nAre you sure? (y/n)\n")
        if double_check == 'y':
            clinical_analysis_folder = 'analysis/clinical_analysis'
            print(f"Deleting analysis/clinical_analysis folder...")
            shutil.rmtree(clinical_analysis_folder)
        else:
            sys.exit()
            
    # Step 1: Create directories.
    print("\n###### STEP 1: CREATE DIRECTORIES ######")
    analysis_folder = 'analysis'
    os.makedirs(analysis_folder, exist_ok=True)
    clinical_analysis_folder = 'analysis/clinical_analysis'
    os.makedirs(clinical_analysis_folder, exist_ok=True)
    figs_folder = 'analysis/clinical_analysis/figs'
    os.makedirs(figs_folder, exist_ok=True)
    print("Directories created.")

    # Step 2: Access eCRF document and extract relevant data into dataframe.
    print("\n###### STEP 2: EXTRACT eCRF DATA ######")
    warnings.simplefilter("ignore", UserWarning)
    df_row_headers = ['dob', 'gender', 'handedness', 'exercise', 'education', 'work_status', 'panic', 'agoraphobia', 'social_anx', 'ocd', 'ptsd', 'gad', 'comorbid_anx', 'msm', 'psi_sociotropy', 'psi_autonomy', 'raads', 'panas_pos_vis_1', 'panas_neg_vis_1', 'qids_vis_1', 'gad_vis_1', 'rosenberg_vis_1', 'madrs_vis_1', 'pre_memory_intensity_guilt_1', 'pre_memory_intensity_guilt_2', 'pre_memory_intensity_indignation_1', 'pre_memory_intensity_indignation_2', 'intervention', 'techniques_guilt', 'techniques_indignation', 'perceived_success_guilt', 'perceived_success_indignation', 'post_memory_intensity_guilt_1', 'post_memory_intensity_guilt_2', 'post_memory_intensity_indignation_1', 'post_memory_intensity_indignation_2', 'rosenberg_vis_2', 'panas_pos_vis_3', 'panas_neg_vis_3', 'qids_vis_3', 'gad_vis_3', 'rosenberg_vis_3', 'madrs_vis_3', 'qids_vis_4', 'rosenberg_vis_4', 'qids_vis_5', 'rosenberg_vis_5']
    ecrf_df = pd.DataFrame(index = df_row_headers)
    ecrf_file_path = '/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/eCRF.xlsx'
    password = 'SussexDepNF22'
    df_values_dict = {}
    location_dict = {
        'P004_vis_1_locations': {'dob': (77, 3), 'gender': (81, 3),  'handedness': (82, 3), 'exercise': (83, 3), 'education': (84, 3), 'work_status': (85, 3), 'panic': (132, 3), 'agoraphobia': (134, 3), 'social_anx': (135, 3), 'ocd': (137, 3), 'ptsd': (140, 3), 'gad': (141, 3), 'comorbid_anx': (142, 3), 'msm': (120, 3), 'psi_sociotropy': (151, 3), 'psi_autonomy': (152, 3), 'raads': (155, 3), 'panas_pos_vis_1': (161, 3), 'panas_neg_vis_1': (162, 3), 'qids_vis_1': (172, 3), 'gad_vis_1': (173, 3), 'rosenberg_vis_1': (174, 3), 'madrs_vis_1': (185, 3)},
        'P004_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 3), 'pre_memory_intensity_guilt_2': (43, 3), 'pre_memory_intensity_indignation_1': (49, 3), 'pre_memory_intensity_indignation_2': (54, 3), 'intervention': (78, 3), 'techniques_guilt': (84, 3), 'techniques_indignation': (85, 3), 'perceived_success_guilt': (86, 3), 'perceived_success_indignation': (87, 3), 'post_memory_intensity_guilt_1': (88, 3), 'post_memory_intensity_guilt 2': (92, 3), 'post_memory_intensity_indignation_1': (97, 3), 'post_memory_indignation_2': (101, 3), 'rosenberg_vis_2': (104, 3)},
        'P004_vis_3_locations': {'panas_pos_vis_3': (36, 3), 'panas_neg_vis_3': (37, 3), 'qids_vis_3': (47, 3), 'gad_vis_3': (48, 3), 'rosenberg_vis_3': (49, 3), 'madrs_vis_3': (60, 3)},
        'P004_vis_4_locations': {'qids_vis_4': (26, 3), 'rosenberg_vis_4': (27, 3)},
        'P004_vis_5_locations': {'qids_vis_5': (28, 3), 'rosenberg_vis_5': (29, 3)},

        'P006_vis_1_locations': {'dob': (77, 4), 'gender': (81, 4),  'handedness': (82, 4), 'exercise': (83, 4), 'education': (84, 4), 'work_status': (85, 4), 'panic': (132, 4), 'agoraphobia': (134, 4), 'social_anx': (135, 4), 'ocd': (137, 4), 'ptsd': (140, 4), 'gad': (141, 4), 'comorbid_anx': (142, 4), 'msm': (120, 4), 'psi_sociotropy': (151, 4), 'psi_autonomy': (152, 4), 'raads': (155, 4), 'panas_pos_vis_1': (161, 4), 'panas_neg_vis_1': (162, 4), 'qids_vis_1': (172, 4), 'gad_vis_1': (173, 4), 'rosenberg_vis_1': (174, 4), 'madrs_vis_1': (185, 4)},
        'P006_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 4), 'pre_memory_intensity_guilt_2': (43, 4), 'pre_memory_intensity_indignation_1': (49, 4), 'pre_memory_intensity_indignation_2': (54, 4), 'intervention': (78, 4), 'techniques_guilt': (84, 4), 'techniques_indignation': (85, 4), 'perceived_success_guilt': (86, 4), 'perceived_success_indignation': (87, 4), 'post_memory_intensity_guilt_1': (88, 4), 'post_memory_intensity_guilt 2': (92, 4), 'post_memory_intensity_indignation_1': (97, 4), 'post_memory_indignation_2': (101, 4), 'rosenberg_vis_2': (104, 4)},
        'P006_vis_3_locations': {'panas_pos_vis_3': (36, 4), 'panas_neg_vis_3': (37, 4), 'qids_vis_3': (47, 4), 'gad_vis_3': (48, 4), 'rosenberg_vis_3': (49, 4), 'madrs_vis_3': (60, 4)},
        'P006_vis_4_locations': {'qids_vis_4': (26, 4), 'rosenberg_vis_4': (27, 4)},
        'P006_vis_5_locations': {'qids_vis_5': (28, 4), 'rosenberg_vis_5': (29, 4)},

        'P020_vis_1_locations': {'dob': (77, 7), 'gender': (81, 7),  'handedness': (82, 7), 'exercise': (83, 7), 'education': (84, 7), 'work_status': (85, 7), 'panic': (132, 7), 'agoraphobia': (134, 7), 'social_anx': (135, 7), 'ocd': (137, 7), 'ptsd': (140, 7), 'gad': (141, 7), 'comorbid_anx': (142, 7), 'msm': (120, 7), 'psi_sociotropy': (151, 7), 'psi_autonomy': (152, 7), 'raads': (155, 7), 'panas_pos_vis_1': (161, 7), 'panas_neg_vis_1': (162, 7), 'qids_vis_1': (172, 7), 'gad_vis_1': (173, 7), 'rosenberg_vis_1': (174, 7), 'madrs_vis_1': (185, 7)},
        'P020_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 6), 'pre_memory_intensity_guilt_2': (43, 6), 'pre_memory_intensity_indignation_1': (49, 6), 'pre_memory_intensity_indignation_2': (54, 6), 'intervention': (78, 6), 'techniques_guilt': (84, 6), 'techniques_indignation': (85, 6), 'perceived_success_guilt': (86, 6), 'perceived_success_indignation': (87, 6), 'post_memory_intensity_guilt_1': (88, 6), 'post_memory_intensity_guilt 2': (92, 6), 'post_memory_intensity_indignation_1': (97, 6), 'post_memory_indignation_2': (101, 6), 'rosenberg_vis_2': (104, 6)},
        'P020_vis_3_locations': {'panas_pos_vis_3': (36, 6), 'panas_neg_vis_3': (37, 6), 'qids_vis_3': (47, 6), 'gad_vis_3': (48, 6), 'rosenberg_vis_3': (49, 6), 'madrs_vis_3': (60, 6)},
        'P020_vis_4_locations': {'qids_vis_4': (26, 6), 'rosenberg_vis_4': (27, 6)},
        'P020_vis_5_locations': {'qids_vis_5': (28, 6), 'rosenberg_vis_5': (29, 6)},

        'P030_vis_1_locations': {'dob': (77, 6), 'gender': (81, 6),  'handedness': (82, 6), 'exercise': (83, 6), 'education': (84, 6), 'work_status': (85, 6), 'panic': (132, 6), 'agoraphobia': (134, 6), 'social_anx': (135, 6), 'ocd': (137, 6), 'ptsd': (140, 6), 'gad': (141, 6), 'comorbid_anx': (142, 6), 'msm': (120, 6), 'psi_sociotropy': (151, 6), 'psi_autonomy': (152, 6), 'raads': (155, 6), 'panas_pos_vis_1': (161, 6), 'panas_neg_vis_1': (162, 6), 'qids_vis_1': (172, 6), 'gad_vis_1': (173, 6), 'rosenberg_vis_1': (174, 6), 'madrs_vis_1': (185, 6)},
        'P030_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 5), 'pre_memory_intensity_guilt_2': (43, 5), 'pre_memory_intensity_indignation_1': (49, 5), 'pre_memory_intensity_indignation_2': (54, 5), 'intervention': (78, 5), 'techniques_guilt': (84, 5), 'techniques_indignation': (85, 5), 'perceived_success_guilt': (86, 5), 'perceived_success_indignation': (87, 5), 'post_memory_intensity_guilt_1': (88, 5), 'post_memory_intensity_guilt 2': (92, 5), 'post_memory_intensity_indignation_1': (97, 5), 'post_memory_indignation_2': (101, 5), 'rosenberg_vis_2': (104, 5)},
        'P030_vis_3_locations': {'panas_pos_vis_3': (36, 5), 'panas_neg_vis_3': (37, 5), 'qids_vis_3': (47, 5), 'gad_vis_3': (48, 5), 'rosenberg_vis_3': (49, 5), 'madrs_vis_3': (60, 5)},
        'P030_vis_4_locations': {'qids_vis_4': (26, 5), 'rosenberg_vis_4': (27, 5)},
        'P030_vis_5_locations': {'qids_vis_5': (28, 5), 'rosenberg_vis_5': (29, 5)},

        'P059_vis_1_locations': {'dob': (77, 23), 'gender': (81, 23),  'handedness': (82, 23), 'exercise': (83, 23), 'education': (84, 23), 'work_status': (85, 23), 'panic': (132, 23), 'agoraphobia': (134, 23), 'social_anx': (135, 23), 'ocd': (137, 23), 'ptsd': (140, 23), 'gad': (141, 23), 'comorbid_anx': (142, 23), 'msm': (120, 23), 'psi_sociotropy': (151, 23), 'psi_autonomy': (152, 23), 'raads': (155, 23), 'panas_pos_vis_1': (161, 23), 'panas_neg_vis_1': (162, 23), 'qids_vis_1': (172, 23), 'gad_vis_1': (173, 23), 'rosenberg_vis_1': (174, 23), 'madrs_vis_1': (185, 23)},
        'P059_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 20), 'pre_memory_intensity_guilt_2': (43, 20), 'pre_memory_intensity_indignation_1': (49, 20), 'pre_memory_intensity_indignation_2': (54, 20), 'intervention': (78, 20), 'techniques_guilt': (84, 20), 'techniques_indignation': (85, 20), 'perceived_success_guilt': (86, 20), 'perceived_success_indignation': (87, 20), 'post_memory_intensity_guilt_1': (88, 20), 'post_memory_intensity_guilt 2': (92, 20), 'post_memory_intensity_indignation_1': (97, 20), 'post_memory_indignation_2': (101, 20), 'rosenberg_vis_2': (104, 20)},
        'P059_vis_3_locations': {'panas_pos_vis_3': (36, 19), 'panas_neg_vis_3': (37, 19), 'qids_vis_3': (47, 19), 'gad_vis_3': (48, 19), 'rosenberg_vis_3': (49, 19), 'madrs_vis_3': (60, 19)},
        'P059_vis_4_locations': {'qids_vis_4': (26, 19), 'rosenberg_vis_4': (27, 19)},
        'P059_vis_5_locations': {'qids_vis_5': (28, 19), 'rosenberg_vis_5': (29, 19)},

        'P078_vis_1_locations': {'dob': (77, 9), 'gender': (81, 9),  'handedness': (82, 9), 'exercise': (83, 9), 'education': (84, 9), 'work_status': (85, 9), 'panic': (132, 9), 'agoraphobia': (134, 9), 'social_anx': (135, 9), 'ocd': (137, 9), 'ptsd': (140, 9), 'gad': (141, 9), 'comorbid_anx': (142, 9), 'msm': (120, 9), 'psi_sociotropy': (151, 9), 'psi_autonomy': (152, 9), 'raads': (155, 9), 'panas_pos_vis_1': (161, 9), 'panas_neg_vis_1': (162, 9), 'qids_vis_1': (172, 9), 'gad_vis_1': (173, 9), 'rosenberg_vis_1': (174, 9), 'madrs_vis_1': (185, 9)},
        'P078_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 7), 'pre_memory_intensity_guilt_2': (43, 7), 'pre_memory_intensity_indignation_1': (49, 7), 'pre_memory_intensity_indignation_2': (54, 7), 'intervention': (78, 7), 'techniques_guilt': (84, 7), 'techniques_indignation': (85, 7), 'perceived_success_guilt': (86, 7), 'perceived_success_indignation': (87, 7), 'post_memory_intensity_guilt_1': (88, 7), 'post_memory_intensity_guilt 2': (92, 7), 'post_memory_intensity_indignation_1': (97, 7), 'post_memory_indignation_2': (101, 7), 'rosenberg_vis_2': (104, 7)},
        'P078_vis_3_locations': {'panas_pos_vis_3': (36, 7), 'panas_neg_vis_3': (37, 7), 'qids_vis_3': (47, 7), 'gad_vis_3': (48, 7), 'rosenberg_vis_3': (49, 7), 'madrs_vis_3': (60, 7)},
        'P078_vis_4_locations': {'qids_vis_4': (26, 7), 'rosenberg_vis_4': (27, 7)},
        'P078_vis_5_locations': {'qids_vis_5': (28, 7), 'rosenberg_vis_5': (29, 7)},

        'P093_vis_1_locations': {'dob': (77, 11), 'gender': (81, 11),  'handedness': (82, 11), 'exercise': (83, 11), 'education': (84, 11), 'work_status': (85, 11), 'panic': (132, 11), 'agoraphobia': (134, 11), 'social_anx': (135, 11), 'ocd': (137, 11), 'ptsd': (140, 11), 'gad': (141, 11), 'comorbid_anx': (142, 11), 'msm': (120, 11), 'psi_sociotropy': (151, 11), 'psi_autonomy': (152, 11), 'raads': (155, 11), 'panas_pos_vis_1': (161, 11), 'panas_neg_vis_1': (162, 11), 'qids_vis_1': (172, 11), 'gad_vis_1': (173, 11), 'rosenberg_vis_1': (174, 11), 'madrs_vis_1': (185, 11)},
        'P093_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 8), 'pre_memory_intensity_guilt_2': (43, 8), 'pre_memory_intensity_indignation_1': (49, 8), 'pre_memory_intensity_indignation_2': (54, 8), 'intervention': (78, 8), 'techniques_guilt': (84, 8), 'techniques_indignation': (85, 8), 'perceived_success_guilt': (86, 8), 'perceived_success_indignation': (87, 8), 'post_memory_intensity_guilt_1': (88, 8), 'post_memory_intensity_guilt 2': (92, 8), 'post_memory_intensity_indignation_1': (97, 8), 'post_memory_indignation_2': (101, 8), 'rosenberg_vis_2': (104, 8)},
        'P093_vis_3_locations': {'panas_pos_vis_3': (36, 8), 'panas_neg_vis_3': (37, 8), 'qids_vis_3': (47, 8), 'gad_vis_3': (48, 8), 'rosenberg_vis_3': (49, 8), 'madrs_vis_3': (60, 8)},
        'P093_vis_4_locations': {'qids_vis_4': (26, 8), 'rosenberg_vis_4': (27, 8)},
        'P093_vis_5_locations': {'qids_vis_5': (28, 8), 'rosenberg_vis_5': (29, 8)},

        'P094_vis_1_locations': {'dob': (77, 12), 'gender': (81, 12),  'handedness': (82, 12), 'exercise': (83, 12), 'education': (84, 12), 'work_status': (85, 12), 'panic': (132, 12), 'agoraphobia': (134, 12), 'social_anx': (135, 12), 'ocd': (137, 12), 'ptsd': (140, 12), 'gad': (141, 12), 'comorbid_anx': (142, 12), 'msm': (120, 12), 'psi_sociotropy': (151, 12), 'psi_autonomy': (152, 12), 'raads': (155, 12), 'panas_pos_vis_1': (161, 12), 'panas_neg_vis_1': (162, 12), 'qids_vis_1': (172, 12), 'gad_vis_1': (173, 12), 'rosenberg_vis_1': (174, 12), 'madrs_vis_1': (185, 12)},
        'P094_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 9), 'pre_memory_intensity_guilt_2': (43, 9), 'pre_memory_intensity_indignation_1': (49, 9), 'pre_memory_intensity_indignation_2': (54, 9), 'intervention': (78, 9), 'techniques_guilt': (84, 9), 'techniques_indignation': (85, 9), 'perceived_success_guilt': (86, 9), 'perceived_success_indignation': (87, 9), 'post_memory_intensity_guilt_1': (88, 9), 'post_memory_intensity_guilt 2': (92, 9), 'post_memory_intensity_indignation_1': (97, 9), 'post_memory_indignation_2': (101, 9), 'rosenberg_vis_2': (104, 9)},
        'P094_vis_3_locations': {'panas_pos_vis_3': (36, 9), 'panas_neg_vis_3': (37, 9), 'qids_vis_3': (47, 9), 'gad_vis_3': (48, 9), 'rosenberg_vis_3': (49, 9), 'madrs_vis_3': (60, 9)},
        'P094_vis_4_locations': {'qids_vis_4': (26, 9), 'rosenberg_vis_4': (27, 9)},
        'P094_vis_5_locations': {'qids_vis_5': (28, 9), 'rosenberg_vis_5': (29, 9)},

        'P100_vis_1_locations': {'dob': (77, 13), 'gender': (81, 13),  'handedness': (82, 13), 'exercise': (83, 13), 'education': (84, 13), 'work_status': (85, 13), 'panic': (132, 13), 'agoraphobia': (134, 13), 'social_anx': (135, 13), 'ocd': (137, 13), 'ptsd': (140, 13), 'gad': (141, 13), 'comorbid_anx': (142, 13), 'msm': (120, 13), 'psi_sociotropy': (151, 13), 'psi_autonomy': (152, 13), 'raads': (155, 13), 'panas_pos_vis_1': (161, 13), 'panas_neg_vis_1': (162, 13), 'qids_vis_1': (172, 13), 'gad_vis_1': (173, 13), 'rosenberg_vis_1': (174, 13), 'madrs_vis_1': (185, 13)},
        'P100_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 10), 'pre_memory_intensity_guilt_2': (43, 10), 'pre_memory_intensity_indignation_1': (49, 10), 'pre_memory_intensity_indignation_2': (54, 10), 'intervention': (78, 10), 'techniques_guilt': (84, 10), 'techniques_indignation': (85, 10), 'perceived_success_guilt': (86, 10), 'perceived_success_indignation': (87, 10), 'post_memory_intensity_guilt_1': (88, 10), 'post_memory_intensity_guilt 2': (92, 10), 'post_memory_intensity_indignation_1': (97, 10), 'post_memory_indignation_2': (101, 10), 'rosenberg_vis_2': (104, 10)},
        'P100_vis_3_locations': {'panas_pos_vis_3': (36, 10), 'panas_neg_vis_3': (37, 10), 'qids_vis_3': (47, 10), 'gad_vis_3': (48, 10), 'rosenberg_vis_3': (49, 10), 'madrs_vis_3': (60, 10)},
        'P100_vis_4_locations': {'qids_vis_4': (26, 10), 'rosenberg_vis_4': (27, 10)},
        'P100_vis_5_locations': {'qids_vis_5': (28, 10), 'rosenberg_vis_5': (29, 10)},

        'P107_vis_1_locations': {'dob': (77, 14), 'gender': (81, 14),  'handedness': (82, 14), 'exercise': (83, 14), 'education': (84, 14), 'work_status': (85, 14), 'panic': (132, 14), 'agoraphobia': (134, 14), 'social_anx': (135, 14), 'ocd': (137, 14), 'ptsd': (140, 14), 'gad': (141, 14), 'comorbid_anx': (142, 14), 'msm': (120, 14), 'psi_sociotropy': (151, 14), 'psi_autonomy': (152, 14), 'raads': (155, 14), 'panas_pos_vis_1': (161, 14), 'panas_neg_vis_1': (162, 14), 'qids_vis_1': (172, 14), 'gad_vis_1': (173, 14), 'rosenberg_vis_1': (174, 14), 'madrs_vis_1': (185, 14)},
        'P107_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 11), 'pre_memory_intensity_guilt_2': (43, 11), 'pre_memory_intensity_indignation_1': (49, 11), 'pre_memory_intensity_indignation_2': (54, 11), 'intervention': (78, 11), 'techniques_guilt': (84, 11), 'techniques_indignation': (85, 11), 'perceived_success_guilt': (86, 11), 'perceived_success_indignation': (87, 11), 'post_memory_intensity_guilt_1': (88, 11), 'post_memory_intensity_guilt 2': (92, 11), 'post_memory_intensity_indignation_1': (97, 11), 'post_memory_indignation_2': (101, 11), 'rosenberg_vis_2': (104, 11)},
        'P107_vis_3_locations': {'panas_pos_vis_3': (36, 11), 'panas_neg_vis_3': (37, 11), 'qids_vis_3': (47, 11), 'gad_vis_3': (48, 11), 'rosenberg_vis_3': (49, 11), 'madrs_vis_3': (60, 11)},
        'P107_vis_4_locations': {'qids_vis_4': (26, 11), 'rosenberg_vis_4': (27, 11)},
        'P107_vis_5_locations': {'qids_vis_5': (28, 11), 'rosenberg_vis_5': (29, 11)},

        'P122_vis_1_locations': {'dob': (77, 17), 'gender': (81, 17),  'handedness': (82, 17), 'exercise': (83, 17), 'education': (84, 17), 'work_status': (85, 17), 'panic': (132, 17), 'agoraphobia': (134, 17), 'social_anx': (135, 17), 'ocd': (137, 17), 'ptsd': (140, 17), 'gad': (141, 17), 'comorbid_anx': (142, 17), 'msm': (120, 17), 'psi_sociotropy': (151, 17), 'psi_autonomy': (152, 17), 'raads': (155, 17), 'panas_pos_vis_1': (161, 17), 'panas_neg_vis_1': (162, 17), 'qids_vis_1': (172, 17), 'gad_vis_1': (173, 17), 'rosenberg_vis_1': (174, 17), 'madrs_vis_1': (185, 17)},
        'P122_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 14), 'pre_memory_intensity_guilt_2': (43, 14), 'pre_memory_intensity_indignation_1': (49, 14), 'pre_memory_intensity_indignation_2': (54, 14), 'intervention': (78, 14), 'techniques_guilt': (84, 14), 'techniques_indignation': (85, 14), 'perceived_success_guilt': (86, 14), 'perceived_success_indignation': (87, 14), 'post_memory_intensity_guilt_1': (88, 14), 'post_memory_intensity_guilt 2': (92, 14), 'post_memory_intensity_indignation_1': (97, 14), 'post_memory_indignation_2': (101, 14), 'rosenberg_vis_2': (104, 14)},
        'P122_vis_3_locations': {'panas_pos_vis_3': (36, 14), 'panas_neg_vis_3': (37, 14), 'qids_vis_3': (47, 14), 'gad_vis_3': (48, 14), 'rosenberg_vis_3': (49, 14), 'madrs_vis_3': (60, 14)},
        'P122_vis_4_locations': {'qids_vis_4': (26, 14), 'rosenberg_vis_4': (27, 14)},
        'P122_vis_5_locations': {'qids_vis_5': (28, 14), 'rosenberg_vis_5': (29, 14)},

        'P125_vis_1_locations': {'dob': (77, 18), 'gender': (81, 18),  'handedness': (82, 18), 'exercise': (83, 18), 'education': (84, 18), 'work_status': (85, 18), 'panic': (132, 18), 'agoraphobia': (134, 18), 'social_anx': (135, 18), 'ocd': (137, 18), 'ptsd': (140, 18), 'gad': (141, 18), 'comorbid_anx': (142, 18), 'msm': (120, 18), 'psi_sociotropy': (151, 18), 'psi_autonomy': (152, 18), 'raads': (155, 18), 'panas_pos_vis_1': (161, 18), 'panas_neg_vis_1': (162, 18), 'qids_vis_1': (172, 18), 'gad_vis_1': (173, 18), 'rosenberg_vis_1': (174, 18), 'madrs_vis_1': (185, 18)},
        'P125_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 15), 'pre_memory_intensity_guilt_2': (43, 15), 'pre_memory_intensity_indignation_1': (49, 15), 'pre_memory_intensity_indignation_2': (54, 15), 'intervention': (78, 15), 'techniques_guilt': (84, 15), 'techniques_indignation': (85, 15), 'perceived_success_guilt': (86, 15), 'perceived_success_indignation': (87, 15), 'post_memory_intensity_guilt_1': (88, 15), 'post_memory_intensity_guilt 2': (92, 15), 'post_memory_intensity_indignation_1': (97, 15), 'post_memory_indignation_2': (101, 15), 'rosenberg_vis_2': (104, 15)},
        'P125_vis_3_locations': {'panas_pos_vis_3': (36, 15), 'panas_neg_vis_3': (37, 15), 'qids_vis_3': (47, 15), 'gad_vis_3': (48, 15), 'rosenberg_vis_3': (49, 15), 'madrs_vis_3': (60, 15)},
        'P125_vis_4_locations': {'qids_vis_4': (26, 15), 'rosenberg_vis_4': (27, 15)},
        'P125_vis_5_locations': {'qids_vis_5': (28, 15), 'rosenberg_vis_5': (29, 15)},

        'P127_vis_1_locations': {'dob': (77, 16), 'gender': (81, 16),  'handedness': (82, 16), 'exercise': (83, 16), 'education': (84, 16), 'work_status': (85, 16), 'panic': (132, 16), 'agoraphobia': (134, 16), 'social_anx': (135, 16), 'ocd': (137, 16), 'ptsd': (140, 16), 'gad': (141, 16), 'comorbid_anx': (142, 16), 'msm': (120, 16), 'psi_sociotropy': (151, 16), 'psi_autonomy': (152, 16), 'raads': (155, 16), 'panas_pos_vis_1': (161, 16), 'panas_neg_vis_1': (162, 16), 'qids_vis_1': (172, 16), 'gad_vis_1': (173, 16), 'rosenberg_vis_1': (174, 16), 'madrs_vis_1': (185, 16)},
        'P127_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 13), 'pre_memory_intensity_guilt_2': (43, 13), 'pre_memory_intensity_indignation_1': (49, 13), 'pre_memory_intensity_indignation_2': (54, 13), 'intervention': (78, 13), 'techniques_guilt': (84, 13), 'techniques_indignation': (85, 13), 'perceived_success_guilt': (86, 13), 'perceived_success_indignation': (87, 13), 'post_memory_intensity_guilt_1': (88, 13), 'post_memory_intensity_guilt 2': (92, 13), 'post_memory_intensity_indignation_1': (97, 13), 'post_memory_indignation_2': (101, 13), 'rosenberg_vis_2': (104, 13)},
        'P127_vis_3_locations': {'panas_pos_vis_3': (36, 13), 'panas_neg_vis_3': (37, 13), 'qids_vis_3': (47, 13), 'gad_vis_3': (48, 13), 'rosenberg_vis_3': (49, 13), 'madrs_vis_3': (60, 13)},
        'P127_vis_4_locations': {'qids_vis_4': (26, 13), 'rosenberg_vis_4': (27, 13)},
        'P127_vis_5_locations': {'qids_vis_5': (28, 13), 'rosenberg_vis_5': (29, 13)},

        'P128_vis_1_locations': {'dob': (77, 15), 'gender': (81, 15),  'handedness': (82, 15), 'exercise': (83, 15), 'education': (84, 15), 'work_status': (85, 15), 'panic': (132, 15), 'agoraphobia': (134, 15), 'social_anx': (135, 15), 'ocd': (137, 15), 'ptsd': (140, 15), 'gad': (141, 15), 'comorbid_anx': (142, 15), 'msm': (120, 15), 'psi_sociotropy': (151, 15), 'psi_autonomy': (152, 15), 'raads': (155, 15), 'panas_pos_vis_1': (161, 15), 'panas_neg_vis_1': (162, 15), 'qids_vis_1': (172, 15), 'gad_vis_1': (173, 15), 'rosenberg_vis_1': (174, 15), 'madrs_vis_1': (185, 15)},
        'P128_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 12), 'pre_memory_intensity_guilt_2': (43, 12), 'pre_memory_intensity_indignation_1': (49, 12), 'pre_memory_intensity_indignation_2': (54, 12), 'intervention': (78, 12), 'techniques_guilt': (84, 12), 'techniques_indignation': (85, 12), 'perceived_success_guilt': (86, 12), 'perceived_success_indignation': (87, 12), 'post_memory_intensity_guilt_1': (88, 12), 'post_memory_intensity_guilt 2': (92, 12), 'post_memory_intensity_indignation_1': (97, 12), 'post_memory_indignation_2': (101, 12), 'rosenberg_vis_2': (104, 12)},
        'P128_vis_3_locations': {'panas_pos_vis_3': (36, 12), 'panas_neg_vis_3': (37, 12), 'qids_vis_3': (47, 12), 'gad_vis_3': (48, 12), 'rosenberg_vis_3': (49, 12), 'madrs_vis_3': (60, 12)},
        'P128_vis_4_locations': {'qids_vis_4': (26, 12), 'rosenberg_vis_4': (27, 12)},
        'P128_vis_5_locations': {'qids_vis_5': (28, 12), 'rosenberg_vis_5': (29, 12)},

        'P136_vis_1_locations': {'dob': (77, 19), 'gender': (81, 19),  'handedness': (82, 19), 'exercise': (83, 19), 'education': (84, 19), 'work_status': (85, 19), 'panic': (132, 19), 'agoraphobia': (134, 19), 'social_anx': (135, 19), 'ocd': (137, 19), 'ptsd': (140, 19), 'gad': (141, 19), 'comorbid_anx': (142, 19), 'msm': (120, 19), 'psi_sociotropy': (151, 19), 'psi_autonomy': (152, 19), 'raads': (155, 19), 'panas_pos_vis_1': (161, 19), 'panas_neg_vis_1': (162, 19), 'qids_vis_1': (172, 19), 'gad_vis_1': (173, 19), 'rosenberg_vis_1': (174, 19), 'madrs_vis_1': (185, 19)},
        'P136_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 16), 'pre_memory_intensity_guilt_2': (43, 16), 'pre_memory_intensity_indignation_1': (49, 16), 'pre_memory_intensity_indignation_2': (54, 16), 'intervention': (78, 16), 'techniques_guilt': (84, 16), 'techniques_indignation': (85, 16), 'perceived_success_guilt': (86, 16), 'perceived_success_indignation': (87, 16), 'post_memory_intensity_guilt_1': (88, 16), 'post_memory_intensity_guilt 2': (92, 16), 'post_memory_intensity_indignation_1': (97, 16), 'post_memory_indignation_2': (101, 16), 'rosenberg_vis_2': (104, 16)},
        'P136_vis_3_locations': {'panas_pos_vis_3': (36, 16), 'panas_neg_vis_3': (37, 16), 'qids_vis_3': (47, 16), 'gad_vis_3': (48, 16), 'rosenberg_vis_3': (49, 16), 'madrs_vis_3': (60, 16)},
        'P136_vis_4_locations': {'qids_vis_4': (26, 16), 'rosenberg_vis_4': (27, 16)},
        'P136_vis_5_locations': {'qids_vis_5': (28, 16), 'rosenberg_vis_5': (29, 16)},

        'P145_vis_1_locations': {'dob': (77, 21), 'gender': (81, 21),  'handedness': (82, 21), 'exercise': (83, 21), 'education': (84, 21), 'work_status': (85, 21), 'panic': (132, 21), 'agoraphobia': (134, 21), 'social_anx': (135, 21), 'ocd': (137, 21), 'ptsd': (140, 21), 'gad': (141, 21), 'comorbid_anx': (142, 21), 'msm': (120, 21), 'psi_sociotropy': (151, 21), 'psi_autonomy': (152, 21), 'raads': (155, 21), 'panas_pos_vis_1': (161, 21), 'panas_neg_vis_1': (162, 21), 'qids_vis_1': (172, 21), 'gad_vis_1': (173, 21), 'rosenberg_vis_1': (174, 21), 'madrs_vis_1': (185, 21)},
        'P145_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 18), 'pre_memory_intensity_guilt_2': (43, 18), 'pre_memory_intensity_indignation_1': (49, 18), 'pre_memory_intensity_indignation_2': (54, 18), 'intervention': (78, 18), 'techniques_guilt': (84, 18), 'techniques_indignation': (85, 18), 'perceived_success_guilt': (86, 18), 'perceived_success_indignation': (87, 18), 'post_memory_intensity_guilt_1': (88, 18), 'post_memory_intensity_guilt 2': (92, 18), 'post_memory_intensity_indignation_1': (97, 18), 'post_memory_indignation_2': (101, 18), 'rosenberg_vis_2': (104, 18)},
        'P145_vis_3_locations': {'panas_pos_vis_3': (36, 17), 'panas_neg_vis_3': (37, 17), 'qids_vis_3': (47, 17), 'gad_vis_3': (48, 17), 'rosenberg_vis_3': (49, 17), 'madrs_vis_3': (60, 17)},
        'P145_vis_4_locations': {'qids_vis_4': (26, 17), 'rosenberg_vis_4': (27, 17)},
        'P145_vis_5_locations': {'qids_vis_5': (28, 17), 'rosenberg_vis_5': (29, 17)},

        'P155_vis_1_locations': {'dob': (77, 22), 'gender': (81, 22),  'handedness': (82, 22), 'exercise': (83, 22), 'education': (84, 22), 'work_status': (85, 22), 'panic': (132, 22), 'agoraphobia': (134, 22), 'social_anx': (135, 22), 'ocd': (137, 22), 'ptsd': (140, 22), 'gad': (141, 22), 'comorbid_anx': (142, 22), 'msm': (120, 22), 'psi_sociotropy': (151, 22), 'psi_autonomy': (152, 22), 'raads': (155, 22), 'panas_pos_vis_1': (161, 22), 'panas_neg_vis_1': (162, 22), 'qids_vis_1': (172, 22), 'gad_vis_1': (173, 22), 'rosenberg_vis_1': (174, 22), 'madrs_vis_1': (185, 22)},
        'P155_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 19), 'pre_memory_intensity_guilt_2': (43, 19), 'pre_memory_intensity_indignation_1': (49, 19), 'pre_memory_intensity_indignation_2': (54, 19), 'intervention': (78, 19), 'techniques_guilt': (84, 19), 'techniques_indignation': (85, 19), 'perceived_success_guilt': (86, 19), 'perceived_success_indignation': (87, 19), 'post_memory_intensity_guilt_1': (88, 19), 'post_memory_intensity_guilt 2': (92, 19), 'post_memory_intensity_indignation_1': (97, 19), 'post_memory_indignation_2': (101, 19), 'rosenberg_vis_2': (104, 19)},
        'P155_vis_3_locations': {'panas_pos_vis_3': (36, 18), 'panas_neg_vis_3': (37, 18), 'qids_vis_3': (47, 18), 'gad_vis_3': (48, 18), 'rosenberg_vis_3': (49, 18), 'madrs_vis_3': (60, 18)},
        'P155_vis_4_locations': {'qids_vis_4': (26, 18), 'rosenberg_vis_4': (27, 18)},
        'P155_vis_5_locations': {'qids_vis_5': (28, 18), 'rosenberg_vis_5': (29, 18)},

        'P199_vis_1_locations': {'dob': (77, 27), 'gender': (81, 27),  'handedness': (82, 27), 'exercise': (83, 27), 'education': (84, 27), 'work_status': (85, 27), 'panic': (132, 27), 'agoraphobia': (134, 27), 'social_anx': (135, 27), 'ocd': (137, 27), 'ptsd': (140, 27), 'gad': (141, 27), 'comorbid_anx': (142, 27), 'msm': (120, 27), 'psi_sociotropy': (151, 27), 'psi_autonomy': (152, 27), 'raads': (155, 27), 'panas_pos_vis_1': (161, 27), 'panas_neg_vis_1': (162, 27), 'qids_vis_1': (172, 27), 'gad_vis_1': (173, 27), 'rosenberg_vis_1': (174, 27), 'madrs_vis_1': (185, 27)},
        'P199_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 22), 'pre_memory_intensity_guilt_2': (43, 22), 'pre_memory_intensity_indignation_1': (49, 22), 'pre_memory_intensity_indignation_2': (54, 22), 'intervention': (78, 22), 'techniques_guilt': (84, 22), 'techniques_indignation': (85, 22), 'perceived_success_guilt': (86, 22), 'perceived_success_indignation': (87, 22), 'post_memory_intensity_guilt_1': (88, 22), 'post_memory_intensity_guilt 2': (92, 22), 'post_memory_intensity_indignation_1': (97, 22), 'post_memory_indignation_2': (101, 22), 'rosenberg_vis_2': (104, 22)},
        'P199_vis_3_locations': {'panas_pos_vis_3': (36, 21), 'panas_neg_vis_3': (37, 21), 'qids_vis_3': (47, 21), 'gad_vis_3': (48, 21), 'rosenberg_vis_3': (49, 21), 'madrs_vis_3': (60, 21)},
        'P199_vis_4_locations': {'qids_vis_4': (26, 21), 'rosenberg_vis_4': (27, 21)},
        'P199_vis_5_locations': {'qids_vis_5': (28, 21), 'rosenberg_vis_5': (29, 21)},

        'P215_vis_1_locations': {'dob': (77, 26), 'gender': (81, 26),  'handedness': (82, 26), 'exercise': (83, 26), 'education': (84, 26), 'work_status': (85, 26), 'panic': (132, 26), 'agoraphobia': (134, 26), 'social_anx': (135, 26), 'ocd': (137, 26), 'ptsd': (140, 26), 'gad': (141, 26), 'comorbid_anx': (142, 26), 'msm': (120, 26), 'psi_sociotropy': (151, 26), 'psi_autonomy': (152, 26), 'raads': (155, 26), 'panas_pos_vis_1': (161, 26), 'panas_neg_vis_1': (162, 26), 'qids_vis_1': (172, 26), 'gad_vis_1': (173, 26), 'rosenberg_vis_1': (174, 26), 'madrs_vis_1': (185, 26)},
        'P215_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 21), 'pre_memory_intensity_guilt_2': (43, 21), 'pre_memory_intensity_indignation_1': (49, 21), 'pre_memory_intensity_indignation_2': (54, 21), 'intervention': (78, 21), 'techniques_guilt': (84, 21), 'techniques_indignation': (85, 21), 'perceived_success_guilt': (86, 21), 'perceived_success_indignation': (87, 21), 'post_memory_intensity_guilt_1': (88, 21), 'post_memory_intensity_guilt 2': (92, 21), 'post_memory_intensity_indignation_1': (97, 21), 'post_memory_indignation_2': (101, 21), 'rosenberg_vis_2': (104, 21)},
        'P215_vis_3_locations': {'panas_pos_vis_3': (36, 20), 'panas_neg_vis_3': (37, 20), 'qids_vis_3': (47, 20), 'gad_vis_3': (48, 20), 'rosenberg_vis_3': (49, 20), 'madrs_vis_3': (60, 20)},
        'P215_vis_4_locations': {'qids_vis_4': (26, 20), 'rosenberg_vis_4': (27, 20)},
        'P215_vis_5_locations': {'qids_vis_5': (28, 20), 'rosenberg_vis_5': (29, 20)},

        'P216_vis_1_locations': {'dob': (77, 28), 'gender': (81, 28),  'handedness': (82, 28), 'exercise': (83, 28), 'education': (84, 28), 'work_status': (85, 28), 'panic': (132, 28), 'agoraphobia': (134, 28), 'social_anx': (135, 28), 'ocd': (137, 28), 'ptsd': (140, 28), 'gad': (141, 28), 'comorbid_anx': (142, 28), 'msm': (120, 28), 'psi_sociotropy': (151, 28), 'psi_autonomy': (152, 28), 'raads': (155, 28), 'panas_pos_vis_1': (161, 28), 'panas_neg_vis_1': (162, 28), 'qids_vis_1': (172, 28), 'gad_vis_1': (173, 28), 'rosenberg_vis_1': (174, 28), 'madrs_vis_1': (185, 28)},
        'P216_vis_2_locations': {'pre_memory_intensity_guilt_1': (38, 23), 'pre_memory_intensity_guilt_2': (43, 23), 'pre_memory_intensity_indignation_1': (49, 23), 'pre_memory_intensity_indignation_2': (54, 23), 'intervention': (78, 23), 'techniques_guilt': (84, 23), 'techniques_indignation': (85, 23), 'perceived_success_guilt': (86, 23), 'perceived_success_indignation': (87, 23), 'post_memory_intensity_guilt_1': (88, 23), 'post_memory_intensity_guilt 2': (92, 23), 'post_memory_intensity_indignation_1': (97, 23), 'post_memory_indignation_2': (101, 23), 'rosenberg_vis_2': (104, 23)},
        'P216_vis_3_locations': {'panas_pos_vis_3': (36, 22), 'panas_neg_vis_3': (37, 22), 'qids_vis_3': (47, 22), 'gad_vis_3': (48, 22), 'rosenberg_vis_3': (49, 22), 'madrs_vis_3': (60, 22)},
        'P216_vis_4_locations': {'qids_vis_4': (26, 22), 'rosenberg_vis_4': (27, 22)},
        'P216_vis_5_locations': {'qids_vis_5': (28, 22), 'rosenberg_vis_5': (29, 22)}
    }
    for x in participants:
        print(f'Extracting {x} data from eCRF.xlsx...')
        decrypted_workbook = io.BytesIO()
        with open(ecrf_file_path, 'rb') as file:
            office_file = msoffcrypto.OfficeFile(file)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted_workbook)
        workbook = openpyxl.load_workbook(decrypted_workbook, data_only = True)
        ecrf_sheet = workbook['Visit 1']
        vis_1_values = [ecrf_sheet.cell(row=row, column=column).value for (row, column) in location_dict[f'{x}_vis_1_locations'].values()]
        ecrf_sheet = workbook['Visit 2']
        vis_2_values = [ecrf_sheet.cell(row=row, column=column).value for (row, column) in location_dict[f'{x}_vis_2_locations'].values()]
        ecrf_sheet = workbook['Visit 3']
        vis_3_values = [ecrf_sheet.cell(row=row, column=column).value for (row, column) in location_dict[f'{x}_vis_3_locations'].values()]
        ecrf_sheet = workbook['Online 1']
        vis_4_values = [ecrf_sheet.cell(row=row, column=column).value for (row, column) in location_dict[f'{x}_vis_4_locations'].values()]
        ecrf_sheet = workbook['Online 2']
        vis_5_values = [ecrf_sheet.cell(row=row, column=column).value for (row, column) in location_dict[f'{x}_vis_5_locations'].values()]
        df_values_dict[f'{x}'] = vis_1_values + vis_2_values + vis_3_values + vis_4_values + vis_5_values
        for key, values in df_values_dict.items():
            ecrf_df[key] = values
        ecrf_data_path = '/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/clinical_analysis/ecrf_data.xlsx'
        ecrf_df.to_excel(ecrf_data_path, index=True)
        workbook.close()
    warnings.resetwarnings()
    print("eCRF data extracted.")

    # Step 3: Run LMMs of Clinical Assessment Scores.
    print("\n###### STEP 3: RUN LMMs OF CLINICAL ASSESSMENT SCORES ######")
    print("Note: The LMMs themselves cannot be performed on server due to complex R interfacing requirements. Please run code instead on local Spyder software.")
    columns = ['p_id', 'visit', 'intervention', 'rosenberg', 'qids', 'madrs', 'gad', 'panas_pos', 'panas_neg']
    rqmgp_df = pd.DataFrame(columns=columns)
    p_id_column = participants * 5
    visit_column = ['1'] * 20 + ['2'] * 20 + ['3'] * 20 + ['4'] * 20 + ['5'] * 20
    intervention_values = ecrf_df.loc['intervention'].tolist()
    intervention_column = list(intervention_values) * 5
    rosenberg_vis_1 = ecrf_df.loc['rosenberg_vis_1'].tolist()
    rosenberg_vis_2 = ecrf_df.loc['rosenberg_vis_2'].tolist()
    rosenberg_vis_3 = ecrf_df.loc['rosenberg_vis_3'].tolist()
    rosenberg_vis_4 = ecrf_df.loc['rosenberg_vis_4'].tolist()
    rosenberg_vis_5 = ecrf_df.loc['rosenberg_vis_5'].tolist()
    rosenberg_column = rosenberg_vis_1 + rosenberg_vis_2 + rosenberg_vis_3 + rosenberg_vis_4 + rosenberg_vis_5
    qids_vis_1 = ecrf_df.loc['qids_vis_1'].tolist()
    qids_vis_2 = [np.nan] * 20
    qids_vis_3 = ecrf_df.loc['qids_vis_3'].tolist()
    qids_vis_4 = ecrf_df.loc['qids_vis_4'].tolist()
    qids_vis_5 = ecrf_df.loc['qids_vis_5'].tolist()
    qids_column = qids_vis_1 + qids_vis_2 + qids_vis_3 + qids_vis_4 + qids_vis_5
    madrs_vis_1 = ecrf_df.loc['madrs_vis_1'].tolist()
    madrs_vis_2 = [np.nan] * 20
    madrs_vis_3 = ecrf_df.loc['madrs_vis_3'].tolist()
    madrs_vis_4 = [np.nan] * 20
    madrs_vis_5 = [np.nan] * 20
    madrs_column = madrs_vis_1 + madrs_vis_2 + madrs_vis_3 + madrs_vis_4 + madrs_vis_5
    gad_vis_1 = ecrf_df.loc['gad_vis_1'].tolist()
    gad_vis_2 = [np.nan] * 20
    gad_vis_3 = ecrf_df.loc['gad_vis_3'].tolist()
    gad_vis_4 = [np.nan] * 20
    gad_vis_5 = [np.nan] * 20
    gad_column = gad_vis_1 + gad_vis_2 + gad_vis_3 + gad_vis_4 + gad_vis_5
    panas_pos_vis_1 = ecrf_df.loc['panas_pos_vis_1'].tolist()
    panas_pos_vis_2 = [np.nan] * 20
    panas_pos_vis_3 = ecrf_df.loc['panas_pos_vis_3'].tolist()
    panas_pos_vis_4 = [np.nan] * 20
    panas_pos_vis_5 = [np.nan] * 20
    panas_pos_column = panas_pos_vis_1 + panas_pos_vis_2 + panas_pos_vis_3 + panas_pos_vis_4 + panas_pos_vis_5
    panas_neg_vis_1 = ecrf_df.loc['panas_neg_vis_1'].tolist()
    panas_neg_vis_2 = [np.nan] * 20
    panas_neg_vis_3 = ecrf_df.loc['panas_neg_vis_3'].tolist()
    panas_neg_vis_4 = [np.nan] * 20
    panas_neg_vis_5 = [np.nan] * 20
    panas_neg_column = panas_neg_vis_1 + panas_neg_vis_2 + panas_neg_vis_3 + panas_neg_vis_4 + panas_neg_vis_5
    rqmgp_df['p_id'] = p_id_column
    rqmgp_df['visit'] = visit_column
    rqmgp_df['intervention'] = intervention_column
    rqmgp_df['rosenberg'] = rosenberg_column
    rqmgp_df['qids'] = qids_column
    rqmgp_df['madrs'] = madrs_column
    rqmgp_df['gad'] = gad_column
    rqmgp_df['panas_pos'] = panas_pos_column
    rqmgp_df['panas_neg'] = panas_neg_column

    rosenberg_df = rqmgp_df.dropna(subset=['rosenberg'])
    visits = ['1', '2', '3', '4', '5']
    interventions = ['a', 'b']
    columns = ['vals', 'mean', 'std_error', 'shap_p']
    index = ['vis1_inta', 'vis2_inta', 'vis3_inta', 'vis4_inta', 'vis5_inta', 'vis1_intb', 'vis2_intb', 'vis3_intb', 'vis4_intb', 'vis5_intb']
    rosenberg_stats_df = pd.DataFrame(columns=columns, index=index)
    for visit in visits:
        for intervention in interventions:
            vals = pd.to_numeric(rosenberg_df[(rosenberg_df['visit'] == visit) & (rosenberg_df['intervention'] == intervention)]['rosenberg'], errors='coerce').dropna().tolist()
            mean = np.mean(vals)
            std_error = np.std(vals) / np.sqrt(len(vals))
            _, shap_p = stats.shapiro(vals)
            rosenberg_stats_df.loc[f'vis{visit}_int{intervention}', 'vals'] = vals
            rosenberg_stats_df.loc[f'vis{visit}_int{intervention}', 'mean'] = mean
            rosenberg_stats_df.loc[f'vis{visit}_int{intervention}', 'std_error'] = std_error
            rosenberg_stats_df.loc[f'vis{visit}_int{intervention}', 'shap_p'] = shap_p

    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(emmeans)
    # rosenberg_df <- as.data.frame(rosenberg_df)
    # model <- lmer(rosenberg~visit*intervention + (1|p_id), data = rosenberg_df)
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("Rosenberg LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # emmeans_results <- emmeans(model, pairwise ~ visit)
    # print(summary(emmeans_results))
    # } else {
    # print("Rosenberg LMM residuals do not meet normality assumptions.")
    # }
    # """
    # ro.globalenv['rosenberg_df'] = pandas2ri.py2rpy(rosenberg_df)
    # result = r(r_script)
    # print(result)

    data = {
        'visit': [1, 2, 3, 4, 5] * 2,
        'intervention': ['A'] * 5 + ['B'] * 5,
        'mean': [
            rosenberg_stats_df.loc['vis1_inta', 'mean'], rosenberg_stats_df.loc['vis2_inta', 'mean'], rosenberg_stats_df.loc['vis3_inta', 'mean'],
            rosenberg_stats_df.loc['vis4_inta', 'mean'], rosenberg_stats_df.loc['vis5_inta', 'mean'], 
            rosenberg_stats_df.loc['vis1_intb', 'mean'], rosenberg_stats_df.loc['vis2_intb', 'mean'], rosenberg_stats_df.loc['vis3_intb', 'mean'], 
            rosenberg_stats_df.loc['vis4_intb', 'mean'], rosenberg_stats_df.loc['vis5_intb', 'mean'], 
        ],
        'std_error': [
            rosenberg_stats_df.loc['vis1_inta', 'std_error'], rosenberg_stats_df.loc['vis2_inta', 'std_error'], rosenberg_stats_df.loc['vis3_inta', 'std_error'], 
            rosenberg_stats_df.loc['vis4_inta', 'std_error'], rosenberg_stats_df.loc['vis5_inta', 'std_error'], 
            rosenberg_stats_df.loc['vis1_intb', 'std_error'], rosenberg_stats_df.loc['vis2_intb', 'std_error'], rosenberg_stats_df.loc['vis3_intb', 'std_error'], 
            rosenberg_stats_df.loc['vis4_intb', 'std_error'], rosenberg_stats_df.loc['vis5_intb', 'std_error'], 
        ]
    }
    plot_data = pd.DataFrame(data)
    rosenberg_plot = (ggplot(plot_data, aes(x='visit', y='mean', color='intervention'))
            + geom_line(size=2)
            + geom_point(size=4)
            + geom_errorbar(aes(ymin='mean - std_error',
                            ymax='mean + std_error'), width=0.2)
            + labs(title='Rosenberg Scores Across Study Visits',
                x='Visit',
                y='Rosenberg Score',
                color='Intervention')
            + theme_classic()
            + scale_y_continuous(expand=(0, 0), limits=[18,28])
            )
    rosenberg_plot = rosenberg_plot + annotate("text", x=2, y=max(plot_data['mean']) + 2.5, label="*", size=16, color="black") + \
        annotate("segment", x=1, xend=3, y=max(plot_data['mean']) +2.25, yend=max(plot_data['mean']) + 2.25, color="black") + \
            annotate("text", x=2.5, y=max(plot_data['mean']) + 3.5, label="*", size=16, color="black") + \
                annotate("segment", x=2, xend=3, y=max(plot_data['mean']) +3.25, yend=max(plot_data['mean']) + 3.25, color="black")
    rosenberg_plot.save('analysis/clinical_analysis/figs/rosenberg_plot.png')

    qids_df = rqmgp_df.dropna(subset=['qids'])
    visits = ['1', '3', '4', '5']
    interventions = ['a', 'b']
    columns = ['vals', 'mean', 'std_error', 'shap_p']
    index = ['vis1_inta', 'vis3_inta', 'vis4_inta', 'vis5_inta', 'vis1_intb', 'vis3_intb', 'vis4_intb', 'vis5_intb']
    qids_stats_df = pd.DataFrame(columns=columns, index=index)
    for visit in visits:
        for intervention in interventions:
            vals = pd.to_numeric(qids_df[(qids_df['visit'] == visit) & (qids_df['intervention'] == intervention)]['qids'], errors='coerce').dropna().tolist()
            mean = np.mean(vals)
            std_error = np.std(vals) / np.sqrt(len(vals))
            _, shap_p = stats.shapiro(vals)
            qids_stats_df.loc[f'vis{visit}_int{intervention}', 'vals'] = vals
            qids_stats_df.loc[f'vis{visit}_int{intervention}', 'mean'] = mean
            qids_stats_df.loc[f'vis{visit}_int{intervention}', 'std_error'] = std_error
            qids_stats_df.loc[f'vis{visit}_int{intervention}', 'shap_p'] = shap_p
    
    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(emmeans)
    # qids_df <- as.data.frame(qids_df)
    # model <- lmer(qids~visit*intervention + (1|p_id), data = qids_df)
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("QIDS LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # emmeans_results <- emmeans(model, pairwise ~ visit)
    # print(summary(emmeans_results))
    # } else {
    # print("QIDS LMM residuals do not meet normality assumptions.")
    # }
    # """
    # ro.globalenv['qids_df'] = pandas2ri.py2rpy(qids_df)
    # result = r(r_script)
    # print(result)

    data = {
        'visit': [1, 2, 3, 4, 5] * 2,
        'intervention': ['A'] * 5 + ['B'] * 5,
        'mean': [
            qids_stats_df.loc['vis1_inta', 'mean'], np.nan, qids_stats_df.loc['vis3_inta', 'mean'], 
            qids_stats_df.loc['vis4_inta', 'mean'], qids_stats_df.loc['vis5_inta', 'mean'], 
            qids_stats_df.loc['vis1_intb', 'mean'], np.nan, qids_stats_df.loc['vis3_intb', 'mean'], 
            qids_stats_df.loc['vis4_intb', 'mean'], qids_stats_df.loc['vis5_intb', 'mean'], 
        ],
        'std_error': [
            qids_stats_df.loc['vis1_inta', 'std_error'], np.nan, qids_stats_df.loc['vis3_inta', 'std_error'], 
            qids_stats_df.loc['vis4_inta', 'std_error'], qids_stats_df.loc['vis5_inta', 'std_error'], 
            qids_stats_df.loc['vis1_intb', 'std_error'], np.nan, qids_stats_df.loc['vis3_intb', 'std_error'], 
            qids_stats_df.loc['vis4_intb', 'std_error'], qids_stats_df.loc['vis5_intb', 'std_error'], 
        ]
    }
    plot_data = pd.DataFrame(data)
    plot_data_line = plot_data[plot_data['visit'] != 2]
    qids_plot = (ggplot(plot_data, aes(x='visit', y='mean', color='intervention'))
            + geom_line(data=plot_data_line, size=2)
            + geom_point(size=4)
            + geom_errorbar(aes(ymin='mean - std_error',
                            ymax='mean + std_error'), width=0.2)
            + labs(title='QIDS Scores Across Study Visits',
                x='Visit',
                y='QIDS Score',
                color='Intervention')
            + theme_classic()
            + scale_y_continuous(expand=(0, 0), limits=[8,20], breaks=[8,10,12,14,16,18,20])
            )
    qids_plot = qids_plot + annotate("text", x=2, y=max(plot_data['mean']) + 2.5, label="**", size=16, color="black") + \
        annotate("segment", x=1, xend=3, y=max(plot_data['mean']) +2.25, yend=max(plot_data['mean']) + 2.25, color="black") + \
            annotate("text", x=2.5, y=max(plot_data['mean']) + 3.5, label="***", size=16, color="black") + \
                annotate("segment", x=1, xend=4, y=max(plot_data['mean']) +3.25, yend=max(plot_data['mean']) + 3.25, color="black") + \
                    annotate("text", x=3, y=max(plot_data['mean']) + 4.5, label="***", size=16, color="black") + \
                        annotate("segment", x=1, xend=5, y=max(plot_data['mean']) +4.25, yend=max(plot_data['mean']) + 4.25, color="black")
    qids_plot.save('analysis/clinical_analysis/figs/qids_plot.png')

    madrs_df = rqmgp_df.dropna(subset=['madrs'])
    visits = ['1', '3']
    interventions = ['a', 'b']
    columns = ['vals', 'mean', 'std_error', 'shap_p']
    index = ['vis1_inta', 'vis3_inta', 'vis1_intb', 'vis3_intb']
    madrs_stats_df = pd.DataFrame(columns=columns, index=index)
    for visit in visits:
        for intervention in interventions:
            vals = pd.to_numeric(madrs_df[(madrs_df['visit'] == visit) & (madrs_df['intervention'] == intervention)]['madrs'], errors='coerce').dropna().tolist()
            mean = np.mean(vals)
            std_error = np.std(vals) / np.sqrt(len(vals))
            _, shap_p = stats.shapiro(vals)
            madrs_stats_df.loc[f'vis{visit}_int{intervention}', 'vals'] = vals
            madrs_stats_df.loc[f'vis{visit}_int{intervention}', 'mean'] = mean
            madrs_stats_df.loc[f'vis{visit}_int{intervention}', 'std_error'] = std_error
            madrs_stats_df.loc[f'vis{visit}_int{intervention}', 'shap_p'] = shap_p

    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(emmeans)
    # madrs_df <- as.data.frame(madrs_df)
    # model <- lmer(madrs~visit*intervention + (1|p_id), data = madrs_df)
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("MADRS LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # } else {
    # print("MADRS LMM residuals do not meet normality assumptions.")
    # }
    # """
    # ro.globalenv['madrs_df'] = pandas2ri.py2rpy(madrs_df)
    # result = r(r_script)
    # print(result)

    data = {
        'visit': [1, 3] * 2,
        'intervention': ['A'] * 2 + ['B'] * 2,
        'mean': [
            madrs_stats_df.loc['vis1_inta', 'mean'], madrs_stats_df.loc['vis3_inta', 'mean'],
            madrs_stats_df.loc['vis1_intb', 'mean'], madrs_stats_df.loc['vis3_intb', 'mean']
        ],
        'std_error': [
            madrs_stats_df.loc['vis1_inta', 'std_error'], madrs_stats_df.loc['vis3_inta', 'std_error'], 
            madrs_stats_df.loc['vis1_intb', 'std_error'], madrs_stats_df.loc['vis3_intb', 'std_error']
        ]
    }
    plot_data = pd.DataFrame(data)
    plot_data['visit'] = pd.Categorical(
        plot_data['visit'], categories=[1, 3], ordered=True)
    madrs_plot = (ggplot(plot_data, aes(x='visit', y='mean', color='intervention', group='intervention'))
            + geom_line(size=2)
            + geom_point(size=4)
            + geom_errorbar(aes(ymin='mean - std_error',
                            ymax='mean + std_error'), width=0.2)
            + scale_x_discrete(name='Visit')
            + labs(title='MADRS Scores Across Study Visits',
                x='Visit',
                y='MADRS Score',
                color='Intervention')
            + theme_classic()
            + scale_y_continuous(expand=(0, 0), limits=[16,32])
            )
    madrs_plot = madrs_plot + annotate("text", x=1.5, y=max(plot_data['mean']) + 3.25, label="***", size=16, color="black") + \
        annotate("segment", x=1, xend=2, y=max(plot_data['mean']) +3, yend=max(plot_data['mean']) + 3, color="black")
    madrs_plot.save('analysis/clinical_analysis/figs/madrs_plot.png')

    gad_df = rqmgp_df.dropna(subset=['gad'])
    visits = ['1', '3']
    interventions = ['a', 'b']
    columns = ['vals', 'mean', 'std_error', 'shap_p']
    index = ['vis1_inta', 'vis3_inta', 'vis1_intb', 'vis3_intb']
    gad_stats_df = pd.DataFrame(columns=columns, index=index)
    for visit in visits:
        for intervention in interventions:
            vals = pd.to_numeric(gad_df[(gad_df['visit'] == visit) & (gad_df['intervention'] == intervention)]['gad'], errors='coerce').dropna().tolist()
            mean = np.mean(vals)
            std_error = np.std(vals) / np.sqrt(len(vals))
            _, shap_p = stats.shapiro(vals)
            gad_stats_df.loc[f'vis{visit}_int{intervention}', 'vals'] = vals
            gad_stats_df.loc[f'vis{visit}_int{intervention}', 'mean'] = mean
            gad_stats_df.loc[f'vis{visit}_int{intervention}', 'std_error'] = std_error
            gad_stats_df.loc[f'vis{visit}_int{intervention}', 'shap_p'] = shap_p

    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(emmeans)
    # gad_df <- as.data.frame(gad_df)
    # model <- lmer(gad~visit*intervention + (1|p_id), data = gad_df)
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("GAD LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # } else {
    # print("GAD LMM residuals do not meet normality assumptions.")
    # }
    # """
    # ro.globalenv['gad_df'] = pandas2ri.py2rpy(gad_df)
    # result = r(r_script)
    # print(result)

    data = {
        'visit': [1, 3] * 2,
        'intervention': ['A'] * 2 + ['B'] * 2,
        'mean': [
            gad_stats_df.loc['vis1_inta', 'mean'], gad_stats_df.loc['vis3_inta', 'mean'],
            gad_stats_df.loc['vis1_intb', 'mean'], gad_stats_df.loc['vis3_intb', 'mean']
        ],
        'std_error': [
            gad_stats_df.loc['vis1_inta', 'std_error'], gad_stats_df.loc['vis3_inta', 'std_error'], 
            gad_stats_df.loc['vis1_intb', 'std_error'], gad_stats_df.loc['vis3_intb', 'std_error']
        ]
    }
    plot_data = pd.DataFrame(data)
    plot_data['visit'] = pd.Categorical(
        plot_data['visit'], categories=[1, 3], ordered=True)
    gad_plot = (ggplot(plot_data, aes(x='visit', y='mean', color='intervention', group='intervention'))
            + geom_line(size=2)
            + geom_point(size=4)
            + geom_errorbar(aes(ymin='mean - std_error',
                            ymax='mean + std_error'), width=0.2)
            + scale_x_discrete(name='Visit')
            + labs(title='GAD Scores Across Study Visits',
                x='Visit',
                y='GAD Score',
                color='Intervention')
            + theme_classic()
            + scale_y_continuous(expand=(0, 0), limits=[4,12])
            )
    gad_plot.save('analysis/clinical_analysis/figs/gad_plot.png')

    panas_pos_df = rqmgp_df.dropna(subset=['panas_pos'])
    visits = ['1', '3']
    interventions = ['a', 'b']
    columns = ['vals', 'mean', 'std_error', 'shap_p']
    index = ['vis1_inta', 'vis3_inta', 'vis1_intb', 'vis3_intb']
    panas_pos_stats_df = pd.DataFrame(columns=columns, index=index)
    for visit in visits:
        for intervention in interventions:
            vals = pd.to_numeric(panas_pos_df[(panas_pos_df['visit'] == visit) & (panas_pos_df['intervention'] == intervention)]['panas_pos'], errors='coerce').dropna().tolist()
            mean = np.mean(vals)
            std_error = np.std(vals) / np.sqrt(len(vals))
            _, shap_p = stats.shapiro(vals)
            panas_pos_stats_df.loc[f'vis{visit}_int{intervention}', 'vals'] = vals
            panas_pos_stats_df.loc[f'vis{visit}_int{intervention}', 'mean'] = mean
            panas_pos_stats_df.loc[f'vis{visit}_int{intervention}', 'std_error'] = std_error
            panas_pos_stats_df.loc[f'vis{visit}_int{intervention}', 'shap_p'] = shap_p

    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(emmeans)
    # panas_pos_df <- as.data.frame(panas_pos_df)
    # model <- lmer(panas_pos~visit*intervention + (1|p_id), data = panas_pos_df)
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("PANAS Positive LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # } else {
    # print("PANAS Positive LMM residuals do not meet normality assumptions.")
    # }
    # """
    # ro.globalenv['panas_pos_df'] = pandas2ri.py2rpy(panas_pos_df)
    # result = r(r_script)
    # print(result)

    data = {
        'visit': [1, 3] * 2,
        'intervention': ['A'] * 2 + ['B'] * 2,
        'mean': [
            panas_pos_stats_df.loc['vis1_inta', 'mean'], panas_pos_stats_df.loc['vis3_inta', 'mean'],
            panas_pos_stats_df.loc['vis1_intb', 'mean'], panas_pos_stats_df.loc['vis3_intb', 'mean']
        ],
        'std_error': [
            panas_pos_stats_df.loc['vis1_inta', 'std_error'], panas_pos_stats_df.loc['vis3_inta', 'std_error'], 
            panas_pos_stats_df.loc['vis1_intb', 'std_error'], panas_pos_stats_df.loc['vis3_intb', 'std_error']
        ]
    }
    plot_data = pd.DataFrame(data)
    plot_data['visit'] = pd.Categorical(
        plot_data['visit'], categories=[1, 3], ordered=True)
    panas_pos_plot = (ggplot(plot_data, aes(x='visit', y='mean', color='intervention', group='intervention'))
            + geom_line(size=2)
            + geom_point(size=4)
            + geom_errorbar(aes(ymin='mean - std_error',
                            ymax='mean + std_error'), width=0.2)
            + scale_x_discrete(name='Visit')
            + labs(title='PANAS Positive Scores Across Study Visits',
                x='Visit',
                y='PANAS Positive Score',
                color='Intervention')
            + theme_classic()
            + scale_y_continuous(expand=(0, 0), limits=[18,28])
            )
    panas_pos_plot = panas_pos_plot + annotate("text", x=1.5, y=max(plot_data['mean']) + 3.25, label="*", size=16, color="black") + \
        annotate("segment", x=1, xend=2, y=max(plot_data['mean']) +3, yend=max(plot_data['mean']) + 3, color="black")
    panas_pos_plot.save('analysis/clinical_analysis/figs/panas_pos_plot.png')

    panas_neg_df = rqmgp_df.dropna(subset=['panas_neg'])
    visits = ['1', '3']
    interventions = ['a', 'b']
    columns = ['vals', 'mean', 'std_error', 'shap_p']
    index = ['vis1_inta', 'vis3_inta', 'vis1_intb', 'vis3_intb']
    panas_neg_stats_df = pd.DataFrame(columns=columns, index=index)
    for visit in visits:
        for intervention in interventions:
            vals = pd.to_numeric(panas_neg_df[(panas_neg_df['visit'] == visit) & (panas_neg_df['intervention'] == intervention)]['panas_neg'], errors='coerce').dropna().tolist()
            mean = np.mean(vals)
            std_error = np.std(vals) / np.sqrt(len(vals))
            _, shap_p = stats.shapiro(vals)
            panas_neg_stats_df.loc[f'vis{visit}_int{intervention}', 'vals'] = vals
            panas_neg_stats_df.loc[f'vis{visit}_int{intervention}', 'mean'] = mean
            panas_neg_stats_df.loc[f'vis{visit}_int{intervention}', 'std_error'] = std_error
            panas_neg_stats_df.loc[f'vis{visit}_int{intervention}', 'shap_p'] = shap_p

    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(emmeans)
    # panas_neg_df <- as.data.frame(panas_neg_df)
    # model <- lmer(panas_neg~visit*intervention + (1|p_id), data = panas_neg_df)
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("PANAS Negative LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # } else {
    # print("PANAS Negative LMM residuals do not meet normality assumptions.")
    # }
    # """
    # ro.globalenv['panas_neg_df'] = pandas2ri.py2rpy(panas_neg_df)
    # result = r(r_script)
    # print(result)

    data = {
        'visit': [1, 3] * 2,
        'intervention': ['A'] * 2 + ['B'] * 2,
        'mean': [
            panas_neg_stats_df.loc['vis1_inta', 'mean'], panas_neg_stats_df.loc['vis3_inta', 'mean'],
            panas_neg_stats_df.loc['vis1_intb', 'mean'], panas_neg_stats_df.loc['vis3_intb', 'mean']
        ],
        'std_error': [
            panas_neg_stats_df.loc['vis1_inta', 'std_error'], panas_neg_stats_df.loc['vis3_inta', 'std_error'], 
            panas_neg_stats_df.loc['vis1_intb', 'std_error'], panas_neg_stats_df.loc['vis3_intb', 'std_error']
        ]
    }
    plot_data = pd.DataFrame(data)
    plot_data['visit'] = pd.Categorical(
        plot_data['visit'], categories=[1, 3], ordered=True)
    panas_neg_plot = (ggplot(plot_data, aes(x='visit', y='mean', color='intervention', group='intervention'))
            + geom_line(size=2)
            + geom_point(size=4)
            + geom_errorbar(aes(ymin='mean - std_error',
                            ymax='mean + std_error'), width=0.2)
            + scale_x_discrete(name='Visit')
            + labs(title='PANAS Negative Scores Across Study Visits',
                x='Visit',
                y='PANAS Negative Score',
                color='Intervention')
            + theme_classic()
            + scale_y_continuous(expand=(0, 0), limits=[20,30])
            )
    panas_neg_plot.save('analysis/clinical_analysis/figs/panas_neg_plot.png')
    print("LMMs run and plots saved.")

    # Step 4: Run LMMs of Memory Intensity Ratings.
    print("\n###### STEP 4: RUN LMMs OF MEMORY INTENSITY RATINGS ######")
    print("Note: The LMMs themselves cannot be performed on server due to complex R interfacing requirements. Please run code instead on local Spyder software.")
    columns = ['p_id', 'time', 'intervention', 'guilt_rating', 'indignation_rating']
    mem_intensity_df = pd.DataFrame(columns=columns)
    p_id_column = participants * 2
    time_column = ['pre'] * 20 + ['post'] * 20
    intervention_values = ecrf_df.loc['intervention'].tolist()
    intervention_column = list(intervention_values) * 2
    pre_memory_intensity_guilt_1 = ecrf_df.loc['pre_memory_intensity_guilt_1'].tolist()
    pre_memory_intensity_guilt_2 = ecrf_df.loc['pre_memory_intensity_guilt_2'].tolist()
    pre_memory_intensity_indignation_1 = ecrf_df.loc['pre_memory_intensity_indignation_1'].tolist()
    pre_memory_intensity_indignation_2 = ecrf_df.loc['pre_memory_intensity_indignation_2'].tolist()
    post_memory_intensity_guilt_1 = ecrf_df.loc['post_memory_intensity_guilt_1'].tolist()
    post_memory_intensity_guilt_2 = ecrf_df.loc['post_memory_intensity_guilt_2'].tolist()
    post_memory_intensity_indignation_1 = ecrf_df.loc['post_memory_intensity_indignation_1'].tolist()
    post_memory_intensity_indignation_2 = ecrf_df.loc['post_memory_intensity_indignation_2'].tolist()
    pre_guilt = [(a + b) / 2 for a, b in zip(pre_memory_intensity_guilt_1, pre_memory_intensity_guilt_2)]
    pre_indignation = [(a + b) / 2 for a, b in zip(pre_memory_intensity_indignation_1, pre_memory_intensity_indignation_2)]
    post_guilt = [(a + b) / 2 for a, b in zip(post_memory_intensity_guilt_1, post_memory_intensity_guilt_2)]
    post_indignation = [(a + b) / 2 for a, b in zip(post_memory_intensity_indignation_1, post_memory_intensity_indignation_2)]
    guilt_rating_column = pre_guilt + post_guilt
    indignation_rating_column = pre_indignation +  post_indignation
    mem_intensity_df['p_id'] = p_id_column
    mem_intensity_df['time'] = time_column
    mem_intensity_df['intervention'] = intervention_column
    mem_intensity_df['guilt_rating'] = guilt_rating_column
    mem_intensity_df['indignation_rating'] = indignation_rating_column

    time = ['pre', 'post']
    interventions = ['a', 'b']
    columns = ['rating', 'mean', 'std_error', 'shap_p']
    index = ['pre_inta', 'post_inta', 'pre_intb', 'post_intb']
    guilt_mem_intensity_stats_df = pd.DataFrame(columns=columns, index=index)
    for item in time:
        for intervention in interventions:
            ratings = mem_intensity_df[(mem_intensity_df['time'] == item) & (mem_intensity_df['intervention'] == intervention)]['guilt_rating'].tolist()
            mean = np.mean(ratings)
            std_error = np.std(ratings) / np.sqrt(len(ratings))
            _, shap_p = stats.shapiro(vals)
            guilt_mem_intensity_stats_df.loc[f'{item}_int{intervention}', 'rating'] = ratings
            guilt_mem_intensity_stats_df.loc[f'{item}_int{intervention}', 'mean'] = mean
            guilt_mem_intensity_stats_df.loc[f'{item}_int{intervention}', 'std_error'] = std_error
            guilt_mem_intensity_stats_df.loc[f'{item}_int{intervention}', 'shap_p'] = shap_p

    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(emmeans)
    # mem_intensity_df <- as.data.frame(mem_intensity_df)
    # model <- lmer(guilt_rating~time*intervention + (1|p_id), data = mem_intensity_df)
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("Guilt Memory Intensity LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # } else {
    # print("Guilt Memory Intensity LMM residuals do not meet normality assumptions.")
    # }
    # """
    # ro.globalenv['mem_intensity_df'] = pandas2ri.py2rpy(mem_intensity_df)
    # result = r(r_script)
    # print(result)

    data = {
        'time': ['Pre-neurofeedback', 'Post-neurofeedback'] * 2,
        'intervention': ['A'] * 2 + ['B'] * 2,
        'mean': [
            guilt_mem_intensity_stats_df.loc['pre_inta', 'mean'], guilt_mem_intensity_stats_df.loc['post_inta', 'mean'],
            guilt_mem_intensity_stats_df.loc['pre_intb', 'mean'], guilt_mem_intensity_stats_df.loc['post_intb', 'mean']
        ],
        'std_error': [
            guilt_mem_intensity_stats_df.loc['pre_inta', 'std_error'], guilt_mem_intensity_stats_df.loc['post_inta', 'std_error'], 
            guilt_mem_intensity_stats_df.loc['pre_intb', 'std_error'], guilt_mem_intensity_stats_df.loc['post_intb', 'std_error']
        ]
    }
    plot_data = pd.DataFrame(data)
    plot_data['time'] = pd.Categorical(
        plot_data['time'], categories=['Pre-neurofeedback', 'Post-neurofeedback'], ordered=True)
    guilt_mem_intensity_plot = (ggplot(plot_data, aes(x='time', y='mean', color='intervention', group='intervention'))
            + geom_line(size=2)
            + geom_point(size=4)
            + geom_errorbar(aes(ymin='mean - std_error',
                            ymax='mean + std_error'), width=0.2)
            + scale_x_discrete(name='Time')
            + labs(title='Guilt Memory Intensity Ratings Before and After Neurofeedback',
                x='Time',
                y='Mean Guilt Memory Intensity Rating',
                color='Intervention')
            + theme_classic()
            + scale_y_continuous(expand=(0, 0), limits=[5,7])
            )
    guilt_mem_intensity_plot.save('analysis/clinical_analysis/figs/guilt_mem_intensity_plot.png')

    time = ['pre', 'post']
    interventions = ['a', 'b']
    columns = ['rating', 'mean', 'std_error', 'shap_p']
    index = ['pre_inta', 'post_inta', 'pre_intb', 'post_intb']
    indignation_mem_intensity_stats_df = pd.DataFrame(columns=columns, index=index)
    for item in time:
        for intervention in interventions:
            ratings = mem_intensity_df[(mem_intensity_df['time'] == item) & (mem_intensity_df['intervention'] == intervention)]['indignation_rating'].tolist()
            mean = np.mean(ratings)
            std_error = np.std(ratings) / np.sqrt(len(ratings))
            _, shap_p = stats.shapiro(vals)
            indignation_mem_intensity_stats_df.loc[f'{item}_int{intervention}', 'rating'] = ratings
            indignation_mem_intensity_stats_df.loc[f'{item}_int{intervention}', 'mean'] = mean
            indignation_mem_intensity_stats_df.loc[f'{item}_int{intervention}', 'std_error'] = std_error
            indignation_mem_intensity_stats_df.loc[f'{item}_int{intervention}', 'shap_p'] = shap_p

    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(emmeans)
    # mem_intensity_df <- as.data.frame(mem_intensity_df)
    # model <- lmer(indignation_rating~time*intervention + (1|p_id), data = mem_intensity_df)
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("Indignation Memory Intensity LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # } else {
    # print("Indignation Memory Intensity LMM residuals do not meet normality assumptions.")
    # }
    # """
    # ro.globalenv['mem_intensity_df'] = pandas2ri.py2rpy(mem_intensity_df)
    # result = r(r_script)
    # print(result)

    data = {
        'time': ['Pre-neurofeedback', 'Post-neurofeedback'] * 2,
        'intervention': ['A'] * 2 + ['B'] * 2,
        'mean': [
            indignation_mem_intensity_stats_df.loc['pre_inta', 'mean'], indignation_mem_intensity_stats_df.loc['post_inta', 'mean'],
            indignation_mem_intensity_stats_df.loc['pre_intb', 'mean'], indignation_mem_intensity_stats_df.loc['post_intb', 'mean']
        ],
        'std_error': [
            indignation_mem_intensity_stats_df.loc['pre_inta', 'std_error'], indignation_mem_intensity_stats_df.loc['post_inta', 'std_error'], 
            indignation_mem_intensity_stats_df.loc['pre_intb', 'std_error'], indignation_mem_intensity_stats_df.loc['post_intb', 'std_error']
        ]
    }
    plot_data = pd.DataFrame(data)
    plot_data['time'] = pd.Categorical(
        plot_data['time'], categories=['Pre-neurofeedback', 'Post-neurofeedback'], ordered=True)
    indignation_mem_intensity_plot = (ggplot(plot_data, aes(x='time', y='mean', color='intervention', group='intervention'))
            + geom_line(size=2)
            + geom_point(size=4)
            + geom_errorbar(aes(ymin='mean - std_error',
                            ymax='mean + std_error'), width=0.2)
            + scale_x_discrete(name='Time')
            + labs(title='Indignation Memory Intensity Ratings Before and After Neurofeedback',
                x='Time',
                y='Mean Indignation Memory Intensity Rating',
                color='Intervention')
            + theme_classic()
            + scale_y_continuous(expand=(0, 0), limits=[5,7])
            )
    indignation_mem_intensity_plot.save('analysis/clinical_analysis/figs/indignation_mem_intensity_plot.png')
    print("LMMs run and plots saved.")

    #%% Step 5: Correlation of Clinical Score Change with Mean Expanded Thermometer Level.
    print("\n###### STEP 5: CORRELATION OF CLINICAL SCORE CHANGE AND EXPANDED THERMOMETER LEVEL ######")
    rqmgp_df['visit'] = pd.to_numeric(rqmgp_df['visit'], errors='coerce')
    columns = ['condition', 'intervention', 'rosenberg_vis_1', 'rosenberg_vis_3', 'rosenberg_diff', 'therm_lvl_exp']
    clin_therm_corr_df = pd.DataFrame(columns=columns)
    condition_column = (['guilt'] * 10 + ['indig'] * 10) * 2
    intervention_column = ['a'] * 20 + ['b'] * 20
    rosenberg_vis_1_inta = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'a')]['rosenberg'].reset_index(drop=True).tolist()
    rosenberg_vis_3_inta = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'a')]['rosenberg'].reset_index(drop=True).tolist()
    rosenberg_vis_1_intb = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'b')]['rosenberg'].reset_index(drop=True).tolist()
    rosenberg_vis_3_intb = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'b')]['rosenberg'].reset_index(drop=True).tolist()
    qids_vis_1_inta = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'a')]['qids'].reset_index(drop=True).tolist()
    qids_vis_3_inta = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'a')]['qids'].reset_index(drop=True).tolist()
    qids_vis_1_intb = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'b')]['qids'].reset_index(drop=True).tolist()
    qids_vis_3_intb = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'b')]['qids'].reset_index(drop=True).tolist()
    madrs_vis_1_inta = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'a')]['madrs'].reset_index(drop=True).tolist()
    madrs_vis_3_inta = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'a')]['madrs'].reset_index(drop=True).tolist()
    madrs_vis_1_intb = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'b')]['madrs'].reset_index(drop=True).tolist()
    madrs_vis_3_intb = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'b')]['madrs'].reset_index(drop=True).tolist()
    gad_vis_1_inta = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'a')]['gad'].reset_index(drop=True).tolist()
    gad_vis_3_inta = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'a')]['gad'].reset_index(drop=True).tolist()
    gad_vis_1_intb = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'b')]['gad'].reset_index(drop=True).tolist()
    gad_vis_3_intb = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'b')]['gad'].reset_index(drop=True).tolist()
    panas_pos_vis_1_inta = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'a')]['panas_pos'].reset_index(drop=True).tolist()
    panas_pos_vis_3_inta = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'a')]['panas_pos'].reset_index(drop=True).tolist()
    panas_pos_vis_1_intb = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'b')]['panas_pos'].reset_index(drop=True).tolist()
    panas_pos_vis_3_intb = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'b')]['panas_pos'].reset_index(drop=True).tolist()
    panas_neg_vis_1_inta = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'a')]['panas_neg'].reset_index(drop=True).tolist()
    panas_neg_vis_3_inta = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'a')]['panas_neg'].reset_index(drop=True).tolist()
    panas_neg_vis_1_intb = rqmgp_df[(rqmgp_df['visit'] == 1) & (rqmgp_df['intervention'] == 'b')]['panas_neg'].reset_index(drop=True).tolist()
    panas_neg_vis_3_intb = rqmgp_df[(rqmgp_df['visit'] == 3) & (rqmgp_df['intervention'] == 'b')]['panas_neg'].reset_index(drop=True).tolist()
    therm_lvl_exp_guilt_a = [-3.791666667, -0.869047619, -2.261904762, 0.69047619, -1.017857143, -3.869047619, 2.970238095, 1.583333333, -1.761904762, -2.791666667]
    therm_lvl_exp_indig_a = [5.648809524, 5.482142857, -5.571428571, -1.047619048, 6.511904762, -3.18452381, -4.553571429, -0.714285714, 2.458333333, 0.880952381]
    therm_lvl_exp_guilt_b = [-6.654761905, 0.017857143, 1.869047619, -0.738095238, -2.130952381, 0.291666667, -0.386904762, -0.928571429, 1.869047619, 0.541666667]
    therm_lvl_exp_indig_b = [0.654761905, 2.55952381, 5.18452381, 0.81547619, 2.982142857, -2.327380952, 3.81547619, 3.029761905, 5.18452381, -1.017857143]
    clin_therm_corr_df['condition'] = condition_column 
    clin_therm_corr_df['intervention'] = intervention_column
    clin_therm_corr_df['rosenberg_vis_1'] = rosenberg_vis_1_inta * 2 + rosenberg_vis_1_intb * 2
    clin_therm_corr_df['rosenberg_vis_3'] = rosenberg_vis_3_inta * 2 + rosenberg_vis_3_intb * 2
    clin_therm_corr_df['rosenberg_diff'] = clin_therm_corr_df['rosenberg_vis_1'] - clin_therm_corr_df['rosenberg_vis_3']
    clin_therm_corr_df['qids_vis_1'] = qids_vis_1_inta * 2 + qids_vis_1_intb * 2
    clin_therm_corr_df['qids_vis_3'] = qids_vis_3_inta * 2 + qids_vis_3_intb * 2
    clin_therm_corr_df['qids_diff'] = clin_therm_corr_df['qids_vis_1'] - clin_therm_corr_df['qids_vis_3']
    clin_therm_corr_df['madrs_vis_1'] = madrs_vis_1_inta * 2 + madrs_vis_1_intb * 2
    clin_therm_corr_df['madrs_vis_3'] = madrs_vis_3_inta * 2 + madrs_vis_3_intb * 2
    clin_therm_corr_df['madrs_diff'] = clin_therm_corr_df['madrs_vis_1'] - clin_therm_corr_df['madrs_vis_3']
    clin_therm_corr_df['gad_vis_1'] = gad_vis_1_inta * 2 + gad_vis_1_intb * 2
    clin_therm_corr_df['gad_vis_3'] = gad_vis_3_inta * 2 + gad_vis_3_intb * 2
    clin_therm_corr_df['gad_diff'] = clin_therm_corr_df['gad_vis_1'] - clin_therm_corr_df['gad_vis_3']
    clin_therm_corr_df['panas_pos_vis_1'] = panas_pos_vis_1_inta * 2 + panas_pos_vis_1_intb * 2
    clin_therm_corr_df['panas_pos_vis_3'] = panas_pos_vis_3_inta * 2 + panas_pos_vis_3_intb * 2
    clin_therm_corr_df['panas_pos_diff'] = clin_therm_corr_df['panas_pos_vis_1'] - clin_therm_corr_df['panas_pos_vis_3']
    clin_therm_corr_df['panas_neg_vis_1'] = panas_neg_vis_1_inta * 2 + panas_neg_vis_1_intb * 2
    clin_therm_corr_df['panas_neg_vis_3'] = panas_neg_vis_3_inta * 2 + panas_neg_vis_3_intb * 2
    clin_therm_corr_df['panas_neg_diff'] = clin_therm_corr_df['panas_neg_vis_1'] - clin_therm_corr_df['panas_neg_vis_3']
    clin_therm_corr_df['therm_lvl_exp'] = therm_lvl_exp_guilt_a + therm_lvl_exp_indig_a + therm_lvl_exp_guilt_b + therm_lvl_exp_indig_b
    combinations = [('guilt', 'a'), ('guilt', 'b'), ('indig', 'a'), ('indig', 'b')]
    for condition, intervention in combinations:
        subset = clin_therm_corr_df[(clin_therm_corr_df['condition'] == condition) & (clin_therm_corr_df['intervention'] == intervention)]
        corr_value = subset['rosenberg_diff'].corr(subset['therm_lvl_exp'])
        print(f"Correlation for {condition} {intervention}: {corr_value:.3f}")
        rosenberg_therm_corr_plot = (ggplot(subset, aes(x='rosenberg_diff', y='therm_lvl_exp'))
            + geom_point()
            + geom_smooth(method='lm', color='blue')
            + theme_classic()
            + labs(title=f'Scatter Plot for {condition}, Intervention {intervention}\nCorrelation: {corr_value:.2f}', x='Rosenberg Difference', y='Thermometer Level Exp'))
        rosenberg_therm_corr_plot.save(f'analysis/clinical_analysis/figs/rosenberg_therm_{condition}_{intervention}_corr_plot.png')
        corr_value = subset['qids_diff'].corr(subset['therm_lvl_exp'])
        print(f"Correlation for {condition} {intervention}: {corr_value:.3f}")
        qids_therm_corr_plot = (ggplot(subset, aes(x='qids_diff', y='therm_lvl_exp'))
            + geom_point()
            + geom_smooth(method='lm', color='blue')
            + theme_classic()
            + labs(title=f'Scatter Plot for {condition}, Intervention {intervention}\nCorrelation: {corr_value:.2f}', x='QIDS Difference', y='Thermometer Level Exp'))
        qids_therm_corr_plot.save(f'analysis/clinical_analysis/figs/qids_therm_{condition}_{intervention}_corr_plot.png')
        corr_value = subset['madrs_diff'].corr(subset['therm_lvl_exp'])
        print(f"Correlation for {condition} {intervention}: {corr_value:.3f}")
        madrs_therm_corr_plot = (ggplot(subset, aes(x='madrs_diff', y='therm_lvl_exp'))
            + geom_point()
            + geom_smooth(method='lm', color='blue')
            + theme_classic()
            + labs(title=f'Scatter Plot for {condition}, Intervention {intervention}\nCorrelation: {corr_value:.2f}', x='MADRS Difference', y='Thermometer Level Exp'))
        madrs_therm_corr_plot.save(f'analysis/clinical_analysis/figs/madrs_therm_{condition}_{intervention}_corr_plot.png')
        corr_value = subset['gad_diff'].corr(subset['therm_lvl_exp'])
        print(f"Correlation for {condition} {intervention}: {corr_value:.3f}")
        gad_therm_corr_plot = (ggplot(subset, aes(x='gad_diff', y='therm_lvl_exp'))
            + geom_point()
            + geom_smooth(method='lm', color='blue')
            + theme_classic()
            + labs(title=f'Scatter Plot for {condition}, Intervention {intervention}\nCorrelation: {corr_value:.2f}', x='GAD Difference', y='Thermometer Level Exp'))
        gad_therm_corr_plot.save(f'analysis/clinical_analysis/figs/gad_therm_{condition}_{intervention}_corr_plot.png')
        corr_value = subset['panas_pos_diff'].corr(subset['therm_lvl_exp'])
        print(f"Correlation for {condition} {intervention}: {corr_value:.3f}")
        panas_pos_therm_corr_plot = (ggplot(subset, aes(x='panas_pos_diff', y='therm_lvl_exp'))
            + geom_point()
            + geom_smooth(method='lm', color='blue')
            + theme_classic()
            + labs(title=f'Scatter Plot for {condition}, Intervention {intervention}\nCorrelation: {corr_value:.2f}', x='PANAS Pos Difference', y='Thermometer Level Exp'))
        panas_pos_therm_corr_plot.save(f'analysis/clinical_analysis/figs/panas_pos_therm_{condition}_{intervention}_corr_plot.png')
        corr_value = subset['panas_neg_diff'].corr(subset['therm_lvl_exp'])
        print(f"Correlation for {condition} {intervention}: {corr_value:.3f}")
        panas_neg_therm_corr_plot = (ggplot(subset, aes(x='panas_neg_diff', y='therm_lvl_exp'))
            + geom_point()
            + geom_smooth(method='lm', color='blue')
            + theme_classic()
            + labs(title=f'Scatter Plot for {condition}, Intervention {intervention}\nCorrelation: {corr_value:.2f}', x='PANAS Neg Difference', y='Thermometer Level Exp'))
        panas_neg_therm_corr_plot.save(f'analysis/clinical_analysis/figs/panas_neg_therm_{condition}_{intervention}_corr_plot.png')
        print("Correlations completed.")
        
    # Step 6: Analysis of baseline factors.
    print("\n###### STEP 6: ANALYSIS OF BASELINE FACTORS ######")
    intervention_column = ecrf_df.loc['intervention'].tolist()
    msm_column = ecrf_df.loc['msm'].tolist()
    psi_sociotropy_column = ecrf_df.loc['psi_sociotropy'].tolist()
    psi_autonomy_column = ecrf_df.loc['psi_autonomy'].tolist()
    raads_column = ecrf_df.loc['raads'].tolist()
    columns = ['participant', 'intervention', 'factor', 'value']
    baseline_factors_df = pd.DataFrame(columns=columns)
    baseline_factors_df['participant'] = participants * 4
    baseline_factors_df['intervention'] = intervention_column * 4
    baseline_factors_df['factor'] = ['msm'] * 20 + ['psi_sociotropy'] * 20 + ['psi_autonomy'] * 20 + ['raads'] * 20
    baseline_factors_df['value'] = msm_column + psi_sociotropy_column + psi_autonomy_column + raads_column
    def perform_tests(df, factor):
        factor_data = df[df['factor'] == factor]
        group_a = factor_data[factor_data['intervention'] == 'a']['value']
        group_b = factor_data[factor_data['intervention'] == 'b']['value']
        shapiro_a = stats.shapiro(group_a)
        shapiro_b = stats.shapiro(group_b)
        if shapiro_a.pvalue > 0.05 and shapiro_b.pvalue > 0.05:
            ttest_result = stats.ttest_ind(group_a, group_b)
            p_value = ttest_result.pvalue
        else:
            mannwhitney_result = stats.mannwhitneyu(group_a, group_b)
            p_value = mannwhitney_result.pvalue
        return p_value
    p_values = {}
    for factor in baseline_factors_df['factor'].unique():
        p_values[factor] = perform_tests(baseline_factors_df, factor)
    baseline_factors_group_df = baseline_factors_df.groupby(['intervention', 'factor']).agg(
        mean_value=('value', 'mean'),
        se_value=('value', lambda x: np.std(x, ddof=1) / np.sqrt(len(x)))
    ).reset_index()
    baseline_factors_group_df = baseline_factors_group_df[['factor', 'intervention', 'mean_value', 'se_value']]
    for factor in baseline_factors_group_df['factor'].unique():
        factor_data = baseline_factors_group_df[baseline_factors_group_df['factor'] == factor]
        factor_plot = (
            ggplot(factor_data, aes(x='intervention', y='mean_value', fill='intervention'))
            + geom_bar(stat='identity', position=position_dodge(width=0.9), width=0.7)
            + geom_errorbar(
                aes(ymin='mean_value - se_value', ymax='mean_value + se_value'),
                width=0.2,
                position=position_dodge(width=0.9))
            + theme_classic()
            + scale_y_continuous(expand=(0, 0))
            + scale_x_discrete(labels={'a': 'A', 'b': 'B'})
            + labs(
                title=f'Mean of {factor} by Intervention',
                x='Intervention',
                y='Mean',
                fill='Intervention'))
        p_value = p_values[factor]
        if p_value < 0.001:
            annotation_text = '***'
        elif p_value < 0.01:
            annotation_text = '**'
        elif p_value < 0.05:
            annotation_text = '*'
        else:
            annotation_text = ''
        if annotation_text:
            max_value = factor_data['mean_value'].max()
            factor_plot = factor_plot + annotate('text', x=1.5, y=max_value + 0.2, label=annotation_text, size=10) + \
                annotate("segment", x=1, xend=2, y=max_value + 0.1, yend=max_value + 0.1, color="black")
        factor_plot.save(f'analysis/clinical_analysis/figs/{factor}_plot.png')
    anxiety_column = ecrf_df.loc['comorbid_anx'].tolist()
    replacement_dict = {'Yes': 'comorbid_anx', 'No': 'anx_depression', 'No Anxiety': 'no_anxiety'}
    anxiety_column = [replacement_dict.get(item, item) for item in anxiety_column]
    columns = ['participant', 'intervention', 'anxiety']
    anxiety_df = pd.DataFrame(columns=columns)
    anxiety_df['participant'] = participants
    anxiety_df['intervention'] = intervention_column
    anxiety_df['anxiety'] = anxiety_column
    anxiety_plot = (
        ggplot(anxiety_df, aes(x='anxiety', fill='intervention')) +
        geom_bar(stat='count', position=position_dodge(width=0.9)) +
        theme_classic() +
        scale_y_continuous(expand=(0, 0)) +
        labs(title='Count of Anxiety Levels by Intervention',
            x='Anxiety Level',
            y='Count') +
        scale_x_discrete(labels={'comorbid_anx': 'Comorbid Anxiety',
                                'anx_depression': 'Anxious Depression',
                                'no_anxiety': 'No Anxiety'}) +
        scale_fill_manual(name='Intervention', labels=['A', 'B'], values=['indianred', 'skyblue']))
    anxiety_plot.save('analysis/clinical_analysis/figs/anxiety_plot.png')
    print("Baseline factors analysis completed.")

    # Step 7: Successful participants analysis.
    print("\n###### STEP 7: SUCCESSFUL PARTICIPANT ANALYSIS ######")
    successful_participants = ['P030', 'P059', 'P078', 'P093', 'P107', 'P127', 'P155']
    baseline_factors_df['successful'] = baseline_factors_df['participant'].isin(successful_participants)
    def perform_tests(df, factor):
        factor_data = df[df['factor'] == factor]
        group_successful = factor_data[factor_data['successful']]['value']
        group_non_successful = factor_data[~factor_data['successful']]['value']
        shapiro_successful = stats.shapiro(group_successful)
        shapiro_non_successful = stats.shapiro(group_non_successful)
        if shapiro_successful.pvalue > 0.05 and shapiro_non_successful.pvalue > 0.05:
            ttest_result = stats.ttest_ind(group_successful, group_non_successful)
            p_value = ttest_result.pvalue
        else:
            mannwhitney_result = stats.mannwhitneyu(group_successful, group_non_successful)
            p_value = mannwhitney_result.pvalue
        return p_value
    baseline_factors_group_df = baseline_factors_df.groupby(['successful', 'factor']).agg(
        mean_value=('value', 'mean'),
        se_value=('value', lambda x: np.std(x, ddof=1) / np.sqrt(len(x)))
    ).reset_index()
    p_values = {}
    for factor in baseline_factors_df['factor'].unique():
        p_values[factor] = perform_tests(baseline_factors_df, factor)
    for factor in baseline_factors_group_df['factor'].unique():
        factor_data = baseline_factors_group_df[baseline_factors_group_df['factor'] == factor]
        p_value = p_values[factor]
        if p_value < 0.001:
            annotation_text = '***'
        elif p_value < 0.01:
            annotation_text = '**'
        elif p_value < 0.05:
            annotation_text = '*'
        else:
            annotation_text = ''
        factor_plot_successful = (
            ggplot(factor_data, aes(x='successful', y='mean_value', fill='successful'))
            + geom_bar(stat='identity', position=position_dodge(width=0.9), width=0.7)
            + geom_errorbar(
                aes(ymin='mean_value - se_value', ymax='mean_value + se_value'),
                width=0.2,
                position=position_dodge(width=0.9))
            + theme_classic()
            + scale_y_continuous(expand=(0, 0))
            + scale_x_discrete(labels={True: 'Successful', False: 'Not Successful'})
            + labs(
                title=f'Mean of {factor} by Success Status',
                x='Success Status',
                y='Mean',
                fill='Success Status'))
        if annotation_text:
            max_value = factor_data['mean_value'].max()
            factor_plot_successful = factor_plot_successful + annotate('text', x=1.5, y=max_value + 0.2, label=annotation_text, size=10) + \
                annotate("segment", x=1, xend=2, y=max_value + 0.1, yend=max_value + 0.1, color="black")
        factor_plot_successful.save(f'analysis/clinical_analysis/figs/{factor}_plot_successful.png')
    anxiety_df['successful'] = anxiety_df['participant'].apply(lambda x: 'Successful' if x in successful_participants else 'Not Successful')
    anxiety_plot_successful = (
        ggplot(anxiety_df, aes(x='anxiety', fill='successful')) +
        geom_bar(stat='count', position=position_dodge(width=0.9)) +
        theme_classic() +
        scale_y_continuous(expand=(0, 0)) +
        labs(title='Count of Anxiety Levels by Participant Success',
            x='Anxiety Level',
            y='Count') +
        scale_x_discrete(labels={'comorbid_anx': 'Comorbid Anxiety',
                                'anx_depression': 'Anxious Depression',
                                'no_anxiety': 'No Anxiety'}) +
        scale_fill_manual(name='Participant Status', values=['indianred', 'skyblue']))
    anxiety_plot_successful.save('analysis/clinical_analysis/figs/anxiety_plot_successful.png')
    print("Successful participant analysis completed.")

#endregion

#region 4) FMRI PREPARATION AND PREPROCESSING.

def fmri_prep_and_preproc():
    participants = ['P004', 'P006', 'P020', 'P030', 'P059', 'P078', 'P093', 'P094', 'P100', 'P107', 'P122', 'P125', 'P127', 'P128', 'P136', 'P145', 'P155', 'P199', 'P215', 'P216']
    runs = ['run01', 'run02', 'run03', 'run04']
    restart = logged_input("\nWould you like to start the fMRI preparation and preprocessing from scratch? This will remove all files from the 'data/bids', 'data/fmriprep_derivatives' and 'data/fully_preproc' folders. (y/n)\n")
    if restart == 'y':
        double_check = logged_input("\nAre you sure? (y/n)\n")
        if double_check == 'y':
            bids_folder = 'data/bids'
            print(f"Deleting data/bids folder...")
            shutil.rmtree(bids_folder)
            fmriprep_derivatives_folder = 'data/fmriprep_derivatives'
            print(f"Deleting data/fmriprep_derivatives folder...")
            shutil.rmtree(fmriprep_derivatives_folder)
            fully_preproc_folder = 'data/fully_preproc'
            print(f"Deleting data/fully_preproc folder...")
            shutil.rmtree(fully_preproc_folder)
        else:
            sys.exit()

    # Step 1: Create Directories.
    print("\n###### STEP 1: CREATE DIRECTORIES ######")
    bids_folder = 'data/bids'
    os.makedirs(bids_folder, exist_ok=True)
    fmriprep_derivatives_folder = 'data/fmriprep_derivatives'
    os.makedirs(fmriprep_derivatives_folder, exist_ok=True)
    fully_preproc_folder = 'data/fully_preproc'
    os.makedirs(fully_preproc_folder, exist_ok=True)
    derivatives_folder = '/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/derivatives'
    os.makedirs(derivatives_folder, exist_ok=True)
    scratch_folder = '/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/scratch'
    os.makedirs(scratch_folder, exist_ok=True)
    logs_folder = '/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/logs'
    os.makedirs(logs_folder, exist_ok=True)
    for p_id in participants:
        p_id_stripped = p_id.replace('P', '')
        fully_preproc_sub_folder = f'data/fully_preproc/sub-{p_id_stripped}'
        os.makedirs(fully_preproc_sub_folder, exist_ok=True)
        fully_preproc_func_folder = f'data/fully_preproc/sub-{p_id_stripped}/func'
        os.makedirs(fully_preproc_func_folder, exist_ok=True)
    print("Directories created.")
        
    # Step 2: Convert DICOMS to BIDS Format.
    print("\n###### STEP 2: CONVERT DICOMS TO BIDS FORMAT ######")
    for p_id in participants:
        path = f'data/raw_data/{p_id}/data/neurofeedback'
        cisc_folder = None
        for folder_name in os.listdir(path):
            if "CISC" in folder_name:
                cisc_folder = folder_name
                break
        if cisc_folder is None:
            print("No 'CISC' folder found in the 'neurofeedback' directory.")
            exit(1)
        p_id_stripped = p_id.replace('P', '')
        if not os.path.exists(f"data/bids/sub-{p_id_stripped}"):
            print(f"Converting DICOMs to BIDS Nifti format for P{p_id_stripped}...")
            subprocess.run(['heudiconv', '-d', f'/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/data/raw_data/P{{subject}}/data/neurofeedback/{cisc_folder}/*.dcm', '-o', '/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/data/bids/', '-f', '/its/home/bsms9pc4/Desktop/cisc2/projects/stone_depnf/Neurofeedback/participant_data/data/bids/code/heuristic.py', '-s', f'{p_id_stripped}', '-c', 'dcm2niix', '-b', '--overwrite'])
        else: 
            print(f"DICOMs already converted to BIDS Nifti format for P{p_id_stripped}. Skipping process.")
    print("BIDS conversion completed.")

    # Step 3: Label Fieldmaps.
    print("\n###### STEP 3: LABEL FIELDMAPS ######")
    good_participants = ['P059', 'P100', 'P107', 'P122', 'P125', 'P127', 'P128', 'P136', 'P145', 'P155', 'P199', 'P215', 'P216']
    for p_id in good_participants:
        print(f"Labelling fieldmap JSON files for {p_id}...")
        p_id_stripped = p_id.replace('P', '')
        func_directory = f"data/bids/sub-{p_id_stripped}/func"
        func_files = []
        for file_name in os.listdir(func_directory):
            if file_name.endswith(".nii.gz"):
                file_path = os.path.join("func", file_name)
                func_files.append(file_path)
        ap_fieldmap_json = f"data/bids/sub-{p_id_stripped}/fmap/sub-{p_id_stripped}_dir-AP_epi.json"
        pa_fieldmap_json = f"data/bids/sub-{p_id_stripped}/fmap/sub-{p_id_stripped}_dir-PA_epi.json"
        fieldmap_json_files = [ap_fieldmap_json, pa_fieldmap_json]
        for fieldmap_json in fieldmap_json_files:
            with open(fieldmap_json, 'r') as file:
                json_data = json.load(file)
            if "IntendedFor" not in json_data:
                items = list(json_data.items())
                intended_for_item = ("IntendedFor", func_files)
                insert_index = next((i for i, (key, _) in enumerate(items) if key > "IntendedFor"), len(items))
                items.insert(insert_index, intended_for_item)
                json_data = dict(items)
                subprocess.run(['chmod', '+w', fieldmap_json], check=True)
                with open(fieldmap_json, 'w') as file:
                    json.dump(json_data, file, indent=2)
            else:
                print(f"{fieldmap_json} already labelled for P{p_id_stripped}. Skipping process.")

    # Step 4: Copy BIDS Niftis and Singularity Image to Cluster.
    print("\n###### STEP 4: COPY BIDS NIFTIS AND SINGULARITY IMAGE TO CLUSTER ######")
    answer = logged_input('Would you like to copy BIDS Niftis and singularity image to the cluster? (y/n)\n')
    if answer == 'y':
        if not os.path.exists('/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/bids'):
            print("Copying BIDS files for all participants to cluster...")
            shutil.copytree('data/bids', '/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/bids')
        if not os.path.exists('/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/fmriprep_24.0.1.simg'):
            print("Copying fmriprep singularity image to cluster...")
            shutil.copy('/research/cisc2/shared/fmriprep_singularity/fmriprep_24.0.1.simg', '/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/fmriprep_24.0.1.simg')
        print("BIDS Niftis and singularity image copied successfully.")
    else:
        print("Skipping process.")

    # Step 5: Run fMRIPrep on Cluster.
    print("\n###### STEP 5: RUN FMRIPREP ON CLUSTER ######")
    fmriprep_cluster_script = r"""
#!/bin/bash
#$ -N bic_fmriprep          # job name #one subject test
#$ -pe openmp 5             # parallel environment #how many CPU cores to use
# # Logging directory o=stdout, e=stderror, -j join them together or not? yes/no (y/n)
# #$ -j y
#$ -o /mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/logs/
#$ -e /mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/logs/
# # Memory assigned soft but makes sure it is available
#$ -l m_mem_free=8G 
# # Memory HARD limit
#$ -l h_vmem=8G
#$ -l 'h=!node001&!node069&!node072&!node076&!node077' # nodes NOT to use 
# # Syntax for task array start-stop:step eg. 1-1000:10 == [1,11,21,31,41...]
#$ -t 1-20  #This sets SGE_TASK_ID! Set it equal to number of subjects #you can put 1 or 1-n
# # Tasks Concurrent (Ie max number of concurrent)
#$ -tc 20 #maximum tasks running simultaneously .
#$ -jc test.long        # Short=2h, test.default= 8h, test.long=7d 21h, verlong.default=30d
module add sge
DATA_DIR=/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/bids
SCRATCH_DIR=/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/scratch
OUT_DIR=/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/derivatives
LICENSE=/research/cisc2/shared/fs_license/license.txt  
cd ${DATA_DIR}
SUBJLIST=$(find sub-* -maxdepth 0  -type d)
len=${#SUBJLIST[@]}
echo Number of subjects  = $len
cd ${HOME}
echo This is the task id $SGE_TASK_ID
i=$(expr $SGE_TASK_ID - 1)
echo this is i $i
arr=($SUBJLIST)
SUBJECT=${arr[i]}
echo $SUBJECT
singularity run --cleanenv \
        -B ${DATA_DIR}:/data \
        -B ${OUT_DIR}/:/out \
        -B ${SCRATCH_DIR}:/wd \
        -B ${LICENSE}:/license \
        /mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/fmriprep_24.0.1.simg \
        --skip_bids_validation \
        --participant-label ${SUBJECT} \
        --omp-nthreads 5 --nthreads 5 --mem_mb 30000 \
        --low-mem \
        --output-spaces MNI152NLin2009cAsym:res-2 \
        --fs-license-file /license \
        --work-dir /wd \
        --cifti-output 91k \
        /data /out/ participant
echo Done
exit
"""
    with open('data/fmriprep_cluster.sh', 'w') as f:
        f.write(fmriprep_cluster_script)
    if not os.path.exists('/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/derivatives/sub-004'):
        answer = logged_input('Would you like to run fMRIPrep? (y/n)\n')
        if answer == 'y':
            subprocess.run(['ssh', '-Y', 'bsms9pc4@apollo2.hpc.susx.ac.uk', 'source /etc/profile; source ~/.bash_profile; qsub /research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/data/fmriprep_cluster.sh'])
            print('Running fMRIPrep on cluster server...')
        else:
            print('Skipping fMRIPrep.')

    # Step 6: Clean up fMRIPrep Files and Move Back to Main Analysis Server.
    print("\n###### STEP 6: CLEAN UP FMRIPREP FILES AND MOVE TO MAIN ANALYSIS SERVER ######")
    if os.path.exists('/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/derivatives/desc-aseg_dseg.tsv'):
        # fmriprepcleanup_sim_folder = '/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/fmriprepcleanup_sim' # for simulation purposes.
        # os.makedirs(fmriprepcleanup_sim_folder, exist_ok=True)
        subprocess.run(['python', 'data/fMRIPrepCleanup.py', '-dir', '/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/derivatives', '-method', 'delete'])
        if not os.path.exists('bids/fmriprep_derivatives'):
            print("Copying fMRIPrep derivatives from cluster server back to analysis server...")
            shutil.copytree('/mnt/lustre/scratch/bsms/bsms9pc4/stone_depnf/fmriprep/derivatives', 'data/fmriprep_derivatives')
        shutil.rmtree(derivatives_folder)
        shutil.rmtree(scratch_folder)
        shutil.rmtree(logs_folder)
    else:
        print("fMRIPrep directory already clean. Skipping process.")

    # Step 7: Assess run integrity with motion outliers.
    runs = ['run-01', 'run-02', 'run-03', 'run-04']
    volumes_per_run = {'run-01': 210, 'run-02': 238, 'run-03': 238, 'run-04': 210}
    for p_id in participants:
        p_id_stripped = p_id.replace('P', '')
        for run in runs:
            confounds_file = f'data/fmriprep_derivatives/sub-{p_id_stripped}/func/sub-{p_id_stripped}_task-nf_{run}_desc-confounds_timeseries.tsv'
            try:
                confounds_df = pd.read_csv(confounds_file, sep='\t')
                motion_outlier_columns = [col for col in confounds_df.columns if 'motion_outlier' in col]
                num_motion_outliers = len(motion_outlier_columns)
                total_volumes = volumes_per_run[run]
                percentage_motion_outliers = (num_motion_outliers / total_volumes) * 100
                if percentage_motion_outliers < 15:
                    print(f"Sub-{p_id_stripped}, {run}: {num_motion_outliers} motion outliers, {percentage_motion_outliers:.2f}% of total volumes. Run passes integrity check.")
                else:
                    print(f"Sub-{p_id_stripped}, {run}: {num_motion_outliers} motion outliers, {percentage_motion_outliers:.2f}% of total volumes. RUN FAILS INTEGRITY CHECK. CONSIDER DISQUALIFYING RUN FROM ANALYSIS.")
            except FileNotFoundError:
                print(f"Confounds file for sub-{p_id_stripped}, {run} not found.")

    # Step 8: Brain Extract and Smooth Images.
    print("\n###### STEP 8: BRAIN EXTRACT AND SMOOTH IMAGES ######") 
    for p_id in participants:
        p_id_stripped = p_id.replace('P', '')
        
        func_directory = f"data/fmriprep_derivatives/sub-{p_id_stripped}/func"
        func_files = []
        for file_name in os.listdir(func_directory):
            if file_name.endswith("bold.nii.gz"):
                file_path = os.path.join(func_directory, file_name)
                func_files.append(file_path)
        func_files = sorted(func_files, key=lambda x: int(x.split('_run-')[1].split('_')[0]))
        mask_files = []
        for file_name in os.listdir(func_directory):
            if file_name.endswith("res-2_desc-brain_mask.nii.gz"):
                file_path = os.path.join(func_directory, file_name)
                mask_files.append(file_path)
        mask_files = sorted(mask_files, key=lambda x: int(x.split('_run-')[1].split('_')[0]))
        output_files = [f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-01_MNI152_func_ss.nii.gz", f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-02_MNI152_func_ss.nii.gz", f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-03_MNI152_func_ss.nii.gz", f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-04_MNI152_func_ss.nii.gz"]
        if not os.path.exists(f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-01_MNI152_func_fully_preproc.nii.gz"):
            for func, mask, outpath in zip(func_files, mask_files, output_files):
                print(f'Performing brain extraction (round 1) on {func} with {mask}, output to {outpath}')
                subprocess.run(['fslmaths', func, '-mul', mask, outpath])

        fully_preproc_func_directory = f"data/fully_preproc/sub-{p_id_stripped}/func"
        func_files = []
        for file_name in os.listdir(fully_preproc_func_directory):
            if file_name.endswith(".nii.gz"):
                file_path = os.path.join(fully_preproc_func_directory, file_name)
                func_files.append(file_path)
        func_files = sorted(func_files, key=lambda x: int(x.split('_run-')[1].split('_')[0]))
        mask_files = []
        for file_name in os.listdir(func_directory):
            if file_name.endswith("res-2_desc-brain_mask.nii.gz"):
                file_path = os.path.join(func_directory, file_name)
                mask_files.append(file_path)
        mask_files = sorted(mask_files, key=lambda x: int(x.split('_run-')[1].split('_')[0]))
        output_files = [f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-01_MNI152_func_ss_smoothed.nii.gz", f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-02_MNI152_func_ss_smoothed.nii.gz", f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-03_MNI152_func_ss_smoothed.nii.gz", f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-04_MNI152_func_ss_smoothed.nii.gz"]
        if not os.path.exists(f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-01_MNI152_func_fully_preproc.nii.gz"):
            for func, mask, outpath in zip(func_files, mask_files, output_files):
                print(f'Performing smoothing on {func} with {mask}, output to {outpath}')
                median = subprocess.check_output(['fslstats', func, '-k', mask, '-p', '50'])
                fwhm = 6
                sigma = str((fwhm / (2 * np.sqrt(2 * np.log(2)))))
                bt =  str((float(median) * 0.75))
                subprocess.run(['susan', func, bt, sigma, '3', '1', '0', outpath])
        
        func_files = []
        for file_name in os.listdir(fully_preproc_func_directory):
            if file_name.endswith("smoothed.nii.gz"):
                file_path = os.path.join(fully_preproc_func_directory, file_name)
                func_files.append(file_path)
        func_files = sorted(func_files, key=lambda x: int(x.split('_run-')[1].split('_')[0]))
        mask_files = []
        for file_name in os.listdir(func_directory):
            if file_name.endswith("res-2_desc-brain_mask.nii.gz"):
                file_path = os.path.join(func_directory, file_name)
                mask_files.append(file_path)
        mask_files = sorted(mask_files, key=lambda x: int(x.split('_run-')[1].split('_')[0]))
        output_files = [f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-01_MNI152_func_fully_preproc.nii.gz", f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-02_MNI152_func_fully_preproc.nii.gz", f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-03_MNI152_func_fully_preproc.nii.gz", f"data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_run-04_MNI152_func_fully_preproc.nii.gz"]
        if not os.path.exists(output_files[0]):
            for func, mask, outpath in zip(func_files, mask_files, output_files):
                print(f'Performing brain extraction (round 2) on {func} with {mask}, output to {outpath}')
                subprocess.run(['fslmaths', func, '-mul', mask, outpath])

        for file_name in os.listdir(fully_preproc_func_directory):
            file_path = os.path.join(fully_preproc_func_directory, file_name)
            if not file_name.endswith("preproc.nii.gz"):
                if os.path.isfile(file_path):
                    os.remove(file_path)
    print("Skull-stripping and smoothing completed. Brain images are fully preprocessed.")     

#endregion

#region 5) FMRI ANALYSIS.

def fmri_analysis():
    print("\n=========================\nF M R I   A N A L Y S I S\n=========================")
    participants = ['P004', 'P006', 'P020', 'P030', 'P059', 'P078', 'P093', 'P094', 'P100', 'P107', 'P122', 'P125', 'P127', 'P128', 'P136', 'P145', 'P155', 'P199', 'P215', 'P216']
    runs = ['run01', 'run02', 'run03', 'run04']
    restart = logged_input("\nWould you like to start the fMRI analysis from scratch? This will remove all files from the 'analysis/fmri_analysis' folder. (y/n)\n")
    if restart == 'y':
        double_check = logged_input("\nAre you sure? (y/n)\n")
        if double_check == 'y':
            fmri_analysis_folder = 'analysis/fmri_analysis'
            print(f"Deleting analysis/fmri_analysis folder...")
            shutil.rmtree(fmri_analysis_folder)
        else:
            sys.exit()
    
    # Step 1: Create Directories.
    print("\n###### STEP 1: CREATE DIRECTORIES ######")
    fmri_analysis_folder = 'analysis/fmri_analysis'
    os.makedirs(fmri_analysis_folder, exist_ok=True)
    analysis_1_folder = 'analysis/fmri_analysis/analysis_1'
    os.makedirs(analysis_1_folder, exist_ok=True)
    plots_folder = 'analysis/fmri_analysis/analysis_1/plots'
    os.makedirs(plots_folder, exist_ok=True)
    analysis_1_first_level_folder = 'analysis/fmri_analysis/analysis_1/first_level'
    os.makedirs(analysis_1_first_level_folder, exist_ok=True)
    analysis_1_first_level_shared_folder = 'analysis/fmri_analysis/analysis_1/first_level/shared'
    os.makedirs(analysis_1_first_level_shared_folder, exist_ok=True)
    analysis_1_onset_files_folder = 'analysis/fmri_analysis/analysis_1/first_level/shared/onset_files'
    os.makedirs(analysis_1_onset_files_folder, exist_ok=True)
    analysis_1_second_level_folder = 'analysis/fmri_analysis/analysis_1/second_level'
    os.makedirs(analysis_1_second_level_folder, exist_ok=True)
    analysis_1_second_level_shared_folder = 'analysis/fmri_analysis/analysis_1/second_level/shared'
    os.makedirs(analysis_1_second_level_shared_folder, exist_ok=True)
    analysis_1_third_level_folder = 'analysis/fmri_analysis/analysis_1/third_level'
    os.makedirs(analysis_1_third_level_folder, exist_ok=True)
    analysis_1_third_level_shared_folder = 'analysis/fmri_analysis/analysis_1/third_level/shared'
    os.makedirs(analysis_1_third_level_shared_folder, exist_ok=True)
    for p_id in participants:
        p_id_stripped = p_id.replace('P', '')
        analysis_1_first_level_participant_folder = f'analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}'
        os.makedirs(analysis_1_first_level_participant_folder, exist_ok=True)
        analysis_1_second_level_participant_folder = f'analysis/fmri_analysis/analysis_1/second_level/sub-{p_id_stripped}'
        os.makedirs(analysis_1_second_level_participant_folder, exist_ok=True)
    print("Directories created.")

    # Step 2: Extract confound regressors [ANALYSIS 1] .
    print("\n###### STEP 2: EXTRACT CONFOUND REGRESSORS [ANALYSIS 1] ######")
    if not os.path.exists('analysis/fmri_analysis/analysis_1/first_level/sub-004/confounds_run01.txt'):
        for p_id in participants:
            p_id_stripped = p_id.replace('P', '')
            confound_dfs = {}
            run01_fmriprep_file = f'data/fmriprep_derivatives/sub-{p_id_stripped}/func/sub-{p_id_stripped}_task-nf_run-01_space-MNI152NLin2009cAsym_res-2_desc-preproc_bold.nii.gz'
            run04_fmriprep_file = f'data/fmriprep_derivatives/sub-{p_id_stripped}/func/sub-{p_id_stripped}_task-nf_run-04_space-MNI152NLin2009cAsym_res-2_desc-preproc_bold.nii.gz'
            nifti_files = [run01_fmriprep_file, run04_fmriprep_file]
            for file in nifti_files:
                try:
                    confound_df = load_confounds_strategy(file, denoise_strategy='compcor', n_compcor=6)[0]
                    confound_df = confound_df.filter(regex='^(?!.*derivative).*$', axis=1)
                    run_number = file.split('_run-')[1].split('_')[0]
                    confound_dfs[f"{p_id_stripped}_run-{run_number}"] = confound_df
                    print(f"Extracted columns for sub-{p_id_stripped}, run-{run_number}: {confound_df.columns.tolist()}") 
                except ValueError as e:
                    print(f"Error processing sub-{p_id_stripped}, file {file}: {e}")
            for key, confound_df in confound_dfs.items():
                p_id_stripped, run = key.split('_run-')
                confounds_file_path = f'analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/confounds_run{run}.txt'
                confound_df.to_csv(confounds_file_path, header=False, index=False, sep='\t')
                print(f"Saved confounds for {key} to {confounds_file_path}")
        print("Confound regressors extracted.")
    else:
        print("Confound regressors already extracted. Skipping process.")

    # Step 3: Create onset timing files [ANALYSIS 1].
    print("\n###### STEP 3: CREATING ONSET TIMING FILES [ANALYSIS 1] ######")
    onsetfile_sub = 'analysis/fmri_analysis/analysis_1/first_level/shared/onset_files/onsetfile_sub.txt'
    with open(onsetfile_sub, 'w') as file:
        data_rows = [
            ['0', '20', '1'],
            ['50', '20', '1'],
            ['100', '20', '1'],
            ['150', '20', '1'],
            ['200', '20', '1'],
            ['250', '20', '1'],
            ['300', '20', '1'],
            ['350', '20', '1'],
            ['400', '20', '1']
        ]
        for row in data_rows:
            formatted_row = "\t".join(row) + "\n"
            file.write(formatted_row)
    onsetfile_guilt = 'analysis/fmri_analysis/analysis_1/first_level/shared/onset_files/onsetfile_guilt.txt'
    with open(onsetfile_guilt, 'w') as file:
        data_rows = [
            ['20', '30', '1'],
            ['120', '30', '1'],
            ['220', '30', '1'],
            ['320', '30', '1']
        ]
        for row in data_rows:
            formatted_row = "\t".join(row) + "\n"
            file.write(formatted_row)
    onsetfile_indig = 'analysis/fmri_analysis/analysis_1/first_level/shared/onset_files/onsetfile_indig.txt'
    with open(onsetfile_indig, 'w') as file:
        data_rows = [
            ['70', '30', '1'],
            ['170', '30', '1'],
            ['270', '30', '1'],
            ['370', '30', '1']
        ]
        for row in data_rows:
            formatted_row = "\t".join(row) + "\n"
            file.write(formatted_row)
    print('Onset timing files created.')

    # Step 4: Trim signal dropout sections of ROIs [ANALYSIS 1].
    print("\n###### STEP 4: TRIM SIGNAL DROPOUT SECTIONS OF ROIS [ANALYSIS 1] ######")     
    runs = ['run-01', 'run-04']
    if not os.path.exists('analysis/fmri_analysis/analysis_1/first_level/sub-004/trimmed_mni_roi_run-01.nii.gz'):
        roi_file = 'data/roi/SCCsphere8_bin_2mm_func.nii.gz'
        for p_id in participants:
            p_id_stripped = p_id.replace('P', '')
            for run in runs:
                mask_file = f'data/fmriprep_derivatives/sub-{p_id_stripped}/func/sub-{p_id_stripped}_task-nf_{run}_space-MNI152NLin2009cAsym_res-2_desc-brain_mask.nii.gz'
                trimmed_roi_file = f'analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/trimmed_mni_roi_{run}.nii.gz'
                try:
                    subprocess.run(['fslmaths', roi_file, '-mul', mask_file, '-bin', trimmed_roi_file])
                    total_voxels_output = subprocess.run(['fslstats', roi_file, '-V'], capture_output=True, text=True)
                    total_voxels = int(total_voxels_output.stdout.split()[0])
                    trimmed_voxels_output = subprocess.run(['fslstats', trimmed_roi_file, '-V'], capture_output=True, text=True)
                    trimmed_voxels = int(trimmed_voxels_output.stdout.split()[0])
                    trimmed_percentage = ((total_voxels - trimmed_voxels) / total_voxels) * 100
                    print(f"Percentage of ROI voxels trimmed for subject {p_id_stripped}, {run}: {trimmed_percentage:.2f}%")
                except subprocess.CalledProcessError as e:
                    print(f"Error occurred while processing {p_id_stripped} for {run}: {e}")
    else:
        print("ROIs already trimmed. Skipping process.")
    
# Step 5: Generate first-level fsf file [ANALYSIS 1].
    print("\n###### STEP 5: GENERATE FIRST-LEVEL FSF FILE [ANALYSIS 1] ######")
    first_level_fsf_template = r"""
# FEAT version number
set fmri(version) 6.00

# Are we in MELODIC?
set fmri(inmelodic) 0

# Analysis level
# 1 : First-level analysis
# 2 : Higher-level analysis
set fmri(level) 1

# Which stages to run
# 0 : No first-level analysis (registration and/or group stats only)
# 7 : Full first-level analysis
# 1 : Pre-processing
# 2 : Statistics
set fmri(analysis) 2

# Use relative filenames
set fmri(relative_yn) 0

# Balloon help
set fmri(help_yn) 1

# Run Featwatcher
set fmri(featwatcher_yn) 1

# Cleanup first-level standard-space images
set fmri(sscleanup_yn) 0

# Output directory
set fmri(outputdir) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/fmri_analysis/analysis_1/first_level/sub-[insert_participant]/[insert_folder_name]"

# TR(s)
set fmri(tr) 2.000000 

# Total volumes
set fmri(npts) 210

# Delete volumes
set fmri(ndelete) 0

# Perfusion tag/control order
set fmri(tagfirst) 1

# Number of first-level analyses
set fmri(multiple) 1

# Higher-level input type
# 1 : Inputs are lower-level FEAT directories
# 2 : Inputs are cope images from FEAT directories
set fmri(inputtype) 2

# Carry out pre-stats processing?
set fmri(filtering_yn) 0

# Brain/background threshold, %
set fmri(brain_thresh) 10

# Critical z for design efficiency calculation
set fmri(critical_z) 5.3

# Noise level
set fmri(noise) 0.66

# Noise AR(1)
set fmri(noisear) 0.34

# Motion correction
# 0 : None
# 1 : MCFLIRT
set fmri(mc) 1

# Spin-history (currently obsolete)
set fmri(sh_yn) 0

# B0 fieldmap unwarping?
set fmri(regunwarp_yn) 0

# GDC Test
set fmri(gdc) ""

# EPI dwell time (ms)
set fmri(dwell) 0.0

# EPI TE (ms)
set fmri(te) 0.0

# % Signal loss threshold
set fmri(signallossthresh) 10

# Unwarp direction
set fmri(unwarp_dir) y-

# Slice timing correction
# 0 : None
# 1 : Regular up (0, 1, 2, 3, ...)
# 2 : Regular down
# 3 : Use slice order file
# 4 : Use slice timings file
# 5 : Interleaved (0, 2, 4 ... 1, 3, 5 ... )
set fmri(st) 0

# Slice timings file
set fmri(st_file) ""

# BET brain extraction
set fmri(bet_yn) 1

# Spatial smoothing FWHM (mm)
set fmri(smooth) 5

# Intensity normalization
set fmri(norm_yn) 0

# Perfusion subtraction
set fmri(perfsub_yn) 0

# Highpass temporal filtering
set fmri(temphp_yn) 1

# Lowpass temporal filtering
set fmri(templp_yn) 0

# MELODIC ICA data exploration
set fmri(melodic_yn) 0

# Carry out main stats?
set fmri(stats_yn) 1

# Carry out prewhitening?
set fmri(prewhiten_yn) 1

# Add motion parameters to model
# 0 : No
# 1 : Yes
set fmri(motionevs) 0
set fmri(motionevsbeta) ""
set fmri(scriptevsbeta) ""

# Robust outlier detection in FLAME?
set fmri(robust_yn) 0

# Higher-level modelling
# 3 : Fixed effects
# 0 : Mixed Effects: Simple OLS
# 2 : Mixed Effects: FLAME 1
# 1 : Mixed Effects: FLAME 1+2
set fmri(mixed_yn) 3

# Higher-level permutations
set fmri(randomisePermutations) 5000

# Number of EVs
set fmri(evs_orig) 2
set fmri(evs_real) 4
set fmri(evs_vox) 0

# Number of contrasts
set fmri(ncon_orig) 4
set fmri(ncon_real) 4

# Number of F-tests
set fmri(nftests_orig) 0
set fmri(nftests_real) 0

# Add constant column to design matrix? (obsolete)
set fmri(constcol) 0

# Carry out post-stats steps?
set fmri(poststats_yn) 1

# Pre-threshold masking?
set fmri(threshmask) ""

# Thresholding
# 0 : None
# 1 : Uncorrected
# 2 : Voxel
# 3 : Cluster
set fmri(thresh) 0

# P threshold
set fmri(prob_thresh) 0.05

# Z threshold
set fmri(z_thresh) 2.3

# Z min/max for colour rendering
# 0 : Use actual Z min/max
# 1 : Use preset Z min/max
set fmri(zdisplay) 0

# Z min in colour rendering
set fmri(zmin) 2

# Z max in colour rendering
set fmri(zmax) 8

# Colour rendering type
# 0 : Solid blobs
# 1 : Transparent blobs
set fmri(rendertype) 1

# Background image for higher-level stats overlays
# 1 : Mean highres
# 2 : First highres
# 3 : Mean functional
# 4 : First functional
# 5 : Standard space template
set fmri(bgimage) 1

# Create time series plots
set fmri(tsplot_yn) 0

# Registration to initial structural
set fmri(reginitial_highres_yn) 0

# Search space for registration to initial structural
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(reginitial_highres_search) 90

# Degrees of Freedom for registration to initial structural
set fmri(reginitial_highres_dof) 3

# Registration to main structural
set fmri(reghighres_yn) 0

# Search space for registration to main structural
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(reghighres_search) 90

# Degrees of Freedom for registration to main structural
set fmri(reghighres_dof) BBR

# Registration to standard image?
set fmri(regstandard_yn) 0

# Use alternate reference images?
set fmri(alternateReference_yn) 0

# Standard image
set fmri(regstandard) "/usr/local/fsl/data/standard/MNI152_T1_2mm_brain"

# Search space for registration to standard space
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(regstandard_search) 90

# Degrees of Freedom for registration to standard space
set fmri(regstandard_dof) 12

# Do nonlinear registration from structural to standard space?
set fmri(regstandard_nonlinear_yn) 0

# Control nonlinear warp field resolution
set fmri(regstandard_nonlinear_warpres) 10

# High pass filter cutoff
set fmri(paradigm_hp) 100

# Total voxels
set fmri(totalVoxels) [insert_total_voxels]


# Number of lower-level copes feeding into higher-level analysis
set fmri(ncopeinputs) 0

# 4D AVW data or FEAT directory (1)
set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/data/fully_preproc/sub-[insert_participant]/func/sub-[insert_participant]_[insert_run]_MNI152_func_fully_preproc"

# Add confound EVs text file
set fmri(confoundevs) 1

# Confound EVs text file for analysis 1
set confoundev_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-[insert_participant]/confounds_[insert_run].txt"

# EV 1 title
set fmri(evtitle1) "guilt"

# Basic waveform shape (EV 1)
# 0 : Square
# 1 : Sinusoid
# 2 : Custom (1 entry per volume)
# 3 : Custom (3 column format)
# 4 : Interaction
# 10 : Empty (all zeros)
set fmri(shape1) 3

# Convolution (EV 1)
# 0 : None
# 1 : Gaussian
# 2 : Gamma
# 3 : Double-Gamma HRF
# 4 : Gamma basis functions
# 5 : Sine basis functions
# 6 : FIR basis functions
# 8 : Alternate Double-Gamma
set fmri(convolve1) 3

# Convolve phase (EV 1)
set fmri(convolve_phase1) 0

# Apply temporal filtering (EV 1)
set fmri(tempfilt_yn1) 1

# Add temporal derivative (EV 1)
set fmri(deriv_yn1) 1

# Custom EV file (EV 1)
set fmri(custom1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/shared/onset_files/onsetfile_guilt.txt"

# Orthogonalise EV 1 wrt EV 0
set fmri(ortho1.0) 0

# Orthogonalise EV 1 wrt EV 1
set fmri(ortho1.1) 0

# Orthogonalise EV 1 wrt EV 2
set fmri(ortho1.2) 0

# EV 2 title
set fmri(evtitle2) "indig"

# Basic waveform shape (EV 2)
# 0 : Square
# 1 : Sinusoid
# 2 : Custom (1 entry per volume)
# 3 : Custom (3 column format)
# 4 : Interaction
# 10 : Empty (all zeros)
set fmri(shape2) 3

# Convolution (EV 2)
# 0 : None
# 1 : Gaussian
# 2 : Gamma
# 3 : Double-Gamma HRF
# 4 : Gamma basis functions
# 5 : Sine basis functions
# 6 : FIR basis functions
# 8 : Alternate Double-Gamma
set fmri(convolve2) 3

# Convolve phase (EV 2)
set fmri(convolve_phase2) 0

# Apply temporal filtering (EV 2)
set fmri(tempfilt_yn2) 1

# Add temporal derivative (EV 2)
set fmri(deriv_yn2) 1

# Custom EV file (EV 2)
set fmri(custom2) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/shared/onset_files/onsetfile_indig.txt"

# Orthogonalise EV 2 wrt EV 0
set fmri(ortho2.0) 0

# Orthogonalise EV 2 wrt EV 1
set fmri(ortho2.1) 0

# Orthogonalise EV 2 wrt EV 2
set fmri(ortho2.2) 0

# Contrast & F-tests mode
# real : control real EVs
# orig : control original EVs
set fmri(con_mode_old) orig
set fmri(con_mode) orig

# Display images for contrast_real 1
set fmri(conpic_real.1) 1

# Title for contrast_real 1
set fmri(conname_real.1) "guilt"

# Real contrast_real vector 1 element 1
set fmri(con_real1.1) 1.0

# Real contrast_real vector 1 element 2
set fmri(con_real1.2) 0

# Real contrast_real vector 1 element 3
set fmri(con_real1.3) 0.0

# Real contrast_real vector 1 element 4
set fmri(con_real1.4) 0

# Display images for contrast_real 2
set fmri(conpic_real.2) 1

# Title for contrast_real 2
set fmri(conname_real.2) "indig"

# Real contrast_real vector 2 element 1
set fmri(con_real2.1) 0.0

# Real contrast_real vector 2 element 2
set fmri(con_real2.2) 0

# Real contrast_real vector 2 element 3
set fmri(con_real2.3) 1.0

# Real contrast_real vector 2 element 4
set fmri(con_real2.4) 0

# Display images for contrast_real 3
set fmri(conpic_real.3) 1

# Title for contrast_real 3
set fmri(conname_real.3) "guilt-indig"

# Real contrast_real vector 3 element 1
set fmri(con_real3.1) 1.0

# Real contrast_real vector 3 element 2
set fmri(con_real3.2) 0

# Real contrast_real vector 3 element 3
set fmri(con_real3.3) -1.0

# Real contrast_real vector 3 element 4
set fmri(con_real3.4) 0

# Display images for contrast_real 4
set fmri(conpic_real.4) 1

# Title for contrast_real 4
set fmri(conname_real.4) "indig-guilt"

# Real contrast_real vector 4 element 1
set fmri(con_real4.1) -1.0

# Real contrast_real vector 4 element 2
set fmri(con_real4.2) 0

# Real contrast_real vector 4 element 3
set fmri(con_real4.3) 1.0

# Real contrast_real vector 4 element 4
set fmri(con_real4.4) 0

# Display images for contrast_orig 1
set fmri(conpic_orig.1) 1

# Title for contrast_orig 1
set fmri(conname_orig.1) "guilt"

# Real contrast_orig vector 1 element 1
set fmri(con_orig1.1) 1.0

# Real contrast_orig vector 1 element 2
set fmri(con_orig1.2) 0.0

# Display images for contrast_orig 2
set fmri(conpic_orig.2) 1

# Title for contrast_orig 2
set fmri(conname_orig.2) "indig"

# Real contrast_orig vector 2 element 1
set fmri(con_orig2.1) 0.0

# Real contrast_orig vector 2 element 2
set fmri(con_orig2.2) 1.0

# Display images for contrast_orig 3
set fmri(conpic_orig.3) 1

# Title for contrast_orig 3
set fmri(conname_orig.3) "guilt-indig"

# Real contrast_orig vector 3 element 1
set fmri(con_orig3.1) 1.0

# Real contrast_orig vector 3 element 2
set fmri(con_orig3.2) -1.0

# Display images for contrast_orig 4
set fmri(conpic_orig.4) 1

# Title for contrast_orig 4
set fmri(conname_orig.4) "indig-guilt"

# Real contrast_orig vector 4 element 1
set fmri(con_orig4.1) -1.0

# Real contrast_orig vector 4 element 2
set fmri(con_orig4.2) 1.0

# Contrast masking - use >0 instead of thresholding?
set fmri(conmask_zerothresh_yn) 0

# Mask real contrast/F-test 1 with real contrast/F-test 2?
set fmri(conmask1_2) 0

# Mask real contrast/F-test 1 with real contrast/F-test 3?
set fmri(conmask1_3) 0

# Mask real contrast/F-test 1 with real contrast/F-test 4?
set fmri(conmask1_4) 0

# Mask real contrast/F-test 2 with real contrast/F-test 1?
set fmri(conmask2_1) 0

# Mask real contrast/F-test 2 with real contrast/F-test 3?
set fmri(conmask2_3) 0

# Mask real contrast/F-test 2 with real contrast/F-test 4?
set fmri(conmask2_4) 0

# Mask real contrast/F-test 3 with real contrast/F-test 1?
set fmri(conmask3_1) 0

# Mask real contrast/F-test 3 with real contrast/F-test 2?
set fmri(conmask3_2) 0

# Mask real contrast/F-test 3 with real contrast/F-test 4?
set fmri(conmask3_4) 0

# Mask real contrast/F-test 4 with real contrast/F-test 1?
set fmri(conmask4_1) 0

# Mask real contrast/F-test 4 with real contrast/F-test 2?
set fmri(conmask4_2) 0

# Mask real contrast/F-test 4 with real contrast/F-test 3?
set fmri(conmask4_3) 0

# Do contrast masking at all?
set fmri(conmask1_1) 0

##########################################################
# Now options that don't appear in the GUI

# Alternative (to BETting) mask image
set fmri(alternative_mask) ""

# Initial structural space registration initialisation transform
set fmri(init_initial_highres) ""

# Structural space registration initialisation transform
set fmri(init_highres) ""

# Standard space registration initialisation transform
set fmri(init_standard) ""

# For full FEAT analysis: overwrite existing .feat output dir?
set fmri(overwrite_yn) 0
"""
    
    first_level_fsf_template_path = f'analysis/fmri_analysis/analysis_1/first_level/shared/first_level_fsf_template.fsf'
    with open(first_level_fsf_template_path, 'w') as f:
        f.write(first_level_fsf_template)
    first_level_fsfs = []
    for p_id in participants:
        p_id_stripped = p_id.replace('P', '')
        for run in runs:
            with open(first_level_fsf_template_path, 'r') as file:
                fsf_data = file.readlines()
            func_file = f'data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_{run}_MNI152_func_fully_preproc.nii.gz'
            fslinfo_output = subprocess.run(["fslinfo", func_file], capture_output=True, text=True)
            lines = fslinfo_output.stdout.splitlines()
            dim1 = int([line.split()[1] for line in lines if "dim1" in line][0])
            dim2 = int([line.split()[1] for line in lines if "dim2" in line][0])
            dim3 = int([line.split()[1] for line in lines if "dim3" in line][0])
            npts = int([line.split()[1] for line in lines if "dim4" in line][0])
            total_voxels = dim1 * dim2 * dim3 * npts
            for i, line in enumerate(fsf_data):
                if "set fmri(outputdir)" in line:
                    fsf_data[i] = f'set fmri(outputdir) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/{run}"\n'
                elif "set fmri(totalVoxels)" in line:
                    fsf_data[i] = f"set fmri(totalVoxels) {total_voxels}\n"
                elif "set feat_files(1)" in line:
                    fsf_data[i] = f'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/data/fully_preproc/sub-{p_id_stripped}/func/sub-{p_id_stripped}_{run}_MNI152_func_fully_preproc"\n'
                elif "set confoundev_files(1)" in line:
                    fsf_data[i] = f'set confoundev_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/confounds_{run}.txt"\n'
            first_level_fsf = f'analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/first_level_fsf_{run}.fsf'
            first_level_fsfs.append(first_level_fsf)
            with open(first_level_fsf, 'w') as file:
                file.writelines(fsf_data)
    print(f'fsf files for first-level GLM generated.')
    
    # Step 6: Run first-level GLM [ANALYSIS 1].
    print("\n###### STEP 6: RUN FIRST-LEVEL GLM [ANALYSIS 1] ######")   
    design_png_paths = []
    for fsf in first_level_fsfs:
        match = re.search(r'sub-(\d{3}).*run-(\d{2})', fsf)
        participant_number = match.group(1)
        run_number = match.group(2)
        if not os.path.exists(f'analysis/fmri_analysis/analysis_1/first_level/sub-{participant_number}/run-{run_number}.feat'):
            print(f'Running first-level GLM of sub-{participant_number} for run{run_number}...')
            subprocess.run(['feat', fsf])
            report_log_path = f'analysis/fmri_analysis/analysis_1/first_level/sub-{participant_number}/run-{run_number}.feat/report_log.html'
            with open(report_log_path, 'r') as file:
                content = file.read()
            if re.search('error', content, re.IGNORECASE):
                print(f"Error found in report_log.html. Investigation required.")
            cope_files = glob.glob(f'analysis/fmri_analysis/analysis_1/first_level/sub-{participant_number}/run-{run_number}.feat/stats/cope*')
            if len(cope_files) != 4:
                print("There are not 4 cope files in the stats folder. Investigation required.")
            design_png_paths.append(f'analysis/fmri_analysis/analysis_1/first_level/sub-{participant_number}/run-{run_number}.feat/design.png')
        else:
            ('First-level GLMs already run. Skipping process.')
    if design_png_paths:
        def hash_image(image_path):
            with Image.open(image_path) as img:
                img_bytes = img.tobytes()
                img_hash = hashlib.sha256(img_bytes).hexdigest()
            return img_hash
        def compare_images(image_paths):
            first_image_hash = hash_image(image_paths[0])
            for image_path in image_paths[1:]:
                if hash_image(image_path) != first_image_hash:
                    print(f"design.png images are not identical: {image_paths[0]} and {image_path}. Investigation required.")
                    return
        compare_images(design_png_paths)
    
    
    # Step 7: Generate second-level fsf filea [ANALYSIS 1].
    print("\n###### STEP 7: GENERATE SECOND-LEVEL FSF FILES [ANALYSIS 1] ######")   
    second_level_fsf_template_nogroup = r"""
# FEAT version number
set fmri(version) 6.00

# Are we in MELODIC?
set fmri(inmelodic) 0

# Analysis level
# 1 : First-level analysis
# 2 : Higher-level analysis
set fmri(level) 2

# Which stages to run
# 0 : No first-level analysis (registration and/or group stats only)
# 7 : Full first-level analysis
# 1 : Pre-processing
# 2 : Statistics
set fmri(analysis) 2

# Use relative filenames
set fmri(relative_yn) 0

# Balloon help
set fmri(help_yn) 1

# Run Featwatcher
set fmri(featwatcher_yn) 1

# Cleanup first-level standard-space images
set fmri(sscleanup_yn) 0

# Output directory
set fmri(outputdir) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/[insert_folder_name]"

# TR(s)
set fmri(tr) 2

# Total volumes
set fmri(npts) 40

# Delete volumes
set fmri(ndelete) 0

# Perfusion tag/control order
set fmri(tagfirst) 1

# Number of first-level analyses
set fmri(multiple) 40

# Higher-level input type
# 1 : Inputs are lower-level FEAT directories
# 2 : Inputs are cope images from FEAT directories
set fmri(inputtype) 1

# Carry out pre-stats processing?
set fmri(filtering_yn) 0

# Brain/background threshold, %
set fmri(brain_thresh) 10

# Critical z for design efficiency calculation
set fmri(critical_z) 5.3

# Noise level
set fmri(noise) 0.66

# Noise AR(1)
set fmri(noisear) 0.34

# Motion correction
# 0 : None
# 1 : MCFLIRT
set fmri(mc) 1

# Spin-history (currently obsolete)
set fmri(sh_yn) 0

# B0 fieldmap unwarping?
set fmri(regunwarp_yn) 0

# GDC Test
set fmri(gdc) ""

# EPI dwell time (ms)
set fmri(dwell) 0.0

# EPI TE (ms)
set fmri(te) 0.0

# % Signal loss threshold
set fmri(signallossthresh) 10

# Unwarp direction
set fmri(unwarp_dir) y-

# Slice timing correction
# 0 : None
# 1 : Regular up (0, 1, 2, 3, ...)
# 2 : Regular down
# 3 : Use slice order file
# 4 : Use slice timings file
# 5 : Interleaved (0, 2, 4 ... 1, 3, 5 ... )
set fmri(st) 0

# Slice timings file
set fmri(st_file) ""

# BET brain extraction
set fmri(bet_yn) 1

# Spatial smoothing FWHM (mm)
set fmri(smooth) 5

# Intensity normalization
set fmri(norm_yn) 0

# Perfusion subtraction
set fmri(perfsub_yn) 0

# Highpass temporal filtering
set fmri(temphp_yn) 1

# Lowpass temporal filtering
set fmri(templp_yn) 0

# MELODIC ICA data exploration
set fmri(melodic_yn) 0

# Carry out main stats?
set fmri(stats_yn) 1

# Carry out prewhitening?
set fmri(prewhiten_yn) 1

# Add motion parameters to model
# 0 : No
# 1 : Yes
set fmri(motionevs) 0
set fmri(motionevsbeta) ""
set fmri(scriptevsbeta) ""

# Robust outlier detection in FLAME?
set fmri(robust_yn) 0

# Higher-level modelling
# 3 : Fixed effects
# 0 : Mixed Effects: Simple OLS
# 2 : Mixed Effects: FLAME 1
# 1 : Mixed Effects: FLAME 1+2
set fmri(mixed_yn) 3

# Higher-level permutations
set fmri(randomisePermutations) 5000

# Number of EVs
set fmri(evs_orig) 3
set fmri(evs_real) 3
set fmri(evs_vox) 0

# Number of contrasts
set fmri(ncon_orig) 1
set fmri(ncon_real) 4

# Number of F-tests
set fmri(nftests_orig) 0
set fmri(nftests_real) 0

# Add constant column to design matrix? (obsolete)
set fmri(constcol) 0

# Carry out post-stats steps?
set fmri(poststats_yn) 1

# Pre-threshold masking?
set fmri(threshmask) ""

# Thresholding
# 0 : None
# 1 : Uncorrected
# 2 : Voxel
# 3 : Cluster
set fmri(thresh) 1

# P threshold
set fmri(prob_thresh) 0.05

# Z threshold
set fmri(z_thresh) 2.3

# Z min/max for colour rendering
# 0 : Use actual Z min/max
# 1 : Use preset Z min/max
set fmri(zdisplay) 0

# Z min in colour rendering
set fmri(zmin) 2

# Z max in colour rendering
set fmri(zmax) 8

# Colour rendering type
# 0 : Solid blobs
# 1 : Transparent blobs
set fmri(rendertype) 1

# Background image for higher-level stats overlays
# 1 : Mean highres
# 2 : First highres
# 3 : Mean functional
# 4 : First functional
# 5 : Standard space template
set fmri(bgimage) 1

# Create time series plots
set fmri(tsplot_yn) 1

# Registration to initial structural
set fmri(reginitial_highres_yn) 0

# Search space for registration to initial structural
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(reginitial_highres_search) 90

# Degrees of Freedom for registration to initial structural
set fmri(reginitial_highres_dof) 3

# Registration to main structural
set fmri(reghighres_yn) 0

# Search space for registration to main structural
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(reghighres_search) 90

# Degrees of Freedom for registration to main structural
set fmri(reghighres_dof) BBR

# Registration to standard image?
set fmri(regstandard_yn) 1

# Use alternate reference images?
set fmri(alternateReference_yn) 0

# Standard image
set fmri(regstandard) "/usr/local/fsl/data/standard/MNI152_T1_2mm_brain"

# Search space for registration to standard space
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(regstandard_search) 90

# Degrees of Freedom for registration to standard space
set fmri(regstandard_dof) 12

# Do nonlinear registration from structural to standard space?
set fmri(regstandard_nonlinear_yn) 0

# Control nonlinear warp field resolution
set fmri(regstandard_nonlinear_warpres) 10

# High pass filter cutoff
set fmri(paradigm_hp) 100

# Number of lower-level copes feeding into higher-level analysis
set fmri(ncopeinputs) 4

# Use lower-level cope 1 for higher-level analysis
set fmri(copeinput.1) 1

# Use lower-level cope 2 for higher-level analysis
set fmri(copeinput.2) 1

# Use lower-level cope 3 for higher-level analysis
set fmri(copeinput.3) 1

# Use lower-level cope 4 for higher-level analysis
set fmri(copeinput.4) 1

# 4D AVW data or FEAT directory (1)
set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-004/run-01.feat"

# 4D AVW data or FEAT directory (2)
set feat_files(2) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-004/run-04.feat"

# 4D AVW data or FEAT directory (3)
set feat_files(3) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-006/run-01.feat"

# 4D AVW data or FEAT directory (4)
set feat_files(4) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-006/run-04.feat"

# 4D AVW data or FEAT directory (5)
set feat_files(5) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-020/run-01.feat"

# 4D AVW data or FEAT directory (6)
set feat_files(6) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-020/run-04.feat"

# 4D AVW data or FEAT directory (7)
set feat_files(7) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-030/run-01.feat"

# 4D AVW data or FEAT directory (8)
set feat_files(8) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-030/run-04.feat"

# 4D AVW data or FEAT directory (9)
set feat_files(9) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-059/run-01.feat"

# 4D AVW data or FEAT directory (10)
set feat_files(10) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-059/run-04.feat"

# 4D AVW data or FEAT directory (11)
set feat_files(11) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-078/run-01.feat"

# 4D AVW data or FEAT directory (12)
set feat_files(12) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-078/run-04.feat"

# 4D AVW data or FEAT directory (13)
set feat_files(13) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-093/run-01.feat"

# 4D AVW data or FEAT directory (14)
set feat_files(14) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-093/run-04.feat"

# 4D AVW data or FEAT directory (15)
set feat_files(15) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-094/run-01.feat"

# 4D AVW data or FEAT directory (16)
set feat_files(16) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-094/run-04.feat"

# 4D AVW data or FEAT directory (17)
set feat_files(17) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-100/run-01.feat"

# 4D AVW data or FEAT directory (18)
set feat_files(18) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-100/run-04.feat"

# 4D AVW data or FEAT directory (19)
set feat_files(19) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-107/run-01.feat"

# 4D AVW data or FEAT directory (20)
set feat_files(20) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-107/run-04.feat"

# 4D AVW data or FEAT directory (21)
set feat_files(21) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-122/run-01.feat"

# 4D AVW data or FEAT directory (22)
set feat_files(22) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-122/run-04.feat"

# 4D AVW data or FEAT directory (23)
set feat_files(23) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-125/run-01.feat"

# 4D AVW data or FEAT directory (24)
set feat_files(24) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-125/run-04.feat"

# 4D AVW data or FEAT directory (25)
set feat_files(25) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-127/run-01.feat"

# 4D AVW data or FEAT directory (26)
set feat_files(26) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-127/run-04.feat"

# 4D AVW data or FEAT directory (27)
set feat_files(27) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-128/run-01.feat"

# 4D AVW data or FEAT directory (28)
set feat_files(28) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-128/run-04.feat"

# 4D AVW data or FEAT directory (29)
set feat_files(29) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-136/run-01.feat"

# 4D AVW data or FEAT directory (30)
set feat_files(30) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-136/run-04.feat"

# 4D AVW data or FEAT directory (31)
set feat_files(31) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-145/run-01.feat"

# 4D AVW data or FEAT directory (32)
set feat_files(32) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-145/run-04.feat"

# 4D AVW data or FEAT directory (33)
set feat_files(33) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-155/run-01.feat"

# 4D AVW data or FEAT directory (34)
set feat_files(34) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-155/run-04.feat"

# 4D AVW data or FEAT directory (35)
set feat_files(35) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-199/run-01.feat"

# 4D AVW data or FEAT directory (36)
set feat_files(36) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-199/run-04.feat"

# 4D AVW data or FEAT directory (37)
set feat_files(37) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-215/run-01.feat"

# 4D AVW data or FEAT directory (38)
set feat_files(38) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-215/run-04.feat"

# 4D AVW data or FEAT directory (39)
set feat_files(39) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-216/run-01.feat"

# 4D AVW data or FEAT directory (40)
set feat_files(40) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-216/run-04.feat"

# Add confound EVs text file
set fmri(confoundevs) 0

# EV 1 title
set fmri(evtitle1) "run01"

# Basic waveform shape (EV 1)
# 0 : Square
# 1 : Sinusoid
# 2 : Custom (1 entry per volume)
# 3 : Custom (3 column format)
# 4 : Interaction
# 10 : Empty (all zeros)
set fmri(shape1) 2

# Convolution (EV 1)
# 0 : None
# 1 : Gaussian
# 2 : Gamma
# 3 : Double-Gamma HRF
# 4 : Gamma basis functions
# 5 : Sine basis functions
# 6 : FIR basis functions
# 8 : Alternate Double-Gamma
set fmri(convolve1) 0

# Convolve phase (EV 1)
set fmri(convolve_phase1) 0

# Apply temporal filtering (EV 1)
set fmri(tempfilt_yn1) 0

# Add temporal derivative (EV 1)
set fmri(deriv_yn1) 0

# Custom EV file (EV 1)
set fmri(custom1) "dummy"

# Orthogonalise EV 1 wrt EV 0
set fmri(ortho1.0) 0

# Orthogonalise EV 1 wrt EV 1
set fmri(ortho1.1) 0

# Orthogonalise EV 1 wrt EV 2
set fmri(ortho1.2) 0

# Orthogonalise EV 1 wrt EV 3
set fmri(ortho1.3) 0

# Higher-level EV value for EV 1 and input 1
set fmri(evg1.1) 1

# Higher-level EV value for EV 1 and input 2
set fmri(evg2.1) 0

# Higher-level EV value for EV 1 and input 3
set fmri(evg3.1) 1.0

# Higher-level EV value for EV 1 and input 4
set fmri(evg4.1) 0

# Higher-level EV value for EV 1 and input 5
set fmri(evg5.1) 1.0

# Higher-level EV value for EV 1 and input 6
set fmri(evg6.1) 0

# Higher-level EV value for EV 1 and input 7
set fmri(evg7.1) 1.0

# Higher-level EV value for EV 1 and input 8
set fmri(evg8.1) 0

# Higher-level EV value for EV 1 and input 9
set fmri(evg9.1) 1.0

# Higher-level EV value for EV 1 and input 10
set fmri(evg10.1) 0

# Higher-level EV value for EV 1 and input 11
set fmri(evg11.1) 1.0

# Higher-level EV value for EV 1 and input 12
set fmri(evg12.1) 0

# Higher-level EV value for EV 1 and input 13
set fmri(evg13.1) 1.0

# Higher-level EV value for EV 1 and input 14
set fmri(evg14.1) 0

# Higher-level EV value for EV 1 and input 15
set fmri(evg15.1) 1.0

# Higher-level EV value for EV 1 and input 16
set fmri(evg16.1) 0

# Higher-level EV value for EV 1 and input 17
set fmri(evg17.1) 1.0

# Higher-level EV value for EV 1 and input 18
set fmri(evg18.1) 0

# Higher-level EV value for EV 1 and input 19
set fmri(evg19.1) 1.0

# Higher-level EV value for EV 1 and input 20
set fmri(evg20.1) 0

# Higher-level EV value for EV 1 and input 21
set fmri(evg21.1) 1.0

# Higher-level EV value for EV 1 and input 22
set fmri(evg22.1) 0

# Higher-level EV value for EV 1 and input 23
set fmri(evg23.1) 1.0

# Higher-level EV value for EV 1 and input 24
set fmri(evg24.1) 0

# Higher-level EV value for EV 1 and input 25
set fmri(evg25.1) 1.0

# Higher-level EV value for EV 1 and input 26
set fmri(evg26.1) 0

# Higher-level EV value for EV 1 and input 27
set fmri(evg27.1) 1.0

# Higher-level EV value for EV 1 and input 28
set fmri(evg28.1) 0

# Higher-level EV value for EV 1 and input 29
set fmri(evg29.1) 1.0

# Higher-level EV value for EV 1 and input 30
set fmri(evg30.1) 0

# Higher-level EV value for EV 1 and input 31
set fmri(evg31.1) 1.0

# Higher-level EV value for EV 1 and input 32
set fmri(evg32.1) 0

# Higher-level EV value for EV 1 and input 33
set fmri(evg33.1) 1.0

# Higher-level EV value for EV 1 and input 34
set fmri(evg34.1) 0

# Higher-level EV value for EV 1 and input 35
set fmri(evg35.1) 1.0

# Higher-level EV value for EV 1 and input 36
set fmri(evg36.1) 0

# Higher-level EV value for EV 1 and input 37
set fmri(evg37.1) 1.0

# Higher-level EV value for EV 1 and input 38
set fmri(evg38.1) 0

# Higher-level EV value for EV 1 and input 39
set fmri(evg39.1) 1.0

# Higher-level EV value for EV 1 and input 40
set fmri(evg40.1) 0

# EV 2 title
set fmri(evtitle2) "run04"

# Basic waveform shape (EV 2)
# 0 : Square
# 1 : Sinusoid
# 2 : Custom (1 entry per volume)
# 3 : Custom (3 column format)
# 4 : Interaction
# 10 : Empty (all zeros)
set fmri(shape2) 2

# Convolution (EV 2)
# 0 : None
# 1 : Gaussian
# 2 : Gamma
# 3 : Double-Gamma HRF
# 4 : Gamma basis functions
# 5 : Sine basis functions
# 6 : FIR basis functions
# 8 : Alternate Double-Gamma
set fmri(convolve2) 0

# Convolve phase (EV 2)
set fmri(convolve_phase2) 0

# Apply temporal filtering (EV 2)
set fmri(tempfilt_yn2) 0

# Add temporal derivative (EV 2)
set fmri(deriv_yn2) 0

# Custom EV file (EV 2)
set fmri(custom2) "dummy"

# Orthogonalise EV 2 wrt EV 0
set fmri(ortho2.0) 0

# Orthogonalise EV 2 wrt EV 1
set fmri(ortho2.1) 0

# Orthogonalise EV 2 wrt EV 2
set fmri(ortho2.2) 0

# Orthogonalise EV 2 wrt EV 3
set fmri(ortho2.3) 0

# Higher-level EV value for EV 2 and input 1
set fmri(evg1.2) 0

# Higher-level EV value for EV 2 and input 2
set fmri(evg2.2) 1.0

# Higher-level EV value for EV 2 and input 3
set fmri(evg3.2) 0

# Higher-level EV value for EV 2 and input 4
set fmri(evg4.2) 1.0

# Higher-level EV value for EV 2 and input 5
set fmri(evg5.2) 0

# Higher-level EV value for EV 2 and input 6
set fmri(evg6.2) 1.0

# Higher-level EV value for EV 2 and input 7
set fmri(evg7.2) 0

# Higher-level EV value for EV 2 and input 8
set fmri(evg8.2) 1.0

# Higher-level EV value for EV 2 and input 9
set fmri(evg9.2) 0

# Higher-level EV value for EV 2 and input 10
set fmri(evg10.2) 1.0

# Higher-level EV value for EV 2 and input 11
set fmri(evg11.2) 0

# Higher-level EV value for EV 2 and input 12
set fmri(evg12.2) 1.0

# Higher-level EV value for EV 2 and input 13
set fmri(evg13.2) 0

# Higher-level EV value for EV 2 and input 14
set fmri(evg14.2) 1.0

# Higher-level EV value for EV 2 and input 15
set fmri(evg15.2) 0

# Higher-level EV value for EV 2 and input 16
set fmri(evg16.2) 1.0

# Higher-level EV value for EV 2 and input 17
set fmri(evg17.2) 0

# Higher-level EV value for EV 2 and input 18
set fmri(evg18.2) 1.0

# Higher-level EV value for EV 2 and input 19
set fmri(evg19.2) 0

# Higher-level EV value for EV 2 and input 20
set fmri(evg20.2) 1.0

# Higher-level EV value for EV 2 and input 21
set fmri(evg21.2) 0

# Higher-level EV value for EV 2 and input 22
set fmri(evg22.2) 1.0

# Higher-level EV value for EV 2 and input 23
set fmri(evg23.2) 0

# Higher-level EV value for EV 2 and input 24
set fmri(evg24.2) 1.0

# Higher-level EV value for EV 2 and input 25
set fmri(evg25.2) 0

# Higher-level EV value for EV 2 and input 26
set fmri(evg26.2) 1.0

# Higher-level EV value for EV 2 and input 27
set fmri(evg27.2) 0

# Higher-level EV value for EV 2 and input 28
set fmri(evg28.2) 1.0

# Higher-level EV value for EV 2 and input 29
set fmri(evg29.2) 0

# Higher-level EV value for EV 2 and input 30
set fmri(evg30.2) 1.0

# Higher-level EV value for EV 2 and input 31
set fmri(evg31.2) 0

# Higher-level EV value for EV 2 and input 32
set fmri(evg32.2) 1.0

# Higher-level EV value for EV 2 and input 33
set fmri(evg33.2) 0

# Higher-level EV value for EV 2 and input 34
set fmri(evg34.2) 1.0

# Higher-level EV value for EV 2 and input 35
set fmri(evg35.2) 0

# Higher-level EV value for EV 2 and input 36
set fmri(evg36.2) 1.0

# Higher-level EV value for EV 2 and input 37
set fmri(evg37.2) 0

# Higher-level EV value for EV 2 and input 38
set fmri(evg38.2) 1.0

# Higher-level EV value for EV 2 and input 39
set fmri(evg39.2) 0

# Higher-level EV value for EV 2 and input 40
set fmri(evg40.2) 1.0

# EV 3 title
set fmri(evtitle3) "fieldmaps"

# Basic waveform shape (EV 3)
# 0 : Square
# 1 : Sinusoid
# 2 : Custom (1 entry per volume)
# 3 : Custom (3 column format)
# 4 : Interaction
# 10 : Empty (all zeros)
set fmri(shape3) 2

# Convolution (EV 3)
# 0 : None
# 1 : Gaussian
# 2 : Gamma
# 3 : Double-Gamma HRF
# 4 : Gamma basis functions
# 5 : Sine basis functions
# 6 : FIR basis functions
# 8 : Alternate Double-Gamma
set fmri(convolve3) 0

# Convolve phase (EV 3)
set fmri(convolve_phase3) 0

# Apply temporal filtering (EV 3)
set fmri(tempfilt_yn3) 0

# Add temporal derivative (EV 3)
set fmri(deriv_yn3) 0

# Custom EV file (EV 3)
set fmri(custom3) "dummy"

# Orthogonalise EV 3 wrt EV 0
set fmri(ortho3.0) 0

# Orthogonalise EV 3 wrt EV 1
set fmri(ortho3.1) 0

# Orthogonalise EV 3 wrt EV 2
set fmri(ortho3.2) 0

# Orthogonalise EV 3 wrt EV 3
set fmri(ortho3.3) 0

# Higher-level EV value for EV 3 and input 1
set fmri(evg1.3) 0

# Higher-level EV value for EV 3 and input 2
set fmri(evg2.3) 0

# Higher-level EV value for EV 3 and input 3
set fmri(evg3.3) 0

# Higher-level EV value for EV 3 and input 4
set fmri(evg4.3) 0

# Higher-level EV value for EV 3 and input 5
set fmri(evg5.3) 0

# Higher-level EV value for EV 3 and input 6
set fmri(evg6.3) 0

# Higher-level EV value for EV 3 and input 7
set fmri(evg7.3) 0

# Higher-level EV value for EV 3 and input 8
set fmri(evg8.3) 0

# Higher-level EV value for EV 3 and input 9
set fmri(evg9.3) 1.0

# Higher-level EV value for EV 3 and input 10
set fmri(evg10.3) 1.0

# Higher-level EV value for EV 3 and input 11
set fmri(evg11.3) 0

# Higher-level EV value for EV 3 and input 12
set fmri(evg12.3) 0

# Higher-level EV value for EV 3 and input 13
set fmri(evg13.3) 0

# Higher-level EV value for EV 3 and input 14
set fmri(evg14.3) 0

# Higher-level EV value for EV 3 and input 15
set fmri(evg15.3) 0

# Higher-level EV value for EV 3 and input 16
set fmri(evg16.3) 0

# Higher-level EV value for EV 3 and input 17
set fmri(evg17.3) 1.0

# Higher-level EV value for EV 3 and input 18
set fmri(evg18.3) 1.0

# Higher-level EV value for EV 3 and input 19
set fmri(evg19.3) 1.0

# Higher-level EV value for EV 3 and input 20
set fmri(evg20.3) 1.0

# Higher-level EV value for EV 3 and input 21
set fmri(evg21.3) 1.0

# Higher-level EV value for EV 3 and input 22
set fmri(evg22.3) 1.0

# Higher-level EV value for EV 3 and input 23
set fmri(evg23.3) 1.0

# Higher-level EV value for EV 3 and input 24
set fmri(evg24.3) 1.0

# Higher-level EV value for EV 3 and input 25
set fmri(evg25.3) 1.0

# Higher-level EV value for EV 3 and input 26
set fmri(evg26.3) 1.0

# Higher-level EV value for EV 3 and input 27
set fmri(evg27.3) 1.0

# Higher-level EV value for EV 3 and input 28
set fmri(evg28.3) 1.0

# Higher-level EV value for EV 3 and input 29
set fmri(evg29.3) 1.0

# Higher-level EV value for EV 3 and input 30
set fmri(evg30.3) 1.0

# Higher-level EV value for EV 3 and input 31
set fmri(evg31.3) 1.0

# Higher-level EV value for EV 3 and input 32
set fmri(evg32.3) 1.0

# Higher-level EV value for EV 3 and input 33
set fmri(evg33.3) 1.0

# Higher-level EV value for EV 3 and input 34
set fmri(evg34.3) 1.0

# Higher-level EV value for EV 3 and input 35
set fmri(evg35.3) 1.0

# Higher-level EV value for EV 3 and input 36
set fmri(evg36.3) 1.0

# Higher-level EV value for EV 3 and input 37
set fmri(evg37.3) 1.0

# Higher-level EV value for EV 3 and input 38
set fmri(evg38.3) 1.0

# Higher-level EV value for EV 3 and input 39
set fmri(evg39.3) 1.0

# Higher-level EV value for EV 3 and input 40
set fmri(evg40.3) 1.0

# Setup Orthogonalisation at higher level?
set fmri(level2orth) 0

# Group membership for input 1
set fmri(groupmem.1) 1

# Group membership for input 2
set fmri(groupmem.2) 1

# Group membership for input 3
set fmri(groupmem.3) 1

# Group membership for input 4
set fmri(groupmem.4) 1

# Group membership for input 5
set fmri(groupmem.5) 1

# Group membership for input 6
set fmri(groupmem.6) 1

# Group membership for input 7
set fmri(groupmem.7) 1

# Group membership for input 8
set fmri(groupmem.8) 1

# Group membership for input 9
set fmri(groupmem.9) 1

# Group membership for input 10
set fmri(groupmem.10) 1

# Group membership for input 11
set fmri(groupmem.11) 1

# Group membership for input 12
set fmri(groupmem.12) 1

# Group membership for input 13
set fmri(groupmem.13) 1

# Group membership for input 14
set fmri(groupmem.14) 1

# Group membership for input 15
set fmri(groupmem.15) 1

# Group membership for input 16
set fmri(groupmem.16) 1

# Group membership for input 17
set fmri(groupmem.17) 1

# Group membership for input 18
set fmri(groupmem.18) 1

# Group membership for input 19
set fmri(groupmem.19) 1

# Group membership for input 20
set fmri(groupmem.20) 1

# Group membership for input 21
set fmri(groupmem.21) 1

# Group membership for input 22
set fmri(groupmem.22) 1

# Group membership for input 23
set fmri(groupmem.23) 1

# Group membership for input 24
set fmri(groupmem.24) 1

# Group membership for input 25
set fmri(groupmem.25) 1

# Group membership for input 26
set fmri(groupmem.26) 1

# Group membership for input 27
set fmri(groupmem.27) 1

# Group membership for input 28
set fmri(groupmem.28) 1

# Group membership for input 29
set fmri(groupmem.29) 1

# Group membership for input 30
set fmri(groupmem.30) 1

# Group membership for input 31
set fmri(groupmem.31) 1

# Group membership for input 32
set fmri(groupmem.32) 1

# Group membership for input 33
set fmri(groupmem.33) 1

# Group membership for input 34
set fmri(groupmem.34) 1

# Group membership for input 35
set fmri(groupmem.35) 1

# Group membership for input 36
set fmri(groupmem.36) 1

# Group membership for input 37
set fmri(groupmem.37) 1

# Group membership for input 38
set fmri(groupmem.38) 1

# Group membership for input 39
set fmri(groupmem.39) 1

# Group membership for input 40
set fmri(groupmem.40) 1

# Contrast & F-tests mode
# real : control real EVs
# orig : control original EVs
set fmri(con_mode_old) real
set fmri(con_mode) real

# Display images for contrast_real 1
set fmri(conpic_real.1) 1

# Title for contrast_real 1
set fmri(conname_real.1) "run01"

# Real contrast_real vector 1 element 1
set fmri(con_real1.1) 1.0

# Real contrast_real vector 1 element 2
set fmri(con_real1.2) 0.0

# Real contrast_real vector 1 element 3
set fmri(con_real1.3) 0.0

# Display images for contrast_real 2
set fmri(conpic_real.2) 1

# Title for contrast_real 2
set fmri(conname_real.2) "run04"

# Real contrast_real vector 2 element 1
set fmri(con_real2.1) 0.0

# Real contrast_real vector 2 element 2
set fmri(con_real2.2) 1.0

# Real contrast_real vector 2 element 3
set fmri(con_real2.3) 0.0

# Display images for contrast_real 3
set fmri(conpic_real.3) 1

# Title for contrast_real 3
set fmri(conname_real.3) "run01-run04"

# Real contrast_real vector 3 element 1
set fmri(con_real3.1) 1.0

# Real contrast_real vector 3 element 2
set fmri(con_real3.2) -1.0

# Real contrast_real vector 3 element 3
set fmri(con_real3.3) 0.0

# Display images for contrast_real 4
set fmri(conpic_real.4) 1

# Title for contrast_real 4
set fmri(conname_real.4) "run04-run01"

# Real contrast_real vector 4 element 1
set fmri(con_real4.1) -1.0

# Real contrast_real vector 4 element 2
set fmri(con_real4.2) 1.0

# Real contrast_real vector 4 element 3
set fmri(con_real4.3) 0.0

# Contrast masking - use >0 instead of thresholding?
set fmri(conmask_zerothresh_yn) 0

# Mask real contrast/F-test 1 with real contrast/F-test 2?
set fmri(conmask1_2) 0

# Mask real contrast/F-test 1 with real contrast/F-test 3?
set fmri(conmask1_3) 0

# Mask real contrast/F-test 1 with real contrast/F-test 4?
set fmri(conmask1_4) 0

# Mask real contrast/F-test 2 with real contrast/F-test 1?
set fmri(conmask2_1) 0

# Mask real contrast/F-test 2 with real contrast/F-test 3?
set fmri(conmask2_3) 0

# Mask real contrast/F-test 2 with real contrast/F-test 4?
set fmri(conmask2_4) 0

# Mask real contrast/F-test 3 with real contrast/F-test 1?
set fmri(conmask3_1) 0

# Mask real contrast/F-test 3 with real contrast/F-test 2?
set fmri(conmask3_2) 0

# Mask real contrast/F-test 3 with real contrast/F-test 4?
set fmri(conmask3_4) 0

# Mask real contrast/F-test 4 with real contrast/F-test 1?
set fmri(conmask4_1) 0

# Mask real contrast/F-test 4 with real contrast/F-test 2?
set fmri(conmask4_2) 0

# Mask real contrast/F-test 4 with real contrast/F-test 3?
set fmri(conmask4_3) 0

# Do contrast masking at all?
set fmri(conmask1_1) 0

##########################################################
# Now options that don't appear in the GUI

# Alternative (to BETting) mask image
set fmri(alternative_mask) ""

# Initial structural space registration initialisation transform
set fmri(init_initial_highres) ""

# Structural space registration initialisation transform
set fmri(init_highres) ""

# Standard space registration initialisation transform
set fmri(init_standard) ""

# For full FEAT analysis: overwrite existing .feat output dir?
set fmri(overwrite_yn) 0
"""
    
    second_level_fsf_template_group = r"""
# FEAT version number
set fmri(version) 6.00

# Are we in MELODIC?
set fmri(inmelodic) 0

# Analysis level
# 1 : First-level analysis
# 2 : Higher-level analysis
set fmri(level) 2

# Which stages to run
# 0 : No first-level analysis (registration and/or group stats only)
# 7 : Full first-level analysis
# 1 : Pre-processing
# 2 : Statistics
set fmri(analysis) 2

# Use relative filenames
set fmri(relative_yn) 0

# Balloon help
set fmri(help_yn) 1

# Run Featwatcher
set fmri(featwatcher_yn) 1

# Cleanup first-level standard-space images
set fmri(sscleanup_yn) 0

# Output directory
set fmri(outputdir) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/[insert_participant]/[insert_folder_name]"

# TR(s)
set fmri(tr) 2

# Total volumes
set fmri(npts) 2

# Delete volumes
set fmri(ndelete) 0

# Perfusion tag/control order
set fmri(tagfirst) 1

# Number of first-level analyses
set fmri(multiple) 2

# Higher-level input type
# 1 : Inputs are lower-level FEAT directories
# 2 : Inputs are cope images from FEAT directories
set fmri(inputtype) 1

# Carry out pre-stats processing?
set fmri(filtering_yn) 0

# Brain/background threshold, %
set fmri(brain_thresh) 10

# Critical z for design efficiency calculation
set fmri(critical_z) 5.3

# Noise level
set fmri(noise) 0.66

# Noise AR(1)
set fmri(noisear) 0.34

# Motion correction
# 0 : None
# 1 : MCFLIRT
set fmri(mc) 1

# Spin-history (currently obsolete)
set fmri(sh_yn) 0

# B0 fieldmap unwarping?
set fmri(regunwarp_yn) 0

# GDC Test
set fmri(gdc) ""

# EPI dwell time (ms)
set fmri(dwell) 0.0

# EPI TE (ms)
set fmri(te) 0.0

# % Signal loss threshold
set fmri(signallossthresh) 10

# Unwarp direction
set fmri(unwarp_dir) y-

# Slice timing correction
# 0 : None
# 1 : Regular up (0, 1, 2, 3, ...)
# 2 : Regular down
# 3 : Use slice order file
# 4 : Use slice timings file
# 5 : Interleaved (0, 2, 4 ... 1, 3, 5 ... )
set fmri(st) 0

# Slice timings file
set fmri(st_file) ""

# BET brain extraction
set fmri(bet_yn) 1

# Spatial smoothing FWHM (mm)
set fmri(smooth) 5

# Intensity normalization
set fmri(norm_yn) 0

# Perfusion subtraction
set fmri(perfsub_yn) 0

# Highpass temporal filtering
set fmri(temphp_yn) 1

# Lowpass temporal filtering
set fmri(templp_yn) 0

# MELODIC ICA data exploration
set fmri(melodic_yn) 0

# Carry out main stats?
set fmri(stats_yn) 1

# Carry out prewhitening?
set fmri(prewhiten_yn) 1

# Add motion parameters to model
# 0 : No
# 1 : Yes
set fmri(motionevs) 0
set fmri(motionevsbeta) ""
set fmri(scriptevsbeta) ""

# Robust outlier detection in FLAME?
set fmri(robust_yn) 0

# Higher-level modelling
# 3 : Fixed effects
# 0 : Mixed Effects: Simple OLS
# 2 : Mixed Effects: FLAME 1
# 1 : Mixed Effects: FLAME 1+2
set fmri(mixed_yn) 3

# Higher-level permutations
set fmri(randomisePermutations) 5000

# Number of EVs
set fmri(evs_orig) 2
set fmri(evs_real) 2
set fmri(evs_vox) 0

# Number of contrasts
set fmri(ncon_orig) 4
set fmri(ncon_real) 4

# Number of F-tests
set fmri(nftests_orig) 0
set fmri(nftests_real) 0

# Add constant column to design matrix? (obsolete)
set fmri(constcol) 0

# Carry out post-stats steps?
set fmri(poststats_yn) 0

# Pre-threshold masking?
set fmri(threshmask) ""

# Thresholding
# 0 : None
# 1 : Uncorrected
# 2 : Voxel
# 3 : Cluster
set fmri(thresh) 1

# P threshold
set fmri(prob_thresh) 0.05

# Z threshold
set fmri(z_thresh) 2.3

# Z min/max for colour rendering
# 0 : Use actual Z min/max
# 1 : Use preset Z min/max
set fmri(zdisplay) 0

# Z min in colour rendering
set fmri(zmin) 2

# Z max in colour rendering
set fmri(zmax) 8

# Colour rendering type
# 0 : Solid blobs
# 1 : Transparent blobs
set fmri(rendertype) 1

# Background image for higher-level stats overlays
# 1 : Mean highres
# 2 : First highres
# 3 : Mean functional
# 4 : First functional
# 5 : Standard space template
set fmri(bgimage) 1

# Create time series plots
set fmri(tsplot_yn) 0

# Registration to initial structural
set fmri(reginitial_highres_yn) 0

# Search space for registration to initial structural
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(reginitial_highres_search) 90

# Degrees of Freedom for registration to initial structural
set fmri(reginitial_highres_dof) 3

# Registration to main structural
set fmri(reghighres_yn) 0

# Search space for registration to main structural
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(reghighres_search) 90

# Degrees of Freedom for registration to main structural
set fmri(reghighres_dof) BBR

# Registration to standard image?
set fmri(regstandard_yn) 0

# Use alternate reference images?
set fmri(alternateReference_yn) 0

# Standard image
set fmri(regstandard) "/usr/local/fsl/data/standard/MNI152_T1_2mm_brain"

# Search space for registration to standard space
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(regstandard_search) 90

# Degrees of Freedom for registration to standard space
set fmri(regstandard_dof) 12

# Do nonlinear registration from structural to standard space?
set fmri(regstandard_nonlinear_yn) 0

# Control nonlinear warp field resolution
set fmri(regstandard_nonlinear_warpres) 10

# High pass filter cutoff
set fmri(paradigm_hp) 100

# Total voxels
set fmri(totalVoxels) 227227350


# Number of lower-level copes feeding into higher-level analysis
set fmri(ncopeinputs) 4

# Use lower-level cope 1 for higher-level analysis
set fmri(copeinput.1) 1

# Use lower-level cope 2 for higher-level analysis
set fmri(copeinput.2) 1

# Use lower-level cope 3 for higher-level analysis
set fmri(copeinput.3) 1

# Use lower-level cope 4 for higher-level analysis
set fmri(copeinput.4) 1

# 4D AVW data or FEAT directory (1)
set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/[insert_participant]/run-01.feat"

# 4D AVW data or FEAT directory (2)
set feat_files(2) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/[insert_participant]/run-04.feat"

# Add confound EVs text file
set fmri(confoundevs) 0

# EV 1 title
set fmri(evtitle1) "run01"

# Basic waveform shape (EV 1)
# 0 : Square
# 1 : Sinusoid
# 2 : Custom (1 entry per volume)
# 3 : Custom (3 column format)
# 4 : Interaction
# 10 : Empty (all zeros)
set fmri(shape1) 2

# Convolution (EV 1)
# 0 : None
# 1 : Gaussian
# 2 : Gamma
# 3 : Double-Gamma HRF
# 4 : Gamma basis functions
# 5 : Sine basis functions
# 6 : FIR basis functions
# 8 : Alternate Double-Gamma
set fmri(convolve1) 0

# Convolve phase (EV 1)
set fmri(convolve_phase1) 0

# Apply temporal filtering (EV 1)
set fmri(tempfilt_yn1) 0

# Add temporal derivative (EV 1)
set fmri(deriv_yn1) 0

# Custom EV file (EV 1)
set fmri(custom1) "dummy"

# Orthogonalise EV 1 wrt EV 0
set fmri(ortho1.0) 1

# Orthogonalise EV 1 wrt EV 1
set fmri(ortho1.1) 0

# Orthogonalise EV 1 wrt EV 2
set fmri(ortho1.2) 0

# Higher-level EV value for EV 1 and input 1
set fmri(evg1.1) 1.0

# Higher-level EV value for EV 1 and input 2
set fmri(evg2.1) 0.0

# EV 2 title
set fmri(evtitle2) "run04"

# Basic waveform shape (EV 2)
# 0 : Square
# 1 : Sinusoid
# 2 : Custom (1 entry per volume)
# 3 : Custom (3 column format)
# 4 : Interaction
# 10 : Empty (all zeros)
set fmri(shape2) 2

# Convolution (EV 2)
# 0 : None
# 1 : Gaussian
# 2 : Gamma
# 3 : Double-Gamma HRF
# 4 : Gamma basis functions
# 5 : Sine basis functions
# 6 : FIR basis functions
# 8 : Alternate Double-Gamma
set fmri(convolve2) 0

# Convolve phase (EV 2)
set fmri(convolve_phase2) 0

# Apply temporal filtering (EV 2)
set fmri(tempfilt_yn2) 0

# Add temporal derivative (EV 2)
set fmri(deriv_yn2) 0

# Custom EV file (EV 2)
set fmri(custom2) "dummy"

# Orthogonalise EV 2 wrt EV 0
set fmri(ortho2.0) 1

# Orthogonalise EV 2 wrt EV 1
set fmri(ortho2.1) 0

# Orthogonalise EV 2 wrt EV 2
set fmri(ortho2.2) 0

# Higher-level EV value for EV 2 and input 1
set fmri(evg1.2) 0.0

# Higher-level EV value for EV 2 and input 2
set fmri(evg2.2) 1.0

# Setup Orthogonalisation at higher level?
set fmri(level2orth) 1

# Group membership for input 1
set fmri(groupmem.1) 1

# Group membership for input 2
set fmri(groupmem.2) 1

# Contrast & F-tests mode
# real : control real EVs
# orig : control original EVs
set fmri(con_mode_old) real
set fmri(con_mode) real

# Display images for contrast_real 1
set fmri(conpic_real.1) 1

# Title for contrast_real 1
set fmri(conname_real.1) "run01"

# Real contrast_real vector 1 element 1
set fmri(con_real1.1) 1

# Real contrast_real vector 1 element 2
set fmri(con_real1.2) 0.0

# Display images for contrast_real 2
set fmri(conpic_real.2) 1

# Title for contrast_real 2
set fmri(conname_real.2) "run04"

# Real contrast_real vector 2 element 1
set fmri(con_real2.1) 0.0

# Real contrast_real vector 2 element 2
set fmri(con_real2.2) 1.0

# Display images for contrast_real 3
set fmri(conpic_real.3) 1

# Title for contrast_real 3
set fmri(conname_real.3) "run01-run04"

# Real contrast_real vector 3 element 1
set fmri(con_real3.1) 1.0

# Real contrast_real vector 3 element 2
set fmri(con_real3.2) -1.0

# Display images for contrast_real 4
set fmri(conpic_real.4) 1

# Title for contrast_real 4
set fmri(conname_real.4) "run04-run01"

# Real contrast_real vector 4 element 1
set fmri(con_real4.1) -1.0

# Real contrast_real vector 4 element 2
set fmri(con_real4.2) 1.0

# Contrast masking - use >0 instead of thresholding?
set fmri(conmask_zerothresh_yn) 0

# Mask real contrast/F-test 1 with real contrast/F-test 2?
set fmri(conmask1_2) 0

# Mask real contrast/F-test 1 with real contrast/F-test 3?
set fmri(conmask1_3) 0

# Mask real contrast/F-test 1 with real contrast/F-test 4?
set fmri(conmask1_4) 0

# Mask real contrast/F-test 2 with real contrast/F-test 1?
set fmri(conmask2_1) 0

# Mask real contrast/F-test 2 with real contrast/F-test 3?
set fmri(conmask2_3) 0

# Mask real contrast/F-test 2 with real contrast/F-test 4?
set fmri(conmask2_4) 0

# Mask real contrast/F-test 3 with real contrast/F-test 1?
set fmri(conmask3_1) 0

# Mask real contrast/F-test 3 with real contrast/F-test 2?
set fmri(conmask3_2) 0

# Mask real contrast/F-test 3 with real contrast/F-test 4?
set fmri(conmask3_4) 0

# Mask real contrast/F-test 4 with real contrast/F-test 1?
set fmri(conmask4_1) 0

# Mask real contrast/F-test 4 with real contrast/F-test 2?
set fmri(conmask4_2) 0

# Mask real contrast/F-test 4 with real contrast/F-test 3?
set fmri(conmask4_3) 0

# Do contrast masking at all?
set fmri(conmask1_1) 0

##########################################################
# Now options that don't appear in the GUI

# Alternative (to BETting) mask image
set fmri(alternative_mask) ""

# Initial structural space registration initialisation transform
set fmri(init_initial_highres) ""

# Structural space registration initialisation transform
set fmri(init_highres) ""

# Standard space registration initialisation transform
set fmri(init_standard) ""

# For full FEAT analysis: overwrite existing .feat output dir?
set fmri(overwrite_yn) 0
"""

    
    group_diffs = logged_input("Select group comparison [1] or no group comparison [2] in the second-level: ")
    if group_diffs == '2':
        group_diffs_fsf_label = 'nogroup'
        pre_thresh_masking = logged_input("Select pre-threshold masking [1] or whole-brain analysis [2] for second-level: ")
        if pre_thresh_masking == '1':
            pre_thresh_masking_fsf_label = 'scc'
        elif pre_thresh_masking == '2':
            pre_thresh_masking_fsf_label = 'wholebrain'
        else:
            print('Invalid response. Please start again.')
            sys.exit()
        liberal_thresholding = logged_input("Select liberal thresholding (0.1) [1] or standard thresholding (0.05) [2] for second-level: ")
        if liberal_thresholding == '1':
            liberal_thresholding_fsf_label = 'liberal'
        elif liberal_thresholding == '2':
            liberal_thresholding_fsf_label = 'stringent'
        else:
            print('Invalid response. Please start again.')
            sys.exit()
        cluster_thresholding = logged_input("Select cluster-corrected thresholding [1] or no correction [2] for second-level: ")
        if cluster_thresholding == '1':
            cluster_thresholding_fsf_label = 'cluster'
        elif cluster_thresholding == '2':
            cluster_thresholding_fsf_label = 'uncorrected'
        else: 
            print('Invalid response. Please start again.')
            sys.exit()
        second_level_fsf_template_path_nogroup = f'analysis/fmri_analysis/analysis_1/second_level/shared/second_level_fsf_template_{pre_thresh_masking_fsf_label}_{liberal_thresholding_fsf_label}_{cluster_thresholding_fsf_label}_{group_diffs_fsf_label}.fsf'
        with open(second_level_fsf_template_path_nogroup, 'w') as f:
            f.write(second_level_fsf_template_nogroup)
        with open(second_level_fsf_template_path_nogroup, 'r') as file:
            fsf_data = file.readlines()
        for i, line in enumerate(fsf_data):
            if "set fmri(outputdir)" in line:
                feat_folder = f'second_level_{pre_thresh_masking_fsf_label}_{liberal_thresholding_fsf_label}_{cluster_thresholding_fsf_label}_{group_diffs_fsf_label}.gfeat'
                fsf_data[i] = f'set fmri(outputdir) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/shared/second_level_{pre_thresh_masking_fsf_label}_{liberal_thresholding_fsf_label}_{cluster_thresholding_fsf_label}_{group_diffs_fsf_label}"\n'
            elif "set fmri(threshmask)" in line:
                if pre_thresh_masking == '1':
                    fsf_data[i] = 'set fmri(threshmask) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/data/roi/SCCsphere8_bin_2mm_func.nii.gz"\n'
            elif "set fmri(thresh)" in line:
                if cluster_thresholding == '1':
                    fsf_data[i] = 'set fmri(thresh) 3\n'
            elif "set fmri(prob_thresh)" in line:
                if liberal_thresholding == '1':
                    fsf_data[i] = 'set fmri(prob_thresh) 0.1\n'
        second_level_fsf = f'analysis/fmri_analysis/analysis_1/second_level/shared/second_level_fsf_{pre_thresh_masking_fsf_label}_{liberal_thresholding_fsf_label}_{cluster_thresholding_fsf_label}_{group_diffs_fsf_label}.fsf'
        with open(second_level_fsf, 'w') as file:
            file.writelines(fsf_data)
        os.remove(second_level_fsf_template_path_nogroup)
        print(f'fsf file for second-level GLM generated and saved to second_level/shared folder.\nGLM details: {pre_thresh_masking_fsf_label}, {liberal_thresholding_fsf_label}, {cluster_thresholding_fsf_label}, {group_diffs_fsf_label}')
    
    elif group_diffs == '1':
        group_diffs_fsf_label = 'group'
        second_level_fsf_template_path_group = f'analysis/fmri_analysis/analysis_1/second_level/shared/second_level_fsf_template_{group_diffs_fsf_label}.fsf'
        with open(second_level_fsf_template_path_group, 'w') as f:
            f.write(second_level_fsf_template_group)
        second_level_group_fsfs = []
        for p_id in participants:
            p_id_stripped = p_id.replace('P', '')
            for run in runs: 
                with open(second_level_fsf_template_path_group, 'r') as file:
                    fsf_data = file.readlines()
                for i, line in enumerate(fsf_data):
                    if "set fmri(outputdir)" in line:
                        fsf_data[i] = f'set fmri(outputdir) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-{p_id_stripped}/sub-{p_id_stripped}_{group_diffs_fsf_label}"\n'
                    elif "set feat_files(1)" in line:
                        fsf_data[i] = f'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/run-01.feat"\n'
                    elif "set feat_files(2)" in line:
                        fsf_data[i] = f'set feat_files(2) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/run-04.feat"\n'
                second_level_group_fsf = f'analysis/fmri_analysis/analysis_1/second_level/sub-{p_id_stripped}/second_level_{group_diffs_fsf_label}_fsf.fsf'
                second_level_group_fsfs.append(second_level_group_fsf)
                with open(second_level_group_fsf, 'w') as file:
                    file.writelines(fsf_data)
        print(f'fsf files for second-level group analysis GLM generated.')

    # Step 8: Run second-level GLM [ANALYSIS 1].
    print("\n###### STEP 8: RUN SECOND-LEVEL GLM [ANALYSIS 1] ######")
    for p_id in participants:
        p_id_stripped = p_id.replace('P', '')
        for run in runs:
            reg_folder = f'analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/{run}.feat/reg'
            os.makedirs(reg_folder, exist_ok=True)
            init_mat_path = '/usr/local/fsl/etc/flirtsch/ident.mat'
            example_func2standard_mat_path = f'analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/{run}.feat/reg/example_func2standard.mat'
            if not os.path.exists(example_func2standard_mat_path):
                shutil.copy(init_mat_path, example_func2standard_mat_path)
            mean_func_path = f'analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/{run}.feat/mean_func.nii.gz'
            standard_path = f'analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/{run}.feat/reg/standard.nii.gz'
            if not os.path.exists(standard_path):
                shutil.copy(mean_func_path, standard_path)
    if group_diffs == '2':
        if not os.path.isdir(f'analysis/fmri_analysis/analysis_1/second_level/shared/{feat_folder}'):
            print('Running second-level GLM (no group comparison) with above parameters...')
            subprocess.run(['feat', second_level_fsf])
            os.remove(second_level_fsf)
            print("Second-level completed. This is the last analysis level, so look at mask.nii.gz overlayed onto bg_image.nii.gz and look at filtered_func_data.nii.gz for QA.")
            report_log_path = f'analysis/fmri_analysis/analysis_1/second_level/shared/{feat_folder}/report_log.html'
            with open(report_log_path, 'r') as file:
                content = file.read()
            if re.search('error', content, re.IGNORECASE):
                print(f"Error found in report_log.html. Investigation required.")
        else:
            print('Second-level already run. Skipping process.')
    elif group_diffs == '1':
        for fsf in second_level_group_fsfs:
            match = re.search(r'sub-(\d{3})', fsf)
            participant_number = match.group(1)
            if not os.path.exists(f'analysis/fmri_analysis/analysis_1/second_level/sub-{participant_number}/sub-{participant_number}_{group_diffs_fsf_label}.gfeat'):
                print(f'Running second-level group analysis GLM for sub-{participant_number}...')
                subprocess.run(['feat', fsf])
                report_log_path = f'analysis/fmri_analysis/analysis_1/second_level/sub-{participant_number}/sub-{participant_number}_{group_diffs_fsf_label}.gfeat/report_log.html'
                with open(report_log_path, 'r') as file:
                    content = file.read()
                if re.search('error', content, re.IGNORECASE):
                    print(f"Error found in report_log.html. Investigation required.")
            else:
                ('Second-level group analysis GLMs already run. Skipping process.')
        
   # Step 9: Generate third-level fsf files [ANALYSIS 1].
    print("\n###### STEP 9: GENERATE THIRD-LEVEL FSF FILES [ANALYSIS 1] ######")   
    third_level_fsf_template = r"""
# FEAT version number
set fmri(version) 6.00

# Are we in MELODIC?
set fmri(inmelodic) 0

# Analysis level
# 1 : First-level analysis
# 2 : Higher-level analysis
set fmri(level) 2

# Which stages to run
# 0 : No first-level analysis (registration and/or group stats only)
# 7 : Full first-level analysis
# 1 : Pre-processing
# 2 : Statistics
set fmri(analysis) 2

# Use relative filenames
set fmri(relative_yn) 0

# Balloon help
set fmri(help_yn) 1

# Run Featwatcher
set fmri(featwatcher_yn) 1

# Cleanup first-level standard-space images
set fmri(sscleanup_yn) 0

# Output directory
set fmri(outputdir) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/third_level/shared/[insert_folder_name]"

# TR(s)
set fmri(tr) 2

# Total volumes
set fmri(npts) 20

# Delete volumes
set fmri(ndelete) 0

# Perfusion tag/control order
set fmri(tagfirst) 1

# Number of first-level analyses
set fmri(multiple) 20

# Higher-level input type
# 1 : Inputs are lower-level FEAT directories
# 2 : Inputs are cope images from FEAT directories
set fmri(inputtype) 2

# Carry out pre-stats processing?
set fmri(filtering_yn) 0

# Brain/background threshold, %
set fmri(brain_thresh) 10

# Critical z for design efficiency calculation
set fmri(critical_z) 5.3

# Noise level
set fmri(noise) 0.66

# Noise AR(1)
set fmri(noisear) 0.34

# Motion correction
# 0 : None
# 1 : MCFLIRT
set fmri(mc) 1

# Spin-history (currently obsolete)
set fmri(sh_yn) 0

# B0 fieldmap unwarping?
set fmri(regunwarp_yn) 0

# GDC Test
set fmri(gdc) ""

# EPI dwell time (ms)
set fmri(dwell) 0.0

# EPI TE (ms)
set fmri(te) 0.0

# % Signal loss threshold
set fmri(signallossthresh) 10

# Unwarp direction
set fmri(unwarp_dir) y-

# Slice timing correction
# 0 : None
# 1 : Regular up (0, 1, 2, 3, ...)
# 2 : Regular down
# 3 : Use slice order file
# 4 : Use slice timings file
# 5 : Interleaved (0, 2, 4 ... 1, 3, 5 ... )
set fmri(st) 0

# Slice timings file
set fmri(st_file) ""

# BET brain extraction
set fmri(bet_yn) 1

# Spatial smoothing FWHM (mm)
set fmri(smooth) 5

# Intensity normalization
set fmri(norm_yn) 0

# Perfusion subtraction
set fmri(perfsub_yn) 0

# Highpass temporal filtering
set fmri(temphp_yn) 1

# Lowpass temporal filtering
set fmri(templp_yn) 0

# MELODIC ICA data exploration
set fmri(melodic_yn) 0

# Carry out main stats?
set fmri(stats_yn) 1

# Carry out prewhitening?
set fmri(prewhiten_yn) 1

# Add motion parameters to model
# 0 : No
# 1 : Yes
set fmri(motionevs) 0
set fmri(motionevsbeta) ""
set fmri(scriptevsbeta) ""

# Robust outlier detection in FLAME?
set fmri(robust_yn) 0

# Higher-level modelling
# 3 : Fixed effects
# 0 : Mixed Effects: Simple OLS
# 2 : Mixed Effects: FLAME 1
# 1 : Mixed Effects: FLAME 1+2
set fmri(mixed_yn) 1

# Higher-level permutations
set fmri(randomisePermutations) 5000

# Number of EVs
set fmri(evs_orig) 2
set fmri(evs_real) 2
set fmri(evs_vox) 0

# Number of contrasts
set fmri(ncon_orig) 1
set fmri(ncon_real) 4

# Number of F-tests
set fmri(nftests_orig) 0
set fmri(nftests_real) 0

# Add constant column to design matrix? (obsolete)
set fmri(constcol) 0

# Carry out post-stats steps?
set fmri(poststats_yn) 1

# Pre-threshold masking?
set fmri(threshmask) ""

# Thresholding
# 0 : None
# 1 : Uncorrected
# 2 : Voxel
# 3 : Cluster
set fmri(thresh) 1

# P threshold
set fmri(prob_thresh) 0.05

# Z threshold
set fmri(z_thresh) 2.3

# Z min/max for colour rendering
# 0 : Use actual Z min/max
# 1 : Use preset Z min/max
set fmri(zdisplay) 0

# Z min in colour rendering
set fmri(zmin) 2

# Z max in colour rendering
set fmri(zmax) 8

# Colour rendering type
# 0 : Solid blobs
# 1 : Transparent blobs
set fmri(rendertype) 1

# Background image for higher-level stats overlays
# 1 : Mean highres
# 2 : First highres
# 3 : Mean functional
# 4 : First functional
# 5 : Standard space template
set fmri(bgimage) 1

# Create time series plots
set fmri(tsplot_yn) 1

# Registration to initial structural
set fmri(reginitial_highres_yn) 0

# Search space for registration to initial structural
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(reginitial_highres_search) 90

# Degrees of Freedom for registration to initial structural
set fmri(reginitial_highres_dof) 3

# Registration to main structural
set fmri(reghighres_yn) 0

# Search space for registration to main structural
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(reghighres_search) 90

# Degrees of Freedom for registration to main structural
set fmri(reghighres_dof) BBR

# Registration to standard image?
set fmri(regstandard_yn) 1

# Use alternate reference images?
set fmri(alternateReference_yn) 0

# Standard image
set fmri(regstandard) "/usr/local/fsl/data/standard/MNI152_T1_2mm_brain"

# Search space for registration to standard space
# 0   : No search
# 90  : Normal search
# 180 : Full search
set fmri(regstandard_search) 90

# Degrees of Freedom for registration to standard space
set fmri(regstandard_dof) 12

# Do nonlinear registration from structural to standard space?
set fmri(regstandard_nonlinear_yn) 0

# Control nonlinear warp field resolution
set fmri(regstandard_nonlinear_warpres) 10

# High pass filter cutoff
set fmri(paradigm_hp) 100

# Total voxels
set fmri(totalVoxels) 1082035


# Number of lower-level copes feeding into higher-level analysis
set fmri(ncopeinputs) 0

# 4D AVW data or FEAT directory (1)
set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (2)
set feat_files(2) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (3)
set feat_files(3) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (4)
set feat_files(4) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (5)
set feat_files(5) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (6)
set feat_files(6) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (7)
set feat_files(7) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (8)
set feat_files(8) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (9)
set feat_files(9) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (10)
set feat_files(10) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (11)
set feat_files(11) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (12)
set feat_files(12) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (13)
set feat_files(13) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (14)
set feat_files(14) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (15)
set feat_files(15) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (16)
set feat_files(16) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (17)
set feat_files(17) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (18)
set feat_files(18) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (19)
set feat_files(19) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# 4D AVW data or FEAT directory (20)
set feat_files(20) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/[insert_folder_name]/[insert_first_level_cope].feat/stats/[insert_second_level_cope]"

# Add confound EVs text file
set fmri(confoundevs) 0

# EV 1 title
set fmri(evtitle1) "groupa"

# Basic waveform shape (EV 1)
# 0 : Square
# 1 : Sinusoid
# 2 : Custom (1 entry per volume)
# 3 : Custom (3 column format)
# 4 : Interaction
# 10 : Empty (all zeros)
set fmri(shape1) 2

# Convolution (EV 1)
# 0 : None
# 1 : Gaussian
# 2 : Gamma
# 3 : Double-Gamma HRF
# 4 : Gamma basis functions
# 5 : Sine basis functions
# 6 : FIR basis functions
# 8 : Alternate Double-Gamma
set fmri(convolve1) 0

# Convolve phase (EV 1)
set fmri(convolve_phase1) 0

# Apply temporal filtering (EV 1)
set fmri(tempfilt_yn1) 0

# Add temporal derivative (EV 1)
set fmri(deriv_yn1) 0

# Custom EV file (EV 1)
set fmri(custom1) "dummy"

# Orthogonalise EV 1 wrt EV 0
set fmri(ortho1.0) 0

# Orthogonalise EV 1 wrt EV 1
set fmri(ortho1.1) 0

# Orthogonalise EV 1 wrt EV 2
set fmri(ortho1.2) 0

# Higher-level EV value for EV 1 and input 1
set fmri(evg1.1) 1.0

# Higher-level EV value for EV 1 and input 2
set fmri(evg2.1) 1.0

# Higher-level EV value for EV 1 and input 3
set fmri(evg3.1) 0.0

# Higher-level EV value for EV 1 and input 4
set fmri(evg4.1) 0.0

# Higher-level EV value for EV 1 and input 5
set fmri(evg5.1) 0.0

# Higher-level EV value for EV 1 and input 6
set fmri(evg6.1) 0

# Higher-level EV value for EV 1 and input 7
set fmri(evg7.1) 0.0

# Higher-level EV value for EV 1 and input 8
set fmri(evg8.1) 0

# Higher-level EV value for EV 1 and input 9
set fmri(evg9.1) 1.0

# Higher-level EV value for EV 1 and input 10
set fmri(evg10.1) 0

# Higher-level EV value for EV 1 and input 11
set fmri(evg11.1) 1.0

# Higher-level EV value for EV 1 and input 12
set fmri(evg12.1) 1.0

# Higher-level EV value for EV 1 and input 13
set fmri(evg13.1) 0.0

# Higher-level EV value for EV 1 and input 14
set fmri(evg14.1) 1.0

# Higher-level EV value for EV 1 and input 15
set fmri(evg15.1) 1.0

# Higher-level EV value for EV 1 and input 16
set fmri(evg16.1) 1.0

# Higher-level EV value for EV 1 and input 17
set fmri(evg17.1) 0.0

# Higher-level EV value for EV 1 and input 18
set fmri(evg18.1) 0

# Higher-level EV value for EV 1 and input 19
set fmri(evg19.1) 1.0

# Higher-level EV value for EV 1 and input 20
set fmri(evg20.1) 1.0

# EV 2 title
set fmri(evtitle2) "groupb"

# Basic waveform shape (EV 2)
# 0 : Square
# 1 : Sinusoid
# 2 : Custom (1 entry per volume)
# 3 : Custom (3 column format)
# 4 : Interaction
# 10 : Empty (all zeros)
set fmri(shape2) 2

# Convolution (EV 2)
# 0 : None
# 1 : Gaussian
# 2 : Gamma
# 3 : Double-Gamma HRF
# 4 : Gamma basis functions
# 5 : Sine basis functions
# 6 : FIR basis functions
# 8 : Alternate Double-Gamma
set fmri(convolve2) 0

# Convolve phase (EV 2)
set fmri(convolve_phase2) 0

# Apply temporal filtering (EV 2)
set fmri(tempfilt_yn2) 0

# Add temporal derivative (EV 2)
set fmri(deriv_yn2) 0

# Custom EV file (EV 2)
set fmri(custom2) "dummy"

# Orthogonalise EV 2 wrt EV 0
set fmri(ortho2.0) 0

# Orthogonalise EV 2 wrt EV 1
set fmri(ortho2.1) 0

# Orthogonalise EV 2 wrt EV 2
set fmri(ortho2.2) 0

# Higher-level EV value for EV 2 and input 1
set fmri(evg1.2) 0

# Higher-level EV value for EV 2 and input 2
set fmri(evg2.2) 0.0

# Higher-level EV value for EV 2 and input 3
set fmri(evg3.2) 1.0

# Higher-level EV value for EV 2 and input 4
set fmri(evg4.2) 1.0

# Higher-level EV value for EV 2 and input 5
set fmri(evg5.2) 1.0

# Higher-level EV value for EV 2 and input 6
set fmri(evg6.2) 1.0

# Higher-level EV value for EV 2 and input 7
set fmri(evg7.2) 1.0

# Higher-level EV value for EV 2 and input 8
set fmri(evg8.2) 1.0

# Higher-level EV value for EV 2 and input 9
set fmri(evg9.2) 0

# Higher-level EV value for EV 2 and input 10
set fmri(evg10.2) 1.0

# Higher-level EV value for EV 2 and input 11
set fmri(evg11.2) 0

# Higher-level EV value for EV 2 and input 12
set fmri(evg12.2) 0.0

# Higher-level EV value for EV 2 and input 13
set fmri(evg13.2) 1.0

# Higher-level EV value for EV 2 and input 14
set fmri(evg14.2) 0.0

# Higher-level EV value for EV 2 and input 15
set fmri(evg15.2) 0

# Higher-level EV value for EV 2 and input 16
set fmri(evg16.2) 0.0

# Higher-level EV value for EV 2 and input 17
set fmri(evg17.2) 1.0

# Higher-level EV value for EV 2 and input 18
set fmri(evg18.2) 1.0

# Higher-level EV value for EV 2 and input 19
set fmri(evg19.2) 0

# Higher-level EV value for EV 2 and input 20
set fmri(evg20.2) 0.0

# Setup Orthogonalisation at higher level?
set fmri(level2orth) 0

# Group membership for input 1
set fmri(groupmem.1) 1

# Group membership for input 2
set fmri(groupmem.2) 1

# Group membership for input 3
set fmri(groupmem.3) 2

# Group membership for input 4
set fmri(groupmem.4) 2

# Group membership for input 5
set fmri(groupmem.5) 2

# Group membership for input 6
set fmri(groupmem.6) 2

# Group membership for input 7
set fmri(groupmem.7) 2

# Group membership for input 8
set fmri(groupmem.8) 2

# Group membership for input 9
set fmri(groupmem.9) 1

# Group membership for input 10
set fmri(groupmem.10) 2

# Group membership for input 11
set fmri(groupmem.11) 1

# Group membership for input 12
set fmri(groupmem.12) 1

# Group membership for input 13
set fmri(groupmem.13) 2

# Group membership for input 14
set fmri(groupmem.14) 1

# Group membership for input 15
set fmri(groupmem.15) 1

# Group membership for input 16
set fmri(groupmem.16) 1

# Group membership for input 17
set fmri(groupmem.17) 2

# Group membership for input 18
set fmri(groupmem.18) 2

# Group membership for input 19
set fmri(groupmem.19) 1

# Group membership for input 20
set fmri(groupmem.20) 1

# Contrast & F-tests mode
# real : control real EVs
# orig : control original EVs
set fmri(con_mode_old) real
set fmri(con_mode) real

# Display images for contrast_real 1
set fmri(conpic_real.1) 1

# Title for contrast_real 1
set fmri(conname_real.1) "groupa"

# Real contrast_real vector 1 element 1
set fmri(con_real1.1) 1.0

# Real contrast_real vector 1 element 2
set fmri(con_real1.2) 0.0

# Display images for contrast_real 2
set fmri(conpic_real.2) 1

# Title for contrast_real 2
set fmri(conname_real.2) "groupb"

# Real contrast_real vector 2 element 1
set fmri(con_real2.1) 0.0

# Real contrast_real vector 2 element 2
set fmri(con_real2.2) 1.0

# Display images for contrast_real 3
set fmri(conpic_real.3) 1

# Title for contrast_real 3
set fmri(conname_real.3) "groupa-groupb"

# Real contrast_real vector 3 element 1
set fmri(con_real3.1) 1.0

# Real contrast_real vector 3 element 2
set fmri(con_real3.2) -1.0

# Display images for contrast_real 4
set fmri(conpic_real.4) 1

# Title for contrast_real 4
set fmri(conname_real.4) "groupb-groupa"

# Real contrast_real vector 4 element 1
set fmri(con_real4.1) -1.0

# Real contrast_real vector 4 element 2
set fmri(con_real4.2) 1.0

# Contrast masking - use >0 instead of thresholding?
set fmri(conmask_zerothresh_yn) 0

# Mask real contrast/F-test 1 with real contrast/F-test 2?
set fmri(conmask1_2) 0

# Mask real contrast/F-test 1 with real contrast/F-test 3?
set fmri(conmask1_3) 0

# Mask real contrast/F-test 1 with real contrast/F-test 4?
set fmri(conmask1_4) 0

# Mask real contrast/F-test 2 with real contrast/F-test 1?
set fmri(conmask2_1) 0

# Mask real contrast/F-test 2 with real contrast/F-test 3?
set fmri(conmask2_3) 0

# Mask real contrast/F-test 2 with real contrast/F-test 4?
set fmri(conmask2_4) 0

# Mask real contrast/F-test 3 with real contrast/F-test 1?
set fmri(conmask3_1) 0

# Mask real contrast/F-test 3 with real contrast/F-test 2?
set fmri(conmask3_2) 0

# Mask real contrast/F-test 3 with real contrast/F-test 4?
set fmri(conmask3_4) 0

# Mask real contrast/F-test 4 with real contrast/F-test 1?
set fmri(conmask4_1) 0

# Mask real contrast/F-test 4 with real contrast/F-test 2?
set fmri(conmask4_2) 0

# Mask real contrast/F-test 4 with real contrast/F-test 3?
set fmri(conmask4_3) 0

# Do contrast masking at all?
set fmri(conmask1_1) 0

##########################################################
# Now options that don't appear in the GUI

# Alternative (to BETting) mask image
set fmri(alternative_mask) ""

# Initial structural space registration initialisation transform
set fmri(init_initial_highres) ""

# Structural space registration initialisation transform
set fmri(init_highres) ""

# Standard space registration initialisation transform
set fmri(init_standard) ""

# For full FEAT analysis: overwrite existing .feat output dir?
set fmri(overwrite_yn) 0
"""
    
    if group_diffs == '1':
        
        pre_thresh_masking = logged_input("Select pre-threshold masking [1] or whole-brain analysis [2] for second-level: ")
        if pre_thresh_masking == '1':
            pre_thresh_masking_fsf_label = 'scc'
        elif pre_thresh_masking == '2':
            pre_thresh_masking_fsf_label = 'wholebrain'
        else:
            print('Invalid response. Please start again.')
            sys.exit()
        liberal_thresholding = logged_input("Select liberal thresholding (0.1) [1] or standard thresholding (0.05) [2] for second-level: ")
        if liberal_thresholding == '1':
            liberal_thresholding_fsf_label = 'liberal'
        elif liberal_thresholding == '2':
            liberal_thresholding_fsf_label = 'stringent'
        else:
            print('Invalid response. Please start again.')
            sys.exit()
        cluster_thresholding = logged_input("Select cluster-corrected thresholding [1] or no correction [2] for second-level: ")
        if cluster_thresholding == '1':
            cluster_thresholding_fsf_label = 'cluster'
        elif cluster_thresholding == '2':
            cluster_thresholding_fsf_label = 'uncorrected'
        else: 
            print('Invalid response. Please start again.')
            sys.exit()
        
        first_level_contrast_of_interest = logged_input("Select first-level contrast to explore at the third-level (guilt [1], indignation [2], guilt-indignation [3], indignation-guilt [4]): ")
        if first_level_contrast_of_interest == '1':
            first_level_contrast_label = 'guilt'
        elif first_level_contrast_of_interest == '2':
            first_level_contrast_label = 'indig'
        elif first_level_contrast_of_interest == '3':
            first_level_contrast_label = 'guilt-indig'
        elif first_level_contrast_of_interest == '4':
            first_level_contrast_label = 'indig-guilt'
        else:
            print('Invalid response. Please start again.')
            sys.exit()

        second_level_contrast_of_interest = logged_input("Select second-level contrast to explore at the third-level (run01 [1], run04 [2], run01-run04 [3], run04-run01 [4]): ")
        if second_level_contrast_of_interest == '1':
            second_level_contrast_label = 'run01'
        elif second_level_contrast_of_interest == '2':
            second_level_contrast_label = 'run04'
        elif second_level_contrast_of_interest == '3':
            second_level_contrast_label = 'run01-run04'
        elif second_level_contrast_of_interest == '4':
            second_level_contrast_label = 'run04-run01'
        else:
            print('Invalid response. Please start again.')
            sys.exit()    

        third_level_fsf_template_path = f'analysis/fmri_analysis/analysis_1/second_level/shared/third_level_fsf_template_{first_level_contrast_label}_{second_level_contrast_label}_{pre_thresh_masking_fsf_label}_{liberal_thresholding_fsf_label}_{cluster_thresholding_fsf_label}_{group_diffs_fsf_label}.fsf'
        with open(third_level_fsf_template_path, 'w') as f:
            f.write(third_level_fsf_template)
        with open(third_level_fsf_template_path, 'r') as file:
            fsf_data = file.readlines()
        for i, line in enumerate(fsf_data):
            if "set fmri(outputdir)" in line:
                feat_folder = f'third_level_{first_level_contrast_label}_{second_level_contrast_label}_{pre_thresh_masking_fsf_label}_{liberal_thresholding_fsf_label}_{cluster_thresholding_fsf_label}_{group_diffs_fsf_label}.gfeat'
                fsf_data[i] = f'set fmri(outputdir) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/third_level/shared/third_level_{first_level_contrast_label}_{second_level_contrast_label}_{pre_thresh_masking_fsf_label}_{liberal_thresholding_fsf_label}_{cluster_thresholding_fsf_label}_{group_diffs_fsf_label}"\n'
            elif "set fmri(threshmask)" in line:
                if pre_thresh_masking == '1':
                    fsf_data[i] = 'set fmri(threshmask) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/data/roi/SCCsphere8_bin_2mm_func.nii.gz"\n'
            elif "set fmri(thresh)" in line:
                if cluster_thresholding == '1':
                    fsf_data[i] = 'set fmri(thresh) 3\n'
            elif "set fmri(prob_thresh)" in line:
                if liberal_thresholding == '1':
                    fsf_data[i] = 'set fmri(prob_thresh) 0.1\n'
            elif "set feat_files(1)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-004/sub-004_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(2)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-006/sub-006_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(3)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-020/sub-020_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(4)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-030/sub-030_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(5)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-059/sub-059_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(6)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-078/sub-078_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(7)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-093/sub-093_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(8)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-094/sub-094_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(9)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-100/sub-100_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(10)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-107/sub-107_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(11)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-122/sub-122_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(12)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-125/sub-125_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(13)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-127/sub-127_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(14)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-128/sub-128_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(15)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-136/sub-136_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(16)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-145/sub-145_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(17)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-155/sub-155_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(18)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-199/sub-199_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(19)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-215/sub-215_group.gfeat/cope4.feat/stats/cope4"\n'
            elif "set feat_files(20)" in line:
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope1.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope1.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope1.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope1.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope2.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope2.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope2.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope2.feat/stats/cope4"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope3.feat/stats/cope1"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope3.feat/stats/cope2"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope3.feat/stats/cope3"\n'
                if first_level_contrast_label == 'guilt-indig' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope3.feat/stats/cope4"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope4.feat/stats/cope1"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope4.feat/stats/cope2"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run01-run04':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope4.feat/stats/cope3"\n'
                if first_level_contrast_label == 'indig-guilt' and second_level_contrast_label == 'run04-run01':
                    fsf_data[i] = 'set feat_files(1) "/research/cisc2/projects/stone_depnf/Neurofeedback/participant_data/analysis/fmri_analysis/analysis_1/second_level/sub-216/sub-216_group.gfeat/cope4.feat/stats/cope4"\n'
        third_level_fsf = f'analysis/fmri_analysis/analysis_1/second_level/shared/second_level_fsf_{first_level_contrast_label}_{second_level_contrast_label}_{pre_thresh_masking_fsf_label}_{liberal_thresholding_fsf_label}_{cluster_thresholding_fsf_label}_{group_diffs_fsf_label}.fsf'
        with open(third_level_fsf, 'w') as file:
            file.writelines(fsf_data)
        os.remove(third_level_fsf_template_path)
        print(f'fsf file for third-level GLM generated and saved to third_level/shared folder.\nGLM details: {first_level_contrast_label}, {second_level_contrast_label}, {pre_thresh_masking_fsf_label}, {liberal_thresholding_fsf_label}, {cluster_thresholding_fsf_label}, {group_diffs_fsf_label}')
    else:
        print('Third-level analysis not selected to be run. Skipping step.')    
    
    # Step 10: Run third-level GLM [ANALYSIS 1].
    print("\n###### STEP 10: RUN THIRD-LEVEL GLM [ANALYSIS 1] ######")
    if group_diffs == '1':
        if not os.path.isdir(f'analysis/fmri_analysis/analysis_1/third_level/shared/{feat_folder}'):
            print('Running third-level GLM with above parameters...')
            subprocess.run(['feat', third_level_fsf])
            os.remove(third_level_fsf)
            print("Third-level completed. This is the last analysis level, so look at mask.nii.gz overlayed onto bg_image.nii.gz and look at filtered_func_data.nii.gz for QA.")
            report_log_path = f'analysis/fmri_analysis/analysis_1/third_level/shared/{feat_folder}/report_log.html'
            with open(report_log_path, 'r') as file:
                content = file.read()
            if re.search('error', content, re.IGNORECASE):
                print(f"Error found in report_log.html. Investigation required.")
        else:
            print('Third-level already run. Skipping process.')
    else:
        print('Third-level analysis not selected to be run. Skipping step.')

    # Step 8: Perform ANOVA on SCC BOLD Data [Analysis 1].
    print("\n###### STEP 8: PERFORM LMMs ON SCC BOLD DATA [ANALYSIS 1] ######")
    print("Note: The LMMs themselves cannot be performed on server due to complex R interfacing requirements. Please run code instead on local Spyder software.")
    participant_column = []
    run_column = []
    contrast_column = []
    roi_mean_column = []
    contrasts = ['1', '2', '3', '4']
    for p_id in participants:
        p_id_stripped = p_id.replace('P', '')
        for run in runs:
            for contrast in contrasts:
                participant_column.append(f'sub-{p_id_stripped}')
                run_column.append(run)
                contrast_column.append(contrast)
                cope_image_path = f'analysis/fmri_analysis/analysis_1/first_level/sub-{p_id_stripped}/{run}.feat/stats/cope{contrast}.nii.gz'
                result = subprocess.run(['fslmeants', '-i', cope_image_path, '-m', 'data/roi/SCCsphere8_bin_2mm_func.nii.gz'], capture_output=True, text=True)
                roi_mean = float(result.stdout.strip())
                roi_mean_column.append(roi_mean)
    a_participants = ['sub-004', 'sub-006', 'sub-100', 'sub-128', 'sub-122', 'sub-125', 'sub-136', 'sub-145', 'sub-215', 'sub-216']
    b_participants = ['sub-020', 'sub-030', 'sub-059', 'sub-078', 'sub-093', 'sub-094', 'sub-107', 'sub-127', 'sub-155', 'sub-199']
    columns = ['participant', 'intervention', 'run', 'contrast', 'roi_mean']
    roi_mean_df = pd.DataFrame(columns=columns)
    roi_mean_df['participant'] = participant_column
    roi_mean_df['intervention'] = roi_mean_df['participant'].apply(lambda x: 'a' if x in a_participants else ('b' if x in b_participants else 'unknown'))
    roi_mean_df['run'] = run_column
    roi_mean_df['contrast'] = contrast_column
    roi_mean_df['roi_mean'] = roi_mean_column
    roi_mean_df_path = 'analysis/fmri_analysis/analysis_1/plots/roi_mean_df.xlsx'
    roi_mean_df.to_excel(roi_mean_df_path, index=False)
    
    roi_mean_df_1 = roi_mean_df.loc[roi_mean_df['contrast'] == 1]
    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # grdevices = importr('grDevices')
    # clubSandwich = importr('clubSandwich')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(nortest)
    # library(clubSandwich)
    # roi_mean_df_1 <- as.data.frame(roi_mean_df_1)
    # model <- lmer(roi_mean~run*intervention + (1|participant), data = roi_mean_df_1)
    # robust_se <- vcovCR(model, type = "CR0")
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # png(filename="analysis/fmri_analysis/analysis_1/plots/lmm_residuals_qqplot_contrast1.png")
    # qqnorm(residuals_model)
    # qqline(residuals_model, col = 'red')
    # dev.off()
    # print("Robust Standard Errors:")
    # print(robust_se)
    # t_crit_run <- qt(0.975, df = 34)
    # t_crit_intervention <- qt(0.975, df = 18)
    # print(t_crit_run)
    # print(t_crit_intervention)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("Contrast 1 LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # print("Model Summary:")
    # print(summary(model))
    # print("Parameter Estimates:")
    # estimates <- fixef(model)
    # print(estimates)
    # } else {
    # print("Contrast 1 LMM residuals do not meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # print("Model Summary:")
    # print(summary(model))
    # print("Parameter Estimates:")
    # estimates <- fixef(model)
    # print(estimates)
    # }
    # """
    # ro.globalenv['roi_mean_df_1'] = pandas2ri.py2rpy(roi_mean_df_1)
    # result = r(r_script)
    # print(result)

    def std_error(x):
        return np.std(x, ddof=1) / np.sqrt(len(x))
    summary_df_1 = roi_mean_df_1.groupby(['intervention', 'run']).agg(
        roi_mean=('roi_mean', 'mean'),
        std_error=('roi_mean', std_error)).reset_index()
    run01_a_rse = 0.02414807
    run04_a_rse = np.sqrt(0.02414807**2 + 0.02991764**2)
    run01_b_rse = np.sqrt(0.02414807**2 + 0.03718549**2)
    run04_b_rse = np.sqrt(0.02414807**2 + 0.02991764**2 + 0.03718549**2 + 0.10379910**2)
    rse_column = [run01_a_rse, run04_a_rse, run01_b_rse, run04_b_rse]
    summary_df_1['rse'] = rse_column
    contrast_1_plot = (ggplot(summary_df_1, aes(x='intervention', y='roi_mean', fill='run')) +
        geom_bar(stat='identity', position='dodge') + 
        geom_errorbar(aes(ymin='roi_mean - rse', ymax='roi_mean + rse'), position=position_dodge(width=0.9), width=0.2) +
        theme_classic() +
        scale_fill_manual(values=['indianred', 'skyblue']) +
        labs(title="Mean SCC BOLD Values for 'Guilt' Contrast", x='Intervention', y='Mean SCC BOLD Value', fill='Run') +
        scale_y_continuous(expand=(0, 0), limits=[-0.5, 0.5], breaks=[-0.5,-0.4,-0.3,-0.2,-0.1,0,0.1,0.2,0.3,0.4,0.5]) +
        scale_x_discrete(labels={'a': 'A', 'b': 'B'}) +
        geom_hline(yintercept=0, linetype='solid', color='black', size=0.5))
    contrast_1_plot.save('analysis/fmri_analysis/analysis_1/plots/contrast_1_plot.png')

    roi_mean_df_2 = roi_mean_df.loc[roi_mean_df['contrast'] == 2]
    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # grdevices = importr('grDevices')
    # clubSandwich = importr('clubSandwich')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # roi_mean_df_2 <- as.data.frame(roi_mean_df_2)
    # model <- lmer(roi_mean~run*intervention + (1|participant), data = roi_mean_df_2)
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # png(filename="analysis/fmri_analysis/analysis_1/plots/lmm_residuals_qqplot_contrast2.png")
    # qqnorm(residuals_model)
    # qqline(residuals_model, col = 'red')
    # dev.off()
    # if (shapiro_test_result$p.value > 0.05) {
    # print("Contrast 2 LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # } else {
    # print("Contrast 2 LMM residuals do not meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # }
    # """
    # ro.globalenv['roi_mean_df_2'] = pandas2ri.py2rpy(roi_mean_df_2)
    # result = r(r_script)
    # print(result)
    def std_error(x):
        return np.std(x, ddof=1) / np.sqrt(len(x))
    summary_df_2 = roi_mean_df_2.groupby(['intervention', 'run']).agg(
        roi_mean=('roi_mean', 'mean'),
        std_error=('roi_mean', std_error)).reset_index()
    contrast_2_plot = (ggplot(summary_df_2, aes(x='intervention', y='roi_mean', fill='run')) +
        geom_bar(stat='identity', position='dodge') + 
        geom_errorbar(aes(ymin='roi_mean - std_error', ymax='roi_mean + std_error'), position=position_dodge(width=0.9), width=0.2) +
        theme_classic() +
        scale_fill_manual(values=['indianred', 'skyblue']) +
        labs(title="Mean SCC BOLD Values for 'Indignation' Contrast", x='Intervention', y='Mean SCC BOLD Value', fill='Run') +
        scale_y_continuous(expand=(0, 0), limits=[-0.5, 0.5], breaks=[-0.5,-0.4,-0.3,-0.2,-0.1,0,0.1,0.2,0.3,0.4,0.5]) +
        scale_x_discrete(labels={'a': 'A', 'b': 'B'}) +
        geom_hline(yintercept=0, linetype='solid', color='black', size=0.5))
    contrast_2_plot.save('analysis/fmri_analysis/analysis_1/plots/contrast_2_plot.png')

    roi_mean_df_3 = roi_mean_df.loc[roi_mean_df['contrast'] == 3]
    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # grdevices = importr('grDevices')
    # clubSandwich = importr('clubSandwich')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(nortest)
    # library(clubSandwich)
    # roi_mean_df_3 <- as.data.frame(roi_mean_df_3)
    # model <- lmer(roi_mean~run*intervention + (1|participant), data = roi_mean_df_3)
    # robust_se <- vcovCR(model, type = "CR0")
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # png(filename="analysis/fmri_analysis/analysis_1/plots/lmm_residuals_qqplot_contrast3.png")
    # qqnorm(residuals_model)
    # qqline(residuals_model, col = 'red')
    # dev.off()
    # print("Robust Standard Errors:")
    # print(robust_se)
    # t_crit_run <- qt(0.975, df = 34)
    # t_crit_intervention <- qt(0.975, df = 18)
    # print(t_crit_run)
    # print(t_crit_intervention)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("Contrast 3 LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # print("Model Summary:")
    # print(summary(model))
    # print("Parameter Estimates:")
    # estimates <- fixef(model)
    # print(estimates)
    # } else {
    # print("Contrast 3 LMM residuals do not meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # print("Model Summary:")
    # print(summary(model))
    # print("Parameter Estimates:")
    # estimates <- fixef(model)
    # print(estimates)
    # }
    # """
    # ro.globalenv['roi_mean_df_3'] = pandas2ri.py2rpy(roi_mean_df_3)
    # result = r(r_script)
    # print(result)
    def std_error(x):
        return np.std(x, ddof=1) / np.sqrt(len(x))
    summary_df_3 = roi_mean_df_3.groupby(['intervention', 'run']).agg(
        roi_mean=('roi_mean', 'mean'),
        std_error=('roi_mean', std_error)).reset_index()
    run01_a_rse = 0.03468101
    run04_a_rse = np.sqrt(0.03468101**2 + 0.05342738**2)
    run01_b_rse = np.sqrt(0.03468101**2 + 0.03849149**2)
    run04_b_rse = np.sqrt(0.03468101**2 + 0.05342738**2 + 0.03849149**2 + 0.10126250**2)
    rse_column = [run01_a_rse, run04_a_rse, run01_b_rse, run04_b_rse]
    summary_df_3['rse'] = rse_column
    contrast_3_plot = (ggplot(summary_df_3, aes(x='intervention', y='roi_mean', fill='run')) +
        geom_bar(stat='identity', position='dodge') + 
        geom_errorbar(aes(ymin='roi_mean - rse', ymax='roi_mean + rse'), position=position_dodge(width=0.9), width=0.2) +
        theme_classic() +
        scale_fill_manual(values=['indianred', 'skyblue']) +
        labs(title="Mean SCC BOLD Values for 'Guilt-Indig' Contrast", x='Intervention', y='Mean SCC BOLD Value', fill='Run') +
        scale_y_continuous(expand=(0, 0), limits=[-0.5, 0.5], breaks=[-0.5,-0.4,-0.3,-0.2,-0.1,0,0.1,0.2,0.3,0.4,0.5]) +
        scale_x_discrete(labels={'a': 'A', 'b': 'B'}) +
        geom_hline(yintercept=0, linetype='solid', color='black', size=0.5))
    contrast_3_plot.save('analysis/fmri_analysis/analysis_1/plots/contrast_3_plot.png')

    roi_mean_df_4 = roi_mean_df.loc[roi_mean_df['contrast'] == 4]
    # os.environ['R_HOME'] = 'C:/Program Files/R/R-4.4.1'
    # pandas2ri.activate()
    # r = ro.r
    # utils = importr('utils')
    # grdevices = importr('grDevices')
    # clubSandwich = importr('clubSandwich')
    # utils.chooseCRANmirror(ind=1)
    # r_script = """
    # library(lme4)
    # library(lmerTest)
    # library(nortest)
    # library(clubSandwich)
    # roi_mean_df_4 <- as.data.frame(roi_mean_df_4)
    # model <- lmer(roi_mean~run*intervention + (1|participant), data = roi_mean_df_4)
    # robust_se <- vcovCR(model, type = "CR0")
    # residuals_model <- residuals(model)
    # shapiro_test_result <- shapiro.test(residuals_model)
    # print(shapiro_test_result)
    # png(filename="analysis/fmri_analysis/analysis_1/plots/lmm_residuals_qqplot_contrast4.png")
    # qqnorm(residuals_model)
    # qqline(residuals_model, col = 'red')
    # dev.off()
    # print("Robust Standard Errors:")
    # print(robust_se)
    # t_crit_run <- qt(0.975, df = 34)
    # t_crit_intervention <- qt(0.975, df = 18)
    # print(t_crit_run)
    # print(t_crit_intervention)
    # if (shapiro_test_result$p.value > 0.05) {
    # print("Contrast 4 LMM residuals meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # print("Model Summary:")
    # print(summary(model))
    # print("Parameter Estimates:")
    # estimates <- fixef(model)
    # print(estimates)
    # } else {
    # print("Contrast 4 LMM residuals do not meet normality assumptions.")
    # anova_result <- anova(model)
    # print(anova_result)
    # print("Model Summary:")
    # print(summary(model))
    # print("Parameter Estimates:")
    # estimates <- fixef(model)
    # print(estimates)
    # }
    # """
    # ro.globalenv['roi_mean_df_4'] = pandas2ri.py2rpy(roi_mean_df_4)
    # result = r(r_script)
    # print(result)
    def std_error(x):
        return np.std(x, ddof=1) / np.sqrt(len(x))
    summary_df_4 = roi_mean_df_4.groupby(['intervention', 'run']).agg(
        roi_mean=('roi_mean', 'mean'),
        std_error=('roi_mean', std_error)).reset_index()
    run01_a_rse = 0.03468101
    run04_a_rse = np.sqrt(0.03468101**2 + 0.05342738**2)
    run01_b_rse = np.sqrt(0.03468101**2 + 0.03849149**2)
    run04_b_rse = np.sqrt(0.03468101**2 + 0.05342738**2 + 0.03849149**2 + 0.10126250**2)
    rse_column = [run01_a_rse, run04_a_rse, run01_b_rse, run04_b_rse]
    summary_df_4['rse'] = rse_column
    contrast_4_plot = (ggplot(summary_df_4, aes(x='intervention', y='roi_mean', fill='run')) +
        geom_bar(stat='identity', position='dodge') + 
        geom_errorbar(aes(ymin='roi_mean - rse', ymax='roi_mean + rse'), position=position_dodge(width=0.9), width=0.2) +
        theme_classic() +
        scale_fill_manual(values=['indianred', 'skyblue']) +
        labs(title="Mean SCC BOLD Values for 'Indig-Guilt' Contrast", x='Intervention', y='Mean SCC BOLD Value', fill='Run') +
        scale_y_continuous(expand=(0, 0), limits=[-0.5, 0.5], breaks=[-0.5,-0.4,-0.3,-0.2,-0.1,0,0.1,0.2,0.3,0.4,0.5]) +
        scale_x_discrete(labels={'a': 'A', 'b': 'B'}) +
        geom_hline(yintercept=0, linetype='solid', color='black', size=0.5))
    contrast_4_plot.save('analysis/fmri_analysis/analysis_1/plots/contrast_4_plot.png')

#endregion

#region 6) SUSCEPTIBILITY ANALYSIS.

def susceptibility_analysis():
    participants = ['P004', 'P006', 'P020', 'P030', 'P059', 'P078', 'P093', 'P094', 'P100', 'P107', 'P122', 'P125', 'P127', 'P128', 'P136', 'P145', 'P155', 'P199', 'P215', 'P216']
    runs = ['run01', 'run02', 'run03', 'run04']
    restart = logged_input("\nWould you like to start the susceptibility artifact analysis from scratch? This will remove all files from the 'analysis/susceptibility' folder. (y/n)\n")
    if restart == 'y':
        double_check = logged_input("\nAre you sure? (y/n)\n")
        if double_check == 'y':
            susceptibility_analysis_folder = 'analysis/susceptibility_analysis'
            print(f"Deleting analysis/susceptibility_analysis folder...")
            shutil.rmtree(susceptibility_analysis_folder)
        else:
            sys.exit()

    # Step 1: Create directories.
    print("\n###### STEP 1: CREATE DIRECTORIES ######")
    analysis_folder = 'analysis'
    os.makedirs(analysis_folder, exist_ok=True)
    susceptibility_analysis_folder = 'analysis/susceptibility_analysis'
    os.makedirs(susceptibility_analysis_folder, exist_ok=True)
    scc_folder = 'analysis/susceptibility_analysis/scc'
    os.makedirs(scc_folder, exist_ok=True)
    data_folder = 'analysis/susceptibility_analysis/data'
    os.makedirs(data_folder, exist_ok=True)
    run_comparisons_folder = 'analysis/susceptibility_analysis/run_comparisons'
    os.makedirs(run_comparisons_folder, exist_ok=True)
    for p_id in participants:
        data_participant_folder = f'analysis/susceptibility_analysis/data/{p_id}'
        os.makedirs(data_participant_folder, exist_ok=True)
        dicoms_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms'
        os.makedirs(dicoms_folder, exist_ok=True)
        dicoms_fieldmaps_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps'
        os.makedirs(dicoms_fieldmaps_folder, exist_ok=True)
        dicoms_run1_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run01'
        os.makedirs(dicoms_run1_folder, exist_ok=True)
        dicoms_run2_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run02'
        os.makedirs(dicoms_run2_folder, exist_ok=True)
        dicoms_run3_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run03'
        os.makedirs(dicoms_run3_folder, exist_ok=True)
        dicoms_run4_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run04'
        os.makedirs(dicoms_run4_folder, exist_ok=True)
        niftis_folder = f'analysis/susceptibility_analysis/data/{p_id}/niftis'
        os.makedirs(niftis_folder, exist_ok=True)
        pngs_folder = f'analysis/susceptibility_analysis/data/{p_id}/pngs'
        os.makedirs(pngs_folder, exist_ok=True)
    print("Directories created.")
    
    # Step 2: Calculate percentage of ROI voxels outside the brain during neurofeedback.
    # print("\n###### STEP 2: CALCULATE PERCENTAGE OF ROI VOXELS OUTSIDE BRAIN DURING NEUROFEEDBACK ######")
    # def get_sequence_numbers(file_name):
    #     parts = file_name.split('_')
    #     return int(parts[1]), int(parts[2].split('.')[0])
    # def copy_files(src_folder, dest_folder, sequence_number):
    #     src_pattern = f'*_{sequence_number:06d}_*.dcm'
    #     matching_files = [f for f in os.listdir(src_folder) if fnmatch.fnmatch(f, src_pattern)]
    #     for file in matching_files:
    #         src_path = os.path.join(src_folder, file)
    #         dest_path = os.path.join(dest_folder, file)
    #         shutil.copy(src_path, dest_path)
    # def read_roi_file(roi_file):
    #     voxel_coordinates = []
    #     with open(roi_file, 'r') as file:
    #         content = file.read()
    #         matches = re.findall(r'(?<=\n)\s*\d+\s+\d+\s+\d+', content)
    #         for match in matches:
    #             coordinates = match.split()
    #             voxel_coordinates.append((int(coordinates[0]), int(coordinates[1]), int(coordinates[2])))
    #     return voxel_coordinates
    # def process_participant(p_id, runs):
    #     path = f'data/raw_data/{p_id}/data/neurofeedback'
    #     cisc_folder = None
    #     for folder_name in os.listdir(path):
    #         if "CISC" in folder_name:
    #             cisc_folder = folder_name
    #             break
    #     if cisc_folder is None:
    #         print(f"No 'CISC' folder found in the 'neurofeedback' directory for participant {p_id}.")
    #         return
    #     src_folder = os.path.join(path, cisc_folder)
    #     analysis_folder = os.path.join(os.getcwd(), p_id, 'analysis', 'susceptibility')
    #     dicoms_folder = os.path.join(analysis_folder, 'susc_scc', 'dicoms')
    #     os.makedirs(dicoms_folder, exist_ok=True)
    #     run_folders = {f"run0{num}_dicoms": os.path.join(dicoms_folder, f"run0{num}_dicoms") for num in range(1, 5)}
    #     for folder in run_folders.values():
    #         os.makedirs(folder, exist_ok=True)
    #     files = [f for f in os.listdir(src_folder) if f.endswith('.dcm')]
    #     seq_vol_counts = {}
    #     for file in files:
    #         sequence_number, volume_number = get_sequence_numbers(file)
    #         if sequence_number not in seq_vol_counts:
    #             seq_vol_counts[sequence_number] = []
    #         seq_vol_counts[sequence_number].append(volume_number)
    #     seq_210 = [seq for seq, vols in seq_vol_counts.items() if len(vols) == 210]
    #     seq_238 = [seq for seq, vols in seq_vol_counts.items() if len(vols) == 238]
    #     min_210, max_210 = min(seq_210), max(seq_210)
    #     min_238, max_238 = min(seq_238), max(seq_238)
    #     if not os.listdir(run_folders["run01_dicoms"]):
    #         print(f"Copying Run01 dicoms for participant {p_id}...")
    #         copy_files(src_folder, run_folders["run01_dicoms"], min_210)
    #     if not os.listdir(run_folders["run02_dicoms"]):
    #         print(f"Copying Run02 dicoms for participant {p_id}...")
    #         copy_files(src_folder, run_folders["run02_dicoms"], min_238)
    #     if not os.listdir(run_folders["run03_dicoms"]):
    #         print(f"Copying Run03 dicoms for participant {p_id}...")
    #         copy_files(src_folder, run_folders["run03_dicoms"], max_238)
    #     if not os.listdir(run_folders["run04_dicoms"]):
    #         print(f"Copying Run04 dicoms for participant {p_id}...")
    #         copy_files(src_folder, run_folders["run04_dicoms"], max_210)
    #     output_folder = os.path.join(analysis_folder, 'susc_scc', 'niftis')
    #     os.makedirs(output_folder, exist_ok=True)
    #     for run in runs:
    #         destination_folder = run_folders[f"{run}_dicoms"]
    #         output_file = os.path.join(output_folder, f"{run}.nii")
    #         if not os.path.exists(output_file):
    #             print(f"Converting {run.upper()} DICOM files to Nifti format for participant {p_id}...")
    #             subprocess.run(['dcm2niix', '-o', output_folder, '-f', f"{run}", '-b', 'n', destination_folder])
    #         averaged_file = os.path.join(output_folder, f"{run}_averaged.nii.gz")
    #         if not os.path.exists(averaged_file):
    #             subprocess.run(['fslmaths', output_file, '-Tmean', averaged_file])
    #     run_num = ['1', '2', '3', '4']
    #     for num in run_num:
    #         roi_file = os.path.join(os.getcwd(), p_id, 'data', 'neurofeedback', cisc_folder, 'depression_neurofeedback', f'target_folder_run-{num}', f'depnf_run-{num}.roi')
    #         voxel_coordinates = read_roi_file(roi_file)
    #         functional_image = f'{p_id}/analysis/susceptibility/susc_scc/niftis/run0{num}_averaged.nii.gz'
    #         functional_image_info = nib.load(functional_image)
    #         functional_dims = functional_image_info.shape
    #         binary_volume = np.zeros(functional_dims)
    #         for voxel in voxel_coordinates:
    #             x, y, z = voxel
    #             binary_volume[x, y, z] = 1
    #         binary_volume = np.flip(binary_volume, axis=1)
    #         functional_affine = functional_image_info.affine
    #         binary_nifti = nib.Nifti1Image(binary_volume, affine=functional_affine)
    #         nib.save(binary_nifti, f'{p_id}/analysis/susceptibility/susc_scc/niftis/run0{num}_subject_space_ROI.nii.gz')
    #     for run in runs:
    #         betted_file = os.path.join(output_folder, f"{run}_averaged_betted.nii.gz")
    #         if not os.path.exists(betted_file):
    #             subprocess.run(['bet', f'{p_id}/analysis/susceptibility/susc_scc/niftis/{run}_averaged.nii.gz', betted_file, '-R'])
    #         functional_image_betted = f'{p_id}/analysis/susceptibility/susc_scc/niftis/{run}_averaged_betted.nii.gz'
    #         binary_nifti_image = f'{p_id}/analysis/susceptibility/susc_scc/niftis/{run}_subject_space_ROI.nii.gz'
    #         screenshot_file = f'{p_id}/analysis/susceptibility/susc_scc/ROI_on_{run}_EPI.png'
    #         binary_img = nib.load(binary_nifti_image)
    #         binary_data = binary_img.get_fdata()
    #         indices = np.nonzero(binary_data)
    #         center_x = int(np.mean(indices[0]))
    #         center_y = int(np.mean(indices[1]))
    #         center_z = int(np.mean(indices[2]))
    #         result = subprocess.run(['fsleyes', 'render', '--scene', 'lightbox', '--voxelLoc', f'{center_x}', f'{center_y}', f'{center_z}', '-hc', '-hl', '-of', screenshot_file, functional_image_betted, binary_nifti_image, '-ot', 'mask', '-mc', '1', '0', '0'], capture_output=True, text=True)
    #         if result.returncode == 0:
    #             print(f"Screenshot saved as {screenshot_file}")
    #         else:
    #             print(f"Error encountered: {result.stderr}")
    #     for run in runs:
    #         bin_file = os.path.join(output_folder, f"{run}_averaged_betted_bin.nii.gz")
    #         threshold = '100'
    #         if not os.path.exists(bin_file):
    #             subprocess.run(['fslmaths', f'{p_id}/analysis/susceptibility/susc_scc/niftis/{run}_averaged_betted.nii.gz', '-thr', threshold, '-bin', bin_file])
    #         inverse_file = os.path.join(output_folder, f"{run}_averaged_betted_bin_inverse.nii.gz")
    #         if not os.path.exists(inverse_file):
    #             subprocess.run(['fslmaths', f'{p_id}/analysis/susceptibility/susc_scc/niftis/{run}_averaged_betted_bin.nii.gz', '-sub', '1', '-abs', inverse_file])
    #         result2 = subprocess.run(['fslstats', f'{p_id}/analysis/susceptibility/susc_scc/niftis/{run}_subject_space_ROI.nii.gz', '-k', f'{p_id}/analysis/susceptibility/susc_scc/niftis/{run}_averaged_betted_bin_inverse.nii.gz', '-V'], capture_output=True, text=True)
    #         if result2.returncode == 0:
    #             result2_output = result2.stdout.strip()
    #         else:
    #             print(f"Error executing second fslstats command for {run}.")
    #             continue
    #         result2_output_values = result2_output.split()
    #         voxels_outside = float(result2_output_values[0])
    #         result3 = subprocess.run(['fslstats', f'{p_id}/analysis/susceptibility/susc_scc/niftis/{run}_subject_space_ROI.nii.gz', '-V'], capture_output=True, text=True)
    #         if result3.returncode == 0:
    #             result3_output = result3.stdout.strip()
    #         else:
    #             print(f"Error executing first fslstats command for {run}.")
    #             continue
    #         result3_output_values = result3_output.split()
    #         total_voxels_in_roi = float(result3_output_values[0])
    #         percentage_outside = (voxels_outside / total_voxels_in_roi) * 100
    #         percentage_outside = round(percentage_outside, 2)
    #         percentage_file = f"{p_id}/analysis/susceptibility/susc_scc/percentage_outside.txt"
    #         if not os.path.exists(percentage_file):
    #             with open(percentage_file, "a") as f:
    #                 f.write("Percentage of ROI voxels in signal dropout regions of merged EPI images.\n\n")
    #                 f.write("run threshold percentage_outside\n")
    #                 f.write(f"{run} {threshold} {percentage_outside}\n")
    #         else:
    #             with open(percentage_file, "r") as f:
    #                 lines = f.readlines()
    #                 matching_lines = [line for line in lines if line.startswith(f"{run}")]
    #                 if matching_lines:
    #                     with open(percentage_file, "w") as f:
    #                         for index, line in enumerate(lines):
    #                             if index not in matching_lines:
    #                                 f.write(line)
    #                         f.write(f"{run} {threshold} {percentage_outside}\n")
    #                 else:
    #                     with open(percentage_file, "a") as f:
    #                         f.write(f"{run} {threshold} {percentage_outside}\n")
    #         print(f"Percentage of ROI voxels in dropout regions saved in percentage_outside.txt file for {run}.")
    # screenshot_file = f'{p_id}/analysis/susceptibility/susc_scc/ROI_on_run01_EPI.png'
    # if not os.path.exists(screenshot_file):
    #     if __name__ == "__main__":
    #         for p_id in participants_to_iterate:
    #             process_participant(p_id, runs)

    # Step 3: Prepare DICOM and Nifti files.
    print("\n###### STEP 3: DICOM AND NIFTI FILE PREPARATION ######")
    for p_id in participants:
        print(f"Preparing DICOM and Nifti files for {p_id}...")
        path = f'data/raw_data/{p_id}/data/neurofeedback'
        cisc_folder = None
        for folder_name in os.listdir(path):
            if "CISC" in folder_name:
                cisc_folder = folder_name
                break
        if cisc_folder is None:
            print("No 'CISC' folder found in the 'neurofeedback' directory.")
            exit(1)
        def get_sequence_numbers(file_name):
            parts = file_name.split('_')
            return int(parts[1]), int(parts[2].split('.')[0])
        def copy_files(src_folder, dest_folder, sequence_number):
            src_pattern = f'*_{sequence_number:06d}_*.dcm'
            matching_files = [f for f in os.listdir(src_folder) if fnmatch.fnmatch(f, src_pattern)]
            for file in matching_files:
                src_path = os.path.join(src_folder, file)
                dest_path = os.path.join(dest_folder, file)
                shutil.copy(src_path, dest_path)
        def main():
            src_folder = os.path.join(path, cisc_folder)
            dicoms_run1_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run01'
            dicoms_run2_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run02'
            dicoms_run3_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run03'
            dicoms_run4_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run04'
            os.makedirs(dicoms_run1_folder, exist_ok=True)
            os.makedirs(dicoms_run2_folder, exist_ok=True)
            os.makedirs(dicoms_run3_folder, exist_ok=True)
            os.makedirs(dicoms_run4_folder, exist_ok=True)
            files = [f for f in os.listdir(src_folder) if f.endswith('.dcm')]
            seq_vol_counts = {}
            for file in files:
                sequence_number, volume_number = get_sequence_numbers(file)
                if sequence_number not in seq_vol_counts:
                    seq_vol_counts[sequence_number] = []
                seq_vol_counts[sequence_number].append(volume_number)
            seq_210 = [sequence_number for sequence_number, volume_numbers in seq_vol_counts.items() if len(volume_numbers) == 210]
            seq_238 = [sequence_number for sequence_number, volume_numbers in seq_vol_counts.items() if len(volume_numbers) == 238]
            min_210 = min(seq_210)
            max_210 = max(seq_210)
            min_238 = min(seq_238)
            max_238 = max(seq_238)
            if not os.listdir(dicoms_run1_folder):
                print(f"Copying Run01 dicoms for {p_id}...")
                copy_files(src_folder, dicoms_run1_folder, min_210)
                print(f"{p_id} Run01 dicoms copied. Number of files:", str(len(os.listdir(dicoms_run1_folder))) + ".", "Sequence number:", min_210)
            if not os.listdir(dicoms_run2_folder):
                print(f"Copying Run02 dicoms for {p_id}...")
                copy_files(src_folder, dicoms_run2_folder, min_238)
                print(f"{p_id} Run02 dicoms copied. Number of files:", str(len(os.listdir(dicoms_run2_folder))) + ".", "Sequence number:", min_238)
            if not os.listdir(dicoms_run3_folder):
                print(f"Copying Run03 dicoms for {p_id}...")
                copy_files(src_folder, dicoms_run3_folder, max_238)
                print(f"{p_id} Run03 dicoms copied. Number of files:", str(len(os.listdir(dicoms_run3_folder))) + ".", "Sequence number:", max_238)
            if not os.listdir(dicoms_run4_folder):
                print(f"Copying Run04 dicoms for {p_id}...")
                copy_files(src_folder, dicoms_run4_folder, max_210)
                print(f"{p_id} Run04 dicoms copied. Number of files:", str(len(os.listdir(dicoms_run4_folder))) + ".", "Sequence number:", max_210)
        if __name__ == "__main__":
            main()
        for run in runs:
            destination_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/{run}'
            output_folder = f'analysis/susceptibility_analysis/data/{p_id}/niftis'
            output_file = os.path.join(output_folder, f'{run}.nii')
            if not os.path.exists(output_file):
                print(f"Converting {run.upper()} DICOM files to Nifti format for {p_id}...")
                subprocess.run(['dcm2niix', '-o', output_folder, '-f', run, '-b', 'n', destination_folder])
                print(f"{p_id} {run.upper()} DICOM files converted to Nifti format.")
            else:
                print(f"{p_id} {run.upper()} Nifti file already exists. Skipping conversion.")
            png_path = f'analysis/susceptibility_analysis/data/{p_id}/pngs/{run}.png'
            nifti_path = f'analysis/susceptibility_analysis/data/{p_id}/niftis/{run}.nii'
            if not os.path.exists(png_path):
                print(f"Saving {p_id} {run} Nifti as PNG...")
                save_png = subprocess.run(['fsleyes', 'render', '--scene', 'ortho', '-of', png_path, nifti_path], capture_output=True, text=True)
                if save_png.returncode == 0:
                    print("Screenshot saved as", png_path)
                else:
                    print("Error encountered:", save_png.stderr)
            else:
                print('PNG files already created. Skipping conversion.')
        print(f"Check PNG files in analysis/susceptibility_analysis/data/{p_id}/pngs/ to see whether Niftis are in correct orientation. Anterior of brain should be facing right in sagittal view, right and left of brain should be swapped in coronal and transverse views, and anterior of the brain should be facing towards the top of the image in the transverse view. Other aspects should be easily viewable. Incorrect orientations can be corrected for using 'fslreorient2std' or 'fslswapdim' commands.")
    bad_participants = ['P004', 'P006', 'P020', 'P030', 'P078', 'P093', 'P094']
    def copy_2_dicoms(source_folder, destination_folder1, destination_folder2, target_volume_count=5):
                sequences2 = defaultdict(list)
                last_two_sets = []
                for filename in os.listdir(source_folder):
                    if filename.endswith('.dcm'):
                        file_parts = filename.split('_')
                        if len(file_parts) == 3:
                            sequence_number = int(file_parts[1])
                            volume_number = int(file_parts[2].split('.')[0])
                            sequences2[sequence_number].append((filename, volume_number))
                for sequence_number, files_info in sequences2.items():
                    if len(files_info) == target_volume_count:
                        last_two_sets.append(files_info)
                        if len(last_two_sets) > 2:
                            last_two_sets.pop(0)
                for idx, files_info in enumerate(last_two_sets):
                    for filename, _ in files_info:
                        if idx == 0:
                            destination_folder = destination_folder1
                        else:
                            destination_folder = destination_folder2
                        source_path = os.path.join(source_folder, filename)
                        destination_path = os.path.join(destination_folder, filename)
                        shutil.copy2(source_path, destination_path)
                        print(f"Copied {filename} to {destination_folder}")
    for p_id in bad_participants:
        print(f"Copying {p_id} fieldmap DICOMS to separate folder...")
        ap_fieldmaps_dicoms_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/ap'  
        pa_fieldmaps_dicoms_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/badpa'
        os.makedirs(ap_fieldmaps_dicoms_folder, exist_ok=True)
        os.makedirs(pa_fieldmaps_dicoms_folder, exist_ok=True)
        path = f'data/raw_data/{p_id}/data/neurofeedback'
        cisc_folder = None
        for folder_name in os.listdir(path):
            if "CISC" in folder_name:
                cisc_folder = folder_name
                break
        if cisc_folder is None:
            print("No 'CISC' folder found in the 'neurofeedback' directory.")
            exit(1)
        source_folder = os.path.join(path, cisc_folder)
        if not os.listdir(ap_fieldmaps_dicoms_folder) or not os.listdir(pa_fieldmaps_dicoms_folder):
            copy_2_dicoms(source_folder, ap_fieldmaps_dicoms_folder, pa_fieldmaps_dicoms_folder, target_volume_count=5)
            print(f"{p_id} fieldmap DICOMS successfully copied.")
        else:
            print(f"{p_id} fieldmap DICOMS already copied. Skipping process.")
        pe_list = ['ap', 'badpa']
        for pe in pe_list:
            destination_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/{pe}'
            output_folder = f'analysis/susceptibility_analysis/data/{p_id}/niftis/fieldmaps'
            output_file = os.path.join(output_folder, f'{pe}_fieldmaps.nii')
            if not os.path.exists(output_file):
                print(f"Converting {p_id} {pe.upper()} fieldmaps DICOM files to Nifti format...")
                subprocess.run(['dcm2niix', '-o', output_folder, '-f', f'{pe}_fieldmaps', '-b', 'n', destination_folder])
                print(f"{p_id} {pe.upper()} fieldmaps DICOM files converted to Nifti format.")
            else:
                print(f"{p_id} {pe.upper()} fieldmaps Nifti file already exists. Skipping conversion.")
    def copy_3_dicoms(source_folder, destination_folder1, destination_folder2, destination_folder3, target_volume_count=5):
            sequences3 = defaultdict(list)
            last_three_sets = [] 
            for filename in os.listdir(source_folder):
                if filename.endswith('.dcm'):
                    file_parts = filename.split('_')
                    if len(file_parts) == 3:
                        sequence_number = int(file_parts[1])
                        volume_number = int(file_parts[2].split('.')[0])
                        sequences3[sequence_number].append((filename, volume_number))
            for sequence_number, files_info in sequences3.items():
                if len(files_info) == target_volume_count:
                    last_three_sets.append(files_info)
                    if len(last_three_sets) > 3:
                        last_three_sets.pop(0)
            for idx, files_info in enumerate(last_three_sets):
                if idx == 0:
                    destination_folder = destination_folder1
                elif idx == 1:
                    destination_folder = destination_folder2
                else:
                    destination_folder = destination_folder3
                for filename, _ in files_info:
                    source_path = os.path.join(source_folder, filename)
                    destination_path = os.path.join(destination_folder, filename)
                    shutil.copy2(source_path, destination_path)
                    print(f"Copied {filename} to {destination_folder}")
    for p_id in participants:
        if p_id not in bad_participants:
            ap_fieldmaps_dicoms_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/ap'
            pa_fieldmaps_dicoms_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/pa'
            rl_fieldmaps_dicoms_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/rl'
            os.makedirs(ap_fieldmaps_dicoms_folder, exist_ok=True)
            os.makedirs(pa_fieldmaps_dicoms_folder, exist_ok=True)
            os.makedirs(rl_fieldmaps_dicoms_folder, exist_ok=True)
            path = f'data/raw_data/{p_id}/data/neurofeedback'
            cisc_folder = None
            for folder_name in os.listdir(path):
                if "CISC" in folder_name:
                    cisc_folder = folder_name
                    break
            if cisc_folder is None:
                print("No 'CISC' folder found in the 'neurofeedback' directory.")
                exit(1)
            source_folder = os.path.join(path, cisc_folder)
            if not os.listdir(ap_fieldmaps_dicoms_folder) or not os.listdir(pa_fieldmaps_dicoms_folder) or not os.listdir(rl_fieldmaps_dicoms_folder):
                copy_3_dicoms(source_folder, ap_fieldmaps_dicoms_folder, pa_fieldmaps_dicoms_folder, rl_fieldmaps_dicoms_folder, target_volume_count=5)
            pe_list = ['ap', 'pa', 'rl']
            for pe in pe_list:
                destination_folder = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/{pe}'
                output_folder = f'analysis/susceptibility_analysis/data/{p_id}/niftis'
                output_file = os.path.join(output_folder, f'{pe}_fieldmaps.nii')
                if not os.path.exists(output_file):
                    print(f"Converting {pe.upper()} fieldmaps DICOM files to Nifti format...")
                    subprocess.run(['dcm2niix', '-o', output_folder, '-f', f'{pe}_fieldmaps', '-b', 'n', destination_folder])
                    print(f"{pe.upper()} fieldmaps DICOM files converted to Nifti format.")
                else:
                    print(f"{pe.upper()} fieldmaps Nifti file already exists. Skipping conversion.")
    for p_id in participants:
        path = f'data/raw_data/{p_id}/data/neurofeedback'
        cisc_folder = None
        for folder_name in os.listdir(path):
            if "CISC" in folder_name:
                cisc_folder = folder_name
                break
        if cisc_folder is None:
            print("No 'CISC' folder found in the 'neurofeedback' directory.")
            exit(1)
        source_folder = os.path.join(path, cisc_folder)
        destination_folder = f'analysis/susceptibility_analysis/data/{p_id}/niftis'
        new_filename = 'structural.nii'
        if not os.path.exists(f'analysis/susceptibility_analysis/data/{p_id}/niftis/structural.nii'):
            nifti_folder = os.path.join(source_folder, 'depression_neurofeedback', 'nifti')
            nii_files = [f for f in os.listdir(nifti_folder) if f.endswith('.nii')]
            if len(nii_files) == 1:
                source_file = os.path.join(nifti_folder, nii_files[0])
                shutil.copy(source_file, destination_folder)
                copied_file_path = os.path.join(destination_folder, os.path.basename(source_file))
                new_file_path = os.path.join(destination_folder, new_filename)
                os.rename(copied_file_path, new_file_path)
                print('T1 Nifti copied and renamed to structural.nii.')
            else:
                print("No .nii file found or multiple .nii files found in the 'nifti' folder.")
        else:
            print('Structural Nifti file already exists. Skipping process.')
        bet_path = f'analysis/susceptibility_analysis/data/{p_id}/niftis/structural_brain.nii'
        structural_path = f'analysis/susceptibility_analysis/data/{p_id}/niftis/structural.nii'
        if not os.path.exists(bet_path):
            print("Performing brain extraction on structural image...")
            subprocess.run(['bet', structural_path, bet_path, '-m', '-R'])
            print("Structural image brain extracted.")
        else:
            print("Structural image already brain extracted. Skipping process.")

    # Step 4: Confirm sequence phase encoding directions for stratification of participants.
    print("\n###### STEP 4: DETERMINING PHASE ENCODING DIRECTIONS ######")
    for p_id in bad_participants:
        ap_fieldmaps = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/ap'
        pa_fieldmaps = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/badpa'
        run01 = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run01'
        run02 = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run02'
        run03 = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run03'
        run04 = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run04'
        folder_list = [ap_fieldmaps, pa_fieldmaps, run01, run02, run03, run04]
        pe_file = f'analysis/susceptibility_analysis/data/{p_id}/pe_axes.txt'
        pe_axes = []
        for folder in folder_list:
            dicom_files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith('.dcm')]
            if len(dicom_files) == 0:
                print("No DICOM files found in the directory.")
            else:
                random_file = random.choice(dicom_files)
                ds = pydicom.dcmread(random_file)
                pe_axis = ds.InPlanePhaseEncodingDirection
                pe_axes.append(pe_axis)
                start_index = folder.rfind('/') + 1  
                end_index = folder.rfind('_')  
                if end_index == -1 or end_index < start_index:
                    folder = folder[start_index:] + "_fieldmaps"
                else:
                    folder = folder[start_index:end_index]
                print(f"Phase Encoding Axis for {folder}: ", pe_axis)
            if not os.path.exists(pe_file):
                with open(pe_file, "a") as f:
                    f.write("sequence pe_axis\n")
                    f.write(f"{folder} {pe_axis}\n")
            else: 
                with open(pe_file, "r") as f:
                    lines = f.readlines()
                matching_lines = [line for line in lines if line.startswith(f"{folder}")]
                if matching_lines:
                    with open(pe_file, "w") as f:
                        for index, line in enumerate(lines):
                            if index not in matching_lines:
                                f.write(line)
                        f.write(f"{folder} {pe_axis}\n")
                else:
                    with open(pe_file, "a") as f:
                        f.write(f"{folder} {pe_axis}\n")
        print("Sequence PE axes saved to pe_axes.txt file.")
        if pe_axes == ['COL', 'ROW', 'ROW', 'ROW', 'ROW', 'COL']:
            print('Sequence PE directions are incorrect as expected (AP, RL, RL, RL, RL, AP) for this participant. Stratification of distortion correction method can now take place.')
        elif pe_axes == ['COL', 'ROW', 'ROW', 'ROW', 'ROW', 'ROW']:
            print('Sequence PE directions are incorrect as expected (AP, RL, RL, RL, RL, RL) for this participant. Stratification of distortion correction method can now take place.')
        else:
            print('Sequence PE directions are not as expected. Please investigate.')
            sys.exit()
    for p_id in participants:
        if p_id not in bad_participants:
            ap_fieldmaps = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/ap'
            pa_fieldmaps = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/pa'
            rl_fieldmaps = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/fieldmaps/rl'
            run01 = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run01'
            run02 = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run02'
            run03 = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run03'
            run04 = f'analysis/susceptibility_analysis/data/{p_id}/dicoms/run04'
            folder_list = [ap_fieldmaps, pa_fieldmaps, rl_fieldmaps, run01, run02, run03, run04]
            pe_file = f'analysis/susceptibility_analysis/data/{p_id}/pe_axes.txt'
            pe_axes = []
            for folder in folder_list:
                dicom_files = [os.path.join(folder, f) for f in os.listdir(folder) if f.endswith('.dcm')]
                if len(dicom_files) == 0:
                    print("No DICOM files found in the directory.")
                else:
                    random_file = random.choice(dicom_files)
                    ds = pydicom.dcmread(random_file)
                    pe_axis = ds.InPlanePhaseEncodingDirection
                    pe_axes.append(pe_axis)
                    start_index = folder.rfind('/') + 1  
                    end_index = folder.rfind('_')  
                    if end_index == -1 or end_index < start_index:
                        folder = folder[start_index:] + "_fieldmaps"
                    else:
                        folder = folder[start_index:end_index]
                    print(f"Phase Encoding Axis for {folder}: ", pe_axis)
                if not os.path.exists(pe_file):
                    with open(pe_file, "a") as f:
                        f.write("sequence pe_axis\n")
                        f.write(f"{folder} {pe_axis}\n")
                else: 
                    with open(pe_file, "r") as f:
                        lines = f.readlines()
                    matching_lines = [line for line in lines if line.startswith(f"{folder}")]
                    if matching_lines:
                        with open(pe_file, "w") as f:
                            for index, line in enumerate(lines):
                                if index not in matching_lines:
                                    f.write(line)
                            f.write(f"{folder} {pe_axis}\n")
                    else:
                        with open(pe_file, "a") as f:
                            f.write(f"{folder} {pe_axis}\n")
            print("Sequence PE axes saved to pe_axes.txt file.")
            if pe_axes == ['COL', 'COL', 'ROW', 'COL', 'COL', 'COL', 'COL']:
                print('Sequence PE directions are correct as expected (AP, PA, RL, PA, PA, PA, PA) for this participant. Stratification of distortion correction method can now take place.')
            else:
                print('Sequence PE directions are not as expected. Please investigate.')
                sys.exit()

    # Step 5: Calculate and apply fieldmaps for relevant participants.
    print("\n###### STEP 5: CALCULATING FIELDMAPS ######")
    for p_id in participants:
        if p_id not in bad_participants:
            ap_fieldmaps = f'analysis/susceptibility_analysis/data/{p_id}/niftis/ap_fieldmaps.nii'
            pa_fieldmaps = f'analysis/susceptibility_analysis/data/{p_id}/niftis/pa_fieldmaps.nii'
            output_file = f'analysis/susceptibility_analysis/data/{p_id}/niftis/merged_fieldmaps.nii'
            if not os.path.exists(output_file):
                print("Merging fieldmap sequences...")
                subprocess.run(['fslmerge', '-t', output_file, ap_fieldmaps, pa_fieldmaps])
                print("Fieldmap sequences merging completed.")
            else:
                print("Fieldmap sequences already merged. Skipping process.")
            fov_phase = 1
            base_res = 64
            phase_res = 1
            echo_spacing = 0.54
            pe_steps = (fov_phase * base_res * phase_res) - 1
            readout_time_s = (pe_steps * echo_spacing) / 1000
            acqparams_file = f'analysis/susceptibility_analysis/data/{p_id}/acqparams.txt'
            if not os.path.exists(acqparams_file):
                with open(acqparams_file, "a") as f:
                    f.write(f"0 -1 0 {readout_time_s}\n")
                    f.write(f"0 -1 0 {readout_time_s}\n")
                    f.write(f"0 -1 0 {readout_time_s}\n")
                    f.write(f"0 -1 0 {readout_time_s}\n")
                    f.write(f"0 -1 0 {readout_time_s}\n")
                    f.write(f"0 1 0 {readout_time_s}\n")
                    f.write(f"0 1 0 {readout_time_s}\n")
                    f.write(f"0 1 0 {readout_time_s}\n")
                    f.write(f"0 1 0 {readout_time_s}\n")
                    f.write(f"0 1 0 {readout_time_s}")
            fieldcoef_output_file = f'analysis/susceptibility_analysis/data/{p_id}/topup_{p_id}_fieldcoef.nii.gz'
            movpar_output_file = f'analysis/susceptibility_analysis/data/{p_id}/topup_{p_id}_movpar.txt'
            if not os.path.exists(fieldcoef_output_file) or not os.path.exists(movpar_output_file):
                print("Calculating fieldmaps...")
                subprocess.run(["topup", f"--imain=analysis/susceptibility_analysis/data/{p_id}/niftis/merged_fieldmaps.nii", f"--datain=analysis/susceptibility_analysis/data/{p_id}/acqparams.txt", "--config=b02b0.cnf", f"--out=analysis/susceptibility_analysis/data/{p_id}/topup_{p_id}", f"--iout=analysis/susceptibility_analysis/data/{p_id}/topup_{p_id}_unwarped"])
                print("Fieldmap calculation completed.")
            else:
                print("Fieldmaps already calculated. Skipping process.")

    # Step 6: Test quality of alternate distortion correction method (Stage 1).
    print("\n###### STEP 6: TESTING ALTERNATE DISTORTION CORRECTION METHOD (STAGE 1) ######")
    good_participants = ['P059', 'P100', 'P107', 'P122', 'P125', 'P127', 'P128', 'P136', 'P145', 'P155', 'P199', 'P215', 'P216']
    perc_outside_pa_values = []
    perc_outside_rl_values = []
    column_headers = ['p_id', 'perc_outside_pa', 'perc_outside_rl']
    group_perc_outside_df = pd.DataFrame(columns = column_headers) 
    for p_id in good_participants:
        print(f"Preparing Stage 1 files for {p_id}...")
        run_comparison_1_folder = 'analysis/susceptibility_analysis/run_comparisons/1'
        os.makedirs(run_comparison_1_folder, exist_ok=True)
        run_comparison_1_participant_folder = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}'
        os.makedirs(run_comparison_1_participant_folder, exist_ok=True)
        run_comparison_1_group_folder = 'analysis/susceptibility_analysis/run_comparisons/1/group'
        os.makedirs(run_comparison_1_group_folder, exist_ok=True)
        pa_fieldmaps = f'analysis/susceptibility_analysis/data/{p_id}/niftis/pa_fieldmaps.nii'
        rl_fieldmaps = f'analysis/susceptibility_analysis/data/{p_id}/niftis/rl_fieldmaps.nii'
        averaged_pa_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/averaged_pa_fieldmaps.nii.gz'
        averaged_rl_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/averaged_rl_fieldmaps.nii.gz'
        if not os.path.exists(averaged_pa_fieldmaps) or not os.path.exists(averaged_rl_fieldmaps):
            subprocess.run(['fslmaths', pa_fieldmaps, '-Tmean', averaged_pa_fieldmaps])
            subprocess.run(['fslmaths', rl_fieldmaps, '-Tmean', averaged_rl_fieldmaps])
        betted_pa_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/betted_pa_fieldmaps.nii.gz'
        betted_rl_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/betted_rl_fieldmaps.nii.gz'
        if not os.path.exists(betted_pa_fieldmaps) or not os.path.exists(betted_rl_fieldmaps):
            subprocess.run(["bet", averaged_pa_fieldmaps, betted_pa_fieldmaps, "-m", "-R"])
            subprocess.run(["bet", averaged_rl_fieldmaps, betted_rl_fieldmaps, "-m", "-R"])
        flirted_pa_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_pa_fieldmaps.nii.gz'
        flirted_rl_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_rl_fieldmaps.nii.gz'
        t1_flirted_pa_fieldmaps_transformation = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/t1_flirted_pa_fieldmaps_transformation.mat'
        t1_flirted_rl_fieldmaps_transformation = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/t1_flirted_rl_fieldmaps_transformation.mat'
        structural_brain = f'analysis/susceptibility_analysis/data/{p_id}/niftis/structural_brain.nii'
        if not os.path.exists(flirted_pa_fieldmaps):
            subprocess.run(["flirt", "-in", betted_pa_fieldmaps, "-ref", structural_brain, "-out", flirted_pa_fieldmaps, "-omat", t1_flirted_pa_fieldmaps_transformation])
            subprocess.run(["flirt", "-in", betted_rl_fieldmaps, "-ref", structural_brain, "-out", flirted_rl_fieldmaps, "-omat", t1_flirted_rl_fieldmaps_transformation])
        def read_roi_file(roi_file):
            voxel_coordinates = []
            with open(roi_file, 'r') as file:
                content = file.read()
                matches = re.findall(r'(?<=\n)\s*\d+\s+\d+\s+\d+', content)
                for match in matches:
                    coordinates = match.split()
                    voxel_coordinates.append(
                        (int(coordinates[0]), int(coordinates[1]), int(coordinates[2])))
            return voxel_coordinates
        path = f'data/raw_data/{p_id}/data/neurofeedback'
        cisc_folder = None
        for folder_name in os.listdir(path):
            if "CISC" in folder_name:
                cisc_folder = folder_name
                break
        if cisc_folder is None:
            print("No 'CISC' folder found in the 'neurofeedback' directory.")
            exit(1)
        roi_file = f'data/raw_data/{p_id}/data/neurofeedback/{cisc_folder}/depression_neurofeedback/target_folder_run-1/depnf_run-1.roi'
        voxel_coordinates = read_roi_file(roi_file)
        averaged_run = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/averaged_run.nii.gz'
        if not os.path.exists(averaged_run):
            run = f'analysis/susceptibility_analysis/data/{p_id}/niftis/run01.nii'
            subprocess.run(['fslmaths', run, '-Tmean', averaged_run])
        functional_image_info = nib.load(averaged_run)
        functional_dims = functional_image_info.shape
        binary_volume = np.zeros(functional_dims)
        for voxel in voxel_coordinates:
            x, y, z = voxel
            binary_volume[x, y, z] = 1
        binary_volume = np.flip(binary_volume, axis=1)
        functional_affine = functional_image_info.affine
        binary_mask = nib.Nifti1Image(binary_volume, affine=functional_affine)
        nib.save(binary_mask, f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/run01_subject_space_ROI.nii.gz')
        roi_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/run01_subject_space_ROI.nii.gz'
        transformed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/transformed_roi_mask.nii.gz'
        temp_file = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/temp_file.nii.gz'
        roi_transformation = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/roi_transformation.mat'
        subprocess.run(['flirt', '-in', averaged_run, '-ref', structural_brain, '-out', temp_file, '-omat', roi_transformation])
        subprocess.run(['flirt', '-in', roi_mask, '-ref', structural_brain, '-applyxfm', '-init', roi_transformation, '-out', transformed_roi_mask, '-interp', 'nearestneighbour'])
        flirted_pa_fieldmaps_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_pa_fieldmaps_bin.nii.gz'
        if not os.path.exists(flirted_pa_fieldmaps_bin):
            subprocess.run(['fslmaths', flirted_pa_fieldmaps, '-thr', '100', '-bin', flirted_pa_fieldmaps_bin])
        flirted_rl_fieldmaps_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_rl_fieldmaps_bin.nii.gz'
        if not os.path.exists(flirted_rl_fieldmaps_bin):
            subprocess.run(['fslmaths', flirted_rl_fieldmaps, '-thr', '100', '-bin', flirted_rl_fieldmaps_bin])
        pa_bin_inv = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_bin_inv.nii.gz'
        if not os.path.exists(pa_bin_inv):
            subprocess.run(['fslmaths', flirted_pa_fieldmaps_bin, '-sub', '1', '-abs', pa_bin_inv])
        rl_bin_inv = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_bin_inv.nii.gz'
        if not os.path.exists(rl_bin_inv):
            subprocess.run(['fslmaths', flirted_rl_fieldmaps_bin, '-sub', '1', '-abs', rl_bin_inv])
        pa_result = subprocess.run(['fslstats', transformed_roi_mask, '-k', pa_bin_inv, '-V'], capture_output=True, text=True)
        if pa_result.returncode == 0:
            pa_result_output = pa_result.stdout.strip()
        else:
            print("Error executing fslstats command.")
        pa_result_output_values = pa_result_output.split()
        pa_voxels_outside = float(pa_result_output_values[0])
        rl_result = subprocess.run(['fslstats', transformed_roi_mask, '-k', rl_bin_inv, '-V'], capture_output=True, text=True)
        if rl_result.returncode == 0:
            rl_result_output = rl_result.stdout.strip()
        else:
            print("Error executing fslstats command.")
        rl_result_output_values = rl_result_output.split()
        rl_voxels_outside = float(rl_result_output_values[0])
        result1 = subprocess.run(['fslstats', transformed_roi_mask, '-V'], capture_output=True, text=True)
        if result1.returncode == 0:
            result1_output = result1.stdout.strip()
        else:
            print("Error executing fslstats command.")
        result1_output_values = result1_output.split()
        total_voxels_in_roi = float(result1_output_values[0])
        perc_outside_pa = (pa_voxels_outside / total_voxels_in_roi) * 100
        perc_outside_pa = round(perc_outside_pa, 2)
        perc_outside_pa_values.append(perc_outside_pa)
        perc_outside_rl = (rl_voxels_outside / total_voxels_in_roi) * 100
        perc_outside_rl = round(perc_outside_rl, 2)
        perc_outside_rl_values.append(perc_outside_rl)
        perc_outside_df = pd.DataFrame({'p_id': [p_id], 'perc_outside_pa': [perc_outside_pa], 'perc_outside_rl': [perc_outside_rl]})
        perc_outside_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/perc_outside_df.txt', sep='\t', index=False)
        group_perc_outside_df = pd.concat([group_perc_outside_df, perc_outside_df], ignore_index=True)
        pa_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_trimmed_roi_mask.nii.gz'
        rl_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_trimmed_roi_mask.nii.gz'
        if not os.path.exists(pa_trimmed_roi_mask) or not os.path.exists(rl_trimmed_roi_mask):
            subprocess.run(['fslmaths', transformed_roi_mask, '-mul', flirted_pa_fieldmaps_bin, pa_trimmed_roi_mask])
            subprocess.run(['fslmaths', transformed_roi_mask, '-mul', flirted_rl_fieldmaps_bin, rl_trimmed_roi_mask])
    group_perc_outside_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/1/group/group_perc_outside_df.txt', sep='\t', index=False)
    plot_data = pd.DataFrame({
        'Participant': good_participants * 2,
        'Perc_Outside': perc_outside_pa_values + perc_outside_rl_values,
        'Sequence': ['PA'] * len(good_participants) + ['RL'] * len(good_participants)
    })
    perc_outside_plot = (
        ggplot(plot_data, aes(x='Participant', y='Perc_Outside', fill='Sequence')) +
        geom_bar(stat='identity', position='dodge') +
        theme_classic() +
        labs(title='Percentage of Voxels in Signal Dropout Regions', x='Participant', y='Percentage of SCC Voxels in Signal Dropout') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='black'), axis_title=element_text(size=12, color='black')) +
        scale_y_continuous(expand=(0, 0))
    )
    perc_outside_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/perc_outside_plot.png')
    perc_outside_pa_overall = np.mean(perc_outside_pa_values)
    perc_outside_rl_overall = np.mean(perc_outside_rl_values)
    pa_std_error = np.std(perc_outside_pa_values) / np.sqrt(len(perc_outside_pa_values))
    rl_std_error = np.std(perc_outside_rl_values) / np.sqrt(len(perc_outside_rl_values))
    _, perc_outside_pa_overall_shap_p = stats.shapiro(perc_outside_pa_values)
    _, perc_outside_rl_overall_shap_p = stats.shapiro(perc_outside_rl_values)
    if perc_outside_pa_overall_shap_p > 0.05 and perc_outside_rl_overall_shap_p > 0.05:
        print(f'Shapiro-Wilk test passed for perc_outside values. Running parametric t-test...')
        _, p_value = stats.ttest_ind(perc_outside_pa_values, perc_outside_rl_values)
        print(f"T-test p-value: {p_value}")
    else:
        print(f'Shapiro-Wilk test failed for perc_outside values. Running non-parametric Mann-Whitney U test...')
        _, p_value = stats.mannwhitneyu(perc_outside_pa_values, perc_outside_rl_values)
        print(f"Mann-Whitney U test p-value: {p_value}")
    plot_data = pd.DataFrame({'Sequence': ['PA', 'RL'], 'Perc_Outside': [perc_outside_pa_overall, perc_outside_rl_overall], 'Std_Error': [pa_std_error, rl_std_error]})
    group_perc_outside_plot = (ggplot(plot_data, aes(x='Sequence', y='Perc_Outside', fill='Sequence')) + 
                        geom_bar(stat='identity', position='dodge') +
                        geom_errorbar(aes(ymin='Perc_Outside - Std_Error', ymax='Perc_Outside + Std_Error'), width=0.2, color='black') +
                        theme_classic() +
                        labs(title='Percentage of Voxels in Signal Dropout Regions', y='Percentage of SCC Voxels in Signal Dropout') +
                        scale_y_continuous(expand=(0, 0), limits=[0,10]) +
                        scale_fill_manual(values={'PA': '#DB5F57', 'RL': '#57D3DB'})
                        )
    if p_value < 0.001:
        group_perc_outside_plot = group_perc_outside_plot + annotate("text", x=1.5, y=max(plot_data['Perc_Outside']) + 3.5, label="***", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Perc_Outside']) + 3, yend=max(plot_data['Perc_Outside']) + 3, color="black")
    elif p_value < 0.01:
        group_perc_outside_plot = group_perc_outside_plot + annotate("text", x=1.5, y=max(plot_data['Perc_Outside']) + 3.5, label="**", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Perc_Outside']) + 3, yend=max(plot_data['Perc_Outside']) + 3, color="black")
    elif p_value < 0.05:
        group_perc_outside_plot = group_perc_outside_plot + annotate("text", x=1.5, y=max(plot_data['Perc_Outside']) + 3.5, label="*", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Perc_Outside']) + 3, yend=max(plot_data['Perc_Outside']) + 3, color="black")    
    group_perc_outside_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/group_perc_outside_plot.png')
    
    column_headers = ['p_id', 'ssim_index', 'voxels_in_bin_ssim_mask', 'perc_roi_voxels_in_bin_ssim_mask']
    group_ssim_df = pd.DataFrame(columns = column_headers) 
    for p_id in good_participants:
        print(f"Running Stage 1 SSIM analysis for {p_id}...")
        def calculate_ssim(image1_path, image2_path, ssim_output_path):
            """Function to calculate SSIM between two Nifti images and save the SSIM map."""
            image1_nii = nib.load(image1_path)
            image2_nii = nib.load(image2_path)
            image1 = image1_nii.get_fdata()
            image2 = image2_nii.get_fdata()
            if image1.shape != image2.shape:
                raise ValueError("Input images must have the same dimensions for SSIM calculation.")
            ssim_index, ssim_map = ssim(image1, image2, full=True, data_range=image1.max() - image1.min())
            ssim_map_nifti = nib.Nifti1Image(ssim_map, affine=image1_nii.affine, header=image1_nii.header)
            nib.save(ssim_map_nifti, ssim_output_path)
            return ssim_index
        ssim_output_path = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/ssim_map.nii.gz'
        flirted_pa_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_pa_fieldmaps.nii.gz'
        flirted_rl_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_rl_fieldmaps.nii.gz'
        if not os.path.exists(ssim_output_path):
            ssim_index = calculate_ssim(flirted_rl_fieldmaps, flirted_pa_fieldmaps, ssim_output_path)
        else:
            df = pd.read_csv(f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/ssim_df.txt', delimiter='\t')
            ssim_index_series = df.loc[df['p_id'] == p_id, 'ssim_index']
            ssim_index = ssim_index_series.iloc[0]
        ssim_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/ssim_bin.nii.gz'
        if not os.path.exists(ssim_bin):
            subprocess.run(["fslmaths", ssim_output_path, "-thr", "0.8", "-binv", ssim_bin])
        combined_pa_rl_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/combined_pa_rl_mask.nii.gz'
        flirted_pa_fieldmaps_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_pa_fieldmaps_bin.nii.gz'
        flirted_rl_fieldmaps_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_rl_fieldmaps_bin.nii.gz'
        if not os.path.exists(combined_pa_rl_mask):
            subprocess.run(['fslmaths', flirted_pa_fieldmaps_bin, '-add', flirted_rl_fieldmaps_bin, combined_pa_rl_mask])
        bin_pa_rl_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/bin_pa_rl_mask.nii.gz'
        if not os.path.exists(bin_pa_rl_mask):
            subprocess.run(['fslmaths', combined_pa_rl_mask, '-bin', bin_pa_rl_mask])
        ssim_bin_trimmed = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/ssim_bin_trimmed.nii.gz'
        if not os.path.exists(ssim_bin_trimmed):
            subprocess.run(['fslmaths', ssim_bin, '-mul', bin_pa_rl_mask, ssim_bin_trimmed])
        voxels_in_whole_mask = subprocess.run(["fslstats", ssim_bin_trimmed, "-V"], capture_output=True, text=True).stdout.split()[0]
        voxels_in_whole_mask = float(voxels_in_whole_mask)
        intersection_mask_path = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/ssim_roi_intersect.nii.gz'
        transformed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/transformed_roi_mask.nii.gz'
        if not os.path.exists(intersection_mask_path):
            subprocess.run(["fslmaths", ssim_bin_trimmed, "-mas", transformed_roi_mask, intersection_mask_path])
        voxels_in_roi_in_mask = subprocess.run(["fslstats", intersection_mask_path, "-V"], capture_output=True, text=True).stdout.split()[0]
        voxels_in_roi_in_mask = float(voxels_in_roi_in_mask)
        perc_roi_voxels_in_mask = (voxels_in_roi_in_mask / total_voxels_in_roi) * 100
        ssim_df = pd.DataFrame({'p_id': [p_id], 'ssim_index': [ssim_index], 'voxels_in_bin_ssim_mask': [voxels_in_whole_mask], 'perc_roi_voxels_in_bin_ssim_mask': [perc_roi_voxels_in_mask]})
        ssim_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/ssim_df.txt', sep='\t', index=False)
        group_ssim_df = pd.concat([group_ssim_df, ssim_df], ignore_index=True)
    group_ssim_df.to_csv('analysis/susceptibility_analysis/run_comparisons/1/group/group_ssim_df.txt', sep='\t', index=False)
    ssim_indexes = group_ssim_df['ssim_index'].tolist()
    ssim_mean = np.mean(ssim_indexes)
    print(f"Mean SSIM index for Stage 1: {ssim_mean}")
    plot_data = pd.DataFrame({
        'Participant': good_participants,
        'SSIM': ssim_indexes,
    })
    ssim_index_plot = (
        ggplot(plot_data, aes(x='Participant', y='SSIM')) +
        geom_bar(stat='identity', position='dodge') +
        geom_hline(yintercept=ssim_mean, linetype='dashed', color='red') +
        theme_classic() +
        labs(title='SSIM Indexes', x='Participant', y='SSIM') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='blue'), axis_title=element_text(size=12, face='bold')) +
        scale_y_continuous(expand=(0, 0), limits=[0.8,1])
    )
    ssim_index_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/ssim_index_plot.png')
    voxels = group_ssim_df['voxels_in_bin_ssim_mask'].tolist()
    voxels_mean = np.mean(voxels)
    plot_data = pd.DataFrame({
        'Participant': good_participants,
        'Voxels': voxels,
    })
    ssim_voxels_plot = (
        ggplot(plot_data, aes(x='Participant', y='Voxels')) +
        geom_bar(stat='identity', position='dodge') +
        geom_hline(yintercept=voxels_mean, linetype='dashed', color='red') +
        theme_classic() +
        labs(title='Number of Voxels in SSIM Mask', x='Participant', y='Voxels') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='blue'), axis_title=element_text(size=12, face='bold')) +
        scale_y_continuous(expand=(0, 0))
    )
    ssim_voxels_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/ssim_voxels_plot.png')
    perc_voxels = group_ssim_df['perc_roi_voxels_in_bin_ssim_mask'].tolist()
    perc_voxels_mean = np.mean(perc_voxels)
    plot_data = pd.DataFrame({
        'Participant': good_participants,
        'Perc_Voxels': perc_voxels,
    })
    ssim_perc_voxels_plot = (
        ggplot(plot_data, aes(x='Participant', y='Perc_Voxels')) +
        geom_bar(stat='identity', position='dodge') +
        geom_hline(yintercept=perc_voxels_mean, linetype='dashed', color='red') +
        theme_classic() +
        labs(title='Percentage of ROI Voxels in SSIM Mask', x='Participant', y='Percentage of SCC Voxels in SSIM Mask') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='black'), axis_title=element_text(size=12)) +
        scale_y_continuous(expand=(0, 0))
    )
    ssim_perc_voxels_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/ssim_perc_voxels_plot.png')

    overlap_perc_av_values = []
    column_headers = ['p_id', 'tissue_type', 'overlap_perc']
    group_overlap_perc_df = pd.DataFrame(columns = column_headers) 
    for p_id in good_participants:
        print(f'Running Stage 1 segmentation analysis for {p_id}...')
        pa_csf_pve_seg = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_seg_pve_0.nii.gz'
        pa_wm_pve_seg = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_seg_pve_1.nii.gz'
        pa_gm_pve_seg = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_seg_pve_2.nii.gz'
        rl_csf_pve_seg = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_seg_pve_0.nii.gz'
        rl_wm_pve_seg = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_seg_pve_1.nii.gz'
        rl_gm_pve_seg = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_seg_pve_2.nii.gz'
        if not os.path.exists(pa_csf_pve_seg):
            pa_seg = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_seg'
            rl_seg = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_seg'
            flirted_pa_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_pa_fieldmaps.nii.gz'
            flirted_rl_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_rl_fieldmaps.nii.gz'
            structural_brain = f'analysis/susceptibility_analysis/data/{p_id}/niftis/structural_brain.nii'
            subprocess.run(["fast", "-n", "3", "-o", pa_seg, structural_brain, flirted_pa_fieldmaps])
            subprocess.run(["fast", "-n", "3", "-o", rl_seg, structural_brain, flirted_rl_fieldmaps])
        pa_csf_pve_seg_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_csf_pve_seg_bin.nii.gz'
        pa_wm_pve_seg_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_wm_pve_seg_bin.nii.gz'
        pa_gm_pve_seg_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_gm_pve_seg_bin.nii.gz'
        if not os.path.exists(pa_csf_pve_seg_bin):
            subprocess.run(['fslmaths', pa_csf_pve_seg, '-thr', '0.5', '-bin', pa_csf_pve_seg_bin])
            subprocess.run(['fslmaths', pa_wm_pve_seg, '-thr', '0.5', '-bin', pa_wm_pve_seg_bin])
            subprocess.run(['fslmaths', pa_gm_pve_seg, '-thr', '0.5', '-bin', pa_gm_pve_seg_bin])
        rl_csf_pve_seg_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_csf_pve_seg_bin.nii.gz'
        rl_wm_pve_seg_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_wm_pve_seg_bin.nii.gz'
        rl_gm_pve_seg_bin = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_gm_pve_seg_bin.nii.gz'
        if not os.path.exists(rl_csf_pve_seg_bin):
            subprocess.run(['fslmaths', rl_csf_pve_seg, '-thr', '0.5', '-bin', rl_csf_pve_seg_bin])
            subprocess.run(['fslmaths', rl_wm_pve_seg, '-thr', '0.5', '-bin', rl_wm_pve_seg_bin])
            subprocess.run(['fslmaths', rl_gm_pve_seg, '-thr', '0.5', '-bin', rl_gm_pve_seg_bin])
        csf_intersect_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/csf_intersect_mask.nii.gz'
        wm_intersect_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/wm_intersect_mask.nii.gz'
        gm_intersect_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/gm_intersect_mask.nii.gz'
        if not os.path.exists(csf_intersect_mask):
            subprocess.run(['fslmaths', pa_csf_pve_seg_bin, '-mul', rl_csf_pve_seg_bin, '-bin', csf_intersect_mask])
            subprocess.run(['fslmaths', pa_wm_pve_seg_bin, '-mul', rl_wm_pve_seg_bin, '-bin', wm_intersect_mask])
            subprocess.run(['fslmaths', pa_gm_pve_seg_bin, '-mul', rl_gm_pve_seg_bin, '-bin', gm_intersect_mask])
        csf_intersect_vol = float(subprocess.run(['fslstats', csf_intersect_mask, '-V'], capture_output=True, text=True).stdout.split()[0])
        wm_intersect_vol = float(subprocess.run(['fslstats', wm_intersect_mask, '-V'], capture_output=True, text=True).stdout.split()[0])
        gm_intersect_vol = float(subprocess.run(['fslstats', gm_intersect_mask, '-V'], capture_output=True, text=True).stdout.split()[0])
        pa_csf_mask_vol = float(subprocess.run(['fslstats', pa_csf_pve_seg_bin, '-V'], capture_output=True, text=True).stdout.split()[0])
        rl_csf_mask_vol = float(subprocess.run(['fslstats', rl_csf_pve_seg_bin, '-V'], capture_output=True, text=True).stdout.split()[0])
        if pa_csf_mask_vol < rl_csf_mask_vol:
            csf_overlap_perc = (csf_intersect_vol / pa_csf_mask_vol) * 100
        else: 
            csf_overlap_perc = (csf_intersect_vol / rl_csf_mask_vol) * 100
        pa_wm_mask_vol = float(subprocess.run(['fslstats', pa_wm_pve_seg_bin, '-V'], capture_output=True, text=True).stdout.split()[0])
        rl_wm_mask_vol = float(subprocess.run(['fslstats', rl_wm_pve_seg_bin, '-V'], capture_output=True, text=True).stdout.split()[0])
        if pa_wm_mask_vol < rl_wm_mask_vol:
            wm_overlap_perc = (wm_intersect_vol / pa_wm_mask_vol) * 100
        else: 
            wm_overlap_perc = (wm_intersect_vol / rl_wm_mask_vol) * 100
        pa_gm_mask_vol = float(subprocess.run(['fslstats', pa_gm_pve_seg_bin, '-V'], capture_output=True, text=True).stdout.split()[0])
        rl_gm_mask_vol = float(subprocess.run(['fslstats', rl_gm_pve_seg_bin, '-V'], capture_output=True, text=True).stdout.split()[0])
        if pa_gm_mask_vol < rl_gm_mask_vol:
            gm_overlap_perc = (gm_intersect_vol / pa_gm_mask_vol) * 100
        else: 
            gm_overlap_perc = (gm_intersect_vol / rl_gm_mask_vol) * 100
        participant_col = []
        tissue_type_col = []
        overlap_perc_col = []
        participant_col.append(p_id)
        participant_col.append(p_id)
        participant_col.append(p_id)
        tissue_type_col.append('csf')
        tissue_type_col.append('wm')
        tissue_type_col.append('gm')
        overlap_perc_col.append(csf_overlap_perc)
        overlap_perc_col.append(wm_overlap_perc)
        overlap_perc_col.append(gm_overlap_perc)
        if p_id == 'P122' or p_id == 'P136':
            values = np.array([wm_overlap_perc, gm_overlap_perc])
            overlap_perc_av = np.mean(values)
            overlap_perc_av_values.append(overlap_perc_av)
        overlap_perc_df = pd.DataFrame({'p_id': participant_col, 'tissue_type': tissue_type_col, 'overlap_perc': overlap_perc_col})
        overlap_perc_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/overlap_perc_df.txt', sep='\t', index=False)
        group_overlap_perc_df = pd.concat([group_overlap_perc_df, overlap_perc_df], ignore_index=True)
    group_overlap_perc_df.to_csv('analysis/susceptibility_analysis/run_comparisons/1/group/group_overlap_perc_df.txt', sep='\t', index=False)
    csf_values = []
    wm_values = []
    gm_values = []
    for p_id in participants:
        if p_id in good_participants:
            filtered_csf = group_overlap_perc_df.loc[(group_overlap_perc_df['tissue_type'] == 'csf') & (group_overlap_perc_df['p_id'] == p_id), 'overlap_perc'].values[0]
            filtered_csf = float(filtered_csf)
            csf_values.append(filtered_csf)
            filtered_wm = group_overlap_perc_df.loc[(group_overlap_perc_df['tissue_type'] == 'wm') & (group_overlap_perc_df['p_id'] == p_id), 'overlap_perc'].values[0]
            filtered_wm = float(filtered_wm)
            wm_values.append(filtered_wm)
            filtered_gm = group_overlap_perc_df.loc[(group_overlap_perc_df['tissue_type'] == 'gm') & (group_overlap_perc_df['p_id'] == p_id), 'overlap_perc'].values[0]
            filtered_gm = float(filtered_gm)
            gm_values.append(filtered_gm)
    plot_data = pd.DataFrame({
        'Participant': good_participants * 3,
        'Overlap_Perc': csf_values + wm_values + gm_values,
        'Tissue_Type': ['CSF'] * len(good_participants) + ['WM'] * len(good_participants) + ['GM'] * len(good_participants)
    })
    overlap_perc_plot = (
        ggplot(plot_data, aes(x='Participant', y='Overlap_Perc', fill='Tissue_Type')) +
        geom_bar(stat='identity', position='dodge') +
        theme_classic() +
        labs(title='Sequence Tissue Type Overlap', x='Participant', y='Tissue Overlap Percentage', fill='Tissue Type') +
        theme(axis_text_x=element_text(rotation=45, hjust=1, color='black'), text=element_text(size=12, color='black'), axis_title=element_text(size=12, color='black')) +
        scale_y_continuous(expand=(0, 0), limits=[0,100]) +
        scale_fill_manual(values={'CSF': '#F9DC5C', 'WM': '#db5f57', 'GM': '#57d3db'})
    )
    overlap_perc_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/overlap_perc_plot.png')
    filtered_csf = group_overlap_perc_df[group_overlap_perc_df['tissue_type'] == 'csf']['overlap_perc'].tolist()
    mean_csf = np.mean(filtered_csf)
    filtered_wm = group_overlap_perc_df[group_overlap_perc_df['tissue_type'] == 'wm']['overlap_perc'].tolist()
    mean_wm = np.mean(filtered_wm)
    filtered_gm = group_overlap_perc_df[group_overlap_perc_df['tissue_type'] == 'gm']['overlap_perc'].tolist()
    mean_gm = np.mean(filtered_gm)
    csf_std_error = np.std(filtered_csf) / np.sqrt(len(filtered_csf))
    wm_std_error = np.std(filtered_wm) / np.sqrt(len(filtered_wm))
    gm_std_error = np.std(filtered_gm) / np.sqrt(len(filtered_gm))
    group_overlap_perc_df['p_id'] = group_overlap_perc_df['p_id'].astype(str)
    group_overlap_perc_df['tissue_type'] = group_overlap_perc_df['tissue_type'].astype(str)
    group_overlap_perc_df['overlap_perc'] = pd.to_numeric(group_overlap_perc_df['overlap_perc'], errors='coerce')
    sphericity_test = rm_anova(data=group_overlap_perc_df, dv='overlap_perc', within='tissue_type', subject='p_id')
    epsilon_value = sphericity_test.loc[sphericity_test['Source'] == 'tissue_type', 'eps'].values[0]
    print(f'Stage 1 segmentation analysis sphericity test epsilon value: {epsilon_value}')
    normality_passed = True
    shapiro_results = group_overlap_perc_df.groupby('tissue_type')['overlap_perc'].apply(stats.shapiro)
    shapiro_p_values = shapiro_results.apply(lambda x: x.pvalue)
    if any(shapiro_p_values < 0.05):
        normality_passed = False
    print(f'Stage 1 segmentation analysis Shapiro-Wilk test of normality passed: {normality_passed}')
    _, p_value_levene = stats.levene(
        group_overlap_perc_df[group_overlap_perc_df['tissue_type'] == 'csf']['overlap_perc'],
        group_overlap_perc_df[group_overlap_perc_df['tissue_type'] == 'wm']['overlap_perc'],
        group_overlap_perc_df[group_overlap_perc_df['tissue_type'] == 'gm']['overlap_perc']
    )
    print(f'Stage 1 segmentation analysis Levene test p-value: {p_value_levene}')
    if normality_passed and p_value_levene > 0.05 and epsilon_value > 0.75:
        print('Stage 1 segmentation analysis parametric assumptions met. Proceeding with two-way ANOVA...')
        anova_result = rm_anova(data=group_overlap_perc_df, dv='overlap_perc', within='tissue_type', subject='p_id')
        print(anova_result)
    else:
        print('Stage 1 segmentation analysis parametric assumptions not met. Two-way ANOVA not run.')
    plot_data = pd.DataFrame({'Tissue_Type': ['CSF', 'WM', 'GM'], 'Overlap_Perc': [mean_csf, mean_wm, mean_gm], 'Std_Error': [csf_std_error, wm_std_error, gm_std_error]})
    group_overlap_perc_plot = (ggplot(plot_data, aes(x='Tissue_Type', y='Overlap_Perc', fill='Tissue_Type')) + 
                        geom_bar(stat='identity', position='dodge') +
                        geom_errorbar(aes(ymin='Overlap_Perc - Std_Error', ymax='Overlap_Perc + Std_Error'), width=0.2, color='black') +
                        theme_classic() +
                        labs(title='Sequence Tissue Type Overlap', x='Tissue Type', y='Tissue Overlap Percentage') +
                        scale_y_continuous(expand=(0, 0), limits=[0,100]) +
                        scale_fill_manual(values={'CSF': '#F9DC5C', 'WM': '#db5f57', 'GM': '#57d3db'})
                        )
    group_overlap_perc_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/group_overlap_perc_plot.png')
    ssim_values = group_ssim_df.loc[group_ssim_df['p_id'].isin(['P122', 'P136']), 'ssim_index'].tolist()
    plot_data = pd.DataFrame({
        'Participant': ['P122', 'P136'],
        'SSIM': ssim_values,
        'Overlap_Perc': overlap_perc_av_values
    })
    plot_data_sorted = plot_data.sort_values(by='Overlap_Perc', ascending=False)
    index_sorted = np.arange(len(plot_data_sorted))
    fig, ax1 = plt.subplots()
    bar_width = 0.35
    bar1 = ax1.bar(index_sorted, plot_data_sorted['SSIM'], bar_width, label='SSIM', color='#db5f57')
    ax1.set_ylabel('SSIM Index', color='#db5f57')
    ax1.tick_params(axis='y', labelcolor='#db5f57')
    ax1.set_ylim(0.94, 0.98)
    ax2 = ax1.twinx()
    bar2 = ax2.bar(index_sorted + bar_width, plot_data_sorted['Overlap_Perc'], bar_width, label='Overlap_Perc', color='#57d3db', alpha=1)
    ax2.set_ylabel(' Tissue Overlap Percentage', color='#57d3db')
    ax2.tick_params(axis='y', labelcolor='#57d3db')
    ax2.set_ylim(75, 95)
    ax1.set_xlabel('Participant')
    ax1.set_xticks(index_sorted + bar_width / 2)
    ax1.set_xticklabels(plot_data_sorted['Participant'], rotation=45, ha='right')
    plt.title('SSIM and Tissue Overlap Percentage Plot')
    fig.legend(loc='upper right', bbox_to_anchor=(1, 1), bbox_transform=ax1.transAxes)
    save_path = 'analysis/susceptibility_analysis/run_comparisons/1/group/ssim_overlap_perc_plot.png'
    plt.tight_layout()
    plt.savefig(save_path, bbox_inches='tight')

    column_headers = ['p_id', 'sequence', 'value']
    group_voxel_intensity_df = pd.DataFrame(columns = column_headers)   
    for p_id in good_participants:
        print(f'Running Stage 1 voxel signal intensity analysis for {p_id}...')
        def extract_voxel_intensities(epi_image_path, mask_image_path):
            epi_img = nib.load(epi_image_path)
            epi_data = epi_img.get_fdata()
            mask_img = nib.load(mask_image_path)
            mask_data = mask_img.get_fdata()
            mask_data = mask_data > 0
            roi_voxel_intensities = epi_data[mask_data]
            voxel_intensity_list = roi_voxel_intensities.tolist()
            return voxel_intensity_list
        flirted_pa_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_pa_fieldmaps.nii.gz'
        flirted_rl_fieldmaps = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/flirted_rl_fieldmaps.nii.gz'
        pa_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/pa_trimmed_roi_mask.nii.gz'
        rl_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/rl_trimmed_roi_mask.nii.gz'
        pa_voxel_intensities = extract_voxel_intensities(flirted_pa_fieldmaps, pa_trimmed_roi_mask)
        rl_voxel_intensities = extract_voxel_intensities(flirted_rl_fieldmaps, rl_trimmed_roi_mask)
        values = pa_voxel_intensities + rl_voxel_intensities
        sequence = ['pa'] * len(pa_voxel_intensities) + ['rl'] * len(rl_voxel_intensities)
        subject = [f'{p_id}'] * len(pa_voxel_intensities) + [f'{p_id}'] * len(rl_voxel_intensities)
        voxel_intensity_df = pd.DataFrame({'p_id': subject, 'sequence': sequence, 'value': values})
        voxel_intensity_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/1/{p_id}/voxel_intensity_df.txt', sep='\t', index=False)
        group_voxel_intensity_df = pd.concat([group_voxel_intensity_df, voxel_intensity_df], ignore_index=True)
    group_voxel_intensity_df.to_csv('analysis/susceptibility_analysis/run_comparisons/1/group/group_voxel_intensity_df.txt', sep='\t', index=False)
    pa_means = []
    rl_means= []
    p_values = []
    pa_std_errors = []
    rl_std_errors = []
    for p_id in good_participants:
        filtered_pa = group_voxel_intensity_df[(group_voxel_intensity_df['p_id'] == f'{p_id}') & (group_voxel_intensity_df['sequence'] == 'pa')]
        mean_value_pa = filtered_pa['value'].mean()
        pa_means.append(mean_value_pa)
        filtered_rl = group_voxel_intensity_df[(group_voxel_intensity_df['p_id'] == f'{p_id}') & (group_voxel_intensity_df['sequence'] == 'rl')]
        mean_value_rl = filtered_rl['value'].mean()
        rl_means.append(mean_value_rl)
        anderson_pa = stats.anderson(filtered_pa['value'])
        anderson_rl = stats.anderson(filtered_rl['value'])
        significance_level = 0.05
        is_pa_normal = anderson_pa.statistic < anderson_pa.critical_values[
            anderson_pa.significance_level.tolist().index(significance_level * 100)]
        is_rl_normal = anderson_rl.statistic < anderson_rl.critical_values[
            anderson_rl.significance_level.tolist().index(significance_level * 100)]
        if is_pa_normal and is_rl_normal:
            print(f'Anderson-Darling test passed for {p_id} voxel intensity values. Running parametric t-test...')
            _, p_value = stats.ttest_ind(filtered_pa['value'], filtered_rl['value'], equal_var=False)
            p_values.append(p_value)
        else:
            print(f'Anderson-Darling test failed for {p_id} voxel intensity values. Running non-parametric Mann Whitney U test...')
            _, p_value = stats.mannwhitneyu(filtered_pa['value'], filtered_rl['value'], alternative='two-sided')
            p_values.append(p_value)
        pa_std_error = np.std(filtered_pa['value']) / np.sqrt(len(filtered_pa['value']))
        pa_std_errors.append(pa_std_error)
        rl_std_error = np.std(filtered_rl['value']) / np.sqrt(len(filtered_rl['value']))
        rl_std_errors.append(rl_std_error)
    plot_data = pd.DataFrame({
        'Participant': good_participants * 2,
        'Mean_Value': pa_means + rl_means,
        'Sequence': ['PA'] * len(good_participants) + ['RL'] * len(good_participants),
        'Significance': ['' for _ in range(len(good_participants) * 2)],
        'Std_Error': pa_std_errors + rl_std_errors
    })
    for idx, p_value in enumerate(p_values):
        if p_value < 0.001:
            plot_data.at[idx, 'Significance'] = '***'
        elif p_value < 0.01:
            plot_data.at[idx, 'Significance'] = '**'
        elif p_value < 0.05:
            plot_data.at[idx, 'Significance'] = '*'
    voxel_intensity_plot = (
        ggplot(plot_data, aes(x='Participant', y='Mean_Value', fill='Sequence')) +
        geom_bar(stat='identity', position='dodge') +
        geom_errorbar(aes(ymin='Mean_Value - Std_Error', ymax='Mean_Value + Std_Error'), position=position_dodge(width=0.9), width=0.2, color='black') +
        theme_classic() +
        labs(title='Mean SCC Voxel Intensity', x='Participant', y='Mean SCC Signal Intensity') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='black'), axis_title=element_text(size=12)) +
        scale_y_continuous(expand=(0, 0), limits=[0,350]) +
        geom_text(
            aes(x='Participant', y='Mean_Value', label='Significance'),
            position=position_dodge(width=0.9),
            color='black',
            size=12,
            ha='center',
            va='bottom',
            show_legend=False))
    voxel_intensity_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/voxel_intensity_plot.png')
    pa_means_overall = np.mean(pa_means)
    rl_means_overall = np.mean(rl_means)
    pa_std_error_overall = np.std(pa_means) / np.sqrt(len(pa_means))
    rl_std_error_overall = np.std(rl_means) / np.sqrt(len(rl_means))
    _, pa_means_overall_shap_p = stats.shapiro(pa_means)
    _, rl_means_overall_shap_p = stats.shapiro(rl_means)
    if pa_means_overall_shap_p > 0.05 and rl_means_overall_shap_p > 0.05:
        print(f'Shapiro-Wilk test passed for voxel intensity values. Running parametric t-test...')
        _, p_value = stats.ttest_rel(pa_means, rl_means)
        print(f"T-test p-value: {p_value}")
    else:
        print(f'Shapiro-Wilk test failed for voxel intensity values. Running non-parametric Wilcoxon test...')
        _, p_value = stats.wilcoxon(pa_means, rl_means)
        print(f"Wilcoxon test p-value: {p_value}")
    plot_data = pd.DataFrame({'p_id': good_participants, 'pa_values': pa_means, 'rl_values': rl_means})
    data_long = pd.melt(plot_data, id_vars=['p_id'], value_vars=['pa_values', 'rl_values'], var_name='sequence', value_name='value')
    data_long['sequence'] = data_long['sequence'].map({'pa_values': 'PA', 'rl_values': 'RL'})
    group_voxel_intensity_ladder_plot = (
        ggplot(data_long, aes(x='sequence', y='value', group='p_id')) +
        geom_line(aes(group='p_id'), color='gray', size=1) +
        geom_point(aes(color='sequence'), size=4) +
        theme_light() +
        theme(
            panel_grid_major=element_blank(), 
            panel_grid_minor=element_blank(), 
            panel_border=element_blank(),
            axis_line_x=element_line(color='black'),  
            axis_line_y=element_line(color='black'),  
        ) +
        labs(title='Ladder Plot of PA and RL Sequences',
            x='Sequence',
            y='Mean SCC Signal Intensity') +
        scale_x_discrete(limits=['PA', 'RL']) +
        scale_y_continuous()
    )
    group_voxel_intensity_ladder_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/group_voxel_intensity_ladder_plot.png')                          
    plot_data = pd.DataFrame({'Sequence': ['PA', 'RL'], 'Mean': [pa_means_overall, rl_means_overall], 'Std_Error': [pa_std_error_overall, rl_std_error_overall]})
    group_voxel_intensity_plot = (ggplot(plot_data, aes(x='Sequence', y='Mean', fill='Sequence')) + 
                        geom_bar(stat='identity', position='dodge') +
                        geom_errorbar(aes(ymin='Mean - Std_Error', ymax='Mean + Std_Error'), width=0.2, color='black') +
                        theme_classic() +
                        labs(title='Mean of Voxel Intensities Across Participants', y='Mean SCC Signal Intensity') +
                        scale_y_continuous(expand=(0, 0), limits=[0,350]) +
                        scale_fill_manual(values={'PA': '#DB5F57', 'RL': '#57D3DB'})
                        )
    if p_value < 0.001:
        group_voxel_intensity_plot = group_voxel_intensity_plot + annotate("text", x=1.5, y=max(plot_data['Mean']) + 40, label="***", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Mean']) +30, yend=max(plot_data['Mean']) + 30, color="black")
    elif p_value < 0.01:
        group_voxel_intensity_plot = group_voxel_intensity_plot + annotate("text", x=1.5, y=max(plot_data['Mean']) + 40, label="**", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Mean']) +30, yend=max(plot_data['Mean']) + 30, color="black")
    elif p_value < 0.05:
        group_voxel_intensity_plot = group_voxel_intensity_plot + annotate("text", x=1.5, y=max(plot_data['Mean']) + 40, label="*", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Mean']) +30, yend=max(plot_data['Mean']) + 30, color="black")    
    group_voxel_intensity_plot.save('analysis/susceptibility_analysis/run_comparisons/1/group/group_voxel_intensity_plot.png')

    # Step 7: Test quality of alternate distortion correction method (Stage 2).
    print("\n###### STEP 7: TESTING ALTERNATE DISTORTION CORRECTION METHOD (STAGE 2) ######")
    perc_outside_corrected_values = []
    perc_outside_uncorrected_values = []
    column_headers = ['p_id', 'perc_outside_corrected', 'perc_outside_uncorrected']
    group_perc_outside_df = pd.DataFrame(columns = column_headers) 
    for p_id in good_participants:  
        print(f'Preparing Stage 2 files for {p_id}...') 
        run_comparison_2_folder = 'analysis/susceptibility_analysis/run_comparisons/2'
        os.makedirs(run_comparison_2_folder, exist_ok=True)
        run_comparison_2_participant_folder = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}'
        os.makedirs(run_comparison_2_participant_folder, exist_ok=True)
        run_comparison_2_group_folder = 'analysis/susceptibility_analysis/run_comparisons/2/group'
        os.makedirs(run_comparison_2_group_folder, exist_ok=True)
        averaged_run = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/averaged_run.nii.gz'
        if not os.path.exists(averaged_run):
            run = f'analysis/susceptibility_analysis/data/{p_id}/niftis/run01.nii'
            subprocess.run(['fslmaths', run, '-Tmean', averaged_run])
        uncorrected_run = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/uncorrected_run.nii.gz'
        if not os.path.exists(uncorrected_run):
            subprocess.run(["bet", averaged_run, uncorrected_run, "-m", "-R"])
        corrected_run = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/corrected_run.nii.gz'
        if not os.path.exists(corrected_run):
            subprocess.run(["applytopup", f"--imain={uncorrected_run}", f"--datain=analysis/susceptibility_analysis/data/{p_id}/acqparams.txt", "--inindex=6", f"--topup=analysis/susceptibility_analysis/data/{p_id}/topup_{p_id}", "--method=jac", f"--out={corrected_run}"])
        flirted_corrected_run = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_corrected_run.nii.gz'
        flirted_uncorrected_run = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_uncorrected_run.nii.gz'
        flirted_corrected_run_transformation = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_corrected_run_transformation.mat'
        flirted_uncorrected_run_transformation = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_uncorrected_run_transformation.mat'
        if not os.path.exists(flirted_corrected_run):
            structural_brain = f'analysis/susceptibility_analysis/data/{p_id}/niftis/structural_brain.nii'
            subprocess.run(["flirt", "-in", corrected_run, "-ref", structural_brain, "-out", flirted_corrected_run, "-omat", flirted_corrected_run_transformation])
            subprocess.run(["flirt", "-in", uncorrected_run, "-ref", structural_brain, "-out", flirted_uncorrected_run, "-omat", flirted_uncorrected_run_transformation])
        def read_roi_file(roi_file):
            voxel_coordinates = []
            with open(roi_file, 'r') as file:
                content = file.read()
                matches = re.findall(r'(?<=\n)\s*\d+\s+\d+\s+\d+', content)
                for match in matches:
                    coordinates = match.split()
                    voxel_coordinates.append(
                        (int(coordinates[0]), int(coordinates[1]), int(coordinates[2])))
            return voxel_coordinates
        path = f'data/raw_data/{p_id}/data/neurofeedback'
        cisc_folder = None
        for folder_name in os.listdir(path):
            if "CISC" in folder_name:
                cisc_folder = folder_name
                break
        if cisc_folder is None:
            print("No 'CISC' folder found in the 'neurofeedback' directory.")
            exit(1)
        roi_file = f'data/raw_data/{p_id}/data/neurofeedback/{cisc_folder}/depression_neurofeedback/target_folder_run-1/depnf_run-1.roi'        
        voxel_coordinates = read_roi_file(roi_file)
        functional_image_info = nib.load(averaged_run)
        functional_dims = functional_image_info.shape
        binary_volume = np.zeros(functional_dims)
        for voxel in voxel_coordinates:
            x, y, z = voxel
            binary_volume[x, y, z] = 1
        binary_volume = np.flip(binary_volume, axis=1)
        functional_affine = functional_image_info.affine
        binary_mask = nib.Nifti1Image(binary_volume, affine=functional_affine)
        nib.save(binary_mask, f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/run01_subject_space_ROI.nii.gz')
        roi_mask = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/run01_subject_space_ROI.nii.gz'
        transformed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/transformed_roi_mask.nii.gz'
        subprocess.run(['flirt', '-in', roi_mask, '-ref', structural_brain, '-applyxfm', '-init', flirted_uncorrected_run_transformation, '-out', transformed_roi_mask, '-interp', 'nearestneighbour'])
        flirted_corrected_bin = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_corrected_bin.nii.gz'
        if not os.path.exists(flirted_corrected_bin):
            subprocess.run(['fslmaths', flirted_corrected_run, '-thr', '100', '-bin', flirted_corrected_bin])
        flirted_uncorrected_bin = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_uncorrected_bin.nii.gz'
        if not os.path.exists(flirted_uncorrected_bin):
            subprocess.run(['fslmaths', flirted_uncorrected_run, '-thr', '100', '-bin', flirted_uncorrected_bin])
        corrected_bin_inv = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/corrected_bin_inv.nii.gz'
        if not os.path.exists(corrected_bin_inv):
            subprocess.run(['fslmaths', flirted_corrected_bin, '-sub', '1', '-abs', corrected_bin_inv])
        uncorrected_bin_inv = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/uncorrected_bin_inv.nii.gz'
        if not os.path.exists(uncorrected_bin_inv):
            subprocess.run(['fslmaths', flirted_uncorrected_bin, '-sub', '1', '-abs', uncorrected_bin_inv])
        corrected_result = subprocess.run(['fslstats', transformed_roi_mask, '-k', corrected_bin_inv, '-V'], capture_output=True, text=True)
        if corrected_result.returncode == 0:
            corrected_result_output = corrected_result.stdout.strip()
        else:
            print("Error executing fslstats command.")
        corrected_result_output_values = corrected_result_output.split()
        corrected_voxels_outside = float(corrected_result_output_values[0])
        uncorrected_result = subprocess.run(['fslstats', transformed_roi_mask, '-k', uncorrected_bin_inv, '-V'], capture_output=True, text=True)
        if uncorrected_result.returncode == 0:
            uncorrected_result_output = uncorrected_result.stdout.strip()
        else:
            print("Error executing fslstats command.")
        uncorrected_result_output_values = uncorrected_result_output.split()
        uncorrected_voxels_outside = float(uncorrected_result_output_values[0])
        result1 = subprocess.run(['fslstats', transformed_roi_mask, '-V'], capture_output=True, text=True)
        if result1.returncode == 0:
            result1_output = result1.stdout.strip()
        else:
            print("Error executing fslstats command.")
        result1_output_values = result1_output.split()
        total_voxels_in_roi = float(result1_output_values[0])
        perc_outside_corrected = (corrected_voxels_outside / total_voxels_in_roi) * 100
        perc_outside_corrected = round(perc_outside_corrected, 2)
        perc_outside_corrected_values.append(perc_outside_corrected)
        perc_outside_uncorrected = (uncorrected_voxels_outside / total_voxels_in_roi) * 100
        perc_outside_uncorrected = round(perc_outside_uncorrected, 2)
        perc_outside_uncorrected_values.append(perc_outside_uncorrected)
        perc_outside_df = pd.DataFrame({'p_id': [p_id], 'perc_outside_corrected': [perc_outside_corrected], 'perc_outside_uncorrected': [perc_outside_uncorrected]})
        perc_outside_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/perc_outside_df.txt', sep='\t', index=False)
        group_perc_outside_df = pd.concat([group_perc_outside_df, perc_outside_df], ignore_index=True)
        corrected_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/corrected_trimmed_roi_mask.nii.gz'
        uncorrected_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/uncorrected_trimmed_roi_mask.nii.gz'
        if not os.path.exists(corrected_trimmed_roi_mask) or not os.path.exists(uncorrected_trimmed_roi_mask):
            subprocess.run(['fslmaths', transformed_roi_mask, '-mul', flirted_corrected_bin, corrected_trimmed_roi_mask])
            subprocess.run(['fslmaths', transformed_roi_mask, '-mul', flirted_uncorrected_bin, uncorrected_trimmed_roi_mask])
    group_perc_outside_df.to_csv('analysis/susceptibility_analysis/run_comparisons/2/group/group_perc_outside_df.txt', sep='\t', index=False)
    plot_data = pd.DataFrame({
        'Participant': good_participants * 2,
        'Perc_Outside': perc_outside_corrected_values + perc_outside_uncorrected_values,
        'Sequence': ['corrected'] * len(good_participants) + ['uncorrected'] * len(good_participants)
    })
    perc_outside_plot = (
        ggplot(plot_data, aes(x='Participant', y='Perc_Outside', fill='Sequence')) +
        geom_bar(stat='identity', position='dodge') +
        theme_classic() +
        labs(title='Percentage of Voxels in Signal Dropout Regions', x='Participant', y='Percentage of SCC Voxels in Signal Dropout') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='black'), axis_title=element_text(size=12)) +
        scale_y_continuous(expand=(0, 0))
    )
    perc_outside_plot.save('analysis/susceptibility_analysis/run_comparisons/2/group/perc_outside_plot.png')
    perc_outside_corrected_overall = np.mean(perc_outside_corrected_values)
    perc_outside_uncorrected_overall = np.mean(perc_outside_uncorrected_values)
    corrected_std_error = np.std(perc_outside_corrected_values) / np.sqrt(len(perc_outside_corrected_values))
    uncorrected_std_error = np.std(perc_outside_uncorrected_values) / np.sqrt(len(perc_outside_uncorrected_values))
    _, perc_outside_corrected_overall_shap_p = stats.shapiro(perc_outside_corrected_values)
    _, perc_outside_uncorrected_overall_shap_p = stats.shapiro(perc_outside_uncorrected_values)
    if perc_outside_corrected_overall_shap_p > 0.05 and perc_outside_uncorrected_overall_shap_p > 0.05:
        print(f'Shapiro-Wilk test passed for perc_outside values. Running parametric t-test...')
        _, p_value = stats.ttest_ind(perc_outside_corrected_values, perc_outside_uncorrected_values)
        print(f"T-test p-value: {p_value}")
    else:
        print(f'Shapiro-Wilk test failed for perc_outside values. Running non-parametric Mann-Whitney U test...')
        _, p_value = stats.mannwhitneyu(perc_outside_corrected_values, perc_outside_uncorrected_values)
        print(f"Mann-Whitney U test p-value: {p_value}")
    plot_data = pd.DataFrame({'Sequence': ['corrected', 'uncorrected'], 'Perc_Outside': [perc_outside_corrected_overall, perc_outside_uncorrected_overall], 'Std_Error': [corrected_std_error, uncorrected_std_error]})
    group_perc_outside_plot = (ggplot(plot_data, aes(x='Sequence', y='Perc_Outside', fill='Sequence')) + 
                        geom_bar(stat='identity', position='dodge') +
                        geom_errorbar(aes(ymin='Perc_Outside - Std_Error', ymax='Perc_Outside + Std_Error'), width=0.2, color='black') +
                        theme_classic() +
                        labs(title='Percentage of Voxels in Signal Dropout Regions', y='Percentage of SCC Voxels in Signal Dropout') +
                        scale_y_continuous(expand=(0, 0)) +
                        scale_fill_manual(values={'PA': '#DB5F57', 'RL': '#57D3DB'})
                        )
    if p_value < 0.001:
        group_perc_outside_plot = group_perc_outside_plot + annotate("text", x=1.5, y=max(plot_data['Perc_Outside']) + 3.5, label="***", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Perc_Outside']) + 3, yend=max(plot_data['Perc_Outside']) + 3, color="black")
    elif p_value < 0.01:
        group_perc_outside_plot = group_perc_outside_plot + annotate("text", x=1.5, y=max(plot_data['Perc_Outside']) + 3.5, label="**", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Perc_Outside']) + 3, yend=max(plot_data['Perc_Outside']) + 3, color="black")
    elif p_value < 0.05:
        group_perc_outside_plot = group_perc_outside_plot + annotate("text", x=1.5, y=max(plot_data['Perc_Outside']) + 3.5, label="*", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Perc_Outside']) + 3, yend=max(plot_data['Perc_Outside']) + 3, color="black")    
    group_perc_outside_plot.save('analysis/susceptibility_analysis/run_comparisons/2/group/group_perc_outside_plot.png')

    column_headers = ['p_id', 'ssim_index', 'voxels_in_bin_ssim_mask', 'perc_roi_voxels_in_bin_ssim_mask']
    group_ssim_df = pd.DataFrame(columns = column_headers)   
    for p_id in good_participants:
        print(f'Running Stage 2 SSIM analysis for {p_id}...')        
        def calculate_ssim(image1_path, image2_path, ssim_output_path):
            """Function to calculate SSIM between two NIfTI images and save the SSIM map."""
            image1_nii = nib.load(image1_path)
            image2_nii = nib.load(image2_path)
            image1 = image1_nii.get_fdata()
            image2 = image2_nii.get_fdata()
            if image1.shape != image2.shape:
                raise ValueError("Input images must have the same dimensions for SSIM calculation.")
            ssim_index, ssim_map = ssim(image1, image2, full=True, data_range=image1.max() - image1.min())
            ssim_map_nifti = nib.Nifti1Image(ssim_map, affine=image1_nii.affine, header=image1_nii.header)
            nib.save(ssim_map_nifti, ssim_output_path)
            return ssim_index
        ssim_output_path = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/ssim_map.nii.gz'
        flirted_corrected_run = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_corrected_run.nii.gz'
        flirted_uncorrected_run = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_uncorrected_run.nii.gz'
        if not os.path.exists(ssim_output_path):
            ssim_index = calculate_ssim(flirted_uncorrected_run, flirted_corrected_run, ssim_output_path)       
        ssim_bin = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/ssim_bin.nii.gz'
        if not os.path.exists(ssim_bin):
            subprocess.run(["fslmaths", ssim_output_path, "-thr", "0.8", "-binv", ssim_bin])
        combined_corr_uncorr_mask = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/combined_corr_uncorr_mask.nii.gz'
        flirted_corrected_bin = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_corrected_bin.nii.gz'
        flirted_uncorrected_bin = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_uncorrected_bin.nii.gz'
        if not os.path.exists(combined_corr_uncorr_mask):
            subprocess.run(['fslmaths', flirted_corrected_bin, '-add', flirted_uncorrected_bin, combined_corr_uncorr_mask])
        bin_corr_uncorr_mask = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/bin_corr_uncorr_mask.nii.gz'
        if not os.path.exists(bin_corr_uncorr_mask):
            subprocess.run(['fslmaths', combined_corr_uncorr_mask, '-bin', bin_corr_uncorr_mask])
        ssim_bin_trimmed = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/ssim_bin_trimmed.nii.gz'
        if not os.path.exists(ssim_bin_trimmed):
            subprocess.run(['fslmaths', ssim_bin, '-mul', bin_corr_uncorr_mask, ssim_bin_trimmed])
        voxels_in_whole_mask = subprocess.run(["fslstats", ssim_bin_trimmed, "-V"], capture_output=True, text=True).stdout.split()[0]
        voxels_in_whole_mask = float(voxels_in_whole_mask)
        intersection_mask_path = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/ssim_roi_intersect.nii.gz'
        transformed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/transformed_roi_mask.nii.gz'
        if not os.path.exists(intersection_mask_path):
            subprocess.run(["fslmaths", ssim_bin_trimmed, "-mas", transformed_roi_mask, intersection_mask_path])
        voxels_in_roi_in_mask = subprocess.run(["fslstats", intersection_mask_path, "-V"], capture_output=True, text=True).stdout.split()[0]
        voxels_in_roi_in_mask = float(voxels_in_roi_in_mask)
        perc_roi_voxels_in_mask = (voxels_in_roi_in_mask / total_voxels_in_roi) * 100
        ssim_df = pd.DataFrame({'p_id': [p_id], 'ssim_index': [ssim_index], 'voxels_in_bin_ssim_mask': [voxels_in_whole_mask], 'perc_roi_voxels_in_bin_ssim_mask': [perc_roi_voxels_in_mask]})
        ssim_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/ssim_df.txt', sep='\t', index=False)
        group_ssim_df = pd.concat([group_ssim_df, ssim_df], ignore_index=True)
    group_ssim_df.to_csv('analysis/susceptibility_analysis/run_comparisons/2/group/group_ssim_df.txt', sep='\t', index=False)
    ssim_indexes = group_ssim_df['ssim_index'].tolist()
    ssim_mean = np.mean(ssim_indexes)
    print(f"Mean SSIM index for Stage 2: {ssim_mean}")
    plot_data = pd.DataFrame({
        'Participant': good_participants,
        'SSIM': ssim_indexes,
    })
    ssim_index_plot = (
        ggplot(plot_data, aes(x='Participant', y='SSIM')) +
        geom_bar(stat='identity', position='dodge') +
        geom_hline(yintercept=ssim_mean, linetype='dashed', color='red') +
        theme_classic() +
        labs(title='SSIM Indexes', x='Participant', y='SSIM') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='blue'), axis_title=element_text(size=12, face='bold')) +
        scale_y_continuous(expand=(0, 0), limits=[0.8,1])
    )
    ssim_index_plot.save('analysis/susceptibility_analysis/run_comparisons/2/group/ssim_index_plot.png')
    voxels = group_ssim_df['voxels_in_bin_ssim_mask'].tolist()
    voxels_mean = np.mean(voxels)
    plot_data = pd.DataFrame({
        'Participant': good_participants,
        'Voxels': voxels,
    })
    ssim_voxels_plot = (
        ggplot(plot_data, aes(x='Participant', y='Voxels')) +
        geom_bar(stat='identity', position='dodge') +
        geom_hline(yintercept=voxels_mean, linetype='dashed', color='red') +
        theme_classic() +
        labs(title='Number of Voxels in SSIM Mask', x='Participant', y='Voxels') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='blue'), axis_title=element_text(size=12, face='bold')) +
        scale_y_continuous(expand=(0, 0))
    )
    ssim_voxels_plot.save('analysis/susceptibility_analysis/run_comparisons/2/group/ssim_voxels_plot.png')
    perc_voxels = group_ssim_df['perc_roi_voxels_in_bin_ssim_mask'].tolist()
    perc_voxels_mean = np.mean(perc_voxels)
    plot_data = pd.DataFrame({
        'Participant': good_participants,
        'Perc_Voxels': perc_voxels,
    })
    ssim_perc_voxels_plot = (
        ggplot(plot_data, aes(x='Participant', y='Perc_Voxels')) +
        geom_bar(stat='identity', position='dodge') +
        geom_hline(yintercept=perc_voxels_mean, linetype='dashed', color='red') +
        theme_classic() +
        labs(title='Percentage of ROI Voxels in SSIM Mask', x='Participant', y='Percentage of SCC Voxels in SSIM Mask') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='black'), axis_title=element_text(size=12)) +
        scale_y_continuous(expand=(0, 0))
    )
    ssim_perc_voxels_plot.save('analysis/susceptibility_analysis/run_comparisons/2/group/ssim_perc_voxels_plot.png')

    column_headers = ['p_id', 'sequence', 'value']
    group_voxel_intensity_df = pd.DataFrame(columns = column_headers)   
    for p_id in good_participants:
        print(f'Running Stage 2 voxel signal intensity analysis for {p_id}...')         
        def extract_voxel_intensities(epi_image_path, mask_image_path):
            epi_img = nib.load(epi_image_path)
            epi_data = epi_img.get_fdata()
            mask_img = nib.load(mask_image_path)
            mask_data = mask_img.get_fdata()
            mask_data = mask_data > 0
            roi_voxel_intensities = epi_data[mask_data]
            voxel_intensity_list = roi_voxel_intensities.tolist()
            return voxel_intensity_list
        flirted_corrected_run = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_corrected_run.nii.gz'
        flirted_uncorrected_run = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/flirted_uncorrected_run.nii.gz'
        corrected_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/corrected_trimmed_roi_mask.nii.gz'
        uncorrected_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/uncorrected_trimmed_roi_mask.nii.gz'
        corrected_voxel_intensities = extract_voxel_intensities(flirted_corrected_run, corrected_trimmed_roi_mask)
        uncorrected_voxel_intensities = extract_voxel_intensities(flirted_uncorrected_run, uncorrected_trimmed_roi_mask)
        values = corrected_voxel_intensities + uncorrected_voxel_intensities
        sequence = ['corrected'] * len(corrected_voxel_intensities) + ['uncorrected'] * len(uncorrected_voxel_intensities)
        subject = [f'{p_id}'] * len(corrected_voxel_intensities) + [f'{p_id}'] * len(uncorrected_voxel_intensities)
        voxel_intensity_df = pd.DataFrame({'p_id': subject, 'sequence': sequence, 'value': values})
        voxel_intensity_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/2/{p_id}/voxel_intensity_df.txt', sep='\t', index=False)
        group_voxel_intensity_df = pd.concat([group_voxel_intensity_df, voxel_intensity_df], ignore_index=True)
    group_voxel_intensity_df.to_csv('analysis/susceptibility_analysis/run_comparisons/2/group/group_voxel_intensity_df.txt', sep='\t', index=False)
    corrected_means = []
    uncorrected_means= []
    p_values = []
    corrected_std_errors = []
    uncorrected_std_errors = []
    for p_id in good_participants:
        filtered_corrected = group_voxel_intensity_df[(group_voxel_intensity_df['p_id'] == f'{p_id}') & (group_voxel_intensity_df['sequence'] == 'corrected')]
        mean_value_corrected = filtered_corrected['value'].mean()
        corrected_means.append(mean_value_corrected)
        filtered_uncorrected = group_voxel_intensity_df[(group_voxel_intensity_df['p_id'] == f'{p_id}') & (group_voxel_intensity_df['sequence'] == 'uncorrected')]
        mean_value_uncorrected = filtered_uncorrected['value'].mean()
        uncorrected_means.append(mean_value_uncorrected)
        anderson_corrected = stats.anderson(filtered_corrected['value'])
        print(f"Anderson-Darling test for corrected values: Statistic={anderson_corrected.statistic}, Critical Values={anderson_corrected.critical_values}, Significance Levels={anderson_corrected.significance_level}")
        anderson_uncorrected = stats.anderson(filtered_uncorrected['value'])
        print(f"Anderson-Darling test for uncorrected values: Statistic={anderson_uncorrected.statistic}, Critical Values={anderson_uncorrected.critical_values}, Significance Levels={anderson_uncorrected.significance_level}")
        significance_level = 0.05
        is_corrected_normal = anderson_corrected.statistic < anderson_corrected.critical_values[
            anderson_corrected.significance_level.tolist().index(significance_level * 100)]
        is_uncorrected_normal = anderson_uncorrected.statistic < anderson_uncorrected.critical_values[
            anderson_uncorrected.significance_level.tolist().index(significance_level * 100)]
        if is_corrected_normal and is_uncorrected_normal:
            print(f'Anderson-Darling test passed for {p_id} voxel intensity values. Running parametric t-test...')
            _, p_value = stats.ttest_ind(filtered_corrected['value'], filtered_uncorrected['value'], equal_var=False)
            p_values.append(p_value)
        else:
            print(f'Anderson-Darling test failed for {p_id} voxel intensity values. Running non-parametric Mann-Whitney U test...')
            _, p_value = stats.mannwhitneyu(filtered_corrected['value'], filtered_uncorrected['value'], alternative='two-sided')
            p_values.append(p_value)
        corrected_std_error = np.std(filtered_corrected['value']) / np.sqrt(len(filtered_corrected['value']))
        corrected_std_errors.append(corrected_std_error)
        uncorrected_std_error = np.std(filtered_uncorrected['value']) / np.sqrt(len(filtered_uncorrected['value']))
        uncorrected_std_errors.append(uncorrected_std_error)
    plot_data = pd.DataFrame({
        'Participant': good_participants * 2,
        'Mean_Value': corrected_means + uncorrected_means,
        'Sequence': ['Corrected'] * len(good_participants) + ['Uncorrected'] * len(good_participants),
        'Significance': ['' for _ in range(len(good_participants) * 2)],
        'Std_Error': corrected_std_errors + uncorrected_std_errors
    })
    for idx, p_value in enumerate(p_values):
        if p_value < 0.001:
            plot_data.at[idx, 'Significance'] = '***'
        elif p_value < 0.01:
            plot_data.at[idx, 'Significance'] = '**'
        elif p_value < 0.05:
            plot_data.at[idx, 'Significance'] = '*'
    voxel_intensity_plot = (
        ggplot(plot_data, aes(x='Participant', y='Mean_Value', fill='Sequence')) +
        geom_bar(stat='identity', position='dodge') +
        geom_errorbar(aes(ymin='Mean_Value - Std_Error', ymax='Mean_Value + Std_Error'), position=position_dodge(width=0.9), width=0.2, color='black') +
        theme_classic() +
        labs(title='Mean SCC Voxel Intensity', x='Participant', y='Mean SCC Signal Intensity') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='black'), axis_title=element_text(size=12)) +
        scale_y_continuous(expand=(0, 0), limits=[0,350]) +
        geom_text(
            aes(x='Participant', y='Mean_Value', label='Significance'),
            position=position_dodge(width=0.9),
            color='black',
            size=12,
            ha='center',
            va='bottom',
            show_legend=False))
    voxel_intensity_plot.save('analysis/susceptibility_analysis/run_comparisons/2/group/voxel_intensity_plot.png')
    corrected_means_overall = np.mean(corrected_means)
    uncorrected_means_overall = np.mean(uncorrected_means)
    corrected_std_error_overall = np.std(corrected_means) / np.sqrt(len(corrected_means))
    uncorrected_std_error_overall = np.std(uncorrected_means) / np.sqrt(len(uncorrected_means))
    _, corrected_means_overall_shap_p = stats.shapiro(corrected_means)
    _, uncorrected_means_overall_shap_p = stats.shapiro(uncorrected_means)
    if corrected_means_overall_shap_p > 0.05 and uncorrected_means_overall_shap_p > 0.05:
        print(f'Shapiro-Wilk test passed for voxel intensity values. Running parametric t-test...')
        _, p_value = stats.ttest_rel(corrected_means, uncorrected_means)
        print(f"T-test p-value: {p_value}")
    else:
        print(f'Shapiro-Wilk test failed for voxel intensity values. Running non-parametric Wilcoxon test...')
        _, p_value = stats.wilcoxon(corrected_means, uncorrected_means)
        print(f"Wilcoxon test p-value: {p_value}")
    plot_data = pd.DataFrame({'p_id': good_participants, 'corrected_values': corrected_means, 'uncorrected_values': uncorrected_means})
    data_long = pd.melt(plot_data, id_vars=['p_id'], value_vars=['corrected_values', 'uncorrected_values'], var_name='sequence', value_name='value')
    data_long['sequence'] = data_long['sequence'].map({'corrected_values': 'CORR', 'uncorrected_values': 'UNCORR'})
    group_voxel_intensity_ladder_plot = (
        ggplot(data_long, aes(x='sequence', y='value', group='p_id')) +
        geom_line(aes(group='p_id'), color='gray', size=1) +
        geom_point(aes(color='sequence'), size=4) +
        theme_light() +
        theme(
            panel_grid_major=element_blank(), 
            panel_grid_minor=element_blank(), 
            panel_border=element_blank(),
            axis_line_x=element_line(color='black'),  
            axis_line_y=element_line(color='black'),  
        ) +
        labs(title='Ladder Plot of Corrected and Uncorrected PA Sequence',
            x='Sequence',
            y='Mean SCC Signal Intensity') +
        scale_x_discrete(limits=['CORR', 'UNCORR']) +
        scale_y_continuous()
    )
    group_voxel_intensity_ladder_plot.save('analysis/susceptibility_analysis/run_comparisons/2/group/group_voxel_intensity_ladder_plot.png')
    plot_data = pd.DataFrame({'Sequence': ['Corrected', 'Uncorrected'], 'Mean': [corrected_means_overall, uncorrected_means_overall], 'Std_Error': [corrected_std_error_overall, uncorrected_std_error_overall]})
    group_voxel_intensity_plot = (ggplot(plot_data, aes(x='Sequence', y='Mean', fill='Sequence')) + 
                        geom_bar(stat='identity', position='dodge') +
                        geom_errorbar(aes(ymin='Mean - Std_Error', ymax='Mean + Std_Error'), width=0.2, color='black') +
                        theme_classic() +
                        labs(title='Mean of Voxel Intensities Across Participants', y='Mean SCC Signal Intensity') +
                        scale_y_continuous(expand=(0, 0), limits=[0,350]) +
                        scale_fill_manual(values={'PA': '#DB5F57', 'RL': '#57D3DB'})
                        )
    if p_value < 0.001:
        group_voxel_intensity_plot = group_voxel_intensity_plot + annotate("text", x=1.5, y=max(plot_data['Mean']) + 40, label="***", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Mean']) +30, yend=max(plot_data['Mean']) + 30, color="black")
    elif p_value < 0.01:
        group_voxel_intensity_plot = group_voxel_intensity_plot + annotate("text", x=1.5, y=max(plot_data['Mean']) + 40, label="**", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Mean']) +30, yend=max(plot_data['Mean']) + 30, color="black")
    elif p_value < 0.05:
        group_voxel_intensity_plot = group_voxel_intensity_plot + annotate("text", x=1.5, y=max(plot_data['Mean']) + 40, label="*", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Mean']) +30, yend=max(plot_data['Mean']) + 30, color="black")    
    group_voxel_intensity_plot.save('analysis/susceptibility_analysis/run_comparisons/2/group/group_voxel_intensity_plot.png')

    # Step 8: Test quality of alternate distortion correction method (Stage 3).
    print("\n###### STEP 8: TESTING ALTERNATE DISTORTION CORRECTION METHOD (STAGE 3) ######")   
    bad_participants = ['P004', 'P006', 'P020', 'P030', 'P078', 'P093', 'P094']
    perc_outside_run01_values = []
    perc_outside_run04_values = []
    column_headers = ['p_id', 'perc_outside_run01', 'perc_outside_run04']
    group_perc_outside_df = pd.DataFrame(columns = column_headers) 
    for p_id in bad_participants:
        print(f"Preparing Stage 3 files for {p_id}...")
        run_comparison_3_folder = 'analysis/susceptibility_analysis/run_comparisons/3'
        os.makedirs(run_comparison_3_folder, exist_ok=True)
        run_comparison_3_participant_folder = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}'
        os.makedirs(run_comparison_3_participant_folder, exist_ok=True)
        run_comparison_3_group_folder = 'analysis/susceptibility_analysis/run_comparisons/3/group'
        os.makedirs(run_comparison_3_group_folder, exist_ok=True)
        run01 = f'analysis/susceptibility_analysis/data/{p_id}/niftis/run01.nii'
        run04 = f'analysis/susceptibility_analysis/data/{p_id}/niftis/run04.nii'
        averaged_run01 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/averaged_run01.nii.gz'
        averaged_run04 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/averaged_run04.nii.gz'
        if not os.path.exists(averaged_run01) or not os.path.exists(averaged_run04):
            subprocess.run(['fslmaths', run01, '-Tmean', averaged_run01])
            subprocess.run(['fslmaths', run04, '-Tmean', averaged_run04])
        betted_run01 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/betted_run01.nii.gz'
        betted_run04 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/betted_run04.nii.gz'
        if not os.path.exists(betted_run01) or not os.path.exists(betted_run04):
            subprocess.run(["bet", averaged_run01, betted_run01, "-m", "-R"])
            subprocess.run(["bet", averaged_run04, betted_run04, "-m", "-R"])
        flirted_run01 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run01.nii.gz'
        flirted_run04 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run04.nii.gz'
        t1_flirted_run01_transformation = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/t1_flirted_run01_transformation.mat'
        t1_flirted_run04_transformation = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/t1_flirted_run04_transformation.mat'
        structural_brain = f'analysis/susceptibility_analysis/data/{p_id}/niftis/structural_brain.nii'
        if not os.path.exists(flirted_run01):
            subprocess.run(["flirt", "-in", betted_run01, "-ref", structural_brain, "-out", flirted_run01, "-omat", t1_flirted_run01_transformation])
            subprocess.run(["flirt", "-in", betted_run04, "-ref", structural_brain, "-out", flirted_run04, "-omat", t1_flirted_run04_transformation])
        def read_roi_file(roi_file):
            voxel_coordinates = []
            with open(roi_file, 'r') as file:
                content = file.read()
                matches = re.findall(r'(?<=\n)\s*\d+\s+\d+\s+\d+', content)
                for match in matches:
                    coordinates = match.split()
                    voxel_coordinates.append(
                        (int(coordinates[0]), int(coordinates[1]), int(coordinates[2])))
            return voxel_coordinates
        path = f'data/raw_data/{p_id}/data/neurofeedback'
        cisc_folder = None
        for folder_name in os.listdir(path):
            if "CISC" in folder_name:
                cisc_folder = folder_name
                break
        if cisc_folder is None:
            print("No 'CISC' folder found in the 'neurofeedback' directory.")
            exit(1)
        roi_file_run01 = f'data/raw_data/{p_id}/data/neurofeedback/{cisc_folder}/depression_neurofeedback/target_folder_run-1/depnf_run-1.roi'
        voxel_coordinates_run01 = read_roi_file(roi_file_run01)
        run01_template = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run01_template.nii.gz'
        if not os.path.exists(run01_template):
            run = f'analysis/susceptibility_analysis/data/{p_id}/niftis/run01.nii'
            subprocess.run(['fslmaths', run, '-Tmean', run01_template])
        functional_image_info = nib.load(run01_template)
        functional_dims = functional_image_info.shape
        binary_volume = np.zeros(functional_dims)
        for voxel in voxel_coordinates:
            x, y, z = voxel
            binary_volume[x, y, z] = 1
        binary_volume = np.flip(binary_volume, axis=1)
        functional_affine = functional_image_info.affine
        binary_mask = nib.Nifti1Image(binary_volume, affine=functional_affine)
        nib.save(binary_mask, f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run01_subject_space_ROI.nii.gz')
        roi_mask_run01 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run01_subject_space_ROI.nii.gz'
        flirted_roi_run01 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_roi_run01.nii.gz'
        if not os.path.exists(flirted_roi_run01):
            subprocess.run(['flirt', '-in', roi_mask_run01, '-ref', structural_brain, '-applyxfm', '-init', t1_flirted_run01_transformation, '-out', flirted_roi_run01, '-interp', 'nearestneighbour'])
        roi_file_run04 = f'data/raw_data/{p_id}/data/neurofeedback/{cisc_folder}/depression_neurofeedback/target_folder_run-4/depnf_run-4.roi'
        voxel_coordinates_run04 = read_roi_file(roi_file_run04)
        run04_template = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run04_template.nii.gz'
        if not os.path.exists(run04_template):
            run = f'analysis/susceptibility_analysis/data/{p_id}/niftis/run04.nii'
            subprocess.run(['fslmaths', run, '-Tmean', run04_template])
        functional_image_info = nib.load(run04_template)
        functional_dims = functional_image_info.shape
        binary_volume = np.zeros(functional_dims)
        for voxel in voxel_coordinates:
            x, y, z = voxel
            binary_volume[x, y, z] = 1
        binary_volume = np.flip(binary_volume, axis=1)
        functional_affine = functional_image_info.affine
        binary_mask = nib.Nifti1Image(binary_volume, affine=functional_affine)
        nib.save(binary_mask, f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run04_subject_space_ROI.nii.gz')
        roi_mask_run04 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run04_subject_space_ROI.nii.gz'
        flirted_roi_run04 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_roi_run04.nii.gz'
        if not os.path.exists(flirted_roi_run04):
            subprocess.run(['flirt', '-in', roi_mask_run04, '-ref', structural_brain, '-applyxfm', '-init', t1_flirted_run04_transformation, '-out', flirted_roi_run04, '-interp', 'nearestneighbour'])
        flirted_run01_bin = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run01_bin.nii.gz'
        if not os.path.exists(flirted_run01_bin):
            subprocess.run(['fslmaths', flirted_run01, '-thr', '100', '-bin', flirted_run01_bin])
        flirted_run04_bin = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run04_bin.nii.gz'
        if not os.path.exists(flirted_run04_bin):
            subprocess.run(['fslmaths', flirted_run04, '-thr', '100', '-bin', flirted_run04_bin])
        run01_bin_inv = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run01_bin_inv.nii.gz'
        if not os.path.exists(run01_bin_inv):
            subprocess.run(['fslmaths', flirted_run01_bin, '-sub', '1', '-abs', run01_bin_inv])
        run04_bin_inv = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run04_bin_inv.nii.gz'
        if not os.path.exists(run04_bin_inv):
            subprocess.run(['fslmaths', flirted_run04_bin, '-sub', '1', '-abs', run04_bin_inv])
        run01_result = subprocess.run(['fslstats', flirted_roi_run01, '-k', run01_bin_inv, '-V'], capture_output=True, text=True)
        if run01_result.returncode == 0:
            run01_result_output = run01_result.stdout.strip()
        else:
            print("Error executing fslstats command.")
        run01_result_output_values = run01_result_output.split()
        run01_voxels_outside = float(run01_result_output_values[0])
        run04_result = subprocess.run(['fslstats', flirted_roi_run04, '-k', run04_bin_inv, '-V'], capture_output=True, text=True)
        if run04_result.returncode == 0:
            run04_result_output = run04_result.stdout.strip()
        else:
            print("Error executing fslstats command.")
        run04_result_output_values = run04_result_output.split()
        run04_voxels_outside = float(run04_result_output_values[0])
        result1 = subprocess.run(['fslstats', flirted_roi_run01, '-V'], capture_output=True, text=True)
        if result1.returncode == 0:
            result1_output = result1.stdout.strip()
        else:
            print("Error executing fslstats command.")
        result1_output_values = result1_output.split()
        total_voxels_in_roi_run01 = float(result1_output_values[0])
        result2 = subprocess.run(['fslstats', flirted_roi_run04, '-V'], capture_output=True, text=True)
        if result2.returncode == 0:
            result2_output = result2.stdout.strip()
        else:
            print("Error executing fslstats command.")
        result2_output_values = result2_output.split()
        total_voxels_in_roi_run04 = float(result2_output_values[0])
        perc_outside_run01 = (run01_voxels_outside / total_voxels_in_roi_run01) * 100
        perc_outside_run01 = round(perc_outside_run01, 2)
        perc_outside_run01_values.append(perc_outside_run01)
        perc_outside_run04 = (run04_voxels_outside / total_voxels_in_roi_run04) * 100
        perc_outside_run04 = round(perc_outside_run04, 2)
        perc_outside_run04_values.append(perc_outside_run04)
        perc_outside_df = pd.DataFrame({'p_id': [p_id], 'perc_outside_run01': [perc_outside_run01], 'perc_outside_run04': [perc_outside_run04]})
        perc_outside_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/perc_outside_df.txt', sep='\t', index=False)
        group_perc_outside_df = pd.concat([group_perc_outside_df, perc_outside_df], ignore_index=True)
        run01_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run01_trimmed_roi_mask.nii.gz'
        run04_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run04_trimmed_roi_mask.nii.gz'
        if not os.path.exists(run01_trimmed_roi_mask) or not os.path.exists(run04_trimmed_roi_mask):
            subprocess.run(['fslmaths', transformed_roi_mask, '-mul', flirted_run01_bin, run01_trimmed_roi_mask])
            subprocess.run(['fslmaths', transformed_roi_mask, '-mul', flirted_run04_bin, run04_trimmed_roi_mask])
    group_perc_outside_df.to_csv('analysis/susceptibility_analysis/run_comparisons/3/group/group_perc_outside_df.txt', sep='\t', index=False)
    plot_data = pd.DataFrame({
        'Participant': bad_participants * 2,
        'Perc_Outside': perc_outside_run01_values + perc_outside_run04_values,
        'Sequence': ['run01'] * len(bad_participants) + ['run04'] * len(bad_participants)
    })
    perc_outside_plot = (
        ggplot(plot_data, aes(x='Participant', y='Perc_Outside', fill='Sequence')) +
        geom_bar(stat='identity', position='dodge') +
        theme_classic() +
        labs(title='Percentage of Voxels in Signal Dropout Regions', x='Participant', y='Percentage of SCC Voxels in Signal Dropout') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='black'), axis_title=element_text(size=12)) +
        scale_y_continuous(expand=(0, 0))
    )
    perc_outside_plot.save('analysis/susceptibility_analysis/run_comparisons/3/group/perc_outside_plot.png')
    perc_outside_run01_overall = np.mean(perc_outside_run01_values)
    perc_outside_run04_overall = np.mean(perc_outside_run04_values)
    run01_std_error = np.std(perc_outside_run01_values) / np.sqrt(len(perc_outside_run01_values))
    run04_std_error = np.std(perc_outside_run04_values) / np.sqrt(len(perc_outside_run04_values))
    _, perc_outside_run01_overall_shap_p = stats.shapiro(perc_outside_run01_values)
    _, perc_outside_run04_overall_shap_p = stats.shapiro(perc_outside_run04_values)
    if perc_outside_run01_overall_shap_p > 0.05 and perc_outside_run04_overall_shap_p > 0.05:
        print(f'Shapiro-Wilk test passed for perc_outside values. Running parametric t-test...')
        _, p_value = stats.ttest_ind(perc_outside_run01_values, perc_outside_run04_values)
        print(f"T-test p-value: {p_value}")
    else:
        print(f'Shapiro-Wilk test failed for perc_outside values. Running non-parametric Mann-Whitney U test...')
        _, p_value = stats.mannwhitneyu(perc_outside_run01_values, perc_outside_run04_values)
        print(f"Mann-Whitney U test p-value: {p_value}")
    plot_data = pd.DataFrame({'Sequence': ['run01', 'run04'], 'Perc_Outside': [perc_outside_run01_overall, perc_outside_run04_overall], 'Std_Error': [run01_std_error, run04_std_error]})
    group_perc_outside_plot = (ggplot(plot_data, aes(x='Sequence', y='Perc_Outside', fill='Sequence')) + 
                        geom_bar(stat='identity', position='dodge') +
                        geom_errorbar(aes(ymin='Perc_Outside - Std_Error', ymax='Perc_Outside + Std_Error'), width=0.2, color='black') +
                        theme_classic() +
                        labs(title='Percentage of Voxels in Signal Dropout Regions', y='Percentage of SCC Voxels in Signal Dropout') +
                        scale_y_continuous(expand=(0, 0)) +
                        scale_fill_manual(values={'PA': '#DB5F57', 'RL': '#57D3DB'})
                        )
    if p_value < 0.001:
        group_perc_outside_plot = group_perc_outside_plot + annotate("text", x=1.5, y=max(plot_data['Perc_Outside']) + 3.5, label="***", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Perc_Outside']) + 3, yend=max(plot_data['Perc_Outside']) + 3, color="black")
    elif p_value < 0.01:
        group_perc_outside_plot = group_perc_outside_plot + annotate("text", x=1.5, y=max(plot_data['Perc_Outside']) + 3.5, label="**", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Perc_Outside']) + 3, yend=max(plot_data['Perc_Outside']) + 3, color="black")
    elif p_value < 0.05:
        group_perc_outside_plot = group_perc_outside_plot + annotate("text", x=1.5, y=max(plot_data['Perc_Outside']) + 3.5, label="*", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Perc_Outside']) + 3, yend=max(plot_data['Perc_Outside']) + 3, color="black")    
    group_perc_outside_plot.save('analysis/susceptibility_analysis/run_comparisons/3/group/group_perc_outside_plot.png')
    
    column_headers = ['p_id', 'ssim_index', 'voxels_in_bin_ssim_mask', 'perc_roi_voxels_in_bin_ssim_mask']
    group_ssim_df = pd.DataFrame(columns = column_headers) 
    for p_id in bad_participants:
        print(f"Running Stage 3 SSIM analysis for {p_id}...")
        def calculate_ssim(image1_path, image2_path, ssim_output_path):
            """Function to calculate SSIM between two NIfTI images and save the SSIM map."""
            image1_nii = nib.load(image1_path)
            image2_nii = nib.load(image2_path)
            image1 = image1_nii.get_fdata()
            image2 = image2_nii.get_fdata()
            if image1.shape != image2.shape:
                raise ValueError("Input images must have the same dimensions for SSIM calculation.")
            ssim_index, ssim_map = ssim(image1, image2, full=True, data_range=image1.max() - image1.min())
            ssim_map_nifti = nib.Nifti1Image(ssim_map, affine=image1_nii.affine, header=image1_nii.header)
            nib.save(ssim_map_nifti, ssim_output_path)
            return ssim_index
        ssim_output_path = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/ssim_map.nii.gz'
        flirted_run01 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run01.nii.gz'
        flirted_run04 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run04.nii.gz'
        if not os.path.exists(ssim_output_path):
            ssim_index = calculate_ssim(flirted_run01, flirted_run04, ssim_output_path)
        ssim_bin = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/ssim_bin.nii.gz'
        if not os.path.exists(ssim_bin):
            subprocess.run(["fslmaths", ssim_output_path, "-thr", "0.8", "-binv", ssim_bin])
        combined_run01_run04_mask = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/combined_run01_run04_mask.nii.gz'
        flirted_run01_bin = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run01_bin.nii.gz'
        flirted_run04_bin = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run04_bin.nii.gz'
        if not os.path.exists(combined_run01_run04_mask):
            subprocess.run(['fslmaths', flirted_run01_bin, '-add', flirted_run04_bin, combined_run01_run04_mask])
        bin_run01_run04_mask = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/bin_run01_run04_mask.nii.gz'
        if not os.path.exists(bin_run01_run04_mask):
            subprocess.run(['fslmaths', combined_run01_run04_mask, '-bin', bin_run01_run04_mask])
        ssim_bin_trimmed = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/ssim_bin_trimmed.nii.gz'
        if not os.path.exists(ssim_bin_trimmed):
            subprocess.run(['fslmaths', ssim_bin, '-mul', bin_run01_run04_mask, ssim_bin_trimmed])
        voxels_in_whole_mask = subprocess.run(["fslstats", ssim_bin_trimmed, "-V"], capture_output=True, text=True).stdout.split()[0]
        voxels_in_whole_mask = float(voxels_in_whole_mask)
        intersection_mask_path_run01 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/ssim_roi_intersect_run01.nii.gz'
        intersection_mask_path_run04 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/ssim_roi_intersect_run04.nii.gz'
        run01_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run01_trimmed_roi_mask.nii.gz'
        run04_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run04_trimmed_roi_mask.nii.gz'
        if not os.path.exists(intersection_mask_path_run01):
            subprocess.run(["fslmaths", ssim_bin_trimmed, "-mas", run01_trimmed_roi_mask, intersection_mask_path_run01])
            subprocess.run(["fslmaths", ssim_bin_trimmed, "-mas", run04_trimmed_roi_mask, intersection_mask_path_run04])
        voxels_in_roi_in_mask_run01 = subprocess.run(["fslstats", intersection_mask_path_run01, "-V"], capture_output=True, text=True).stdout.split()[0]
        voxels_in_roi_in_mask_run01 = float(voxels_in_roi_in_mask_run01)
        perc_roi_voxels_in_mask_run01 = (voxels_in_roi_in_mask_run01 / total_voxels_in_roi_run01) * 100
        voxels_in_roi_in_mask_run04 = subprocess.run(["fslstats", intersection_mask_path_run04, "-V"], capture_output=True, text=True).stdout.split()[0]
        voxels_in_roi_in_mask_run04 = float(voxels_in_roi_in_mask_run04)
        perc_roi_voxels_in_mask_run04 = (voxels_in_roi_in_mask_run04 / total_voxels_in_roi_run04) * 100
        perc_roi_voxels_in_mask_av = (perc_roi_voxels_in_mask_run01 + perc_roi_voxels_in_mask_run04) / 2
        ssim_df = pd.DataFrame({'p_id': [p_id], 'ssim_index': [ssim_index], 'voxels_in_bin_ssim_mask': [voxels_in_whole_mask], 'perc_roi_voxels_in_bin_ssim_mask': [perc_roi_voxels_in_mask_av]})
        ssim_df.to_csv(f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/ssim_df.txt', sep='\t', index=False)
        group_ssim_df = pd.concat([group_ssim_df, ssim_df], ignore_index=True)
    group_ssim_df.to_csv('analysis/susceptibility_analysis/run_comparisons/3/group/group_ssim_df.txt', sep='\t', index=False)
    ssim_indexes = group_ssim_df['ssim_index'].tolist()
    ssim_mean = np.mean(ssim_indexes)
    print(f"Mean SSIM index for Stage 3: {ssim_mean}")
    plot_data = pd.DataFrame({
        'Participant': bad_participants,
        'SSIM': ssim_indexes,
    })
    ssim_index_plot = (
        ggplot(plot_data, aes(x='Participant', y='SSIM')) +
        geom_bar(stat='identity', position='dodge') +
        geom_hline(yintercept=ssim_mean, linetype='dashed', color='red') +
        theme_classic() +
        labs(title='SSIM Indexes', x='Participant', y='SSIM') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='blue'), axis_title=element_text(size=12, face='bold')) +
        scale_y_continuous(expand=(0, 0), limits=[0.8,1])
    )
    ssim_index_plot.save('analysis/susceptibility_analysis/run_comparisons/3/group/ssim_index_plot.png')
    voxels = group_ssim_df['voxels_in_bin_ssim_mask'].tolist()
    voxels_mean = np.mean(voxels)
    plot_data = pd.DataFrame({
        'Participant': bad_participants,
        'Voxels': voxels,
    })
    ssim_voxels_plot = (
        ggplot(plot_data, aes(x='Participant', y='Voxels')) +
        geom_bar(stat='identity', position='dodge') +
        geom_hline(yintercept=voxels_mean, linetype='dashed', color='red') +
        theme_classic() +
        labs(title='Number of Voxels in SSIM Mask', x='Participant', y='Voxels') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='blue'), axis_title=element_text(size=12, face='bold')) +
        scale_y_continuous(expand=(0, 0))
    )
    ssim_voxels_plot.save('analysis/susceptibility_analysis/run_comparisons/3/group/ssim_voxels_plot.png')
    perc_voxels = group_ssim_df['perc_roi_voxels_in_bin_ssim_mask'].tolist()
    perc_voxels_mean = np.mean(perc_voxels)
    plot_data = pd.DataFrame({
        'Participant': bad_participants,
        'Perc_Voxels': perc_voxels,
    })
    ssim_perc_voxels_plot = (
        ggplot(plot_data, aes(x='Participant', y='Perc_Voxels')) +
        geom_bar(stat='identity', position='dodge') +
        geom_hline(yintercept=perc_voxels_mean, linetype='dashed', color='red') +
        theme_classic() +
        labs(title='Percentage of ROI Voxels in SSIM Mask', x='Participant', y='Percentage of SCC Voxels in SSIM Mask') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='black'), axis_title=element_text(size=12)) +
        scale_y_continuous(expand=(0, 0))
    )
    ssim_perc_voxels_plot.save('analysis/susceptibility_analysis/run_comparisons/3/group/ssim_perc_voxels_plot.png')

    column_headers = ['p_id', 'sequence', 'value']
    group_voxel_intensity_df = pd.DataFrame(columns = column_headers)   
    for p_id in bad_participants:
        print(f'Running Stage 3 voxel signal intensity analysis for {p_id}...')
        def extract_voxel_intensities(epi_image_path, mask_image_path):
            epi_img = nib.load(epi_image_path)
            epi_data = epi_img.get_fdata()
            mask_img = nib.load(mask_image_path)
            mask_data = mask_img.get_fdata()
            mask_data = mask_data > 0
            roi_voxel_intensities = epi_data[mask_data]
            voxel_intensity_list = roi_voxel_intensities.tolist()
            return voxel_intensity_list
        flirted_run01 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run01.nii.gz'
        flirted_run04 = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/flirted_run04.nii.gz'
        run01_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run01_trimmed_roi_mask.nii.gz'
        run04_trimmed_roi_mask = f'analysis/susceptibility_analysis/run_comparisons/3/{p_id}/run04_trimmed_roi_mask.nii.gz'
        run01_voxel_intensities = extract_voxel_intensities(flirted_run01, run01_trimmed_roi_mask)
        run04_voxel_intensities = extract_voxel_intensities(flirted_run04, run04_trimmed_roi_mask)
        values = run01_voxel_intensities + run04_voxel_intensities
        sequence = ['run01'] * len(run01_voxel_intensities) + ['run04'] * len(run04_voxel_intensities)
        subject = [f'{p_id}'] * len(run01_voxel_intensities) + [f'{p_id}'] * len(run04_voxel_intensities)
        voxel_intensity_df = pd.DataFrame({'p_id': subject, 'sequence': sequence, 'value': values})
        voxel_intensity_df.to_csv(f'{p_id}/analysis/susceptibility/fnirt_test/3/voxel_intensity_df.txt', sep='\t', index=False)
        group_voxel_intensity_df = pd.concat([group_voxel_intensity_df, voxel_intensity_df], ignore_index=True)
    group_voxel_intensity_df.to_csv('analysis/susceptibility_analysis/run_comparisons/3/group/group_voxel_intensity_df.txt', sep='\t', index=False)
    run01_means = []
    run04_means= []
    p_values = []
    run01_std_errors = []
    run04_std_errors = []
    for p_id in bad_participants:
        filtered_run01 = group_voxel_intensity_df[(group_voxel_intensity_df['p_id'] == f'{p_id}') & (group_voxel_intensity_df['sequence'] == 'run01')]
        mean_value_run01 = filtered_run01['value'].mean()
        run01_means.append(mean_value_run01)
        filtered_run04 = group_voxel_intensity_df[(group_voxel_intensity_df['p_id'] == f'{p_id}') & (group_voxel_intensity_df['sequence'] == 'run04')]
        mean_value_run04 = filtered_run04['value'].mean()
        run04_means.append(mean_value_run04)
        anderson_run01 = stats.anderson(filtered_run01['value'])
        anderson_run04 = stats.anderson(filtered_run04['value'])
        significance_level = 0.05
        is_run01_normal = anderson_run01.statistic < anderson_pa.critical_values[
            anderson_pa.significance_level.tolist().index(significance_level * 100)]
        is_run04_normal = anderson_run04.statistic < anderson_rl.critical_values[
            anderson_rl.significance_level.tolist().index(significance_level * 100)]
        if is_run01_normal and is_run04_normal:
            print(f'Anderson-Darling test passed for {p_id} voxel intensity values. Running parametric t-test...')
            _, p_value = stats.ttest_ind(filtered_run01['value'], filtered_run04['value'], equal_var=False)
            p_values.append(p_value)
        else:
            print(f'Anderson-Darling test failed for {p_id} voxel intensity values. Running non-parametric Mann Whitney U test...')
            _, p_value = stats.mannwhitneyu(filtered_run01['value'], filtered_run04['value'], alternative='two-sided')
            p_values.append(p_value)
        run01_std_error = np.std(filtered_run01['value']) / np.sqrt(len(filtered_run01['value']))
        run01_std_errors.append(run01_std_error)
        run04_std_error = np.std(filtered_run04['value']) / np.sqrt(len(filtered_run04['value']))
        run04_std_errors.append(run04_std_error)
    plot_data = pd.DataFrame({
        'Participant': bad_participants * 2,
        'Mean_Value': run01_means + run04_means,
        'Sequence': ['RUN01'] * len(bad_participants) + ['RUN04'] * len(bad_participants),
        'Significance': ['' for _ in range(len(bad_participants) * 2)],
        'Std_Error': run01_std_errors + run04_std_errors
    })
    for idx, p_value in enumerate(p_values):
        if p_value < 0.001:
            plot_data.at[idx, 'Significance'] = '***'
        elif p_value < 0.01:
            plot_data.at[idx, 'Significance'] = '**'
        elif p_value < 0.05:
            plot_data.at[idx, 'Significance'] = '*'
    voxel_intensity_plot = (
        ggplot(plot_data, aes(x='Participant', y='Mean_Value', fill='Sequence')) +
        geom_bar(stat='identity', position='dodge') +
        geom_errorbar(aes(ymin='Mean_Value - Std_Error', ymax='Mean_Value + Std_Error'), position=position_dodge(width=0.9), width=0.2, color='black') +
        theme_classic() +
        labs(title='Mean SCC Voxel Intensity', x='Participant', y='Mean SCC Signal Intensity') +
        theme(axis_text_x=element_text(rotation=45, hjust=1), text=element_text(size=12, color='black'), axis_title=element_text(size=12)) +
        scale_y_continuous(expand=(0, 0), limits=[0,350]) +
        geom_text(
            aes(x='Participant', y='Mean_Value', label='Significance'),
            position=position_dodge(width=0.9),
            color='black',
            size=12,
            ha='center',
            va='bottom',
            show_legend=False))
    voxel_intensity_plot.save('analysis/susceptibility_analysis/run_comparisons/3/group/voxel_intensity_plot.png')
    run01_means_overall = np.mean(run01_means)
    run04_means_overall = np.mean(run04_means)
    run01_std_error_overall = np.std(run01_means) / np.sqrt(len(run01_means))
    run04_std_error_overall = np.std(run04_means) / np.sqrt(len(run04_means))
    _, run01_means_overall_shap_p = stats.shapiro(run01_means)
    _, run04_means_overall_shap_p = stats.shapiro(run04_means)
    if run01_means_overall_shap_p > 0.05 and run04_means_overall_shap_p > 0.05:
        print(f'Shapiro-Wilk test passed for voxel intensity values. Running parametric t-test...')
        _, p_value = stats.ttest_rel(run01_means, run04_means)
        print(f"T-test p-value: {p_value}")
    else:
        print(f'Shapiro-Wilk test failed for voxel intensity values. Running non-parametric Wilcoxon test...')
        _, p_value = stats.wilcoxon(run01_means, run04_means)
        print(f"Wilcoxon test p-value: {p_value}")
    plot_data = pd.DataFrame({'p_id': bad_participants, 'run01_values': run01_means, 'run04_values': run04_means})
    data_long = pd.melt(plot_data, id_vars=['p_id'], value_vars=['run01_values', 'run04_values'], var_name='sequence', value_name='value')
    data_long['sequence'] = data_long['sequence'].map({'run01_values': 'RUN01', 'run04_values': 'RUN04'})
    group_voxel_intensity_ladder_plot = (
        ggplot(data_long, aes(x='sequence', y='value', group='p_id')) +
        geom_line(aes(group='p_id'), color='gray', size=1) +
        geom_point(aes(color='sequence'), size=4) +
        theme_light() +
        theme(
            panel_grid_major=element_blank(), 
            panel_grid_minor=element_blank(), 
            panel_border=element_blank(),
            axis_line_x=element_line(color='black'),  
            axis_line_y=element_line(color='black'),  
        ) +
        labs(title='Ladder Plot of RUN01 and RUN04 Sequences',
            x='Sequence',
            y='Mean SCC Signal Intensity') +
        scale_x_discrete(limits=['RUN01', 'RUN04']) +
        scale_y_continuous()
    )
    group_voxel_intensity_ladder_plot.save('analysis/susceptibility_analysis/run_comparisons/3/group/group_voxel_intensity_ladder_plot.png')
    plot_data = pd.DataFrame({'Sequence': ['RUN01', 'RUN04'], 'Mean': [run01_means_overall, run04_means_overall], 'Std_Error': [run01_std_error_overall, run04_std_error_overall]})
    group_voxel_intensity_plot = (ggplot(plot_data, aes(x='Sequence', y='Mean', fill='Sequence')) + 
                        geom_bar(stat='identity', position='dodge') +
                        geom_errorbar(aes(ymin='Mean - Std_Error', ymax='Mean + Std_Error'), width=0.2, color='black') +
                        theme_classic() +
                        labs(title='Mean of Voxel Intensities Across Participants', y='Mean SCC Signal Intensity') +
                        scale_y_continuous(expand=(0, 0), limits=[0,350]) +
                        scale_fill_manual(values={'PA': '#DB5F57', 'RL': '#57D3DB'})
                        )
    if p_value < 0.001:
        group_voxel_intensity_plot = group_voxel_intensity_plot + annotate("text", x=1.5, y=max(plot_data['Mean']) + 40, label="***", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Mean']) +30, yend=max(plot_data['Mean']) + 30, color="black")
    elif p_value < 0.01:
        group_voxel_intensity_plot = group_voxel_intensity_plot + annotate("text", x=1.5, y=max(plot_data['Mean']) + 40, label="**", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Mean']) +30, yend=max(plot_data['Mean']) + 30, color="black")
    elif p_value < 0.05:
        group_voxel_intensity_plot = group_voxel_intensity_plot + annotate("text", x=1.5, y=max(plot_data['Mean']) + 40, label="*", size=16, color="black") + \
            annotate("segment", x=1, xend=2, y=max(plot_data['Mean']) +30, yend=max(plot_data['Mean']) + 30, color="black")    
    group_voxel_intensity_plot.save('analysis/susceptibility_analysis/run_comparisons/3/group/group_voxel_intensity_plot.png')

    # Step 9: Testing FNIRT parameters.
    print("\n###### STEP 9: TESTING FNIRT PARAMETERS ######")   
    bad_participants = ['P004', 'P006', 'P020', 'P030', 'P078', 'P093', 'P094']
    for p_id in bad_participants:
        fnirt_test_folder = 'analysis/susceptibility_analysis/fnirt_test'
        os.makedirs(fnirt_test_folder, exist_ok=True)
        fnirt_test_1_folder = 'analysis/susceptibility_analysis/fnirt_test/1'
        os.makedirs(fnirt_test_1_folder, exist_ok=True)
        fnirt_test_2_folder = 'analysis/susceptibility_analysis/fnirt_test/2'
        os.makedirs(fnirt_test_2_folder, exist_ok=True)
        fnirt_test_3_folder = 'analysis/susceptibility_analysis/fnirt_test/3'
        os.makedirs(fnirt_test_3_folder, exist_ok=True)
        fnirt_test_4_folder = 'analysis/susceptibility_analysis/fnirt_test/4'
        os.makedirs(fnirt_test_4_folder, exist_ok=True)
        fnirt_test_5_folder = 'analysis/susceptibility_analysis/fnirt_test/5'
        os.makedirs(fnirt_test_5_folder, exist_ok=True)
        fnirt_test_6_folder = 'analysis/susceptibility_analysis/fnirt_test/6'
        os.makedirs(fnirt_test_6_folder, exist_ok=True)
        fnirt_test_1_participant_folder = f'analysis/susceptibility_analysis/fnirt_test/1/{p_id}'
        os.makedirs(fnirt_test_1_participant_folder, exist_ok=True)
        fnirt_test_2_participant_folder = f'analysis/susceptibility_analysis/fnirt_test/2/{p_id}'
        os.makedirs(fnirt_test_2_participant_folder, exist_ok=True)
        fnirt_test_3_participant_folder = f'analysis/susceptibility_analysis/fnirt_test/3/{p_id}'
        os.makedirs(fnirt_test_3_participant_folder, exist_ok=True)
        fnirt_test_4_participant_folder = f'analysis/susceptibility_analysis/fnirt_test/4/{p_id}'
        os.makedirs(fnirt_test_4_participant_folder, exist_ok=True)
        fnirt_test_5_participant_folder = f'analysis/susceptibility_analysis/fnirt_test/5/{p_id}'
        os.makedirs(fnirt_test_5_participant_folder, exist_ok=True)
        fnirt_test_6_participant_folder = f'analysis/susceptibility_analysis/fnirt_test/6/{p_id}'
        os.makedirs(fnirt_test_6_participant_folder, exist_ok=True)

        averaged_run01 = f'data/raw_data/{p_id}/data/neurofeedback/fnirt_test_files/averaged_run01.nii.gz'
        averaged_run01_brain = f'data/raw_data/{p_id}/data/neurofeedback/fnirt_test_files/averaged_run01_brain.nii.gz'
        structural = f'data/raw_data/{p_id}/data/neurofeedback/fnirt_test_files/structural.nii'
        structural_brain = f'data/raw_data/{p_id}/data/neurofeedback/fnirt_test_files/structural_brain.nii.gz'
        structural_downsampled = f'data/raw_data/{p_id}/data/neurofeedback/fnirt_test_files/structural_downsampled.nii.gz'
        structural_brain_downsampled = f'data/raw_data/{p_id}/data/neurofeedback/fnirt_test_files/structural_brain_downsampled.nii.gz'
        structural_brain_downsampled_mask = f'data/raw_data/{p_id}/data/neurofeedback/fnirt_test_files/structural_brain_downsampled_mask.nii.gz'

        print(f"Running FLIRT betted, FNIRT unbetted, applywarp unbetted, downsampled, for {p_id}...")
        flirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/1/{p_id}/flirted_run01.nii.gz'
        flirted_run01_matrix = f'analysis/susceptibility_analysis/fnirt_test/1/{p_id}/flirted_run01_matrix.mat'
        if not os.path.exists(flirted_run01):
            subprocess.run(['flirt', '-in', averaged_run01_brain, '-ref', structural_brain_downsampled, '-out', flirted_run01, '-omat', flirted_run01_matrix, '-dof', '6'])
        warp_run01 = f'analysis/susceptibility_analysis/fnirt_test/1/{p_id}/warp_run01'
        if not os.path.exists(warp_run01):
            subprocess.run(['fnirt', f'--in={averaged_run01}', f'--ref={structural_downsampled}', f'--aff={flirted_run01_matrix}', f'--cout={warp_run01}'])
        fnirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/1/{p_id}/fnirted_run01.nii.gz'
        if not os.path.exists(fnirted_run01):
            subprocess.run(['applywarp', f'--in={averaged_run01}', f'--ref={structural_downsampled}', f'--warp={warp_run01}', f'--out={fnirted_run01}'])

        print(f"Running FLIRT betted, FNIRT betted, applywarp betted, downsampled, for {p_id}...")
        flirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/2/{p_id}/flirted_run01.nii.gz'
        flirted_run01_matrix = f"{p_id}/2/flirted_run01_matrix.mat"
        if not os.path.exists(flirted_run01):
            subprocess.run(['flirt', '-in', averaged_run01_brain, '-ref', structural_brain_downsampled, '-out', flirted_run01, '-omat', flirted_run01_matrix, '-dof', '6'])
        warp_run01 = f'analysis/susceptibility_analysis/fnirt_test/2/{p_id}/warp_run01'
        if not os.path.exists(warp_run01):
            subprocess.run(['fnirt', f'--in={averaged_run01_brain}', f'--ref={structural_brain_downsampled}', f'--aff={flirted_run01_matrix}', f'--cout={warp_run01}'])
        fnirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/2/{p_id}/fnirted_run01.nii.gz'
        if not os.path.exists(fnirted_run01):
            subprocess.run(['applywarp', f'--in={averaged_run01_brain}', f'--ref={structural_brain_downsampled}', f'--warp={warp_run01}', f'--out={fnirted_run01}'])

        print(f"Running FLIRT betted, FNIRT unbetted, applywarp unbetted, downsampled, with refmask, for {p_id}...")
        flirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/3/{p_id}/flirted_run01.nii.gz'
        flirted_run01_matrix = f'analysis/susceptibility_analysis/fnirt_test/3/{p_id}/flirted_run01_matrix.mat'
        if not os.path.exists(flirted_run01):
            subprocess.run(['flirt', '-in', averaged_run01_brain, '-ref', structural_brain_downsampled, '-out', flirted_run01, '-omat', flirted_run01_matrix, '-dof', '6'])
        warp_run01 = f'analysis/susceptibility_analysis/fnirt_test/3/{p_id}/warp_run01'
        if not os.path.exists(warp_run01):
            subprocess.run(['fnirt', f'--in={averaged_run01}', f'--ref={structural_downsampled}', f'--refmask={structural_brain_downsampled_mask}', f'--aff={flirted_run01_matrix}', f'--cout={warp_run01}'])
        fnirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/3/{p_id}/fnirted_run01.nii.gz'
        if not os.path.exists(fnirted_run01):
            subprocess.run(['applywarp', f'--in={averaged_run01}', f'--ref={structural_downsampled}', f'--mask={structural_brain_downsampled_mask}', f'--warp={warp_run01}', f'--out={fnirted_run01}'])

        print(f"Running FLIRT betted, FNIRT betted, applywarp betted, downsampled, with refmask, for {p_id}...")
        flirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/4/{p_id}/flirted_run01.nii.gz'
        flirted_run01_matrix = f'analysis/susceptibility_analysis/fnirt_test/4/{p_id}/flirted_run01_matrix.mat'
        if not os.path.exists(flirted_run01):
            subprocess.run(['flirt', '-in', averaged_run01_brain, '-ref', structural_brain_downsampled, '-out', flirted_run01, '-omat', flirted_run01_matrix, '-dof', '6'])
        warp_run01 = f'analysis/susceptibility_analysis/fnirt_test/4/{p_id}/warp_run01'
        if not os.path.exists(warp_run01):
            subprocess.run(['fnirt', f'--in={averaged_run01_brain}', f'--ref={structural_brain_downsampled}', f'--refmask={structural_brain_downsampled_mask}', f'--aff={flirted_run01_matrix}', f'--cout={warp_run01}'])
        fnirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/4/{p_id}/fnirted_run01.nii.gz'
        if not os.path.exists(fnirted_run01):
            subprocess.run(['applywarp', f'--in={averaged_run01_brain}', f'--ref={structural_brain_downsampled}', f'--mask={structural_brain_downsampled_mask}', f'--warp={warp_run01}', f'--out={fnirted_run01}'])

        print(f"Running FLIRT betted, FNIRT unbetted, applywarp unbetted, downsampled, with lambda configuration 1, for {p_id}...")
        flirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/5/{p_id}/flirted_run01.nii.gz'
        flirted_run01_matrix = f'analysis/susceptibility_analysis/fnirt_test/5/{p_id}/flirted_run01_matrix.mat'
        if not os.path.exists(flirted_run01):
            subprocess.run(['flirt', '-in', averaged_run01_brain, '-ref', structural_brain_downsampled, '-out', flirted_run01, '-omat', flirted_run01_matrix, '-dof', '6'])
        warp_run01 = f'analysis/susceptibility_analysis/fnirt_test/5/{p_id}/warp_run01'
        if not os.path.exists(warp_run01):
            subprocess.run(['fnirt', f'--in={averaged_run01}', f'--ref={structural_downsampled}', f'--aff={flirted_run01_matrix}', f'--cout={warp_run01}', '--lambda=200,75,40,20'])
        fnirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/5/{p_id}/fnirted_run01.nii.gz'
        if not os.path.exists(fnirted_run01):
            subprocess.run(['applywarp', f'--in={averaged_run01}', f'--ref={structural_downsampled}', f'--warp={warp_run01}', f'--out={fnirted_run01}'])

        print(f"Running FLIRT betted, FNIRT unbetted, applywarp unbetted, downsampled, with lambda configuration 2 for {p_id}...")
        flirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/6/{p_id}/flirted_run01.nii.gz'
        flirted_run01_matrix = f'analysis/susceptibility_analysis/fnirt_test/6/{p_id}/flirted_run01_matrix.mat'
        if not os.path.exists(flirted_run01):
            subprocess.run(['flirt', '-in', averaged_run01_brain, '-ref', structural_brain_downsampled, '-out', flirted_run01, '-omat', flirted_run01_matrix, '-dof', '6'])
        warp_run01 = f'analysis/susceptibility_analysis/fnirt_test/6/{p_id}/warp_run01'
        if not os.path.exists(warp_run01):
            subprocess.run(['fnirt', f'--in={averaged_run01}', f'--ref={structural_downsampled}', f'--aff={flirted_run01_matrix}', f'--cout={warp_run01}', '--lambda=400,200,150,75,60,45'])
        fnirted_run01 = f'analysis/susceptibility_analysis/fnirt_test/6/{p_id}/fnirted_run01.nii.gz'
        if not os.path.exists(fnirted_run01):
            subprocess.run(['applywarp', f'--in={averaged_run01}', f'--ref={structural_downsampled}', f'--warp={warp_run01}', f'--out={fnirted_run01}'])
#endregion

#region MAIN MENU.
print("\nWelcome to the neurofeedback analyser script. Please ensure that the following steps are complete before proceeding:\n")
print("1. Upload the participant's data to Box.\n")
print("2. In the Bash terminal, change the working directory to the participant_data folder within the cisc2 drive.")

MAX_LOGS = 5  
class StreamToLogger:
    def __init__(self, logger, log_level=logging.INFO):
        self.logger = logger
        self.log_level = log_level
        self.linebuf = ''
    def write(self, buf):
        for line in buf.rstrip().splitlines():
            self.logger.log(self.log_level, line)
    def flush(self):
        pass
def rotate_logs(log_file_path):
    if os.path.exists(log_file_path):
        with open(log_file_path, "r") as f:
            logs = f.read().split("\n##### NEW SESSION #####\n")
        if len(logs) > MAX_LOGS:
            with open(log_file_path, "w") as f:
                f.write("\n##### NEW SESSION #####\n".join(logs[-MAX_LOGS:]))
def setup_logging(log_file_path):
    rotate_logs(log_file_path)
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    file_handler = logging.FileHandler(log_file_path)
    file_handler.setLevel(logging.DEBUG)
    console_handler = logging.StreamHandler(sys.stdout)
    console_handler.setLevel(logging.INFO)
    file_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    console_formatter = logging.Formatter('%(message)s')
    file_handler.setFormatter(file_formatter)
    console_handler.setFormatter(console_formatter)
    logger.addHandler(file_handler)
    logger.addHandler(console_handler)
    sys.stdout = StreamToLogger(logger, logging.INFO)
    sys.stderr = StreamToLogger(logger, logging.ERROR)
    with open(log_file_path, "a") as log_file:
        log_file.write("\n##### NEW SESSION #####\n")
        log_file.write(f"Script started on: {datetime.now()}\n")
def logged_input(prompt=""):
    logger = logging.getLogger()
    user_input = input(prompt)
    logging.info(f"User input: {user_input}")
    return user_input

def main_menu():
    while True:
        print("\n##### MAIN MENU #####")
        print("1. DOWNLOAD BOX FILES TO SERVER (needs updating)")
        print("2. THERMOMETER ANALYSIS")
        print("3. CLINICAL ANALYSIS")
        print("4. FMRI PREPARATION AND PREPROCESSING")
        print("5. FMRI ANALYSIS")
        print("6. SUSCEPTIBILITY ANALYSIS")
        print("7. EXIT")

        choice = logged_input("\nPlease choose an option (1-7): ")

        if choice == '1':
            download_box_files()
        elif choice == '2':
            thermometer_analysis()
        elif choice == '3':
            clinical_analysis()
        elif choice == '4':
            fmri_prep_and_preproc()
        elif choice == '5':
            fmri_analysis()
        elif choice == '6':
            susceptibility_analysis()
        elif choice == '7':
            print("Exiting... Goodbye!")
            sys.exit()
        else:
            print("Invalid choice. Please select a number between 1 and 7.")

if __name__ == "__main__":
    log_dir = os.getcwd()
    log_file_name = "nf_analyser_log.txt"
    log_file_path = os.path.join(log_dir, log_file_name)
    setup_logging(log_file_path)
    main_menu()
        
#endregion